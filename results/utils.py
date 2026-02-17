from django.db import transaction

from core.models import CombinationSubject
from .models import ExamSession, StudentExamMetrics, StudentExamPosition, StudentResult
# results/utils.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

def recalculate_all_metrics_for_session(exam_session_id):
    """
    Manually recalculate all metrics and positions for an exam session.
    Useful for fixing data or after major changes.
    """
    from .signals import calculate_student_metrics, calculate_all_positions
    
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
        
        with transaction.atomic():
            # Clear existing metrics and positions
            StudentExamMetrics.objects.filter(exam_session=exam_session).delete()
            StudentExamPosition.objects.filter(exam_session=exam_session).delete()
            
            # Get all students with results in this session
            student_ids = exam_session.results.values_list(
                'student_id', flat=True
            ).distinct()
            
            from students.models import Student
            students = Student.objects.filter(id__in=student_ids)
            
            education_level = exam_session.class_level.educational_level
            
            # Calculate metrics for each student
            for student in students:
                calculate_student_metrics(exam_session, student, education_level)
            
            # Calculate positions
            calculate_all_positions(exam_session)
            
            return True, f"Recalculated metrics for {students.count()} students"
            
    except Exception as e:
        return False, f"Error recalculating metrics: {str(e)}"


def bulk_recalculate_metrics(exam_session_ids):
    """
    Recalculate metrics for multiple exam sessions.
    """
    results = []
    for session_id in exam_session_ids:
        success, message = recalculate_all_metrics_for_session(session_id)
        results.append({
            'exam_session_id': session_id,
            'success': success,
            'message': message
        })
    return results





import openpyxl
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal

def export_student_sessions_to_excel(student, exam_sessions, session_metrics, filters=None):
    """
    Export student exam results to Excel with subject-wise aggregation across sessions.
    
    Args:
        student: Student object
        exam_sessions: QuerySet of ExamSession objects
        session_metrics: Dictionary mapping session_id -> {'metrics': metrics_obj, 'positions': position_obj}
        filters: Optional filter parameters
    
    Returns:
        openpyxl.Workbook object
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Report"

    # ============================================
    # 1. STYLES DEFINITION
    # ============================================
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)

    # ============================================
    # 2. HEADER INFORMATION
    # ============================================
    ws.merge_cells('A1:I1')
    ws['A1'] = "STUDENT EXAM RESULTS REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:C2')
    ws['A2'] = f"Student Name: {student.full_name}"
    ws['D2'] = f"Gender: {student.get_gender_display() if student.gender else 'N/A'}"
    ws.merge_cells('F2:I2')
    ws['F2'] = f"Admission: {student.registration_number or 'N/A'}"

    if filters:
        filter_row = 3
        ws.merge_cells(f'A{filter_row}:I{filter_row}')
        filter_text = "Filters Applied: "
        filter_parts = []
        if filters.get('class_level'):
            filter_parts.append(f"Class Level: {filters['class_level']}")
        if filters.get('academic_year'):
            filter_parts.append(f"Academic Year: {filters['academic_year']}")
        if filters.get('term'):
            filter_parts.append(f"Term: {filters['term']}")
        if filters.get('date_from'):
            filter_parts.append(f"From: {filters['date_from']}")
        if filters.get('date_to'):
            filter_parts.append(f"To: {filters['date_to']}")
        
        ws[f'A{filter_row}'] = filter_text + ", ".join(filter_parts)
        ws[f'A{filter_row}'].font = Font(italic=True)

    # ============================================
    # 3. DATA COLLECTION AND PROCESSING
    # ============================================
    from results.models import StudentResult, GradingScale, DivisionScale
    
    # Get all results for this student across selected sessions
    results = StudentResult.objects.filter(
        student=student, 
        exam_session__in=exam_sessions
    ).select_related('subject', 'exam_session')
    
    # Organize data by subject
    subjects_data = {}
    exam_sessions_list = list(exam_sessions)
    exam_sessions_by_id = {s.id: s for s in exam_sessions_list}
    
    # Determine education level
    education_level = None
    if student.class_level and student.class_level.educational_level:
        education_level = student.class_level.educational_level
        level_code = education_level.code.upper() if education_level.code else ''
    else:
        level_code = ''
    
    # First, collect all subjects with their results
    for result in results:
        subject_id = result.subject_id
        exam_session_id = result.exam_session_id
        
        if subject_id not in subjects_data:
            subjects_data[subject_id] = {
                'id': subject_id,
                'name': result.subject.name,
                'code': result.subject.code,
                'marks': {},  # session_id -> marks
                'grades': {},  # session_id -> grade
                'percentages': {},  # session_id -> percentage
                'points': {},  # session_id -> grade_points
                'has_marks_count': 0,
                'total_marks': Decimal('0'),
                'total_points': Decimal('0'),
                'sessions_with_data': set(),
            }
        
        # Store session-specific data
        if result.marks_obtained is not None:
            subjects_data[subject_id]['marks'][exam_session_id] = result.marks_obtained
            subjects_data[subject_id]['grades'][exam_session_id] = result.grade or '-'
            subjects_data[subject_id]['percentages'][exam_session_id] = result.percentage
            subjects_data[subject_id]['points'][exam_session_id] = result.grade_point or Decimal('0')
            subjects_data[subject_id]['has_marks_count'] += 1
            subjects_data[subject_id]['total_marks'] += result.marks_obtained
            subjects_data[subject_id]['total_points'] += (result.grade_point or Decimal('0'))
            subjects_data[subject_id]['sessions_with_data'].add(exam_session_id)
        else:
            subjects_data[subject_id]['marks'][exam_session_id] = None
            subjects_data[subject_id]['grades'][exam_session_id] = '-'
            subjects_data[subject_id]['percentages'][exam_session_id] = None
            subjects_data[subject_id]['points'][exam_session_id] = None
    
    # Convert to list for ordered processing
    subjects_list = sorted(
        subjects_data.values(), 
        key=lambda x: x['name']
    )
    
    # ============================================
    # 4. CALCULATE SUBJECT AGGREGATES
    # ============================================
    for subject in subjects_list:
        # Calculate average marks across sessions with marks
        if subject['has_marks_count'] > 0:
            subject['average_marks'] = subject['total_marks'] / subject['has_marks_count']
            subject['average_points'] = subject['total_points'] / subject['has_marks_count']
            
            # Determine overall grade based on average marks
            # Get grading scale for the student's education level
            if education_level:
                grade_scale = GradingScale.objects.filter(
                    education_level=education_level,
                    min_mark__lte=subject['average_marks'],
                    max_mark__gte=subject['average_marks']
                ).first()
                
                if grade_scale:
                    subject['overall_grade'] = grade_scale.grade
                    subject['overall_grade_description'] = grade_scale.get_grade_display()
                else:
                    subject['overall_grade'] = '-'
                    subject['overall_grade_description'] = '-'
            else:
                subject['overall_grade'] = '-'
                subject['overall_grade_description'] = '-'
        else:
            subject['average_marks'] = None
            subject['average_points'] = None
            subject['overall_grade'] = '-'
            subject['overall_grade_description'] = '-'
    
    # ============================================
    # 5. CALCULATE SUBJECT POSITIONS
    # ============================================
    # Filter subjects with valid average marks for ranking
    subjects_with_avg = [
        s for s in subjects_list 
        if s['average_marks'] is not None
    ]
    
    # Sort by average marks (descending) for ranking
    subjects_with_avg.sort(key=lambda x: x['average_marks'], reverse=True)
    
    # Assign positions (handling ties)
    position_map = {}
    current_position = 1
    previous_avg = None
    skip_count = 0
    
    for i, subject in enumerate(subjects_with_avg):
        if previous_avg is not None and subject['average_marks'] == previous_avg:
            # Same position as previous
            position_map[subject['id']] = current_position - 1
            skip_count += 1
        else:
            # New position
            position_map[subject['id']] = current_position
            current_position += 1 + skip_count
            skip_count = 0
        
        previous_avg = subject['average_marks']
    
    # Add position to subjects
    for subject in subjects_list:
        subject['position'] = position_map.get(subject['id'])
    
    # ============================================
    # 6. CALCULATE GRAND TOTALS AND AGGREGATES
    # ============================================
    
    # Calculate grand total marks across all subjects and sessions
    grand_total_marks = sum(float(s['total_marks']) for s in subjects_list if s['total_marks'] > 0)
    
    # Calculate grand average (average of all subject averages)
    valid_averages = [float(s['average_marks']) for s in subjects_list if s['average_marks'] is not None]
    grand_average = sum(valid_averages) / len(valid_averages) if valid_averages else 0
    
    # Calculate grand average percentage
    grand_average_percentage = (grand_average / 100) * 100  # Assuming 100 is max
    
    # EDUCATION LEVEL SPECIFIC CALCULATIONS
    grand_grade = '-'
    grand_division = '-'
    grand_total_points = Decimal('0')
    grand_average_points = Decimal('0')
    
    if level_code == 'O_LEVEL':
        # For O-Level: Need at least 7 subjects with valid grade points
        # Take best 7 grade points for calculation
        all_grade_points = []
        for subject in subjects_list:
            if subject['average_points'] is not None and subject['average_points'] > 0:
                all_grade_points.append(float(subject['average_points']))
        
        # Also collect from session metrics as backup
        for session_id, data in session_metrics.items():
            metrics = data.get('metrics')
            if metrics and metrics.total_grade_points:
                # This is already the best 7 from that session
                all_grade_points.append(float(metrics.total_grade_points))
        
        if len(all_grade_points) >= 7:
            # Sort and take best 7
            all_grade_points.sort(reverse=True)
            best_7_points = all_grade_points[:7]
            grand_total_points = Decimal(str(sum(best_7_points)))
            
            # Determine division based on total grade points
            if education_level:
                try:
                    points_int = int(grand_total_points)
                    division_scale = DivisionScale.objects.filter(
                        education_level=education_level,
                        min_points__lte=points_int,
                        max_points__gte=points_int
                    ).first()
                    if division_scale:
                        grand_division = division_scale.division
                except (ValueError, TypeError):
                    pass
    
    elif level_code == 'A_LEVEL':
        # For A-Level: Need core subjects from combination
        if student.combination:
            # Get core and subsidiary subjects
            combo_subjects = CombinationSubject.objects.filter(
                combination=student.combination
            )
            core_subject_ids = list(combo_subjects.filter(role='CORE').values_list('subject_id', flat=True))
            subsidiary_subject_ids = list(combo_subjects.filter(role='SUB').values_list('subject_id', flat=True))
            
            # Collect core subject points
            core_points = []
            for subject in subjects_list:
                if subject['id'] in core_subject_ids and subject['average_points'] is not None:
                    core_points.append(float(subject['average_points']))
            
            # Need at least 3 core subjects
            if len(core_points) >= 3:
                # Take best 3 core subjects
                core_points.sort(reverse=True)
                best_3_core = core_points[:3]
                grand_total_points = Decimal(str(sum(best_3_core)))
                
                # Add best subsidiary if available
                subsidiary_points = []
                for subject in subjects_list:
                    if subject['id'] in subsidiary_subject_ids and subject['average_points'] is not None:
                        subsidiary_points.append(float(subject['average_points']))
                
                if subsidiary_points:
                    subsidiary_points.sort(reverse=True)
                    grand_total_points += Decimal(str(subsidiary_points[0]))
                
                # Determine division
                if education_level:
                    try:
                        points_int = int(grand_total_points)
                        division_scale = DivisionScale.objects.filter(
                            education_level=education_level,
                            min_points__lte=points_int,
                            max_points__gte=points_int
                        ).first()
                        if division_scale:
                            grand_division = division_scale.division
                    except (ValueError, TypeError):
                        pass
    
    else:
        # Primary/Nursery - no grade points or division
        # Just calculate basic stats
        pass
    
    # Calculate grand average grade points
    subjects_with_points = [s for s in subjects_list if s['average_points'] is not None]
    grand_average_points = sum(float(s['average_points']) for s in subjects_with_points) / len(subjects_with_points) if subjects_with_points else 0
    
    # Determine grand grade based on grand average
    if education_level and grand_average:
        grade_scale = GradingScale.objects.filter(
            education_level=education_level,
            min_mark__lte=grand_average,
            max_mark__gte=grand_average
        ).first()
        if grade_scale:
            grand_grade = grade_scale.grade
    
    # Calculate grand exam position (average of session positions)
    valid_positions = []
    for session_id, data in session_metrics.items():
        positions = data.get('positions')
        if positions and positions.class_position:
            valid_positions.append(positions.class_position)
    
    grand_exam_position = round(sum(valid_positions) / len(valid_positions), 1) if valid_positions else '-'
    
    # Calculate grand remark
    grand_remark = "Passed" if grand_average_percentage >= 50 else "Fail" if grand_average_percentage > 0 else "-"
    
    # Create a dictionary of grand aggregates for easy access
    grand_aggregates = {
        'total_marks': round(grand_total_marks, 2),
        'average_percentage': round(grand_average_percentage, 2),
        'average_grade': grand_grade,
        'division': grand_division,
        'class_position': grand_exam_position,
        'remark': grand_remark,
        'grand_average': round(grand_average, 2),
        'grand_points': round(float(grand_total_points), 2) if grand_total_points > 0 else 0,
        'average_points': round(float(grand_average_points), 2),
        'level_code': level_code,
    }
    
    # ============================================
    # 7. COLUMN MAPPING
    # ============================================
    data_start_row = 5
    col_idx = 3  # Start after S/N and SUBJECT columns
    session_col_map = {}

    # Static Subject Headers (S/N and SUBJECT)
    for col, val in [(1, "S/N"), (2, "SUBJECT")]:
        cell = ws.cell(row=data_start_row, column=col, value=val)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.merge_cells(
            start_row=data_start_row, 
            start_column=col, 
            end_row=data_start_row+1, 
            end_column=col
        )

    # Add exam session headers (MARK | GRADE)
    for session in exam_sessions_list:
        cell = ws.cell(row=data_start_row, column=col_idx, value=session.name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        ws.merge_cells(
            start_row=data_start_row, 
            start_column=col_idx, 
            end_row=data_start_row, 
            end_column=col_idx+1
        )
        
        for i, sub in enumerate(["MARK", "GRADE"]):
            sub_cell = ws.cell(row=data_start_row+1, column=col_idx + i, value=sub)
            sub_cell.font = header_font_white
            sub_cell.fill = header_fill
            sub_cell.border = thin_border
            sub_cell.alignment = center_align
        
        session_col_map[session.id] = col_idx
        col_idx += 2

    # Trailing static columns for subject aggregates
    final_headers = ["TOTAL MARKS", "AVERAGE", "GRADE", "POSITION", "REMARK"]
    aggregate_cols = {}
    
    for header_text in final_headers:
        cell = ws.cell(row=data_start_row, column=col_idx, value=header_text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.merge_cells(
            start_row=data_start_row, 
            start_column=col_idx, 
            end_row=data_start_row+1, 
            end_column=col_idx
        )
        aggregate_cols[header_text] = col_idx
        col_idx += 1

    # ============================================
    # 8. POPULATE SUBJECT DATA ROWS
    # ============================================
    current_row = data_start_row + 2
    
    for idx, subject in enumerate(subjects_list, 1):
        # S/N and Subject Name
        ws.cell(row=current_row, column=1, value=idx).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        
        subject_cell = ws.cell(row=current_row, column=2, value=subject['name'])
        subject_cell.alignment = left_align
        subject_cell.border = thin_border
        
        # Session-wise marks and grades
        for session_id, col_start in session_col_map.items():
            # Marks column
            marks = subject['marks'].get(session_id)
            marks_cell = ws.cell(
                row=current_row, 
                column=col_start, 
                value=float(marks) if marks is not None else '-'
            )
            marks_cell.alignment = center_align
            marks_cell.border = thin_border
            
            # Grade column
            grade = subject['grades'].get(session_id, '-')
            grade_cell = ws.cell(row=current_row, column=col_start+1, value=grade)
            grade_cell.alignment = center_align
            grade_cell.border = thin_border
        
        # Subject aggregate columns
        # TOTAL MARKS
        total_marks_col = aggregate_cols.get("TOTAL MARKS")
        if total_marks_col and subject['total_marks'] > 0:
            total_cell = ws.cell(
                row=current_row, 
                column=total_marks_col, 
                value=float(subject['total_marks'])
            )
        else:
            total_cell = ws.cell(row=current_row, column=total_marks_col, value='-')
        total_cell.alignment = center_align
        total_cell.border = thin_border
        
        # AVERAGE
        avg_col = aggregate_cols.get("AVERAGE")
        if avg_col and subject['average_marks'] is not None:
            avg_cell = ws.cell(
                row=current_row, 
                column=avg_col, 
                value=round(float(subject['average_marks']), 2)
            )
        else:
            avg_cell = ws.cell(row=current_row, column=avg_col, value='-')
        avg_cell.alignment = center_align
        avg_cell.border = thin_border
        
        # GRADE
        grade_col = aggregate_cols.get("GRADE")
        if grade_col:
            grade_cell = ws.cell(
                row=current_row, 
                column=grade_col, 
                value=subject['overall_grade']
            )
            grade_cell.alignment = center_align
            grade_cell.border = thin_border
        
        # POSITION
        pos_col = aggregate_cols.get("POSITION")
        if pos_col and subject['position']:
            pos_cell = ws.cell(
                row=current_row, 
                column=pos_col, 
                value=subject['position']
            )
        else:
            pos_cell = ws.cell(row=current_row, column=pos_col, value='-')
        pos_cell.alignment = center_align
        pos_cell.border = thin_border
        
        # REMARK
        remark_col = aggregate_cols.get("REMARK")
        if remark_col:
            if subject['average_marks'] is not None:
                remark = "Pass" if subject['average_marks'] >= 40 else "Fail"
            else:
                remark = "-"
            remark_cell = ws.cell(row=current_row, column=remark_col, value=remark)
            remark_cell.alignment = center_align
            remark_cell.border = thin_border
        
        current_row += 1

    # ============================================
    # 9. AGGREGATE FOOTERS (SESSION SUMMARY) - MERGED CELLS
    # ============================================
    
    # Footer labels for session-level aggregates
    footer_labels = [
        ("Total Marks", "total_marks"),
        ("Average", "average_percentage"),
        ("Grade", "average_grade"),
        ("Division", "division"),
        ("Exam Position", "class_position"),
        ("Remark", "remark")
    ]

    for label_text, metric_key in footer_labels:
        # S/N column - empty with border
        ws.cell(row=current_row, column=1).border = thin_border
        
        # Subject column - label
        label_cell = ws.cell(row=current_row, column=2, value=label_text)
        label_cell.font = bold_font
        label_cell.alignment = left_align
        label_cell.border = thin_border

        # Iterate through sessions - merge MARK and GRADE columns for each session
        for session_id, col_start in session_col_map.items():
            session_data = session_metrics.get(session_id, {})
            metrics = session_data.get('metrics')
            pos_obj = session_data.get('positions')

            # Get the value for this metric
            val = "-"
            if metrics:
                if metric_key == "total_marks":
                    val = round(float(metrics.total_marks), 2) if metrics.total_marks else "-"
                elif metric_key == "average_percentage":
                    val = round(float(metrics.average_percentage), 2) if metrics.average_percentage else "-"
                elif metric_key == "average_grade":
                    val = metrics.average_grade if metrics.average_grade else "-"
                elif metric_key == "division":
                    val = metrics.division.division if metrics and metrics.division else "-"
                elif metric_key == "class_position":
                    val = pos_obj.class_position if pos_obj and pos_obj.class_position else "-"
                elif metric_key == "remark":
                    # Determine remark based on average percentage
                    avg_pct = float(metrics.average_percentage) if metrics.average_percentage else 0
                    val = "Passed" if avg_pct >= 50 else "Fail"

            # Merge the MARK and GRADE columns for this session
            merge_start = col_start
            merge_end = col_start + 1
            
            # Set value in the first cell of the merged range
            merged_cell = ws.cell(row=current_row, column=merge_start, value=val)
            merged_cell.alignment = center_align
            merged_cell.border = thin_border
            
            # Merge the cells
            ws.merge_cells(
                start_row=current_row,
                start_column=merge_start,
                end_row=current_row,
                end_column=merge_end
            )

        # Trailing aggregate columns - add borders to all remaining columns
        for c in range(aggregate_cols.get("TOTAL MARKS"), col_idx):
            ws.cell(row=current_row, column=c).border = thin_border

        current_row += 1

    # ============================================
    # 10. GRAND AGGREGATE FOOTER (GRAND TOTALS)
    # ============================================
    
    # Add a blank row for separation
    current_row += 1
    
    # Grand Total header
    ws.cell(row=current_row, column=1).border = thin_border
    grand_label = ws.cell(row=current_row, column=2, value="GRAND TOTAL")
    grand_label.font = bold_font
    grand_label.alignment = left_align
    grand_label.border = thin_border
    
    # Grand Total values - merge MARK and GRADE columns for each session
    for session_id, col_start in session_col_map.items():
        # Merge the MARK and GRADE columns for this session
        merge_start = col_start
        merge_end = col_start + 1
        
        # For grand total, we could put a combined value or leave empty
        # Here we're leaving it empty but with borders
        merged_cell = ws.cell(row=current_row, column=merge_start, value="-")
        merged_cell.alignment = center_align
        merged_cell.border = thin_border
        
        ws.merge_cells(
            start_row=current_row,
            start_column=merge_start,
            end_row=current_row,
            end_column=merge_end
        )
    
    # Grand Total aggregate columns
    # TOTAL MARKS
    total_marks_col = aggregate_cols.get("TOTAL MARKS")
    if total_marks_col:
        total_cell = ws.cell(row=current_row, column=total_marks_col, value=grand_aggregates['total_marks'])
        total_cell.alignment = center_align
        total_cell.border = thin_border
        total_cell.font = bold_font
    
    # AVERAGE
    avg_col = aggregate_cols.get("AVERAGE")
    if avg_col:
        avg_cell = ws.cell(row=current_row, column=avg_col, value=grand_aggregates['grand_average'])
        avg_cell.alignment = center_align
        avg_cell.border = thin_border
        avg_cell.font = bold_font
    
    # GRADE
    grade_col = aggregate_cols.get("GRADE")
    if grade_col:
        grade_cell = ws.cell(row=current_row, column=grade_col, value=grand_aggregates['average_grade'])
        grade_cell.alignment = center_align
        grade_cell.border = thin_border
        grade_cell.font = bold_font
    
    # POSITION
    pos_col = aggregate_cols.get("POSITION")
    if pos_col:
        pos_cell = ws.cell(row=current_row, column=pos_col, value=grand_aggregates['class_position'])
        pos_cell.alignment = center_align
        pos_cell.border = thin_border
        pos_cell.font = bold_font
    
    # REMARK
    remark_col = aggregate_cols.get("REMARK")
    if remark_col:
        remark_cell = ws.cell(row=current_row, column=remark_col, value=grand_aggregates['remark'])
        remark_cell.alignment = center_align
        remark_cell.border = thin_border
        remark_cell.font = bold_font
    
    current_row += 1
    
    # ============================================
    # 11. GRAND DIVISION AND POINTS ROW
    # ============================================
    
    # Add a row for grand division and points
    ws.cell(row=current_row, column=1).border = thin_border
    
    if level_code in ['O_LEVEL', 'A_LEVEL']:
        division_label = ws.cell(row=current_row, column=2, value=f"Grand Division / Points ({level_code})")
    else:
        division_label = ws.cell(row=current_row, column=2, value="Grand Division / Points")
    
    division_label.font = bold_font
    division_label.alignment = left_align
    division_label.border = thin_border
    
    for session_id, col_start in session_col_map.items():
        merge_start = col_start
        merge_end = col_start + 1
        
        merged_cell = ws.cell(row=current_row, column=merge_start, value="-")
        merged_cell.alignment = center_align
        merged_cell.border = thin_border
        
        ws.merge_cells(
            start_row=current_row,
            start_column=merge_start,
            end_row=current_row,
            end_column=merge_end
        )
    
    # Add grand division and points in aggregate columns
    if level_code in ['O_LEVEL', 'A_LEVEL']:
        # Show division in first aggregate column
        div_col = aggregate_cols.get("TOTAL MARKS")
        if div_col:
            div_cell = ws.cell(row=current_row, column=div_col, value=grand_aggregates['division'])
            div_cell.alignment = center_align
            div_cell.border = thin_border
            div_cell.font = bold_font
        
        # Show total points in second aggregate column
        points_col = aggregate_cols.get("AVERAGE")
        if points_col:
            points_cell = ws.cell(row=current_row, column=points_col, value=grand_aggregates['grand_points'])
            points_cell.alignment = center_align
            points_cell.border = thin_border
            points_cell.font = bold_font
    
    current_row += 1
    
    # ============================================
    # 12. GRAND SUMMARY ROW (Additional metrics)
    # ============================================
    
    # Add a row for grand average percentage and points
    ws.cell(row=current_row, column=1).border = thin_border
    summary_label = ws.cell(row=current_row, column=2, value="Grand Average % / Points")
    summary_label.font = bold_font
    summary_label.alignment = left_align
    summary_label.border = thin_border
    
    for session_id, col_start in session_col_map.items():
        merge_start = col_start
        merge_end = col_start + 1
        
        merged_cell = ws.cell(row=current_row, column=merge_start, value="-")
        merged_cell.alignment = center_align
        merged_cell.border = thin_border
        
        ws.merge_cells(
            start_row=current_row,
            start_column=merge_start,
            end_row=current_row,
            end_column=merge_end
        )
    
    # Add additional metrics in aggregate columns
    # We can show grand average percentage and grand points in the first two aggregate columns
    avg_col = aggregate_cols.get("TOTAL MARKS")
    if avg_col:
        avg_cell = ws.cell(row=current_row, column=avg_col, value=f"{grand_aggregates['average_percentage']}%")
        avg_cell.alignment = center_align
        avg_cell.border = thin_border
    
    points_col = aggregate_cols.get("AVERAGE")
    if points_col:
        points_cell = ws.cell(row=current_row, column=points_col, value=grand_aggregates['average_points'])
        points_cell.alignment = center_align
        points_cell.border = thin_border
    
    current_row += 1
    
    # ============================================
    # 13. FINAL FORMATTING
    # ============================================
    # Set column widths
    ws.column_dimensions['A'].width = 5   # S/N
    ws.column_dimensions['B'].width = 30  # Subject Name
    
    # Set widths for session columns and aggregate columns
    for i in range(3, col_idx):
        ws.column_dimensions[get_column_letter(i)].width = 12
    
    # Add report generation timestamp
    current_row += 1
    ws.cell(
        row=current_row, 
        column=1, 
        value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    
    return wb


def calculate_subject_position(subject_id, subjects_list):
    """
    Helper function to calculate position of a subject based on average marks.
    Handles tie-breaking consistently.
    """
    # Filter subjects with valid averages
    subjects_with_avg = [
        s for s in subjects_list 
        if s.get('average_marks') is not None
    ]
    
    # Sort by average marks descending
    subjects_with_avg.sort(key=lambda x: x['average_marks'], reverse=True)
    
    # Find position of target subject
    for idx, subject in enumerate(subjects_with_avg, 1):
        if subject['id'] == subject_id:
            return idx
    
    return None