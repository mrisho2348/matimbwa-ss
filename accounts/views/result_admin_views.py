# results/views/grading_scales.py
import csv
import json
import math
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import os
from django.conf import settings
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import openpyxl
import pandas as pd
from django.template.loader import render_to_string
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.core.paginator import EmptyPage, Paginator
from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, Max, Min, Q, Sum
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.timesince import timesince
from django.views.decorators.http import (require_GET, require_http_methods,
                                          require_POST)

from openpyxl.utils import get_column_letter
from weasyprint import HTML

from core.models import (AcademicYear, ClassLevel, EducationalLevel,
                         StreamClass, Subject, Term)
from results.models import (DivisionScale, ExamSession, ExamType, GradingScale,
                            StudentExamMetrics, StudentExamPosition,
                            StudentResult)
from students.models import Student


@login_required
def grading_scales_list(request):
    """Display grading scales management page"""
    grading_scales = GradingScale.objects.select_related('education_level').all().order_by(
        'education_level__name', '-min_mark'
    )
    
    education_levels = EducationalLevel.objects.filter(is_active=True)
    
    # Get GRADE_CHOICES from model
    grade_choices = GradingScale.GRADE_CHOICES
    
    context = {
        'grading_scales': grading_scales,
        'education_levels': education_levels,
        'grade_choices': grade_choices,
        'page_title': 'Grading Scales Management',
    }
    
    return render(request, 'admin/results/grading_scales_list.html', context)


@login_required
@require_POST
def grading_scales_crud(request):
    """Handle AJAX CRUD operations for grading scales"""
    action = request.POST.get('action', '').lower()
    
    try:
        if action == 'create':
            return create_grading_scale(request)
        elif action == 'update':
            return update_grading_scale(request)
        elif action == 'delete':
            return delete_grading_scale(request)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid action specified.'
            })
            
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'An error occurred: {str(e)}'
        })


def create_grading_scale(request):
    """Create a new grading scale"""
    education_level_id = request.POST.get('education_level')
    grade = request.POST.get('grade', '').strip()
    min_mark = request.POST.get('min_mark', '').strip()
    max_mark = request.POST.get('max_mark', '').strip()
    points = request.POST.get('points', '0').strip()
    description = request.POST.get('description', '').strip()
    
    # Validate required fields
    if not all([education_level_id, grade, min_mark, max_mark]):
        return JsonResponse({
            'success': False,
            'message': 'Education level, grade, min mark, and max mark are required.'
        })
    
    # Validate grade choice
    valid_grades = [choice[0] for choice in GradingScale.GRADE_CHOICES]
    if grade not in valid_grades:
        return JsonResponse({
            'success': False,
            'message': 'Invalid grade selected.'
        })
    
    # Convert to appropriate types
    try:
        min_mark_val = Decimal(min_mark)
        max_mark_val = Decimal(max_mark)
        points_val = Decimal(points) if points else Decimal('0.0')
    except (ValueError, TypeError):
        return JsonResponse({
            'success': False,
            'message': 'Invalid number format for marks or points.'
        })
    
    # Validate mark ranges
    if min_mark_val < 0 or min_mark_val > 100:
        return JsonResponse({
            'success': False,
            'message': 'Min mark must be between 0 and 100.'
        })
    
    if max_mark_val < 0 or max_mark_val > 100:
        return JsonResponse({
            'success': False,
            'message': 'Max mark must be between 0 and 100.'
        })
    
    if min_mark_val > max_mark_val:
        return JsonResponse({
            'success': False,
            'message': 'Min mark cannot exceed max mark.'
        })
    
    # Get education level
    try:
        education_level = EducationalLevel.objects.get(id=education_level_id)
    except EducationalLevel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected education level does not exist.'
        })
    
    # Check for duplicate grade in same education level
    if GradingScale.objects.filter(
        education_level=education_level,
        grade=grade
    ).exists():
        return JsonResponse({
            'success': False,
            'message': f'Grade "{grade}" already exists for {education_level.name}.'
        })
    
    # Check for overlapping mark ranges
    overlapping_scales = GradingScale.objects.filter(
        education_level=education_level,
        min_mark__lt=max_mark_val,
        max_mark__gt=min_mark_val
    )
    
    if overlapping_scales.exists():
        overlapping = overlapping_scales.first()
        return JsonResponse({
            'success': False,
            'message': f'Mark range overlaps with existing grade "{overlapping.grade}" ({overlapping.min_mark}-{overlapping.max_mark}).'
        })
    
    try:
        with transaction.atomic():
            # Create the grading scale
            grading_scale = GradingScale.objects.create(
                education_level=education_level,
                grade=grade,
                min_mark=min_mark_val,
                max_mark=max_mark_val,
                points=points_val,
                description=description
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Grading scale "{grade}" created successfully for {education_level.name}.',
                'grading_scale': {
                    'id': grading_scale.id,
                    'education_level_id': grading_scale.education_level.id,
                    'education_level_name': grading_scale.education_level.name,
                    'grade': grading_scale.grade,
                    'grade_display': grading_scale.get_grade_display(),
                    'min_mark': str(grading_scale.min_mark),
                    'max_mark': str(grading_scale.max_mark),
                    'points': str(grading_scale.points),
                    'description': grading_scale.description
                }
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating grading scale: {str(e)}'
        })


def update_grading_scale(request):
    """Update an existing grading scale"""
    grading_scale_id = request.POST.get('id')
    if not grading_scale_id:
        return JsonResponse({
            'success': False,
            'message': 'Grading scale ID is required.'
        })
    
    try:
        grading_scale = GradingScale.objects.get(id=grading_scale_id)
    except GradingScale.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Grading scale not found.'
        })
    
    education_level_id = request.POST.get('education_level')
    grade = request.POST.get('grade', '').strip()
    min_mark = request.POST.get('min_mark', '').strip()
    max_mark = request.POST.get('max_mark', '').strip()
    points = request.POST.get('points', '').strip()
    description = request.POST.get('description', '').strip()
    
    # Validate required fields
    if not all([education_level_id, grade, min_mark, max_mark]):
        return JsonResponse({
            'success': False,
            'message': 'Education level, grade, min mark, and max mark are required.'
        })
    
    # Validate grade choice
    valid_grades = [choice[0] for choice in GradingScale.GRADE_CHOICES]
    if grade not in valid_grades:
        return JsonResponse({
            'success': False,
            'message': 'Invalid grade selected.'
        })
    
    # Convert to appropriate types
    try:
        min_mark_val = Decimal(min_mark)
        max_mark_val = Decimal(max_mark)
        points_val = Decimal(points) if points else grading_scale.points
    except (ValueError, TypeError):
        return JsonResponse({
            'success': False,
            'message': 'Invalid number format for marks or points.'
        })
    
    # Validate mark ranges
    if min_mark_val < 0 or min_mark_val > 100:
        return JsonResponse({
            'success': False,
            'message': 'Min mark must be between 0 and 100.'
        })
    
    if max_mark_val < 0 or max_mark_val > 100:
        return JsonResponse({
            'success': False,
            'message': 'Max mark must be between 0 and 100.'
        })
    
    if min_mark_val > max_mark_val:
        return JsonResponse({
            'success': False,
            'message': 'Min mark cannot exceed max mark.'
        })
    
    # Get education level
    try:
        education_level = EducationalLevel.objects.get(id=education_level_id)
    except EducationalLevel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected education level does not exist.'
        })
    
    # Check for duplicate grade in same education level (excluding current)
    if GradingScale.objects.filter(
        education_level=education_level,
        grade=grade
    ).exclude(id=grading_scale.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'Grade "{grade}" already exists for {education_level.name}.'
        })
    
    # Check for overlapping mark ranges (excluding current)
    overlapping_scales = GradingScale.objects.filter(
        education_level=education_level,
        min_mark__lt=max_mark_val,
        max_mark__gt=min_mark_val
    ).exclude(id=grading_scale.id)
    
    if overlapping_scales.exists():
        overlapping = overlapping_scales.first()
        if overlapping:
            return JsonResponse({
                'success': False,
                'message': f'Mark range overlaps with existing grade "{overlapping.grade}" ({overlapping.min_mark}-{overlapping.max_mark}).'
            })
    
    try:
        with transaction.atomic():
            # Update the grading scale
            grading_scale.education_level = education_level
            grading_scale.grade = grade
            grading_scale.min_mark = min_mark_val
            grading_scale.max_mark = max_mark_val
            grading_scale.points = points_val
            grading_scale.description = description
            grading_scale.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Grading scale "{grade}" updated successfully.',
                'grading_scale': {
                    'id': grading_scale.id,
                    'education_level_id': grading_scale.education_level.id,
                    'education_level_name': grading_scale.education_level.name,
                    'grade': grading_scale.grade,
                    'grade_display': grading_scale.get_grade_display(),
                    'min_mark': str(grading_scale.min_mark),
                    'max_mark': str(grading_scale.max_mark),
                    'points': str(grading_scale.points),
                    'description': grading_scale.description
                }
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating grading scale: {str(e)}'
        })


def delete_grading_scale(request):
    """Delete a grading scale"""
    grading_scale_id = request.POST.get('id')
    if not grading_scale_id:
        return JsonResponse({
            'success': False,
            'message': 'Grading scale ID is required.'
        })
    
    try:
        grading_scale = GradingScale.objects.get(id=grading_scale_id)
        grade_info = f'{grading_scale.grade} ({grading_scale.education_level.name})'
        grading_scale.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Grading scale "{grade_info}" deleted successfully.',
            'id': grading_scale_id
        })
        
    except GradingScale.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Grading scale not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting grading scale: {str(e)}'
        })




@login_required
def division_scales_list(request):
    """Display division scales management page"""
    division_scales = DivisionScale.objects.select_related('education_level').all().order_by('education_level', 'min_points')
    education_levels = EducationalLevel.objects.filter(is_active=True)
    
    context = {
        'division_scales': division_scales,
        'education_levels': education_levels,
        'DIVISION_CHOICES': DivisionScale.DIVISION_CHOICES,
        'page_title': 'Division Scales Management',
    }
    
    return render(request, 'admin/results/division_scales_list.html', context)

@login_required
def division_scales_crud(request):
    """Handle AJAX CRUD operations for division scales"""
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_division_scale(request)
            elif action == 'update':
                return update_division_scale(request)
            elif action == 'delete':
                return delete_division_scale(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })

def create_division_scale(request):
    """Create a new division scale"""
    # Get and validate required fields
    education_level_id = request.POST.get('education_level')
    if not education_level_id:
        return JsonResponse({
            'success': False,
            'message': 'Education level is required.'
        })
    
    division = request.POST.get('division', '').strip()
    if not division:
        return JsonResponse({
            'success': False,
            'message': 'Division is required.'
        })
    
    min_points_str = request.POST.get('min_points', '').strip()
    max_points_str = request.POST.get('max_points', '').strip()
    
    if not min_points_str or not max_points_str:
        return JsonResponse({
            'success': False,
            'message': 'Both minimum and maximum points are required.'
        })
    
    try:
        min_points = int(min_points_str)
        max_points = int(max_points_str)
        
        if min_points < 0 or max_points < 0:
            return JsonResponse({
                'success': False,
                'message': 'Points cannot be negative.'
            })
        
        if min_points > max_points:
            return JsonResponse({
                'success': False,
                'message': 'Minimum points cannot exceed maximum points.'
            })
        
        # Validate points range
        if max_points > 999:
            return JsonResponse({
                'success': False,
                'message': 'Points cannot exceed 999.'
            })
            
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Points must be valid numbers.'
        })
    
    # Get education level
    try:
        education_level = EducationalLevel.objects.get(id=education_level_id)
    except EducationalLevel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected education level does not exist.'
        })
    
    # Validate division choice
    valid_divisions = [choice[0] for choice in DivisionScale.DIVISION_CHOICES]
    if division not in valid_divisions:
        return JsonResponse({
            'success': False,
            'message': 'Invalid division selected.'
        })
    
    # Check for duplicate division within the same education level
    if DivisionScale.objects.filter(
        education_level=education_level,
        division=division
    ).exists():
        return JsonResponse({
            'success': False,
            'message': f'Division "{division}" already exists for {education_level.name}.'
        })
    
    # Check for overlapping point ranges
    overlapping_scales = DivisionScale.objects.filter(
        education_level=education_level
    ).filter(
        min_points__lte=max_points,
        max_points__gte=min_points
    )
    
    if overlapping_scales.exists():
        return JsonResponse({
            'success': False,
            'message': f'Point range overlaps with existing division scale(s).'
        })
    
    try:
        # Create the division scale
        division_scale = DivisionScale.objects.create(
            education_level=education_level,
            division=division,
            min_points=min_points,
            max_points=max_points
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Division scale "{division}" created successfully for {education_level.name}.',
            'division_scale': {
                'id': division_scale.id,
                'education_level_id': division_scale.education_level.id,
                'education_level_name': division_scale.education_level.name,
                'division': division_scale.division,
                'division_display': division_scale.get_division_display(),
                'min_points': division_scale.min_points,
                'max_points': division_scale.max_points,
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'A division scale with these details already exists.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating division scale: {str(e)}'
        })

def update_division_scale(request):
    """Update an existing division scale"""
    scale_id = request.POST.get('id')
    if not scale_id:
        return JsonResponse({
            'success': False,
            'message': 'Division scale ID is required.'
        })
    
    try:
        division_scale = DivisionScale.objects.get(id=scale_id)
    except DivisionScale.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Division scale not found.'
        })
    
    # Get and validate required fields
    education_level_id = request.POST.get('education_level')
    division = request.POST.get('division', '').strip()
    min_points_str = request.POST.get('min_points', '').strip()
    max_points_str = request.POST.get('max_points', '').strip()
    
    if not all([education_level_id, division, min_points_str, max_points_str]):
        return JsonResponse({
            'success': False,
            'message': 'All fields are required.'
        })
    
    try:
        min_points = int(min_points_str)
        max_points = int(max_points_str)
        
        if min_points < 0 or max_points < 0:
            return JsonResponse({
                'success': False,
                'message': 'Points cannot be negative.'
            })
        
        if min_points > max_points:
            return JsonResponse({
                'success': False,
                'message': 'Minimum points cannot exceed maximum points.'
            })
        
        # Validate points range
        if max_points > 999:
            return JsonResponse({
                'success': False,
                'message': 'Points cannot exceed 999.'
            })
            
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Points must be valid numbers.'
        })
    
    # Get education level
    try:
        education_level = EducationalLevel.objects.get(id=education_level_id)
    except EducationalLevel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected education level does not exist.'
        })
    
    # Validate division choice
    valid_divisions = [choice[0] for choice in DivisionScale.DIVISION_CHOICES]
    if division not in valid_divisions:
        return JsonResponse({
            'success': False,
            'message': 'Invalid division selected.'
        })
    
    # Check for duplicate division within the same education level (excluding current)
    if DivisionScale.objects.filter(
        education_level=education_level,
        division=division
    ).exclude(id=division_scale.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'Division "{division}" already exists for {education_level.name}.'
        })
    
    # Check for overlapping point ranges (excluding current)
    overlapping_scales = DivisionScale.objects.filter(
        education_level=education_level
    ).filter(
        min_points__lte=max_points,
        max_points__gte=min_points
    ).exclude(id=division_scale.id)
    
    if overlapping_scales.exists():
        return JsonResponse({
            'success': False,
            'message': f'Point range overlaps with existing division scale(s).'
        })
    
    try:
        # Update the division scale
        division_scale.education_level = education_level
        division_scale.division = division
        division_scale.min_points = min_points
        division_scale.max_points = max_points
        division_scale.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Division scale "{division}" updated successfully.',
            'division_scale': {
                'id': division_scale.id,
                'education_level_id': division_scale.education_level.id,
                'education_level_name': division_scale.education_level.name,
                'division': division_scale.division,
                'division_display': division_scale.get_division_display(),
                'min_points': division_scale.min_points,
                'max_points': division_scale.max_points,
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'A division scale with these details already exists.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating division scale: {str(e)}'
        })

def delete_division_scale(request):
    """Delete a division scale"""
    scale_id = request.POST.get('id')
    if not scale_id:
        return JsonResponse({
            'success': False,
            'message': 'Division scale ID is required.'
        })
    
    try:
        division_scale = DivisionScale.objects.get(id=scale_id)
    except DivisionScale.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Division scale not found.'
        })
    
    # Check if division scale is referenced by any student exam metrics
    if division_scale.studentexammetrics_set.exists():
        return JsonResponse({
            'success': False,
            'message': 'Cannot delete division scale that is referenced by student exam metrics.'
        })
    
    scale_info = f'{division_scale.get_division_display()} ({division_scale.education_level.name})'
    
    try:
        division_scale.delete()
        return JsonResponse({
            'success': True,
            'message': f'Division scale "{scale_info}" deleted successfully.',
            'id': scale_id
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting division scale: {str(e)}'
        })

@login_required
def get_division_scales_by_level(request):
    """AJAX endpoint to get division scales by education level"""
    education_level_id = request.GET.get('education_level_id')
    
    if not education_level_id:
        return JsonResponse({
            'success': False,
            'message': 'Education level ID is required.'
        })
    
    try:
        division_scales = DivisionScale.objects.filter(
            education_level_id=education_level_id
        ).order_by('min_points')
        
        scales_list = []
        for scale in division_scales:
            scales_list.append({
                'id': scale.id,
                'division': scale.division,
                'division_display': scale.get_division_display(),
                'min_points': scale.min_points,
                'max_points': scale.max_points,
                'range': f"{scale.min_points} - {scale.max_points}"
            })
        
        return JsonResponse({
            'success': True,
            'division_scales': scales_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
def calculate_division(request):
    """Calculate division for given points and education level"""
    if request.method == 'POST':
        education_level_id = request.POST.get('education_level_id')
        total_points = request.POST.get('total_points')
        
        if not education_level_id or not total_points:
            return JsonResponse({
                'success': False,
                'message': 'Education level and total points are required.'
            })
        
        try:
            total_points = float(total_points)
            education_level = EducationalLevel.objects.get(id=education_level_id)
            
            # Get division scales for this education level
            division_scale = DivisionScale.objects.filter(
                education_level=education_level,
                min_points__lte=total_points,
                max_points__gte=total_points
            ).first()
            
            if division_scale:
                return JsonResponse({
                    'success': True,
                    'division': division_scale.division,
                    'division_display': division_scale.get_division_display(),
                    'points': total_points,
                    'education_level': education_level.name
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'No division scale found for {total_points} points in {education_level.name}.'
                })
                
        except EducationalLevel.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Education level not found.'
            })
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid points value.'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error calculating division: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.'
    })

@login_required
def validate_division_scales(request):
    """Validate division scales for an education level"""
    education_level_id = request.GET.get('education_level_id')
    
    if not education_level_id:
        return JsonResponse({
            'success': False,
            'message': 'Education level ID is required.'
        })
    
    try:
        education_level = EducationalLevel.objects.get(id=education_level_id)
        division_scales = DivisionScale.objects.filter(
            education_level=education_level
        ).order_by('min_points')
        
        # Check for gaps in point ranges
        validation_results = []
        has_gaps = False
        previous_max = -1
        
        for i, scale in enumerate(division_scales):
            # Check if there's a gap
            if scale.min_points > previous_max + 1:
                has_gaps = True
                validation_results.append({
                    'type': 'gap',
                    'message': f'Gap between {previous_max} and {scale.min_points} points',
                    'severity': 'warning'
                })
            
            # Check for overlapping ranges
            for j, other_scale in enumerate(division_scales):
                if i != j and scale.min_points <= other_scale.max_points and scale.max_points >= other_scale.min_points:
                    validation_results.append({
                        'type': 'overlap',
                        'message': f'Overlap with {other_scale.get_division_display()}',
                        'severity': 'error'
                    })
            
            previous_max = scale.max_points
        
        # Check if covers all possible points (0-999)
        if division_scales:
            first_scale = division_scales.first()
            last_scale = division_scales.last()
            
            if first_scale.min_points > 0:
                validation_results.append({
                    'type': 'coverage',
                    'message': f'No division for points 0-{first_scale.min_points - 1}',
                    'severity': 'warning'
                })
            
            if last_scale.max_points < 999:
                validation_results.append({
                    'type': 'coverage',
                    'message': f'No division for points {last_scale.max_points + 1}-999',
                    'severity': 'warning'
                })
        
        return JsonResponse({
            'success': True,
            'validation_results': validation_results,
            'has_errors': any(result['severity'] == 'error' for result in validation_results),
            'has_warnings': any(result['severity'] == 'warning' for result in validation_results),
            'total_scales': division_scales.count()
        })
        
    except EducationalLevel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Education level not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error validating division scales: {str(e)}'
        })        


# ============================================================================
# EXAM TYPE MANAGEMENT VIEWS
# ============================================================================

@login_required
def exam_types_list(request):
    """Display exam types management page"""
    exam_types = ExamType.objects.all().order_by('name')
    
    # Statistics
    total_exam_types = exam_types.count()
    
    context = {
        'exam_types': exam_types,
        'total_exam_types': total_exam_types,
        'page_title': 'Exam Types Management',
    }
    
    return render(request, 'admin/results/exam_types_list.html', context)


@login_required
def exam_types_crud(request):
    """Handle AJAX CRUD operations for exam types"""
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_exam_type(request)
            elif action == 'update':
                return update_exam_type(request)
            elif action == 'delete':
                return delete_exam_type(request)
            elif action == 'bulk_delete':
                return bulk_delete_exam_types(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })


def create_exam_type(request):
    """Create a new exam type"""
    # Get and validate required fields
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({
            'success': False,
            'message': 'Exam type name is required.'
        })
    
    code = request.POST.get('code', '').strip()
    if not code:
        return JsonResponse({
            'success': False,
            'message': 'Exam type code is required.'
        })
    
    weight_str = request.POST.get('weight', '').strip()
    max_score_str = request.POST.get('max_score', '100').strip()
    
    # Validate name and code length
    if len(name) > 100:
        return JsonResponse({
            'success': False,
            'message': 'Exam type name cannot exceed 100 characters.'
        })
    
    if len(code) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Exam type code cannot exceed 20 characters.'
        })
    
    # Validate code format (alphanumeric with hyphens/underscores)
    if not code.replace('-', '').replace('_', '').isalnum():
        return JsonResponse({
            'success': False,
            'message': 'Exam type code can only contain letters, numbers, hyphens, and underscores.'
        })
    
    # Validate weight
    try:
        weight = float(weight_str)
        if weight < 0 or weight > 100:
            return JsonResponse({
                'success': False,
                'message': 'Weight must be between 0 and 100.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Weight must be a valid number.'
        })
    
    # Validate max score
    try:
        max_score = float(max_score_str)
        if max_score <= 0:
            return JsonResponse({
                'success': False,
                'message': 'Maximum score must be greater than 0.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Maximum score must be a valid number.'
        })
    
    # Get optional field
    description = request.POST.get('description', '').strip()
    
    # Validate description length if provided
    if description and len(description) > 500:
        return JsonResponse({
            'success': False,
            'message': 'Description cannot exceed 500 characters.'
        })
    
    # Check for duplicate code
    if ExamType.objects.filter(code__iexact=code).exists():
        return JsonResponse({
            'success': False,
            'message': f'Exam type with code "{code}" already exists.'
        })
    
    try:
        # Create the exam type
        exam_type = ExamType.objects.create(
            name=name,
            code=code.upper(),  # Store code in uppercase for consistency
            weight=weight,
            max_score=max_score,
            description=description
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Exam type "{name}" created successfully.',
            'exam_type': {
                'id': exam_type.id,
                'name': exam_type.name,
                'code': exam_type.code,
                'weight': float(exam_type.weight),
                'max_score': float(exam_type.max_score),
                'description': exam_type.description
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'An exam type with code "{code}" already exists.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating exam type: {str(e)}'
        })


def update_exam_type(request):
    """Update an existing exam type"""
    exam_type_id = request.POST.get('id')
    if not exam_type_id:
        return JsonResponse({
            'success': False,
            'message': 'Exam type ID is required.'
        })
    
    try:
        exam_type = ExamType.objects.get(id=exam_type_id)
    except ExamType.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam type not found.'
        })
    
    # Get and validate required fields
    name = request.POST.get('name', '').strip()
    code = request.POST.get('code', '').strip()
    weight_str = request.POST.get('weight', '').strip()
    max_score_str = request.POST.get('max_score', '').strip()
    
    if not all([name, code, weight_str, max_score_str]):
        return JsonResponse({
            'success': False,
            'message': 'All fields are required.'
        })
    
    # Validate name and code length
    if len(name) > 100:
        return JsonResponse({
            'success': False,
            'message': 'Exam type name cannot exceed 100 characters.'
        })
    
    if len(code) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Exam type code cannot exceed 20 characters.'
        })
    
    # Validate code format
    if not code.replace('-', '').replace('_', '').isalnum():
        return JsonResponse({
            'success': False,
            'message': 'Exam type code can only contain letters, numbers, hyphens, and underscores.'
        })
    
    # Validate weight
    try:
        weight = float(weight_str)
        if weight < 0 or weight > 100:
            return JsonResponse({
                'success': False,
                'message': 'Weight must be between 0 and 100.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Weight must be a valid number.'
        })
    
    # Validate max score
    try:
        max_score = float(max_score_str)
        if max_score <= 0:
            return JsonResponse({
                'success': False,
                'message': 'Maximum score must be greater than 0.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Maximum score must be a valid number.'
        })
    
    # Get optional field
    description = request.POST.get('description', '').strip()
    
    # Validate description length
    if description and len(description) > 500:
        return JsonResponse({
            'success': False,
            'message': 'Description cannot exceed 500 characters.'
        })
    
    # Check for duplicate code (excluding current)
    if ExamType.objects.filter(code__iexact=code).exclude(id=exam_type.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'Exam type with code "{code}" already exists.'
        })
    
    try:
        # Update the exam type
        exam_type.name = name
        exam_type.code = code.upper()
        exam_type.weight = weight
        exam_type.max_score = max_score
        exam_type.description = description
        exam_type.full_clean()  # Run model validation
        exam_type.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Exam type "{name}" updated successfully.',
            'exam_type': {
                'id': exam_type.id,
                'name': exam_type.name,
                'code': exam_type.code,
                'weight': float(exam_type.weight),
                'max_score': float(exam_type.max_score),
                'description': exam_type.description
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'An exam type with code "{code}" already exists.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except ValidationError as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating exam type: {str(e)}'
        })


def delete_exam_type(request):
    """Delete an exam type"""
    exam_type_id = request.POST.get('id')
    if not exam_type_id:
        return JsonResponse({
            'success': False,
            'message': 'Exam type ID is required.'
        })
    
    try:
        exam_type = ExamType.objects.get(id=exam_type_id)
    except ExamType.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam type not found.'
        })
    
    # Check if exam type is used in any exam sessions
    if exam_type.examsession_set.exists():
        session_count = exam_type.examsession_set.count()
        return JsonResponse({
            'success': False,
            'message': f'Cannot delete exam type "{exam_type.name}". It is used in {session_count} exam session(s).'
        })
    
    exam_type_name = exam_type.name
    
    try:
        exam_type.delete()
        return JsonResponse({
            'success': True,
            'message': f'Exam type "{exam_type_name}" deleted successfully.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting exam type: {str(e)}'
        })


def bulk_delete_exam_types(request):
    """Bulk delete multiple exam types"""
    try:
        exam_type_ids = request.POST.getlist('exam_type_ids[]')
        
        if not exam_type_ids:
            return JsonResponse({
                'success': False,
                'message': 'No exam types selected.'
            })
        
        exam_types = ExamType.objects.filter(id__in=exam_type_ids)
        deletable_types = []
        non_deletable_types = []
        
        for exam_type in exam_types:
            # Check if exam type is used in any exam sessions
            if exam_type.examsession_set.exists():
                non_deletable_types.append(exam_type.name)
            else:
                deletable_types.append(exam_type)
        
        # Delete deletable exam types
        if deletable_types:
            deleted_count = len(deletable_types)
            ExamType.objects.filter(id__in=[et.id for et in deletable_types]).delete()
            
            message = f'Successfully deleted {deleted_count} exam type(s).'
            if non_deletable_types:
                message += f' Could not delete: {", ".join(non_deletable_types[:3])}'
                if len(non_deletable_types) > 3:
                    message += f' and {len(non_deletable_types) - 3} more (used in exam sessions).'
                else:
                    message += ' (used in exam sessions).'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'deleted_count': deleted_count,
                'failed_count': len(non_deletable_types)
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f'None of the selected exam types can be deleted (all are used in exam sessions).'
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error bulk deleting exam types: {str(e)}'
        })        




@login_required
def exam_sessions_list(request):
    """Display exam sessions management page"""
    # Get all exam sessions with related data
    exam_sessions = ExamSession.objects.select_related(
        'exam_type',
        'academic_year',
        'term',
        'class_level',
        'class_level__educational_level',
        'stream_class'
    ).prefetch_related(
        'papers__subject'
    ).all()
    
    # Get filter parameters
    academic_year_filter = request.GET.get('academic_year', '')
    term_filter = request.GET.get('term', '')
    class_level_filter = request.GET.get('class_level', '')
    stream_filter = request.GET.get('stream', '')
    status_filter = request.GET.get('status', '')
    
    # Apply filters
    if academic_year_filter:
        exam_sessions = exam_sessions.filter(academic_year_id=academic_year_filter)
    
    if term_filter:
        exam_sessions = exam_sessions.filter(term_id=term_filter)
    
    if class_level_filter:
        exam_sessions = exam_sessions.filter(class_level_id=class_level_filter)
    
    if stream_filter:
        exam_sessions = exam_sessions.filter(stream_class_id=stream_filter)
    
    if status_filter:
        exam_sessions = exam_sessions.filter(status=status_filter)
    
    # Count statistics
    total_sessions = exam_sessions.count()
    draft_sessions = exam_sessions.filter(status='draft').count()
    submitted_sessions = exam_sessions.filter(status='submitted').count()
    verified_sessions = exam_sessions.filter(status='verified').count()
    published_sessions = exam_sessions.filter(status='published').count()
    
    # Get data for filters
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    # Only show terms for the selected academic year if filter is applied
    if academic_year_filter:
        terms = Term.objects.filter(academic_year_id=academic_year_filter).order_by('term_number')
    else:
        terms = Term.objects.none()
    
    class_levels = ClassLevel.objects.filter(is_active=True).select_related('educational_level').order_by('order')
    # Only show streams for the selected class if filter is applied
    if class_level_filter:
        streams = StreamClass.objects.filter(class_level_id=class_level_filter, is_active=True)
    else:
        streams = StreamClass.objects.none()
    
    status_choices = ExamSession.STATUS_CHOICES
    
    # Get exam types for form
    exam_types = ExamType.objects.all().order_by('name')
    
    context = {
        'exam_sessions': exam_sessions,
        'academic_years': academic_years,
        'terms': terms,
        'class_levels': class_levels,
        'streams': streams,
        'status_choices': status_choices,
        'exam_types': exam_types,
        'total_sessions': total_sessions,
        'draft_sessions': draft_sessions,
        'submitted_sessions': submitted_sessions,
        'verified_sessions': verified_sessions,
        'published_sessions': published_sessions,
        'page_title': 'Exam Sessions Management',
        'today': timezone.now().date(),
    }
    
    return render(request, 'admin/results/exam_sessions_list.html', context)


@login_required
def get_terms_for_academic_year(request):
    """API endpoint to get terms for a specific academic year"""
    if request.method == 'GET':
        academic_year_id = request.GET.get('academic_year_id')
        
        if academic_year_id:
            try:
                terms = Term.objects.filter(
                    academic_year_id=academic_year_id
                ).order_by('term_number')
                
                terms_data = [{
                    'id': term.id,
                    'text': term.get_term_number_display()
                } for term in terms]
                
                return JsonResponse({
                    'success': True,
                    'terms': terms_data
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Academic year ID is required'
            }, status=400)


@login_required
def get_streams_for_exam_session(request):
    """API endpoint to get streams for a specific class level"""
    if request.method == 'GET':
        class_level_id = request.GET.get('class_level_id')
        
        if class_level_id:
            try:
                streams = StreamClass.objects.filter(
                    class_level_id=class_level_id,
                    is_active=True
                ).order_by('stream_letter')
                
                streams_data = [{
                    'id': stream.id,
                    'text': stream.stream_letter
                } for stream in streams]
                
                return JsonResponse({
                    'success': True,
                    'streams': streams_data
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': str(e)
                }, status=400)
        else:
            return JsonResponse({
                'success': False,
                'message': 'Class level ID is required'
            }, status=400)


@login_required
def exam_sessions_crud(request):
    """Handle AJAX CRUD operations for exam sessions"""
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_exam_session(request)
            elif action == 'update':
                return update_exam_session(request)
            elif action == 'toggle_status':
                return toggle_exam_session_status(request)
            elif action == 'delete':
                return delete_exam_session(request)
            elif action == 'update_status':
                return update_exam_session_status(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })

def create_exam_session(request):
    """Create a new exam session"""
    # Get and validate required fields
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({
            'success': False,
            'message': 'Exam session name is required.'
        })
    
    exam_type_id = request.POST.get('exam_type')
    if not exam_type_id:
        return JsonResponse({
            'success': False,
            'message': 'Exam type is required.'
        })
    
    academic_year_id = request.POST.get('academic_year')
    if not academic_year_id:
        return JsonResponse({
            'success': False,
            'message': 'Academic year is required.'
        })
    
    term_id = request.POST.get('term')
    if not term_id:
        return JsonResponse({
            'success': False,
            'message': 'Term is required.'
        })
    
    class_level_id = request.POST.get('class_level')
    if not class_level_id:
        return JsonResponse({
            'success': False,
            'message': 'Class level is required.'
        })
    
    exam_date_str = request.POST.get('exam_date')
    if not exam_date_str:
        return JsonResponse({
            'success': False,
            'message': 'Exam date is required.'
        })
    
    # Parse exam date
    try:
        exam_date = datetime.strptime(exam_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid exam date format. Please use YYYY-MM-DD.'
        })
    
    # Get optional fields
    stream_class_id = request.POST.get('stream_class', '').strip()
    stream_class = None
    if stream_class_id:
        try:
            stream_class = StreamClass.objects.get(id=stream_class_id)
        except StreamClass.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Selected stream class does not exist.'
            })
    
    status = request.POST.get('status', 'draft')
    
    # Validate name length
    if len(name) < 5 or len(name) > 200:
        return JsonResponse({
            'success': False,
            'message': 'Exam session name must be between 5 and 200 characters.'
        })
    
    # Validate date is not in the future for non-draft statuses
    if status != 'draft' and exam_date > date.today():
        return JsonResponse({
            'success': False,
            'message': 'Exam date cannot be in the future for non-draft sessions.'
        })
    
    # Get related objects
    try:
        exam_type = ExamType.objects.get(id=exam_type_id)
        academic_year = AcademicYear.objects.get(id=academic_year_id)
        term = Term.objects.get(id=term_id)
        class_level = ClassLevel.objects.get(id=class_level_id)
    except (ExamType.DoesNotExist, AcademicYear.DoesNotExist, 
            Term.DoesNotExist, ClassLevel.DoesNotExist) as e:
        return JsonResponse({
            'success': False,
            'message': f'Selected item does not exist: {str(e)}'
        })
    
    # Check for duplicate exam session name in same academic year and term
    duplicate_check = ExamSession.objects.filter(
        name__iexact=name,
        academic_year=academic_year,
        term=term,
        class_level=class_level
    )
    
    if stream_class:
        duplicate_check = duplicate_check.filter(stream_class=stream_class)
    else:
        duplicate_check = duplicate_check.filter(stream_class__isnull=True)
    
    if duplicate_check.exists():
        return JsonResponse({
            'success': False,
            'message': f'An exam session with name "{name}" already exists for the selected parameters.'
        })
    
    try:
        # Create the exam session
        exam_session = ExamSession.objects.create(
            name=name,
            exam_type=exam_type,
            academic_year=academic_year,
            term=term,
            class_level=class_level,
            stream_class=stream_class,
            exam_date=exam_date,
            status=status
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Exam session "{name}" created successfully.',
            'exam_session': {
                'id': exam_session.id,
                'name': exam_session.name,
                'exam_type_name': exam_session.exam_type.name,
                'academic_year_name': exam_session.academic_year.name,
                'term_name': f'{exam_session.term.get_term_number_display()}',
                'class_level_name': exam_session.class_level.name,
                'stream_class_name': str(exam_session.stream_class) if exam_session.stream_class else 'All Streams',
                'exam_date': exam_session.exam_date.strftime('%Y-%m-%d'),
                'status': exam_session.status,
                'status_display': exam_session.get_status_display()
            }
        })
        
    except IntegrityError as e:
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating exam session: {str(e)}'
        })

def update_exam_session(request):
    """Update an existing exam session"""
    exam_session_id = request.POST.get('id')
    if not exam_session_id:
        return JsonResponse({
            'success': False,
            'message': 'Exam session ID is required.'
        })
    
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
    except ExamSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam session not found.'
        })
    
    # Check if exam session can be edited (not published or verified)
    if exam_session.status in ['published', 'verified']:
        return JsonResponse({
            'success': False,
            'message': f'Cannot edit a {exam_session.get_status_display().lower()} exam session.'
        })
    
    # Get and validate required fields
    name = request.POST.get('name', '').strip()
    exam_type_id = request.POST.get('exam_type')
    academic_year_id = request.POST.get('academic_year')
    term_id = request.POST.get('term')
    class_level_id = request.POST.get('class_level')
    exam_date_str = request.POST.get('exam_date')
    
    if not all([name, exam_type_id, academic_year_id, term_id, class_level_id, exam_date_str]):
        return JsonResponse({
            'success': False,
            'message': 'All required fields must be filled.'
        })
    
    # Parse exam date
    try:
        exam_date = datetime.strptime(exam_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid exam date format.'
        })
    
    # Get optional fields
    stream_class_id = request.POST.get('stream_class', '').strip()
    stream_class = None
    if stream_class_id:
        try:
            stream_class = StreamClass.objects.get(id=stream_class_id)
        except StreamClass.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Selected stream class does not exist.'
            })
    
    # Validate name length
    if len(name) < 5 or len(name) > 200:
        return JsonResponse({
            'success': False,
            'message': 'Exam session name must be between 5 and 200 characters.'
        })
    
    # Get related objects
    try:
        exam_type = ExamType.objects.get(id=exam_type_id)
        academic_year = AcademicYear.objects.get(id=academic_year_id)
        term = Term.objects.get(id=term_id)
        class_level = ClassLevel.objects.get(id=class_level_id)
    except (ExamType.DoesNotExist, AcademicYear.DoesNotExist, 
            Term.DoesNotExist, ClassLevel.DoesNotExist) as e:
        return JsonResponse({
            'success': False,
            'message': f'Selected item does not exist: {str(e)}'
        })
    
    # Check for duplicate exam session name (excluding current)
    duplicate_check = ExamSession.objects.filter(
        name__iexact=name,
        academic_year=academic_year,
        term=term,
        class_level=class_level
    ).exclude(id=exam_session.id)
    
    if stream_class:
        duplicate_check = duplicate_check.filter(stream_class=stream_class)
    else:
        duplicate_check = duplicate_check.filter(stream_class__isnull=True)
    
    if duplicate_check.exists():
        return JsonResponse({
            'success': False,
            'message': f'Another exam session with name "{name}" already exists for the selected parameters.'
        })
    
    try:
        # Update the exam session
        exam_session.name = name
        exam_session.exam_type = exam_type
        exam_session.academic_year = academic_year
        exam_session.term = term
        exam_session.class_level = class_level
        exam_session.stream_class = stream_class
        exam_session.exam_date = exam_date
        exam_session.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Exam session "{name}" updated successfully.',
            'exam_session': {
                'id': exam_session.id,
                'name': exam_session.name,
                'exam_type_name': exam_session.exam_type.name,
                'academic_year_name': exam_session.academic_year.name,
                'term_name': f'{exam_session.term.get_term_number_display()}',
                'class_level_name': exam_session.class_level.name,
                'stream_class_name': str(exam_session.stream_class) if exam_session.stream_class else 'All Streams',
                'exam_date': exam_session.exam_date.strftime('%Y-%m-%d'),
                'status': exam_session.status,
                'status_display': exam_session.get_status_display()
            }
        })
        
    except IntegrityError as e:
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating exam session: {str(e)}'
        })

def toggle_exam_session_status(request):
    """Toggle exam session status through workflow"""
    exam_session_id = request.POST.get('id')
    new_status = request.POST.get('status', '').lower()
    
    if not exam_session_id:
        return JsonResponse({
            'success': False,
            'message': 'Exam session ID is required.'
        })
    
    valid_statuses = ['draft', 'submitted', 'verified', 'published']
    if new_status not in valid_statuses:
        return JsonResponse({
            'success': False,
            'message': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
        })
    
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
    except ExamSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam session not found.'
        })
    
    # Validate status transition
    current_status = exam_session.status
    allowed_transitions = {
        'draft': ['submitted', 'draft'],
        'submitted': ['verified', 'draft'],
        'verified': ['published', 'submitted'],
        'published': ['published']  # Once published, cannot change
    }
    
    if new_status not in allowed_transitions.get(current_status, []):
        return JsonResponse({
            'success': False,
            'message': f'Cannot change status from {current_status} to {new_status}.'
        })
    
    # Additional validation for publishing
    if new_status == 'published':
        # Check if exam date is in the past
        if exam_session.exam_date > date.today():
            return JsonResponse({
                'success': False,
                'message': 'Cannot publish an exam session with future date.'
            })
        
        # Check if there are papers added
        if not exam_session.results.exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot publish an exam session without any papers.'
            })
    
    try:
        # Update status
        exam_session.status = new_status
        exam_session.save()
        
        action_text = {
            'draft': 'returned to draft',
            'submitted': 'submitted for review',
            'verified': 'verified',
            'published': 'published'
        }.get(new_status, 'updated')
        
        return JsonResponse({
            'success': True,
            'message': f'Exam session "{exam_session.name}" {action_text} successfully.',
            'status': exam_session.status,
            'status_display': exam_session.get_status_display()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating status: {str(e)}'
        })

def delete_exam_session(request):
    """Delete an exam session"""
    exam_session_id = request.POST.get('id')
    if not exam_session_id:
        return JsonResponse({
            'success': False,
            'message': 'Exam session ID is required.'
        })
    
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
    except ExamSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam session not found.'
        })
    
    # Check if exam session can be deleted
    if exam_session.status in ['published', 'verified']:
        return JsonResponse({
            'success': False,
            'message': f'Cannot delete a {exam_session.get_status_display().lower()} exam session.'
        })
    
    # Check if there are any papers or results
    if exam_session.results.exists():
        return JsonResponse({
            'success': False,
            'message': 'Cannot delete exam session with papers. Delete papers first.'
        })
    
    exam_session_name = exam_session.name
    
    try:
        exam_session.delete()
        return JsonResponse({
            'success': True,
            'message': f'Exam session "{exam_session_name}" deleted successfully.',
            'id': exam_session_id
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting exam session: {str(e)}'
        })

def update_exam_session_status(request):
    """Update exam session status with custom message"""
    exam_session_id = request.POST.get('id')
    status = request.POST.get('status', '').lower()
    
    if not exam_session_id or not status:
        return JsonResponse({
            'success': False,
            'message': 'Exam session ID and status are required.'
        })
    
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
    except ExamSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam session not found.'
        })
    
    # Validate status
    valid_statuses = [choice[0] for choice in ExamSession.STATUS_CHOICES]
    if status not in valid_statuses:
        return JsonResponse({
            'success': False,
            'message': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
        })
    
    try:
        old_status = exam_session.status
        exam_session.status = status
        exam_session.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Exam session status changed from {old_status} to {status}.',
            'status': exam_session.status,
            'status_display': exam_session.get_status_display()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating status: {str(e)}'
        })

@login_required
def exam_session_detail(request, exam_session_id):
    """View exam session details"""
    exam_session = get_object_or_404(
        ExamSession.objects.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'class_level__educational_level',
            'stream_class'
        ),
        id=exam_session_id
    )
    
    # Get papers for this exam session
    papers = exam_session.papers.select_related('subject').all()
    
    # Get student count for this class/stream
    if exam_session.stream_class:
        student_count = Student.objects.filter(
            class_level=exam_session.class_level,
            stream_class=exam_session.stream_class,
            is_active=True
        ).count()
    else:
        student_count = Student.objects.filter(
            class_level=exam_session.class_level,
            is_active=True
        ).count()
    
    # Get statistics
    total_papers = papers.count()
    papers_with_results = papers.filter(results__isnull=False).distinct().count()
    
    context = {
        'exam_session': exam_session,
        'papers': papers,
        'student_count': student_count,
        'total_papers': total_papers,
        'papers_with_results': papers_with_results,
        'page_title': f'Exam Session: {exam_session.name}',
    }
    
    return render(request, 'admin/results/exam_session_detail.html', context)



@login_required
def get_streams_for_exam_session(request):
    """AJAX endpoint to get streams for a class level in exam session"""
    class_level_id = request.GET.get('class_level_id')
    
    if not class_level_id:
        return JsonResponse({'success': False, 'message': 'Class level ID required'})
    
    try:
        streams = StreamClass.objects.filter(
            class_level_id=class_level_id,
            is_active=True
        ).select_related('class_level').order_by('stream_letter')
        
        stream_list = [{'id': '', 'text': 'All Streams'}]  # Add option for all streams
        for stream in streams:
            stream_list.append({
                'id': stream.id,
                'text': f"{stream.class_level.name}{stream.stream_letter}"
            })
        
        return JsonResponse({
            'success': True,
            'streams': stream_list
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def get_exam_session_stats(request):
    """AJAX endpoint to get exam session statistics"""
    exam_session_id = request.GET.get('exam_session_id')
    
    if not exam_session_id:
        return JsonResponse({'success': False, 'message': 'Exam session ID required'})
    
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
        
        # Get paper count
        paper_count = exam_session.papers.count()
        
        # Get student count
        if exam_session.stream_class:
            student_count = Student.objects.filter(
                class_level=exam_session.class_level,
                stream_class=exam_session.stream_class,
                is_active=True
            ).count()
        else:
            student_count = Student.objects.filter(
                class_level=exam_session.class_level,
                is_active=True
            ).count()
        
        # Get result statistics
        result_stats = StudentResult.objects.filter(
            exam_paper__exam_session=exam_session
        ).aggregate(
            total_students=Count('student', distinct=True),
            papers_marked=Count('exam_paper', distinct=True),
            average_percentage=Avg('percentage'),
            highest_percentage=Max('percentage'),
            lowest_percentage=Min('percentage')
        )
        
        return JsonResponse({
            'success': True,
            'stats': {
                'paper_count': paper_count,
                'student_count': student_count,
                **result_stats
            }
        })
    except ExamSession.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Exam session not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})        
    
def admin_exam_sessions_data(request):
    """Return exam sessions data as JSON for DataTable"""
    try:
        # Get DataTable parameters
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        
        # Get filter parameters
        academic_year = request.GET.get('academic_year')
        term = request.GET.get('term')
        class_level = request.GET.get('class_level')
        stream = request.GET.get('stream')
        status = request.GET.get('status')
        
        # Log the filters for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.debug(f"Filters - Academic Year: {academic_year}, Term: {term}, Class Level: {class_level}, Stream: {stream}, Status: {status}")
        
        # Build base queryset
        queryset = ExamSession.objects.select_related(
            'exam_type', 'academic_year', 'term', 
            'class_level', 'stream_class'
        ).order_by('-exam_date')
        
        # Apply filters
        if academic_year:
            queryset = queryset.filter(academic_year_id=academic_year)
        if term:
            queryset = queryset.filter(term_id=term)
        if class_level:
            queryset = queryset.filter(class_level_id=class_level)
        if stream:
            queryset = queryset.filter(stream_class_id=stream)
        if status:
            queryset = queryset.filter(status=status)
        
        # Apply search
        if search_value:
            queryset = queryset.filter(
                Q(name__icontains=search_value) |
                Q(exam_type__name__icontains=search_value) |
                Q(academic_year__name__icontains=search_value) |
                Q(class_level__name__icontains=search_value)
            )
        
        # Get total count
        total_records = queryset.count()
        logger.debug(f"Total records found: {total_records}")
        
        # If no records, return empty response immediately
        if total_records == 0:
            return JsonResponse({
                'draw': draw,
                'recordsTotal': 0,
                'recordsFiltered': 0,
                'data': []
            })
        
        # Paginate
        paginator = Paginator(queryset, length)
        page_number = (start // length) + 1
        try:
            page_obj = paginator.get_page(page_number)
        except EmptyPage:
            # If page is out of range, return empty results
            return JsonResponse({
                'draw': draw,
                'recordsTotal': total_records,
                'recordsFiltered': total_records,
                'data': []
            })
        
        # Prepare data
        data = []
        today = date.today()
        
        for session in page_obj:
            # Determine exam date status
            if session.exam_date < today:
                exam_date_status = 'past'
            elif session.exam_date == today:
                exam_date_status = 'today'
            else:
                exam_date_status = 'future'
            
            data.append({
                'id': session.id,
                'name': session.name,
                'exam_type': session.exam_type.id,
                'exam_type_name': session.exam_type.name,
                'academic_year': session.academic_year.id,
                'academic_year_name': session.academic_year.name,
                'term': session.term.id,
                'term_display': session.term.get_term_number_display(),
                'class_level': session.class_level.id,
                'class_level_name': session.class_level.name,
                'stream_class': session.stream_class.id if session.stream_class else None,
                'stream_class_stream_letter': session.stream_class.stream_letter if session.stream_class else None,
                'exam_date': session.exam_date.isoformat(),
                'exam_date_formatted': session.exam_date.strftime('%Y-%m-%d'),
                'exam_date_status': exam_date_status,
                'status': session.status,
                'status_display': session.get_status_display(),
                'created_at': session.created_at.isoformat(),
                'created_at_formatted': session.created_at.strftime('%Y-%m-%d'),
                'created_at_relative': timesince(session.created_at),
            })
        
        logger.debug(f"Returning {len(data)} records")
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': total_records,
            'data': data
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in admin_exam_sessions_data: {str(e)}")
        return JsonResponse({
            'draw': 1,
            'recordsTotal': 0,
            'recordsFiltered': 0,
            'data': [],
            'error': str(e)
        }, status=500)


@login_required
def get_exam_session_by_id(request):
    """Get single exam session data for editing"""
    try:
        exam_session_id = request.GET.get('exam_session_id')
        if not exam_session_id:
            return JsonResponse({
                'success': False,
                'message': 'Exam session ID is required.'
            })
        
        session = ExamSession.objects.select_related(
            'exam_type', 'academic_year', 'term', 
            'class_level', 'stream_class'
        ).get(id=exam_session_id)
        
        data = {
            'id': session.id,
            'name': session.name,
            'exam_type': session.exam_type.id,
            'exam_type_name': session.exam_type.name,
            'academic_year': session.academic_year.id,
            'academic_year_name': session.academic_year.name,
            'term': session.term.id,
            'term_name': session.term.get_term_number_display(),
            'class_level': session.class_level.id,
            'class_level_name': session.class_level.name,
            'stream_class': session.stream_class.id if session.stream_class else '',
            'stream_class_name': str(session.stream_class) if session.stream_class else '',
            'exam_date': session.exam_date.strftime('%Y-%m-%d'),
            'status': session.status,
            'status_display': session.get_status_display(),
        }
        
        return JsonResponse({
            'success': True,
            'exam_session': data
        })
    except ExamSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Exam session not found.'
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in get_exam_session_by_id: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error loading exam session: {str(e)}'
        })




@login_required
def manage_results(request, exam_session_id):
    exam_session = get_object_or_404(
        ExamSession.objects.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'class_level__educational_level',
            'stream_class'
        ),
        id=exam_session_id
    )

    subjects = Subject.objects.filter(
        educational_level=exam_session.class_level.educational_level,
        is_active=True
    ).order_by('code')

    if exam_session.stream_class:
        students = Student.objects.filter(
            class_level=exam_session.class_level,
            stream_class=exam_session.stream_class,
            is_active=True
        )
    else:
        students = Student.objects.filter(
            class_level=exam_session.class_level,
            is_active=True
        )

    students = students.order_by('first_name')

    results = StudentResult.objects.filter(
        exam_session=exam_session
    ).select_related('student', 'subject')

    student_results_data = []

    for student in students:
        student_results = results.filter(student=student)

        subject_results = {}
        total_subjects = 0

        for subject in subjects:
            result = student_results.filter(subject=subject).first()

            if result:
                subject_results[subject.id] = {
                    'marks': result.marks_obtained,
                    'grade': result.grade,
                    'grade_point': result.grade_point,
                    'position': result.position_in_paper,
                    'result_id': result.id
                }
                if result.marks_obtained is not None:
                    total_subjects += 1
            else:
                subject_results[subject.id] = {
                    'marks': None,
                    'grade': '',
                    'grade_point': None,
                    'position': None,
                    'result_id': None
                }

        # 🔹 Pull pre-calculated exam metrics
        try:
            metrics = StudentExamMetrics.objects.get(
                student=student,
                exam_session=exam_session
            )
        except StudentExamMetrics.DoesNotExist:
            metrics = None

        # 🔹 Pull pre-calculated positions
        try:
            position = StudentExamPosition.objects.get(
                student=student,
                exam_session=exam_session
            )
        except StudentExamPosition.DoesNotExist:
            position = None

        student_results_data.append({
            'student': student,
            'student_id': student.id,
            'registration_number': student.registration_number or f"S{student.id:04d}",
            'full_name': student.full_name,
            'gender': student.get_gender_display(),

            # Metrics (already computed elsewhere)
            'total_marks': metrics.total_marks if metrics else None,
            'average_marks': metrics.average_marks if metrics else None,
            'average_percentage': metrics.average_percentage if metrics else None,
            'average_grade': metrics.average_grade if metrics else '',
            'average_remark': metrics.average_remark if metrics else '',
            'total_grade_points': metrics.total_grade_points if metrics else None,
            'division': metrics.division if metrics else '',

            # Positions
            'class_position': position.class_position if position else None,
            'stream_position': position.stream_position if position else None,

            'subject_results': subject_results,
            'has_results': total_subjects > 0
        })

    context = {
        'exam_session': exam_session,
        'subjects': subjects,
        'students': students,
        'student_results_data': student_results_data,
        'page_title': f'Manage Results - {exam_session.name}',
    }

    return render(request, 'admin/results/manage_results.html', context)



@login_required
@require_POST
def save_student_results(request):
    """Save multiple student results with strict marks validation (0–100 or empty)"""

    try:
        data = json.loads(request.body)

        exam_session_id = data.get('exam_session_id')
        results_data = data.get('results', [])
        is_auto_save = data.get('is_auto_save', False)

        if not exam_session_id or not isinstance(results_data, list):
            return JsonResponse({
                'success': False,
                'message': 'Invalid data provided.'
            }, status=400)

        exam_session = get_object_or_404(ExamSession, id=exam_session_id)

        saved_results = []
        saved_count = 0
        skipped_count = 0

        with transaction.atomic():
            for result_data in results_data:
                student_id = result_data.get('student_id')
                subject_id = result_data.get('subject_id')
                marks_obtained = result_data.get('marks_obtained')

                # Required IDs
                if not student_id or not subject_id:
                    skipped_count += 1
                    continue

                # Allow empty marks (skip save)
                if marks_obtained in ("", None):
                    skipped_count += 1
                    continue

                # Ensure numeric (use Decimal to match model DecimalFields)
                try:
                    marks_obtained = Decimal(str(marks_obtained))
                except (TypeError, ValueError, InvalidOperation):
                    skipped_count += 1
                    continue

                # Enforce range 0–100
                if marks_obtained < 0 or marks_obtained > 100:
                    skipped_count += 1
                    continue

                # Save or update result
                result, created = StudentResult.objects.get_or_create(
                    exam_session=exam_session,
                    student_id=student_id,
                    subject_id=subject_id,
                    defaults={'marks_obtained': marks_obtained}
                )

                if not created:
                    result.marks_obtained = marks_obtained

                # Percentage (use Decimal arithmetic)
                max_score = exam_session.exam_type.max_score
                if max_score and max_score > 0:
                    try:
                        max_score_decimal = Decimal(str(max_score))
                        result.percentage = (marks_obtained / max_score_decimal) * Decimal('100')
                    except (InvalidOperation, TypeError):
                        result.percentage = None

                # Grade calculation
                education_level = exam_session.class_level.educational_level
                grade_scale = GradingScale.objects.filter(
                    education_level=education_level,
                    min_mark__lte=marks_obtained,
                    max_mark__gte=marks_obtained
                ).first()

                if grade_scale:
                    result.grade = grade_scale.grade
                    result.grade_point = Decimal(str(grade_scale.points))
                else:
                    result.grade = None
                    result.grade_point = None

                result.save()

                saved_results.append({
                    'student_id': student_id,
                    'subject_id': subject_id,
                    'result_id': result.id
                })

                saved_count += 1

        # Calculate positions only for manual save
        if saved_count > 0 and not is_auto_save:
            calculate_subject_positions(exam_session)

        return JsonResponse({
            'success': True,
            'message': f'Saved {saved_count} results. Skipped {skipped_count}.',
            'saved_count': saved_count,
            'skipped_count': skipped_count,
            'saved_results': saved_results
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON payload.'
        }, status=400)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error saving results: {str(e)}'
        }, status=500)



@login_required
@require_POST
def save_multiple_results(request):
    """Save multiple new results from modal with marks validation (0–100 or empty)"""

    try:
        exam_session_id = request.POST.get('exam_session_id')
        results_json = request.POST.get('results')

        if not exam_session_id or not results_json:
            return JsonResponse({
                'success': False,
                'message': 'Invalid data provided.'
            }, status=400)

        try:
            results_data = json.loads(results_json)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid results JSON.'
            }, status=400)

        exam_session = get_object_or_404(ExamSession, id=exam_session_id)

        saved_count = 0
        skipped_count = 0
        errors = []

        with transaction.atomic():
            for result_data in results_data:
                student_id = result_data.get('studentId')
                subject_id = result_data.get('subjectId')
                marks = result_data.get('marks')

                # Validate IDs
                if not student_id or not subject_id:
                    skipped_count += 1
                    continue

                # Allow empty marks → skip
                if marks in ("", None):
                    skipped_count += 1
                    continue

                # Ensure numeric marks (use Decimal)
                try:
                    marks = Decimal(str(marks))
                except (TypeError, ValueError, InvalidOperation):
                    errors.append(
                        f"Invalid marks for student {student_id}, subject {subject_id}"
                    )
                    continue

                # Enforce range 0–100
                if marks < 0 or marks > 100:
                    errors.append(
                        f"Marks out of range (0–100) for student {student_id}, subject {subject_id}"
                    )
                    continue

                # Prevent duplicate result
                if StudentResult.objects.filter(
                    exam_session=exam_session,
                    student_id=student_id,
                    subject_id=subject_id
                ).exists():
                    errors.append(
                        f"Result already exists for student {student_id}, subject {subject_id}"
                    )
                    continue

                # Create result
                result = StudentResult.objects.create(
                    exam_session=exam_session,
                    student_id=student_id,
                    subject_id=subject_id,
                    marks_obtained=marks
                )

                # Percentage (use Decimal arithmetic)
                max_score = exam_session.exam_type.max_score
                if max_score and max_score > 0:
                    try:
                        max_score_decimal = Decimal(str(max_score))
                        result.percentage = (marks / max_score_decimal) * Decimal('100')
                    except (InvalidOperation, TypeError):
                        result.percentage = None

                # Grade
                education_level = exam_session.class_level.educational_level
                grade_scale = GradingScale.objects.filter(
                    education_level=education_level,
                    min_mark__lte=marks,
                    max_mark__gte=marks
                ).first()

                if grade_scale:
                    result.grade = grade_scale.grade
                    result.grade_point = grade_scale.points

                result.save()
                saved_count += 1

        # Calculate positions only if something was saved
        if saved_count > 0:
            calculate_subject_positions(exam_session)

        response_data = {
            'success': True,
            'message': f'Saved {saved_count} results. Skipped {skipped_count}.',
            'saved_count': saved_count,
            'skipped_count': skipped_count
        }

        if errors:
            response_data['errors'] = errors[:5]  # limit noise

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error saving results: {str(e)}'
        }, status=500)

def calculate_subject_positions(exam_session):
    """Calculate positions for all subjects in an exam session"""
    try:
        # Get all subjects in this exam session
        subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        )
        
        for subject in subjects:
            calculate_subject_position(exam_session, subject)
            
    except Exception as e:
        print(f"Error calculating positions: {str(e)}")

def calculate_subject_position(exam_session, subject):
    """Calculate positions for a specific subject"""
    try:
        # Get all results for this subject ordered by marks (descending)
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject,
            marks_obtained__isnull=False
        ).select_related('student').order_by('-marks_obtained')
        
        position = 1
        previous_marks = None
        skip_position = 0
        
        for index, result in enumerate(results):
            current_marks = result.marks_obtained
            
            # Handle equal marks (same position)
            if previous_marks is not None and current_marks == previous_marks:
                # Same position as previous student
                result.position_in_paper = position - 1
                skip_position += 1
            else:
                # New position
                result.position_in_paper = position
                position += 1 + skip_position
                skip_position = 0
            
            result.save(update_fields=['position_in_paper'])
            previous_marks = current_marks
            
    except Exception as e:
        print(f"Error calculating position for subject {subject.id}: {str(e)}")

@login_required
def subject_results_entry(request, exam_session_id, subject_id):
    """Entry page for entering marks for all students in a specific subject"""
    exam_session = get_object_or_404(
        ExamSession.objects.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'class_level__educational_level',
            'stream_class'
        ),
        id=exam_session_id
    )
    
    subject = get_object_or_404(Subject, id=subject_id)
    
    # Get students based on stream
    if exam_session.stream_class:
        students = Student.objects.filter(
            class_level=exam_session.class_level,
            stream_class=exam_session.stream_class,
            is_active=True
        )
    else:
        students = Student.objects.filter(
            class_level=exam_session.class_level,
            is_active=True
        )
    
    students = students.order_by('first_name', 'last_name')
    
    # Get existing results for this subject
    existing_results = StudentResult.objects.filter(
        exam_session=exam_session,
        subject=subject
    ).select_related('student')
    
    # Create a dictionary of existing results for quick lookup
    results_dict = {result.student_id: result for result in existing_results}
    
    # Prepare student data with existing marks
    students_data = []
    for student in students:
        result = results_dict.get(student.id)
        students_data.append({
            'id': student.id,
            'registration_number': student.registration_number or f"S{student.id:04d}",
            'full_name': student.full_name,
            'gender': student.get_gender_display(),
            'existing_marks': result.marks_obtained if result else None,
            'existing_grade': result.grade if result else '',
            'result_id': result.id if result else None,
            'has_result': bool(result)
        })
    
    # Get statistics
    total_students = students.count()
    with_marks = len([s for s in students_data if s['existing_marks'] is not None])
    without_marks = total_students - with_marks
    
    context = {
        'exam_session': exam_session,
        'subject': subject,
        'students_data': students_data,
        'total_students': total_students,
        'with_marks': with_marks,
        'without_marks': without_marks,
        'progress_percentage': (with_marks / total_students * 100) if total_students > 0 else 0,
        'page_title': f'Enter Marks - {subject.name}',
    }
    
    return render(request, 'admin/results/subject_results_entry.html', context)        




@login_required
def download_excel_template(request, exam_session_id, subject_id):
    """Generate and download Excel template for marks entry"""
    try:
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type', 'class_level', 'stream_class',
                'academic_year', 'term'
            ),
            id=exam_session_id
        )
        subject = get_object_or_404(Subject, id=subject_id)
        
        # Get students for this exam session
        if exam_session.stream_class:
            students = Student.objects.filter(
                class_level=exam_session.class_level,
                stream_class=exam_session.stream_class,
                is_active=True
            )
        else:
            students = Student.objects.filter(
                class_level=exam_session.class_level,
                is_active=True
            )
        
        students = students.order_by('first_name', 'last_name')
        
        # Get existing results
        existing_results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject
        )
        results_dict = {r.student_id: r.marks_obtained for r in existing_results}
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marks Entry"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        info_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        info_font = Font(name="Arial", size=10, bold=True)
        
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Add header information
        ws['A1'] = "MARKS ENTRY TEMPLATE"
        ws['A1'].font = Font(name="Arial", size=14, bold=True)
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Exam Session: {exam_session.name}"
        ws['A2'].font = info_font
        ws.merge_cells('A2:D2')
        
        ws['A3'] = f"Subject: {subject.name} ({subject.code})"
        ws['A3'].font = info_font
        ws.merge_cells('A3:D3')
        
        ws['A4'] = f"Class: {exam_session.class_level.name}"
        if exam_session.stream_class:
            ws['A4'] += f" {exam_session.stream_class.stream_letter}"
        ws['A4'].font = info_font
        ws.merge_cells('A4:D4')
        
        ws['A5'] = f"Academic Year: {exam_session.academic_year.name}"
        ws['A5'].font = info_font
        ws.merge_cells('A5:D5')
        
        ws['A6'] = f"Term: {exam_session.term.get_term_number_display()}"
        ws['A6'].font = info_font
        ws.merge_cells('A6:D6')
        
        ws['A7'] = f"Maximum Score: {exam_session.exam_type.max_score}"
        ws['A7'].font = info_font
        ws.merge_cells('A7:D7')
        
        # Add instruction rows
        ws['A9'] = "INSTRUCTIONS:"
        ws['A9'].font = Font(name="Arial", size=11, bold=True, color="FF0000")
        ws.merge_cells('A9:D9')
        
        instructions = [
            "1. Fill marks in the 'Marks' column (Column D).",
            "2. Leave blank for absent students.",
            "3. Enter 0 for zero marks (it's a valid mark).",
            "4. Do not modify Student ID, Registration No., or Student Name columns.",
            f"5. Marks range: 0 - {exam_session.exam_type.max_score}",
            "6. Save file and upload using the upload feature."
        ]
        
        for i, instruction in enumerate(instructions, start=10):
            cell = ws[f'A{i}']
            cell.value = instruction
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{i}:D{i}')
        
        # Add headers for data
        header_row = 16
        headers = ["Student ID", "Registration No.", "Student Name", "Marks"]
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = center_alignment
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Add student data
        data_start_row = header_row + 1
        
        for i, student in enumerate(students, start=data_start_row):
            # Student ID
            ws.cell(row=i, column=1, value=student.id)
            ws.cell(row=i, column=1).border = cell_border
            ws.cell(row=i, column=1).alignment = center_alignment
            
            # Registration Number
            ws.cell(row=i, column=2, value=student.registration_number or f"S{student.id:04d}")
            ws.cell(row=i, column=2).border = cell_border
            ws.cell(row=i, column=2).alignment = left_alignment
            
            # Student Name
            ws.cell(row=i, column=3, value=student.full_name)
            ws.cell(row=i, column=3).border = cell_border
            ws.cell(row=i, column=3).alignment = left_alignment
            
            # Marks (pre-filled if exists)
            existing_marks = results_dict.get(student.id)
            ws.cell(row=i, column=4, value=existing_marks)
            ws.cell(row=i, column=4).border = cell_border
            ws.cell(row=i, column=4).alignment = center_alignment
            
            # Add data validation for marks
            if exam_session.exam_type.max_score:
                dv = openpyxl.worksheet.datavalidation.DataValidation(
                    type="decimal",
                    operator="between",
                    formula1=0,
                    formula2=exam_session.exam_type.max_score,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Invalid Marks",
                    error="Marks must be between 0 and " + str(exam_session.exam_type.max_score)
                )
                ws.add_data_validation(dv)
                dv.add(f'D{i}:D{i}')
        
        # Add summary row
        summary_row = data_start_row + len(students) + 2
        ws.cell(row=summary_row, column=1, value="Summary:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value="Total Students:")
        ws.cell(row=summary_row + 1, column=2, value=len(students))
        ws.cell(row=summary_row + 1, column=2).font = Font(bold=True)
        
        ws.cell(row=summary_row + 2, column=1, value="With Existing Marks:")
        ws.cell(row=summary_row + 2, column=2, 
                value=f"=COUNTIF(D{data_start_row}:D{data_start_row + len(students) - 1},\"<>\"&\"\")")
        ws.cell(row=summary_row + 2, column=2).font = Font(bold=True)
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Marks_Template_{subject.code}_{exam_session.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Error generating template: {str(e)}")
        return redirect('manage_results', exam_session_id=exam_session_id)

@login_required
@require_POST
def upload_excel_marks(request):
    """Handle Excel file upload for bulk marks entry"""
    try:
        excel_file = request.FILES.get('excel_file')
        exam_session_id = request.POST.get('exam_session_id')
        subject_id = request.POST.get('subject_id')
        
        if not excel_file:
            return JsonResponse({
                'success': False,
                'message': 'No file uploaded'
            })
        
        if not exam_session_id or not subject_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing exam session or subject ID'
            })
        
        exam_session = get_object_or_404(ExamSession, id=exam_session_id)
        subject = get_object_or_404(Subject, id=subject_id)
        
        # Read Excel file
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Parse data
        data = []
        errors = []
        processed_count = 0
        
        # Find header row
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
            if row and row[0] and isinstance(row[0], str) and "Student ID" in str(row[0]):
                header_row = row
                break
        
        if not header_row:
            return JsonResponse({
                'success': False,
                'message': 'Could not find header row in Excel file'
            })
        
        # Find start of data
        data_start_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), start=1):
            if row and isinstance(row[0], (int, float)):
                data_start_row = i
                break
        
        if not data_start_row:
            return JsonResponse({
                'success': False,
                'message': 'No data found in Excel file'
            })
        
        # Process each row
        for i, row in enumerate(ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=True), start=data_start_row):
            if not row or not row[0]:
                continue
            
            try:
                student_id = int(row[0])
                marks = row[3] if len(row) > 3 else None
                
                # Validate student exists
                try:
                    student = Student.objects.get(id=student_id, is_active=True)
                except Student.DoesNotExist:
                    errors.append(f"Row {i}: Student ID {student_id} not found")
                    continue
                
                # Validate marks
                if marks is not None and marks != '':
                    try:
                        marks = Decimal(str(marks))
                        max_score = exam_session.exam_type.max_score
                        
                        if marks < 0 or marks > max_score:
                            errors.append(f"Row {i}: Marks {marks} out of range (0-{max_score})")
                            continue
                    except (ValueError, InvalidOperation):
                        errors.append(f"Row {i}: Invalid marks format '{marks}'")
                        continue
                else:
                    marks = None
                
                data.append({
                    'student_id': student_id,
                    'marks': marks
                })
                
            except Exception as e:
                errors.append(f"Row {i}: Error processing row - {str(e)}")
        
        # Save marks in transaction
        with transaction.atomic():
            for item in data:
                try:
                    # Get or create result
                    result, created = StudentResult.objects.get_or_create(
                        exam_session=exam_session,
                        student_id=item['student_id'],
                        subject=subject,
                        defaults={'marks_obtained': item['marks']}
                    )
                    
                    if not created:
                        result.marks_obtained = item['marks']
                    
                    # Calculate percentage
                    if item['marks'] is not None and exam_session.exam_type.max_score > 0:
                        result.percentage = (Decimal(str(item['marks'])) / 
                                           Decimal(str(exam_session.exam_type.max_score))) * Decimal('100')
                    
                    # Calculate grade
                    if item['marks'] is not None:
                        education_level = exam_session.class_level.educational_level
                        grade_scale = GradingScale.objects.filter(
                            education_level=education_level,
                            min_mark__lte=item['marks'],
                            max_mark__gte=item['marks']
                        ).first()
                        
                        if grade_scale:
                            result.grade = grade_scale.grade
                            result.grade_point = grade_scale.points
                    
                    result.save()
                    processed_count += 1
                    
                except Exception as e:
                    errors.append(f"Student ID {item['student_id']}: Error saving - {str(e)}")
        
        # Calculate positions after upload
        calculate_subject_positions(exam_session)
        
        response_data = {
            'success': True,
            'message': f'Successfully processed {processed_count} marks',
            'processed_count': processed_count,
            'total_rows': len(data)
        }
        
        if errors:
            response_data['errors'] = errors[:10]  # Limit errors to first 10
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        })




@login_required
def download_pdf_report(request, exam_session_id, subject_id):
    """Generate and download PDF report with gender + subject performance analytics"""
    try:
        # ==================================================
        # 1. LOAD CORE OBJECTS
        # ==================================================
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type',
                'class_level',
                'stream_class',
                'academic_year',
                'term',
                'class_level__educational_level'
            ),
            id=exam_session_id
        )

        subject = get_object_or_404(Subject, id=subject_id)

        # ==================================================
        # 2. LOAD STUDENTS
        # ==================================================
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }

        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class

        students = Student.objects.filter(**student_filters).order_by(
            'first_name', 'last_name'
        )

        # ==================================================
        # 3. LOAD RESULTS (THIS SUBJECT)
        # ==================================================
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject
        ).select_related('student')

        results_by_student = {r.student_id: r for r in results}

        # ==================================================
        # 4. PREPARE STUDENT DATA + GENDER STATS
        # ==================================================
        students_data = []
        total_marks = 0
        students_with_marks = 0

        gender_totals = {}
        gender_marks_sum = {}
        gender_with_marks = {}

        for student in students:
            result = results_by_student.get(student.id)
            marks = result.marks_obtained if result else None
            grade = ''

            gender = student.get_gender_display() or 'Unknown'

            gender_totals.setdefault(gender, 0)
            gender_marks_sum.setdefault(gender, 0)
            gender_with_marks.setdefault(gender, 0)

            gender_totals[gender] += 1

            if marks is not None:
                total_marks += marks
                students_with_marks += 1

                gender_marks_sum[gender] += marks
                gender_with_marks[gender] += 1

                grade_scale = GradingScale.objects.filter(
                    education_level=exam_session.class_level.educational_level,
                    min_mark__lte=marks,
                    max_mark__gte=marks
                ).first()

                if grade_scale:
                    grade = grade_scale.grade

            students_data.append({
                'reg_number': student.registration_number or f"S{student.id:04d}",
                'name': student.full_name,
                'gender': gender,
                'marks': marks,
                'grade': grade,
                'has_marks': marks is not None
            })

        # ==================================================
        # 5. OVERALL STATISTICS
        # ==================================================
        statistics = {
            'total_students': students.count(),
            'students_with_marks': students_with_marks,
            'students_without_marks': students.count() - students_with_marks,
            'percentage_completed': (
                (students_with_marks / students.count()) * 100
                if students.exists() else 0
            ),
            'average_marks': (
                total_marks / students_with_marks
                if students_with_marks > 0 else 0
            ),
            'highest_marks': max(
                [s['marks'] for s in students_data if s['marks'] is not None],
                default=0
            ),
            'lowest_marks': min(
                [s['marks'] for s in students_data if s['marks'] is not None],
                default=0
            ),
        }

        # ==================================================
        # 6. GRADE DISTRIBUTION
        # ==================================================
        grade_distribution = {}
        for s in students_data:
            if s['grade']:
                grade_distribution[s['grade']] = (
                    grade_distribution.get(s['grade'], 0) + 1
                )

        # ==================================================
        # 7. GENDER STATISTICS
        # ==================================================
        gender_statistics = []

        for gender, total in gender_totals.items():
            with_marks = gender_with_marks.get(gender, 0)
            avg = (
                gender_marks_sum[gender] / with_marks
                if with_marks > 0 else 0
            )

            gender_statistics.append({
                'gender': gender,
                'total': total,
                'with_marks': with_marks,
                'average': round(avg, 2)
            })

        # ==================================================
        # 8. SUBJECT PERFORMANCE (THIS SUBJECT)
        # ==================================================
        subject_performance = results.aggregate(
            average=Avg('marks_obtained'),
            highest=Max('marks_obtained'),
            lowest=Min('marks_obtained')
        )

        # ==================================================
        # 9. SUBJECT POSITION IN EXAM SESSION
        # ==================================================
        subject_averages = (
            StudentResult.objects
            .filter(exam_session=exam_session)
            .values('subject')
            .annotate(avg_marks=Avg('marks_obtained'))
            .order_by('-avg_marks')
        )

        subject_position = None
        total_subjects = len(subject_averages)

        for index, row in enumerate(subject_averages, start=1):
            if row['subject'] == subject.id:
                subject_position = index
                break

        subject_ranking = {
            'position': subject_position,
            'total_subjects': total_subjects
        }

        # ==================================================
        # 10. CONTEXT
        # ==================================================
        context = {
            'exam_session': exam_session,
            'subject': subject,
            'students_data': students_data,
            'statistics': statistics,
            'grade_distribution': grade_distribution,
            'gender_statistics': gender_statistics,
            'subject_performance': subject_performance,
            'subject_ranking': subject_ranking,
            'today': timezone.now().date(),
            'generated_by': request.user.get_full_name() or request.user.username,
        }

        # ==================================================
        # 11. RENDER PDF
        # ==================================================
        html_string = render_to_string(
            'admin/results/pdf_subject_report.html',
            context
        )

        pdf = HTML(
            string=html_string,
            base_url=request.build_absolute_uri()
        ).write_pdf()

        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename='
            f'"Marks_Report_{subject.code}_{exam_session.name}.pdf"'
        )

        return response

    except Exception as e:
        messages.error(request, f"Error generating PDF: {e}")
        return redirect('manage_results', exam_session_id=exam_session_id)


# Add these new views to the existing file

@login_required
def download_session_excel_template(request, exam_session_id):
    """Generate and download Excel template for entire exam session"""
    try:
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type', 'class_level', 'stream_class',
                'academic_year', 'term'
            ),
            id=exam_session_id
        )
        
        subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('code')
        
        # Get students for this exam session
        if exam_session.stream_class:
            students = Student.objects.filter(
                class_level=exam_session.class_level,
                stream_class=exam_session.stream_class,
                is_active=True
            )
        else:
            students = Student.objects.filter(
                class_level=exam_session.class_level,
                is_active=True
            )
        
        students = students.order_by('first_name', 'last_name')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Session Marks Entry"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        info_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        info_font = Font(name="Arial", size=10, bold=True)
        
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Add header information
        ws['A1'] = "EXAM SESSION MARKS ENTRY TEMPLATE"
        ws['A1'].font = Font(name="Arial", size=14, bold=True)
        ws['A1'].alignment = center_alignment
        ws.merge_cells(f'A1:{get_column_letter(3 + len(subjects))}1')
        
        ws['A2'] = f"Exam Session: {exam_session.name}"
        ws['A2'].font = info_font
        ws.merge_cells(f'A2:{get_column_letter(3 + len(subjects))}2')
        
        ws['A3'] = f"Class: {exam_session.class_level.name}"
        if exam_session.stream_class:
            ws['A3'] += f" {exam_session.stream_class.stream_letter}"
        ws['A3'].font = info_font
        ws.merge_cells(f'A3:{get_column_letter(3 + len(subjects))}3')
        
        ws['A4'] = f"Academic Year: {exam_session.academic_year.name} - Term: {exam_session.term.get_term_number_display()}"
        ws['A4'].font = info_font
        ws.merge_cells(f'A4:{get_column_letter(3 + len(subjects))}4')
        
        ws['A5'] = f"Maximum Score per Subject: {exam_session.exam_type.max_score}"
        ws['A5'].font = info_font
        ws.merge_cells(f'A5:{get_column_letter(3 + len(subjects))}5')
        
        # Add instruction rows
        ws['A7'] = "INSTRUCTIONS:"
        ws['A7'].font = Font(name="Arial", size=11, bold=True, color="FF0000")
        ws.merge_cells(f'A7:{get_column_letter(3 + len(subjects))}7')
        
        instructions = [
            "1. Fill marks in the subject columns (Column D onwards).",
            "2. Leave cells blank for absent students or unmarked subjects.",
            "3. Enter 0 for zero marks (it's a valid mark).",
            "4. Do not modify Student ID, Registration No., or Student Name columns.",
            f"5. Marks range: 0 - {exam_session.exam_type.max_score} for each subject",
            "6. Save file and upload using the upload feature.",
            "7. All subjects are included in this template."
        ]
        
        for i, instruction in enumerate(instructions, start=8):
            cell = ws[f'A{i}']
            cell.value = instruction
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{i}:{get_column_letter(3 + len(subjects))}{i}')
        
        # Add headers for data
        header_row = 15
        headers = ["Student ID", "Registration No.", "Student Name"]
        
        # Add subject headers
        for subject in subjects:
            headers.append(f"{subject.name}")
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = center_alignment
            # Set column widths
            if col_num <= 3:
                ws.column_dimensions[get_column_letter(col_num)].width = 20
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Add student data
        data_start_row = header_row + 1
        
        for i, student in enumerate(students, start=data_start_row):
            # Student ID
            ws.cell(row=i, column=1, value=student.id)
            ws.cell(row=i, column=1).border = cell_border
            ws.cell(row=i, column=1).alignment = center_alignment
            
            # Registration Number
            ws.cell(row=i, column=2, value=student.registration_number or f"S{student.id:04d}")
            ws.cell(row=i, column=2).border = cell_border
            ws.cell(row=i, column=2).alignment = left_alignment
            
            # Student Name
            ws.cell(row=i, column=3, value=student.full_name)
            ws.cell(row=i, column=3).border = cell_border
            ws.cell(row=i, column=3).alignment = left_alignment
            
            # Get existing results for this student
            existing_results = StudentResult.objects.filter(
                exam_session=exam_session,
                student=student
            ).select_related('subject')
            
            results_dict = {r.subject_id: r.marks_obtained for r in existing_results}
            
            # Add marks columns for each subject
            for col_num, subject in enumerate(subjects, start=4):
                existing_marks = results_dict.get(subject.id)
                cell = ws.cell(row=i, column=col_num, value=existing_marks)
                cell.border = cell_border
                cell.alignment = center_alignment
                
                # Add data validation for marks
                if exam_session.exam_type.max_score:
                    dv = openpyxl.worksheet.datavalidation.DataValidation(
                        type="decimal",
                        operator="between",
                        formula1=0,
                        formula2=exam_session.exam_type.max_score,
                        allow_blank=True,
                        showErrorMessage=True,
                        errorTitle="Invalid Marks",
                        error=f"Marks must be between 0 and {exam_session.exam_type.max_score}"
                    )
                    ws.add_data_validation(dv)
                    dv.add(f'{get_column_letter(col_num)}{i}:{get_column_letter(col_num)}{i}')
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Session_Template_{exam_session.name.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating template: {str(e)}", status=500)

@login_required
@require_POST
def upload_session_excel(request):
    excel_file = request.FILES.get('excel_file')
    exam_session_id = request.POST.get('exam_session_id')

    if not excel_file or not exam_session_id:
        return JsonResponse({'success': False, 'message': 'Missing file or exam session ID'})

    exam_session = get_object_or_404(ExamSession, id=exam_session_id)

    if exam_session.status == 'published':
        return JsonResponse({'success': False, 'message': 'Cannot modify a published exam session'})

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active

    subjects = Subject.objects.filter(
        educational_level=exam_session.class_level.educational_level,
        is_active=True
    )

    subject_map = {
        normalize(subject.name): subject
        for subject in subjects
    }

    # ----------------------------
    # FIND HEADER ROW
    # ----------------------------
    header_row_idx = None
    header_values = []

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if row and "Student ID" in [str(c) for c in row if c]:
            header_row_idx = i
            header_values = [normalize(str(c)) for c in row]
            break

    if not header_row_idx:
        return JsonResponse({'success': False, 'message': 'Header row not found'})

    # ----------------------------
    # MAP SUBJECT COLUMNS
    # ----------------------------
    subject_columns = {}

    for idx, header in enumerate(header_values):
        if header in subject_map:
            subject_columns[idx] = subject_map[header]

    if not subject_columns:
        return JsonResponse({'success': False, 'message': 'No valid subject columns found'})

    processed = 0
    errors = []

    with transaction.atomic():
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1
        ):
            if not row or not row[0]:
                continue

            try:
                student = Student.objects.get(id=int(row[0]), is_active=True)
            except Student.DoesNotExist:
                errors.append(f"Row {row_idx}: Invalid student ID")
                continue

            for col_idx, subject in subject_columns.items():
                if col_idx >= len(row):
                    continue

                marks = row[col_idx]

                if marks is None or marks == "":
                    continue

                try:
                    marks = Decimal(str(marks))
                except:
                    errors.append(f"Row {row_idx}, {subject.name}: Invalid mark")
                    continue

                max_score = exam_session.exam_type.max_score
                if marks < 0 or marks > max_score:
                    errors.append(f"Row {row_idx}, {subject.name}: Out of range")
                    continue

                result, _ = StudentResult.objects.get_or_create(
                    exam_session=exam_session,
                    student=student,
                    subject=subject
                )

                result.marks_obtained = marks
                result.percentage = (marks / Decimal(max_score)) * 100

                grade_scale = GradingScale.objects.filter(
                    education_level=exam_session.class_level.educational_level,
                    min_mark__lte=marks,
                    max_mark__gte=marks
                ).first()

                if grade_scale:
                    result.grade = grade_scale.grade
                    result.grade_point = grade_scale.points

                result.save()
                processed += 1

    return JsonResponse({
        'success': True,
        'message': f'Processed {processed} marks successfully',
        'errors': errors[:10]
    })

def normalize(text):
    return text.strip().lower() if text else ""


@login_required
def download_session_excel_report(request, exam_session_id):
    """Generate comprehensive Excel report for entire exam session with subject-wise performance"""
    try:
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type',
                'class_level',
                'stream_class',
                'academic_year',
                'term',
                'class_level__educational_level'
            ),
            id=exam_session_id
        )

        subjects = list(
            Subject.objects.filter(
                educational_level=exam_session.class_level.educational_level,
                is_active=True
            ).order_by('code')
        )

        # Get students
        student_qs = Student.objects.filter(
            class_level=exam_session.class_level,
            is_active=True
        )
        if exam_session.stream_class:
            student_qs = student_qs.filter(stream_class=exam_session.stream_class)
        
        students = list(student_qs.order_by('first_name', 'last_name'))

        # Get all results
        results = StudentResult.objects.filter(
            exam_session=exam_session
        ).select_related('student', 'subject')

        # Get metrics and positions
        metrics = StudentExamMetrics.objects.filter(
            exam_session=exam_session
        ).select_related('student', 'division')

        positions = StudentExamPosition.objects.filter(
            exam_session=exam_session
        ).select_related('student')

        # Create lookup dictionaries
        results_map = {}
        for r in results:
            results_map.setdefault(r.student_id, {})[r.subject_id] = r

        metrics_map = {m.student_id: m for m in metrics}
        position_map = {p.student_id: p for p in positions}

        # ============================================
        # CREATE EXCEL WORKBOOK
        # ============================================
        wb = openpyxl.Workbook()
        
        # ============================================
        # SHEET 1: COMPREHENSIVE RESULTS
        # ============================================
        ws_main = wb.active
        ws_main.title = "Comprehensive Results"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Arial", size=14, bold=True)
        info_font = Font(name="Arial", size=10, bold=True)
        normal_font = Font(name="Arial", size=10)
        bold_font = Font(name="Arial", size=10, bold=True)
        
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        cell_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Add header information
        ws_main['A1'] = "EXAM SESSION COMPREHENSIVE REPORT"
        ws_main['A1'].font = title_font
        ws_main['A1'].alignment = center_alignment
        ws_main.merge_cells(f'A1:{get_column_letter(7 + len(subjects))}1')
        
        ws_main['A2'] = f"Exam Session: {exam_session.name}"
        ws_main['A2'].font = info_font
        ws_main.merge_cells(f'A2:{get_column_letter(7 + len(subjects))}2')
        
        ws_main['A3'] = f"Class: {exam_session.class_level.name}"
        if exam_session.stream_class:
            ws_main['A3'] += f" {exam_session.stream_class.stream_letter}"
        ws_main['A3'].font = info_font
        ws_main.merge_cells(f'A3:{get_column_letter(7 + len(subjects))}3')
        
        ws_main['A4'] = f"Academic Year: {exam_session.academic_year.name} | Term: {exam_session.term.get_term_number_display()}"
        ws_main['A4'].font = info_font
        ws_main.merge_cells(f'A4:{get_column_letter(7 + len(subjects))}4')
        
        ws_main['A5'] = f"Exam Date: {exam_session.exam_date.strftime('%B %d, %Y')}"
        ws_main['A5'].font = info_font
        ws_main.merge_cells(f'A5:{get_column_letter(7 + len(subjects))}5')
        
        # Column headers
        header_row = 7
        headers = ["S.No", "Registration No.", "Student Name", "Gender"]
        
        # Add subject headers
        for subject in subjects:
            headers.append(f"{subject.name}")
        
        headers.extend(["Total Marks", "Average", "Grade Points", "Division", "Position"])
        
        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws_main.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = center_alignment
            
            # Set column widths
            if col_num == 1:  # S.No
                ws_main.column_dimensions[get_column_letter(col_num)].width = 8
            elif col_num == 2:  # Reg No
                ws_main.column_dimensions[get_column_letter(col_num)].width = 15
            elif col_num == 3:  # Name
                ws_main.column_dimensions[get_column_letter(col_num)].width = 25
            elif col_num == 4:  # Gender
                ws_main.column_dimensions[get_column_letter(col_num)].width = 10
            elif col_num <= 4 + len(subjects):  # Subject columns
                ws_main.column_dimensions[get_column_letter(col_num)].width = 12
            else:  # Summary columns
                ws_main.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Add student data
        data_start_row = header_row + 1
        
        for i, student in enumerate(students, start=data_start_row):
            # Basic info
            ws_main.cell(row=i, column=1, value=i - header_row).border = cell_border
            ws_main.cell(row=i, column=2, value=student.registration_number or f"S{student.id:04d}").border = cell_border
            ws_main.cell(row=i, column=3, value=student.full_name).border = cell_border
            ws_main.cell(row=i, column=4, value=student.get_gender_display()).border = cell_border
            
            # Subject marks
            total_marks = 0
            subjects_with_marks = 0
            
            for col_num, subject in enumerate(subjects, start=5):
                result = results_map.get(student.id, {}).get(subject.id)
                if result and result.marks_obtained is not None:
                    marks = result.marks_obtained
                    grade = result.grade or ""
                    cell_value = f"{marks:.1f}"
                    if grade:
                        cell_value += f" ({grade})"
                    
                    ws_main.cell(row=i, column=col_num, value=cell_value)
                    
                    total_marks += float(marks)
                    subjects_with_marks += 1
                else:
                    ws_main.cell(row=i, column=col_num, value="-")
                
                ws_main.cell(row=i, column=col_num).border = cell_border
                ws_main.cell(row=i, column=col_num).alignment = center_alignment
            
            # Summary columns
            summary_col = 5 + len(subjects)
            
            # Total Marks
            ws_main.cell(row=i, column=summary_col, value=total_marks if subjects_with_marks > 0 else "-")
            ws_main.cell(row=i, column=summary_col).border = cell_border
            ws_main.cell(row=i, column=summary_col).alignment = center_alignment
            
            # Average
            average = total_marks / subjects_with_marks if subjects_with_marks > 0 else None
            ws_main.cell(row=i, column=summary_col + 1, value=f"{average:.2f}" if average else "-")
            ws_main.cell(row=i, column=summary_col + 1).border = cell_border
            ws_main.cell(row=i, column=summary_col + 1).alignment = center_alignment
            
            # Grade Points (from metrics)
            metrics = metrics_map.get(student.id)
            ws_main.cell(row=i, column=summary_col + 2, 
                        value=f"{metrics.total_grade_points:.1f}" if metrics and metrics.total_grade_points else "-")
            ws_main.cell(row=i, column=summary_col + 2).border = cell_border
            ws_main.cell(row=i, column=summary_col + 2).alignment = center_alignment
            
            # Division
            division = metrics.division.division if metrics and metrics.division else "-"
            ws_main.cell(row=i, column=summary_col + 3, value=division)
            ws_main.cell(row=i, column=summary_col + 3).border = cell_border
            ws_main.cell(row=i, column=summary_col + 3).alignment = center_alignment
            
            # Position
            position = position_map.get(student.id)
            ws_main.cell(row=i, column=summary_col + 4, 
                        value=position.class_position if position and position.class_position else "-")
            ws_main.cell(row=i, column=summary_col + 4).border = cell_border
            ws_main.cell(row=i, column=summary_col + 4).alignment = center_alignment
        
        # Add summary statistics at the bottom
        summary_start = data_start_row + len(students) + 2
        
        ws_main.cell(row=summary_start, column=1, value="SUMMARY STATISTICS").font = title_font
        ws_main.merge_cells(f'A{summary_start}:C{summary_start}')
        
        summary_rows = [
            ("Total Students", len(students)),
            ("Total Subjects", len(subjects)),
            ("Exam Type", exam_session.exam_type.name),
            ("Max Score per Subject", f"{exam_session.exam_type.max_score:.1f}"),
            ("Report Generated", timezone.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Generated By", request.user.get_full_name() or request.user.username),
        ]
        
        for idx, (label, value) in enumerate(summary_rows, start=1):
            ws_main.cell(row=summary_start + idx, column=1, value=label).font = bold_font
            ws_main.cell(row=summary_start + idx, column=2, value=value).font = normal_font
        
        # ============================================
        # SHEET 2: SUBJECT-WISE PERFORMANCE
        # ============================================
        ws_performance = wb.create_sheet(title="Subject Performance")
        
        # Title
        ws_performance['A1'] = "SUBJECT-WISE PERFORMANCE ANALYSIS"
        ws_performance['A1'].font = title_font
        ws_performance.merge_cells('A1:G1')
        ws_performance['A1'].alignment = center_alignment
        
        # Headers for subject performance
        perf_headers = ["Subject Code", "Subject Name", "Students with Marks", 
                        "Average Marks", "Highest Marks", "Lowest Marks", 
                        "Pass Rate (%)", "Grade Distribution"]
        
        for col_num, header in enumerate(perf_headers, start=1):
            cell = ws_performance.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = center_alignment
            
            # Set column widths
            if col_num == 1:  # Code
                ws_performance.column_dimensions[get_column_letter(col_num)].width = 12
            elif col_num == 2:  # Name
                ws_performance.column_dimensions[get_column_letter(col_num)].width = 30
            elif col_num in [3, 7]:  # Count and Pass Rate
                ws_performance.column_dimensions[get_column_letter(col_num)].width = 18
            elif col_num == 8:  # Grade Distribution
                ws_performance.column_dimensions[get_column_letter(col_num)].width = 25
            else:  # Marks columns
                ws_performance.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Calculate subject performance
        perf_start_row = 4
        
        for idx, subject in enumerate(subjects, start=perf_start_row):
            subject_results = results.filter(subject=subject, marks_obtained__isnull=False)
            marks_list = [float(r.marks_obtained) for r in subject_results if r.marks_obtained is not None]
            
            # Basic info
            ws_performance.cell(row=idx, column=1, value=subject.code).border = cell_border
            ws_performance.cell(row=idx, column=2, value=subject.name).border = cell_border
            
            # Students with marks
            students_with_marks = len(marks_list)
            ws_performance.cell(row=idx, column=3, value=students_with_marks).border = cell_border
            ws_performance.cell(row=idx, column=3).alignment = center_alignment
            
            if marks_list:
                # Average marks
                avg_marks = sum(marks_list) / len(marks_list)
                ws_performance.cell(row=idx, column=4, value=f"{avg_marks:.2f}").border = cell_border
                ws_performance.cell(row=idx, column=4).alignment = center_alignment
                
                # Highest marks
                highest = max(marks_list)
                ws_performance.cell(row=idx, column=5, value=f"{highest:.1f}").border = cell_border
                ws_performance.cell(row=idx, column=5).alignment = center_alignment
                
                # Lowest marks
                lowest = min(marks_list)
                ws_performance.cell(row=idx, column=6, value=f"{lowest:.1f}").border = cell_border
                ws_performance.cell(row=idx, column=6).alignment = center_alignment
                
                # Pass rate (assuming pass mark is 40)
                pass_mark = 40
                passed = sum(1 for m in marks_list if m >= pass_mark)
                pass_rate = (passed / len(marks_list)) * 100 if marks_list else 0
                ws_performance.cell(row=idx, column=7, value=f"{pass_rate:.1f}%").border = cell_border
                ws_performance.cell(row=idx, column=7).alignment = center_alignment
                
                # Grade distribution
                grades = {}
                for r in subject_results:
                    if r.grade:
                        grades[r.grade] = grades.get(r.grade, 0) + 1
                
                grade_dist = ", ".join([f"{grade}:{count}" for grade, count in grades.items()])
                ws_performance.cell(row=idx, column=8, value=grade_dist).border = cell_border
            else:
                # No marks for this subject
                for col in range(4, 9):
                    ws_performance.cell(row=idx, column=col, value="-").border = cell_border
                    ws_performance.cell(row=idx, column=col).alignment = center_alignment
        
        # Add subject performance summary
        perf_summary_row = perf_start_row + len(subjects) + 2
        ws_performance.cell(row=perf_summary_row, column=1, value="PERFORMANCE SUMMARY").font = title_font
        ws_performance.merge_cells(f'A{perf_summary_row}:B{perf_summary_row}')
        
        # ============================================
        # SHEET 3: GRADE DISTRIBUTION
        # ============================================
        ws_grades = wb.create_sheet(title="Grade Distribution")
        
        # Title
        ws_grades['A1'] = "GRADE DISTRIBUTION ACROSS ALL SUBJECTS"
        ws_grades['A1'].font = title_font
        ws_grades.merge_cells('A1:D1')
        ws_grades['A1'].alignment = center_alignment
        
        # Grade headers
        grade_headers = ["Grade", "Description", "Marks Range", "Count", "Percentage (%)"]
        
        for col_num, header in enumerate(grade_headers, start=1):
            cell = ws_grades.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = center_alignment
            
            # Set column widths
            if col_num == 2:  # Description
                ws_grades.column_dimensions[get_column_letter(col_num)].width = 20
            else:
                ws_grades.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Grade definitions and counts
        grade_definitions = [
            ("A", "Excellent", "80-100"),
            ("B", "Very Good", "70-79"),
            ("C", "Good", "60-69"),
            ("D", "Satisfactory", "50-59"),
            ("E", "Fair", "40-49"),
            ("F", "Fail", "0-39"),
            ("S", "Subsidiary", "N/A"),
        ]
        
        # Count grades across all results
        grade_counts = {}
        total_grades = 0
        
        for result in results:
            if result.grade:
                grade_counts[result.grade] = grade_counts.get(result.grade, 0) + 1
                total_grades += 1
        
        # Write grade data
        grade_start_row = 4
        
        for idx, (grade, description, range_text) in enumerate(grade_definitions, start=grade_start_row):
            ws_grades.cell(row=idx, column=1, value=grade).border = cell_border
            ws_grades.cell(row=idx, column=2, value=description).border = cell_border
            ws_grades.cell(row=idx, column=3, value=range_text).border = cell_border
            
            count = grade_counts.get(grade, 0)
            ws_grades.cell(row=idx, column=4, value=count).border = cell_border
            ws_grades.cell(row=idx, column=4).alignment = center_alignment
            
            percentage = (count / total_grades * 100) if total_grades > 0 else 0
            ws_grades.cell(row=idx, column=5, value=f"{percentage:.1f}%").border = cell_border
            ws_grades.cell(row=idx, column=5).alignment = center_alignment
        
        # Add total row
        total_row = grade_start_row + len(grade_definitions)
        ws_grades.cell(row=total_row, column=3, value="TOTAL").font = bold_font
        ws_grades.cell(row=total_row, column=3).border = cell_border
        
        ws_grades.cell(row=total_row, column=4, value=total_grades).font = bold_font
        ws_grades.cell(row=total_row, column=4).border = cell_border
        ws_grades.cell(row=total_row, column=4).alignment = center_alignment
        
        ws_grades.cell(row=total_row, column=5, value="100.0%").font = bold_font
        ws_grades.cell(row=total_row, column=5).border = cell_border
        ws_grades.cell(row=total_row, column=5).alignment = center_alignment
        
        # ============================================
        # SHEET 4: TOP PERFORMERS
        # ============================================
        ws_top = wb.create_sheet(title="Top Performers")
        
        # Title
        ws_top['A1'] = "TOP PERFORMERS - CLASS RANKING"
        ws_top['A1'].font = title_font
        ws_top.merge_cells('A1:F1')
        ws_top['A1'].alignment = center_alignment
        
        # Headers
        top_headers = ["Position", "Registration No.", "Student Name", 
                      "Total Marks", "Average", "Division"]
        
        for col_num, header in enumerate(top_headers, start=1):
            cell = ws_top.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = center_alignment
            
            # Set column widths
            if col_num == 3:  # Name
                ws_top.column_dimensions[get_column_letter(col_num)].width = 25
            else:
                ws_top.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Get top 10 performers based on position
        top_students = []
        for position in positions.order_by('class_position')[:20]:  # Top 20
            metrics = metrics_map.get(position.student_id)
            if metrics and position.class_position:
                student = next((s for s in students if s.id == position.student_id), None)
                if student:
                    top_students.append({
                        'position': position.class_position,
                        'reg_no': student.registration_number or f"S{student.id:04d}",
                        'name': student.full_name,
                        'total_marks': metrics.total_marks,
                        'average': metrics.average_marks,
                        'division': metrics.division.division if metrics.division else "-"
                    })
        
        # Write top performers
        top_start_row = 4
        
        for idx, student_data in enumerate(top_students, start=top_start_row):
            ws_top.cell(row=idx, column=1, value=student_data['position']).border = cell_border
            ws_top.cell(row=idx, column=1).alignment = center_alignment
            
            ws_top.cell(row=idx, column=2, value=student_data['reg_no']).border = cell_border
            
            ws_top.cell(row=idx, column=3, value=student_data['name']).border = cell_border
            
            ws_top.cell(row=idx, column=4, 
                       value=f"{student_data['total_marks']:.1f}" if student_data['total_marks'] else "-").border = cell_border
            ws_top.cell(row=idx, column=4).alignment = center_alignment
            
            ws_top.cell(row=idx, column=5, 
                       value=f"{student_data['average']:.2f}" if student_data['average'] else "-").border = cell_border
            ws_top.cell(row=idx, column=5).alignment = center_alignment
            
            ws_top.cell(row=idx, column=6, value=student_data['division']).border = cell_border
            ws_top.cell(row=idx, column=6).alignment = center_alignment
        
        # ============================================
        # CREATE RESPONSE
        # ============================================
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Session_Report_{exam_session.name.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating Excel report: {str(e)}", status=500)
    

@login_required
def download_session_summary(request, exam_session_id):
    """Download session summary as CSV"""
    try:
        exam_session = get_object_or_404(ExamSession, id=exam_session_id)
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Session_Summary_{exam_session.name}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow(['Exam Session Summary', exam_session.name])
        writer.writerow(['Academic Year', exam_session.academic_year.name])
        writer.writerow(['Term', exam_session.term.get_term_number_display()])
        writer.writerow(['Class', exam_session.class_level.name])
        writer.writerow(['Stream', exam_session.stream_class.stream_letter if exam_session.stream_class else 'All'])
        writer.writerow(['Exam Date', exam_session.exam_date])
        writer.writerow([])
        
        # Add summary statistics
        writer.writerow(['Summary Statistics'])
        # You can add more statistics here
        
        return response
        
    except Exception as e:
        return HttpResponse(f"Error generating summary: {str(e)}", status=500)


# Add this view function
@login_required
def student_sessions_list(request, student_id):
    """Display all exam sessions for a specific student where they have results"""
    try:
        student = get_object_or_404(
            Student.objects.select_related(
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=student_id,
            is_active=True
        )

        sessions_with_results = ExamSession.objects.filter(
            results__student=student
        ).distinct()

        exam_sessions = sessions_with_results.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'class_level__educational_level',
            'stream_class'
        ).annotate(
            total_subjects=Count('results', filter=Q(results__student=student)),
            average_marks=Avg('results__marks_obtained', filter=Q(results__student=student)),
            average_percentage=Avg('results__percentage', filter=Q(results__student=student)),
            total_marks=Sum('results__marks_obtained', filter=Q(results__student=student))
        ).order_by('-exam_date', '-created_at')

        session_metrics = {}
        for session in exam_sessions:
            metrics = StudentExamMetrics.objects.filter(
                student=student,
                exam_session=session
            ).first()

            positions = StudentExamPosition.objects.filter(
                student=student,
                exam_session=session
            ).first()

            session_metrics[session.id] = {
                'metrics': metrics,
                'positions': positions
            }

        current_class_info = {
            'class_level': student.class_level.name if student.class_level else 'Not Assigned',
            'stream': student.stream_class.stream_letter if student.stream_class else 'N/A',
            'registration_number': student.registration_number or f'S{student.id:04d}',
            'academic_year': student.academic_year.name if student.academic_year else 'N/A'
        }

        context = {
            'student': student,
            'exam_sessions': exam_sessions,
            'session_metrics': session_metrics,
            'current_class_info': current_class_info,
            'total_sessions': exam_sessions.count(),
            'page_title': f'{student.full_name} - Exam Sessions',
            'breadcrumb_title': 'Student Exam History',
        }

        return render(request, 'admin/results/student_sessions_list.html', context)

    except Exception as e:
        messages.error(request, f"Error loading student sessions: {str(e)}")
        return redirect('admin_students_list')



@login_required
def student_session_results(request, student_id, exam_session_id):
    """Display detailed results for a student in a specific exam session with subject positions"""
    try:
        student = get_object_or_404(
            Student.objects.select_related(
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=student_id,
            is_active=True
        )

        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type',
                'academic_year',
                'term',
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=exam_session_id
        )

        if not StudentResult.objects.filter(
            exam_session=exam_session,
            student=student
        ).exists():
            messages.warning(request, "No results found for this student in the selected exam session.")
            return redirect('student_sessions_list', student_id=student_id)

        subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('code')

        results = StudentResult.objects.filter(
            exam_session=exam_session,
            student=student
        ).select_related('subject')

        results_dict = {result.subject_id: result for result in results}

        # Get all students in the same class/stream for position calculation
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }
        
        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class
        
        all_students = Student.objects.filter(**student_filters)
        total_students_count = all_students.count()

        subject_results = []
        total_marks = 0
        total_grade_points = 0
        subjects_with_marks = 0

        for subject in subjects:
            result = results_dict.get(subject.id)

            if result and result.marks_obtained is not None:
                marks = float(result.marks_obtained)
                percentage = float(result.percentage) if result.percentage else 0
                grade_point = float(result.grade_point) if result.grade_point else 0

                total_marks += marks
                total_grade_points += grade_point
                subjects_with_marks += 1

                # Calculate subject position on the fly
                subject_position = get_student_subject_position(
                    student_id, subject.id, exam_session_id, exam_session
                )
                
                # Get total students with marks in this subject
                students_with_marks_in_subject = StudentResult.objects.filter(
                    exam_session=exam_session,
                    subject=subject,
                    marks_obtained__isnull=False
                ).count()

                subject_results.append({
                    'subject': subject,
                    'result': result,
                    'marks': marks,
                    'percentage': percentage,
                    'grade': result.grade,
                    'grade_point': grade_point,
                    'position': subject_position,
                    'total_students_in_subject': students_with_marks_in_subject,
                    'has_result': True
                })
            else:
                subject_results.append({
                    'subject': subject,
                    'result': None,
                    'marks': None,
                    'percentage': None,
                    'grade': '-',
                    'grade_point': None,
                    'position': None,
                    'total_students_in_subject': 0,
                    'has_result': False
                })

        metrics = StudentExamMetrics.objects.filter(
            student=student,
            exam_session=exam_session
        ).select_related('division').first()

        positions = StudentExamPosition.objects.filter(
            student=student,
            exam_session=exam_session
        ).first()

        overall_stats = {
            'total_subjects': len(subjects),
            'subjects_with_marks': subjects_with_marks,
            'completion_rate': (subjects_with_marks / len(subjects) * 100) if subjects else 0,
            'total_marks': total_marks,
            'total_grade_points': total_grade_points,
            'average_marks': total_marks / subjects_with_marks if subjects_with_marks else 0,
            'average_grade_points': total_grade_points / subjects_with_marks if subjects_with_marks else 0,
        }

        if exam_session.stream_class:
            class_students_count = Student.objects.filter(
                class_level=exam_session.class_level,
                stream_class=exam_session.stream_class,
                is_active=True
            ).count()
        else:
            class_students_count = Student.objects.filter(
                class_level=exam_session.class_level,
                is_active=True
            ).count()

        class_ranking = (
            f"{positions.class_position} of {class_students_count}"
            if positions and positions.class_position
            else "Not ranked"
        )

        context = {
            'student': student,
            'exam_session': exam_session,
            'subject_results': subject_results,
            'metrics': metrics,
            'positions': positions,
            'overall_stats': overall_stats,
            'class_ranking': class_ranking,
            'class_students_count': class_students_count,
            'total_students_count': total_students_count,
            'page_title': f'{student.full_name} - {exam_session.name}',
            'breadcrumb_title': 'Student Results Detail',
        }

        return render(request, 'admin/results/student_session_results.html', context)

    except Exception as e:
        messages.error(request, f"Error loading student results: {str(e)}")
        return redirect('student_sessions_list', student_id=student_id)


def get_student_subject_position(student_id, subject_id, exam_session_id, exam_session=None):
    """
    Helper function to calculate a student's position in a specific subject
    based on marks, with tie-breaking by registration number and name.
    Returns the position as an integer or None if not found.
    """
    try:
        if not exam_session:
            exam_session = ExamSession.objects.get(id=exam_session_id)
        
        # Get all students in the same class/stream
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }
        
        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class
        
        students = Student.objects.filter(**student_filters)
        
        # Get results for this subject
        results = StudentResult.objects.filter(
            exam_session_id=exam_session_id,
            subject_id=subject_id,
            marks_obtained__isnull=False
        ).select_related('student')
        
        # Build list of students with marks
        students_with_marks = []
        for result in results:
            if result.marks_obtained is not None:
                try:
                    marks_float = float(result.marks_obtained)
                    students_with_marks.append({
                        'student_id': result.student_id,
                        'marks': marks_float,
                        'registration_number': result.student.registration_number or f"S{result.student_id:04d}",
                        'full_name': result.student.full_name
                    })
                except (TypeError, ValueError):
                    continue
        
        # Sort by marks (descending) and then by tie-breakers
        students_with_marks.sort(key=lambda x: (
            -x['marks'],  # Primary: Higher marks first
            x['registration_number'],  # Secondary: Registration number (alphabetical)
            x['full_name']  # Tertiary: Full name (alphabetical)
        ))
        
        # Find the target student's position
        for position, item in enumerate(students_with_marks, start=1):
            if item['student_id'] == student_id:
                return position
        
        return None
        
    except Exception as e:
        print(f"Error calculating subject position: {str(e)}")
        return None


@login_required
def download_student_pdf_report(request, student_id, exam_session_id):
    """Generate and download PDF report for student exam results with subject positions"""
    try:
        # Get student
        student = get_object_or_404(
            Student.objects.select_related(
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=student_id,
            is_active=True
        )

        # Get exam session
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type',
                'academic_year',
                'term',
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=exam_session_id
        )

        # Get all students count for ranking context
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }
        
        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class
        
        all_students = Student.objects.filter(**student_filters)
        total_students_count = all_students.count()

        # Subjects
        subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('code')

        # Results
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            student=student
        ).select_related('subject')

        results_dict = {result.subject_id: result for result in results}

        subject_results = []
        total_marks = 0
        total_grade_points = 0
        subjects_with_marks = 0

        for subject in subjects:
            result = results_dict.get(subject.id)

            if result and result.marks_obtained is not None:
                marks = float(result.marks_obtained)
                percentage = float(result.percentage) if result.percentage else 0
                grade_point = float(result.grade_point) if result.grade_point else 0

                total_marks += marks
                total_grade_points += grade_point
                subjects_with_marks += 1

                # Calculate subject position on the fly
                subject_position = get_student_subject_position(
                    student_id, subject.id, exam_session_id, exam_session
                )
                
                # Get total students with marks in this subject
                students_with_marks_in_subject = StudentResult.objects.filter(
                    exam_session=exam_session,
                    subject=subject,
                    marks_obtained__isnull=False
                ).count()

                subject_results.append({
                    'subject': subject,
                    'result': result,
                    'marks': marks,
                    'percentage': percentage,
                    'grade': result.grade,
                    'grade_point': grade_point,
                    'position': subject_position,
                    'total_students_in_subject': students_with_marks_in_subject,
                    'has_result': True
                })
            else:
                subject_results.append({
                    'subject': subject,
                    'result': None,
                    'marks': None,
                    'percentage': None,
                    'grade': '-',
                    'grade_point': None,
                    'position': None,
                    'total_students_in_subject': 0,
                    'has_result': False
                })

        # Metrics
        metrics = StudentExamMetrics.objects.filter(
            student=student,
            exam_session=exam_session
        ).select_related('division').first()

        # Positions
        positions = StudentExamPosition.objects.filter(
            student=student,
            exam_session=exam_session
        ).first()

        overall_stats = {
            'total_subjects': len(subjects),
            'subjects_with_marks': subjects_with_marks,
            'completion_rate': (subjects_with_marks / len(subjects) * 100) if subjects else 0,
            'total_marks': total_marks,
            'total_grade_points': total_grade_points,
            'average_marks': total_marks / subjects_with_marks if subjects_with_marks else 0,
            'average_grade_points': total_grade_points / subjects_with_marks if subjects_with_marks else 0,
        }

        # Class count
        if exam_session.stream_class:
            class_students_count = Student.objects.filter(
                class_level=exam_session.class_level,
                stream_class=exam_session.stream_class,
                is_active=True
            ).count()
        else:
            class_students_count = Student.objects.filter(
                class_level=exam_session.class_level,
                is_active=True
            ).count()

        class_ranking = (
            f"{positions.class_position} of {class_students_count}"
            if positions and positions.class_position
            else "Not ranked"
        )

        context = {
            'student': student,
            'exam_session': exam_session,
            'subject_results': subject_results,
            'metrics': metrics,
            'positions': positions,
            'overall_stats': overall_stats,
            'class_ranking': class_ranking,
            'class_students_count': class_students_count,
            'total_students_count': total_students_count,
            'generated_date': timezone.now(),
            'generated_by': request.user.get_full_name() or request.user.username,
            'school_name': getattr(settings, 'SCHOOL_NAME', 'Your School Name'),
            'school_address': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'school_logo': os.path.join(settings.MEDIA_ROOT, 'school_logo.png') if os.path.exists(os.path.join(settings.MEDIA_ROOT, 'school_logo.png')) else None,
        }

        html_string = render_to_string(
            'admin/results/student_pdf_report.html',
            context
        )

        html = HTML(
            string=html_string,
            base_url=request.build_absolute_uri()
        )

        pdf_file = html.write_pdf()

        filename = (
            f"Results_{student.registration_number or student.id}_"
            f"{exam_session.name.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
        )

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        messages.error(request, f"Error generating PDF report: {str(e)}")
        return redirect(
            'student_session_results',
            student_id=student_id,
            exam_session_id=exam_session_id
        )


@login_required
def exam_sessions_list_view(request):
    """Display exam sessions management page with DataTable"""
    # Get all exam sessions with related data
    exam_sessions = ExamSession.objects.select_related(
        'exam_type',
        'academic_year',
        'term',
        'class_level',
        'class_level__educational_level',
        'stream_class'
    ).all()
    
    # Get filter parameters from request
    academic_year_filter = request.GET.get('academic_year', '')
    term_filter = request.GET.get('term', '')
    class_level_filter = request.GET.get('class_level', '')
    stream_filter = request.GET.get('stream', '')
    
    # Apply filters if provided
    if academic_year_filter:
        exam_sessions = exam_sessions.filter(academic_year_id=academic_year_filter)
    
    if term_filter:
        exam_sessions = exam_sessions.filter(term_id=term_filter)
    
    if class_level_filter:
        exam_sessions = exam_sessions.filter(class_level_id=class_level_filter)
    
    if stream_filter:
        exam_sessions = exam_sessions.filter(stream_class_id=stream_filter)
    
    # Get data for filters
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    class_levels = ClassLevel.objects.filter(is_active=True).select_related('educational_level').order_by('order')
    
    # Only show terms if academic year is selected
    terms = Term.objects.none()
    if academic_year_filter:
        terms = Term.objects.filter(academic_year_id=academic_year_filter).order_by('term_number')
    
    # Only show streams if class level is selected
    streams = StreamClass.objects.none()
    if class_level_filter:
        streams = StreamClass.objects.filter(
            class_level_id=class_level_filter,
            is_active=True
        ).order_by('stream_letter')
    
    # Calculate statistics
    total_sessions = exam_sessions.count()
    draft_sessions = exam_sessions.filter(status='draft').count()
    submitted_sessions = exam_sessions.filter(status='submitted').count()
    verified_sessions = exam_sessions.filter(status='verified').count()
    published_sessions = exam_sessions.filter(status='published').count()
    
    context = {
        'exam_sessions': exam_sessions,
        'academic_years': academic_years,
        'terms': terms,
        'class_levels': class_levels,
        'streams': streams,
        
        # Statistics for cards
        'total_sessions': total_sessions,
        'draft_sessions': draft_sessions,
        'submitted_sessions': submitted_sessions,
        'verified_sessions': verified_sessions,
        'published_sessions': published_sessions,
        
        'page_title': 'Exam Sessions Management',
        'breadcrumb_title': 'Exam Sessions',
    }
    
    return render(request, 'admin/results/exam_sessions_list_view.html', context)


@login_required
def exam_session_report_view(request, exam_session_id):
    """View exam session report"""
    exam_session = get_object_or_404(
        ExamSession.objects.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'stream_class'
        ),
        id=exam_session_id
    )
    
    context = {
        'exam_session': exam_session,
        'page_title': f'Report - {exam_session.name}',
    }
    
    return render(request, 'admin/results/exam_session_report.html', context)


@login_required
def exam_session_analysis_view(request, exam_session_id):
    """Enhanced exam session analysis with divisions/grades, gender, rankings, filters and cross-analysis matrix"""
    exam_session = get_object_or_404(
        ExamSession.objects.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'class_level__educational_level',
            'stream_class'
        ),
        id=exam_session_id
    )
    
    # Check if education level is Primary or Nursery
    is_primary_nursery = exam_session.class_level.educational_level.code in ['PRIMARY', 'NURSERY']
    
    # Get all metrics for this exam session
    metrics = StudentExamMetrics.objects.filter(
        exam_session=exam_session
    ).select_related('student', 'division').order_by('-average_marks')
    
    # Get positions
    positions = StudentExamPosition.objects.filter(
        exam_session=exam_session
    ).select_related('student')
    
    # Get all possible divisions from DivisionScale for this education level (only for non-primary/nursery)
    all_possible_divisions = []
    division_code_to_display = {}
    division_display_to_code = {}
    
    if not is_primary_nursery:
        if exam_session.class_level and exam_session.class_level.educational_level:
            division_scales = DivisionScale.objects.filter(
                education_level=exam_session.class_level.educational_level
            ).order_by('min_points')
            
            for scale in division_scales:
                display_name = scale.get_division_display()
                code = scale.division
                all_possible_divisions.append(display_name)
                division_code_to_display[code] = display_name
                division_display_to_code[display_name] = code
    
    # Get all possible genders
    all_possible_genders = ['Male', 'Female']
    other_genders = set()
    for metric in metrics:
        gender = metric.student.get_gender_display()
        if gender not in all_possible_genders:
            other_genders.add(gender)
    all_possible_genders.extend(sorted(list(other_genders)))
    
    # Get all possible grades for Primary/Nursery
    all_possible_grades = []
    grade_distribution_all = {}
    if is_primary_nursery:
        # Get all grades from GradingScale for this education level
        grading_scales = GradingScale.objects.filter(
            education_level=exam_session.class_level.educational_level
        ).order_by('min_mark')
        all_possible_grades = [scale.grade for scale in grading_scales]
        
        # Count grade distribution for all students
        for metric in metrics:
            if metric.average_grade:
                grade_distribution_all[metric.average_grade] = grade_distribution_all.get(metric.average_grade, 0) + 1
    
    # Combine data
    all_student_data = []
    for metric in metrics:
        position = positions.filter(student=metric.student).first()
        
        # Get division display name (only for non-primary/nursery)
        division_code = metric.division.division if metric.division else None
        division_display = metric.division.get_division_display() if metric.division else 'Not Assigned'
        
        student_info = {
            'id': metric.student.id,
            'registration_number': metric.student.registration_number or f"S{metric.student.id:04d}",
            'full_name': metric.student.full_name,
            'gender': metric.student.get_gender_display(),
            'total_marks': metric.total_marks,
            'average_marks': metric.average_marks,
            'average_percentage': metric.average_percentage,
            'average_grade': metric.average_grade,
            'total_grade_points': metric.total_grade_points,
            'division_code': division_code,
            'division_display': division_display,
            'division': division_code,
            'class_position': position.class_position if position else None,
            'stream_position': position.stream_position if position else None,
            'rank': position.class_position if position else None,
        }
        all_student_data.append(student_info)
    
    # Get filter parameters
    division_filter = request.GET.get('division', '')
    grade_filter = request.GET.get('grade', '')
    gender_filter = request.GET.get('gender', '')
    rank_filter = request.GET.get('rank_filter', '')
    top_n = request.GET.get('top_n', '10')
    bottom_n = request.GET.get('bottom_n', '10')
    
    # Apply filters to get filtered student_data
    student_data = all_student_data.copy()
    
    # Apply division filter (only for non-primary/nursery)
    if division_filter and not is_primary_nursery:
        filtered_students = []
        for student in student_data:
            if division_filter == 'Not Assigned':
                if not student['division']:
                    filtered_students.append(student)
            elif student['division_display'] == division_filter or student['division_code'] == division_filter:
                filtered_students.append(student)
        student_data = filtered_students
    
    # Apply grade filter (only for primary/nursery)
    if grade_filter and is_primary_nursery:
        filtered_students = []
        for student in student_data:
            if grade_filter == 'No Grade':
                if not student['average_grade']:
                    filtered_students.append(student)
            elif student['average_grade'] == grade_filter:
                filtered_students.append(student)
        student_data = filtered_students
    
    if gender_filter:
        student_data = [s for s in student_data if s['gender'].lower() == gender_filter.lower()]
    
    if rank_filter == 'top':
        try:
            n = int(top_n)
            student_data = sorted(
                [s for s in student_data if s['rank']], 
                key=lambda x: x['rank']
            )[:n]
        except ValueError:
            pass
    elif rank_filter == 'bottom':
        try:
            n = int(bottom_n)
            sorted_data = sorted(
                [s for s in student_data if s['rank']], 
                key=lambda x: x['rank'],
                reverse=True
            )
            student_data = sorted_data[:n]
        except ValueError:
            pass
    
    # Calculate statistics for filtered data
    total_students = len(student_data)
    students_with_division = len([s for s in student_data if s['division']]) if not is_primary_nursery else 0
    students_without_division = total_students - students_with_division if not is_primary_nursery else 0
    
    # Division distribution (only divisions with students) - for non-primary/nursery
    division_distribution = {}
    if not is_primary_nursery:
        for student in student_data:
            if student['division_display'] and student['division_display'] != 'Not Assigned':
                division_distribution[student['division_display']] = division_distribution.get(
                    student['division_display'], 0) + 1
    
    # Grade distribution (only grades with students) - for primary/nursery
    grade_distribution = {}
    if is_primary_nursery:
        for student in student_data:
            if student['average_grade']:
                grade_distribution[student['average_grade']] = grade_distribution.get(student['average_grade'], 0) + 1
    
    # Gender distribution
    gender_distribution = {}
    for student in student_data:
        gender = student['gender']
        gender_distribution[gender] = gender_distribution.get(gender, 0) + 1
    
    # Average marks by gender
    gender_averages = {}
    gender_counts = {}
    for student in student_data:
        gender = student['gender']
        if student['average_marks']:
            gender_averages[gender] = gender_averages.get(gender, 0) + float(student['average_marks'])
            gender_counts[gender] = gender_counts.get(gender, 0) + 1
    
    for gender in gender_averages:
        if gender_counts[gender] > 0:
            gender_averages[gender] = round(gender_averages[gender] / gender_counts[gender], 2)
    
    # Average marks by division (for non-primary/nursery)
    division_averages = {}
    division_counts = {}
    if not is_primary_nursery:
        for student in student_data:
            division = student['division_display']
            if division and division != 'Not Assigned' and student['average_marks']:
                division_averages[division] = division_averages.get(division, 0) + float(student['average_marks'])
                division_counts[division] = division_counts.get(division, 0) + 1
        
        for division in division_averages:
            if division_counts[division] > 0:
                division_averages[division] = round(division_averages[division] / division_counts[division], 2)
    
    # ---------------------------------------------------------
    # CROSS-ANALYSIS MATRIX - CONDITIONAL BASED ON EDUCATION LEVEL
    # ---------------------------------------------------------
    
    if is_primary_nursery:
        # ---------------------------------------------------------
        # GRADE × GENDER CROSS-ANALYSIS MATRIX - FOR PRIMARY/NURSERY
        # ---------------------------------------------------------
        
        # Use all possible grades from GradingScale
        matrix_grades = all_possible_grades.copy()
        matrix_genders = all_possible_genders.copy()
        
        # Initialize matrix with zeros
        grade_gender_matrix = {gender: {grade: 0 for grade in matrix_grades} for gender in matrix_genders}
        grade_totals = {grade: 0 for grade in matrix_grades}
        gender_totals = {gender: 0 for gender in matrix_genders}
        grand_total = 0
        
        # Populate matrix with actual data from ALL students (not filtered)
        for student in all_student_data:
            gen = student['gender']
            grade = student['average_grade']
            
            if grade and grade in matrix_grades:
                grade_gender_matrix[gen][grade] += 1
                grade_totals[grade] += 1
                gender_totals[gen] += 1
                grand_total += 1
        
        # Add 'No Grade' category if needed
        students_without_grade_all = len([s for s in all_student_data if not s['average_grade']])
        if students_without_grade_all > 0:
            matrix_grades_with_na = matrix_grades + ['No Grade']
            
            for gender in matrix_genders:
                grade_gender_matrix[gender]['No Grade'] = 0
            grade_totals['No Grade'] = 0
            
            for student in all_student_data:
                gen = student['gender']
                grade = student['average_grade']
                if not grade:
                    grade_gender_matrix[gen]['No Grade'] += 1
                    grade_totals['No Grade'] += 1
                    gender_totals[gen] += 1
                    grand_total += 1
            
            matrix_grades = matrix_grades_with_na
        
        matrix_columns = matrix_grades
        matrix_data = grade_gender_matrix
        column_totals = grade_totals
        
    else:
        # ---------------------------------------------------------
        # DIVISION × GENDER CROSS-ANALYSIS MATRIX - FOR SECONDARY
        # ---------------------------------------------------------
        
        # Use all possible divisions from DivisionScale
        matrix_divisions = all_possible_divisions.copy()
        matrix_genders = all_possible_genders.copy()
        
        # Initialize matrix with zeros
        division_gender_matrix = {gender: {div: 0 for div in matrix_divisions} for gender in matrix_genders}
        division_totals = {div: 0 for div in matrix_divisions}
        gender_totals = {gender: 0 for gender in matrix_genders}
        grand_total = 0
        
        # Populate matrix with actual data from ALL students (not filtered)
        for student in all_student_data:
            gen = student['gender']
            div = student['division_display']
            
            if div and div != 'Not Assigned' and div in matrix_divisions:
                division_gender_matrix[gen][div] += 1
                division_totals[div] += 1
                gender_totals[gen] += 1
                grand_total += 1
        
        # Add 'Not Assigned' category if needed
        students_without_division_all = len([s for s in all_student_data if not s['division']])
        if students_without_division_all > 0:
            matrix_divisions_with_na = matrix_divisions + ['Not Assigned']
            
            for gender in matrix_genders:
                division_gender_matrix[gender]['Not Assigned'] = 0
            division_totals['Not Assigned'] = 0
            
            for student in all_student_data:
                gen = student['gender']
                div = student['division_display']
                if not div or div == 'Not Assigned':
                    division_gender_matrix[gen]['Not Assigned'] += 1
                    division_totals['Not Assigned'] += 1
                    gender_totals[gen] += 1
                    grand_total += 1
            
            matrix_divisions = matrix_divisions_with_na
        
        matrix_columns = matrix_divisions
        matrix_data = division_gender_matrix
        column_totals = division_totals
    
    # Top performers (from all students, not filtered)
    top_performers = sorted(
        [s for s in all_student_data if s['rank']],
        key=lambda x: x['rank']
    )[:10]
    
    # Bottom performers (from all students, not filtered)
    bottom_performers = sorted(
        [s for s in all_student_data if s['rank']],
        key=lambda x: x['rank'],
        reverse=True
    )[:10]
    
    # Filter options
    if is_primary_nursery:
        unique_divisions_for_filter = []  # No division filter for primary/nursery
        unique_divisions_display = []
        unique_grades_for_filter = all_possible_grades + ['No Grade'] if all_possible_grades else []
    else:
        unique_divisions_for_filter = ['Division I', 'Division II', 'Division III', 'Division IV', 'Division 0', 'Not Assigned']
        unique_divisions_display = matrix_divisions
        unique_grades_for_filter = []
    
    # Get unique genders for filter dropdowns
    unique_genders = sorted(set([s['gender'] for s in all_student_data if s['gender']]))
    
    context = {
        'exam_session': exam_session,
        'is_primary_nursery': is_primary_nursery,
        
        # Student data
        'student_data': student_data,
        'all_student_data': all_student_data,
        'total_students': len(student_data),
        'total_students_all': len(all_student_data),
        
        # Division/Grade stats
        'students_with_division': students_with_division if not is_primary_nursery else 0,
        'students_with_division_all': len([s for s in all_student_data if s['division']]) if not is_primary_nursery else 0,
        'students_without_division': students_without_division if not is_primary_nursery else 0,
        'students_without_division_all': len([s for s in all_student_data if not s['division']]) if not is_primary_nursery else 0,
        
        # Distributions
        'division_distribution': division_distribution if not is_primary_nursery else {},
        'grade_distribution': grade_distribution if is_primary_nursery else {},
        'gender_distribution': gender_distribution,
        
        # Averages
        'gender_averages': gender_averages,
        'division_averages': division_averages if not is_primary_nursery else {},
        
        # Top/Bottom performers
        'top_performers': top_performers,
        'bottom_performers': bottom_performers,
        
        # Cross-analysis matrix data
        'matrix_data': matrix_data,
        'matrix_columns': matrix_columns,
        'column_totals': column_totals,
        'gender_totals': gender_totals,
        'grand_total': grand_total,
        
        # Education level specific data
        'all_possible_grades': all_possible_grades if is_primary_nursery else [],
        'grade_distribution_all': grade_distribution_all if is_primary_nursery else {},
        
        # Division specific data (for secondary)
        'division_gender_matrix': division_gender_matrix if not is_primary_nursery else {},
        'division_columns': matrix_divisions if not is_primary_nursery else [],
        'division_totals': division_totals if not is_primary_nursery else {},
        
        # Grade specific data (for primary/nursery)
        'grade_gender_matrix': grade_gender_matrix if is_primary_nursery else {},
        'grade_columns': matrix_grades if is_primary_nursery else [],
        'grade_totals': grade_totals if is_primary_nursery else {},
        
        # Filter options
        'unique_divisions': unique_divisions_for_filter,
        'unique_divisions_display': unique_divisions_display,
        'unique_grades': unique_grades_for_filter,
        'unique_genders': unique_genders,
        
        # Current filter values
        'division_filter': division_filter if not is_primary_nursery else '',
        'grade_filter': grade_filter if is_primary_nursery else '',
        'gender_filter': gender_filter,
        'rank_filter': rank_filter,
        'top_n': top_n,
        'bottom_n': bottom_n,
        
        # Division display mapping
        'division_code_to_display': division_code_to_display if not is_primary_nursery else {},
        'division_display_to_code': division_display_to_code if not is_primary_nursery else {},
        
        # Rank filter options
        'rank_filter_options': [
            ('', 'All Students'),
            ('top', 'Top N Students'),
            ('bottom', 'Bottom N Students'),
        ],
        
        'page_title': f'Analysis - {exam_session.name}',
    }
    
    return render(request, 'admin/results/exam_session_analysis.html', context)


@login_required
def exam_session_analysis_pdf(request, exam_session_id):
    """Generate PDF report for exam session analysis with sections and filter support"""
    try:
        print(f"[ANALYSIS_PDF] Start | exam_session_id={exam_session_id}")

        # Get exam session
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type',
                'academic_year',
                'term',
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=exam_session_id
        )
        print(f"[ANALYSIS_PDF] Loaded session: {exam_session.name}")

        # Check if education level is Primary or Nursery
        is_primary_nursery = exam_session.class_level.educational_level.code in ['PRIMARY', 'NURSERY']
        print(f"[ANALYSIS_PDF] Is Primary/Nursery: {is_primary_nursery}")

        # Section parameter
        section = request.GET.get('section', 'full')
        print(f"[ANALYSIS_PDF] Section requested: {section}")

        # ============= GET FILTER PARAMETERS =============
        division_filter = request.GET.get('division', '')
        grade_filter = request.GET.get('grade', '')
        gender_filter = request.GET.get('gender', '')
        rank_filter = request.GET.get('rank_filter', '')
        top_n = request.GET.get('top_n', '10')
        bottom_n = request.GET.get('bottom_n', '10')
        
        print(f"[ANALYSIS_PDF] Filters - Division: '{division_filter}', Grade: '{grade_filter}', Gender: '{gender_filter}', Rank: '{rank_filter}', TopN: {top_n}, BottomN: {bottom_n}")

        # Metrics
        metrics = StudentExamMetrics.objects.filter(
            exam_session=exam_session
        ).select_related('student', 'division').order_by('-average_marks')
        print(f"[ANALYSIS_PDF] Metrics count: {metrics.count()}")

        # Positions
        positions = StudentExamPosition.objects.filter(
            exam_session=exam_session
        ).select_related('student')
        print(f"[ANALYSIS_PDF] Positions count: {positions.count()}")

        # Get all possible divisions from DivisionScale for this education level (only for non-primary/nursery)
        all_possible_divisions = []
        division_code_to_display = {}
        division_display_to_code = {}
        
        if not is_primary_nursery:
            if exam_session.class_level and exam_session.class_level.educational_level:
                division_scales = DivisionScale.objects.filter(
                    education_level=exam_session.class_level.educational_level
                ).order_by('min_points')
                
                for scale in division_scales:
                    display_name = scale.get_division_display()
                    code = scale.division
                    all_possible_divisions.append(display_name)
                    division_code_to_display[code] = display_name
                    division_display_to_code[display_name] = code
            
            print(f"[ANALYSIS_PDF] All possible divisions: {all_possible_divisions}")
        
        # Get all possible grades from GradingScale for this education level (only for primary/nursery)
        all_possible_grades = []
        grade_distribution_all = {}
        if is_primary_nursery:
            grading_scales = GradingScale.objects.filter(
                education_level=exam_session.class_level.educational_level
            ).order_by('min_mark')
            all_possible_grades = [scale.grade for scale in grading_scales]
            print(f"[ANALYSIS_PDF] All possible grades: {all_possible_grades}")

        # Get all possible genders from the Student model
        all_possible_genders = ['Male', 'Female']
        other_genders = set()
        for metric in metrics:
            gender = metric.student.get_gender_display()
            if gender not in all_possible_genders:
                other_genders.add(gender)
        all_possible_genders.extend(sorted(list(other_genders)))
        print(f"[ANALYSIS_PDF] All possible genders: {all_possible_genders}")

        # ============= BUILD ALL STUDENT DATA =============
        all_student_data = []
        
        for metric in metrics:
            position = positions.filter(student=metric.student).first()
            
            # Get division information (only for non-primary/nursery)
            division_code = metric.division.division if metric.division else None
            division_display = metric.division.get_division_display() if metric.division else 'Not Assigned'
            
            student_info = {
                'id': metric.student.id,
                'registration_number': metric.student.registration_number or f"S{metric.student.id:04d}",
                'full_name': metric.student.full_name,
                'gender': metric.student.get_gender_display(),
                'total_marks': metric.total_marks,
                'average_marks': metric.average_marks,
                'average_percentage': metric.average_percentage,
                'average_grade': metric.average_grade,
                'total_grade_points': metric.total_grade_points,
                'division_code': division_code,
                'division_display': division_display,
                'division': division_code,
                'class_position': position.class_position if position else None,
                'stream_position': position.stream_position if position else None,
                'rank': position.class_position if position else None,
            }
            all_student_data.append(student_info)

        print(f"[ANALYSIS_PDF] Total students processed: {len(all_student_data)}")

        # ============= APPLY FILTERS =============
        filtered_student_data = all_student_data.copy()
        
        # Apply Division Filter (only for non-primary/nursery)
        if division_filter and not is_primary_nursery:
            division_filtered = []
            for student in filtered_student_data:
                if division_filter == 'Not Assigned':
                    if not student['division_code']:
                        division_filtered.append(student)
                elif student['division_display'] == division_filter or student['division_code'] == division_filter:
                    division_filtered.append(student)
            
            filtered_student_data = division_filtered
            print(f"[ANALYSIS_PDF] After division filter '{division_filter}': {len(filtered_student_data)} students")
        
        # Apply Grade Filter (only for primary/nursery)
        if grade_filter and is_primary_nursery:
            grade_filtered = []
            for student in filtered_student_data:
                if grade_filter == 'No Grade':
                    if not student['average_grade']:
                        grade_filtered.append(student)
                elif student['average_grade'] == grade_filter:
                    grade_filtered.append(student)
            
            filtered_student_data = grade_filtered
            print(f"[ANALYSIS_PDF] After grade filter '{grade_filter}': {len(filtered_student_data)} students")
        
        # Apply Gender Filter
        if gender_filter:
            filtered_student_data = [
                s for s in filtered_student_data 
                if s['gender'].lower() == gender_filter.lower()
            ]
            print(f"[ANALYSIS_PDF] After gender filter '{gender_filter}': {len(filtered_student_data)} students")
        
        # Apply Rank Filter
        if rank_filter == 'top':
            try:
                n = int(top_n)
                ranked_students = [s for s in filtered_student_data if s['rank']]
                filtered_student_data = sorted(
                    ranked_students,
                    key=lambda x: x['rank']
                )[:n]
                print(f"[ANALYSIS_PDF] After top {n} filter: {len(filtered_student_data)} students")
            except ValueError:
                pass
        elif rank_filter == 'bottom':
            try:
                n = int(bottom_n)
                ranked_students = [s for s in filtered_student_data if s['rank']]
                filtered_student_data = sorted(
                    ranked_students,
                    key=lambda x: x['rank'],
                    reverse=True
                )[:n]
                print(f"[ANALYSIS_PDF] After bottom {n} filter: {len(filtered_student_data)} students")
            except ValueError:
                pass

        # Statistics for filtered data
        total_students_filtered = len(filtered_student_data)
        
        if not is_primary_nursery:
            students_with_division_filtered = len([s for s in filtered_student_data if s['division_code']])
            students_without_division_filtered = total_students_filtered - students_with_division_filtered
        else:
            students_with_division_filtered = 0
            students_without_division_filtered = 0
        
        # Division distribution (filtered) - for non-primary/nursery
        division_distribution_filtered = {}
        if not is_primary_nursery:
            for student in filtered_student_data:
                if student['division_display'] and student['division_display'] != 'Not Assigned':
                    division_distribution_filtered[student['division_display']] = division_distribution_filtered.get(
                        student['division_display'], 0) + 1
        
        # Grade distribution (filtered) - for primary/nursery
        grade_distribution_filtered = {}
        if is_primary_nursery:
            for student in filtered_student_data:
                if student['average_grade']:
                    grade_distribution_filtered[student['average_grade']] = grade_distribution_filtered.get(
                        student['average_grade'], 0) + 1
        
        # Gender distribution (filtered)
        gender_distribution_filtered = {}
        for student in filtered_student_data:
            gender = student['gender']
            gender_distribution_filtered[gender] = gender_distribution_filtered.get(gender, 0) + 1
        
        # Gender averages (filtered)
        gender_averages_filtered = {}
        gender_counts_filtered = {}
        for student in filtered_student_data:
            gender = student['gender']
            if student['average_marks']:
                gender_averages_filtered[gender] = gender_averages_filtered.get(gender, 0) + float(student['average_marks'])
                gender_counts_filtered[gender] = gender_counts_filtered.get(gender, 0) + 1
        
        for gender in gender_averages_filtered:
            if gender_counts_filtered[gender] > 0:
                gender_averages_filtered[gender] = round(gender_averages_filtered[gender] / gender_counts_filtered[gender], 2)
        
        # Division averages (filtered) - for non-primary/nursery
        division_averages_filtered = {}
        division_counts_filtered = {}
        if not is_primary_nursery:
            for student in filtered_student_data:
                division = student['division_display']
                if division and division != 'Not Assigned' and student['average_marks']:
                    division_averages_filtered[division] = division_averages_filtered.get(division, 0) + float(student['average_marks'])
                    division_counts_filtered[division] = division_counts_filtered.get(division, 0) + 1
            
            for division in division_averages_filtered:
                if division_counts_filtered[division] > 0:
                    division_averages_filtered[division] = round(
                        division_averages_filtered[division] / division_counts_filtered[division], 2
                    )

        # ============= CROSS-ANALYSIS MATRIX (ALL STUDENTS - UNFILTERED) =============
        
        if is_primary_nursery:
            # ---------------------------------------------------------
            # GRADE × GENDER CROSS-ANALYSIS MATRIX - FOR PRIMARY/NURSERY
            # ---------------------------------------------------------
            
            matrix_grades = all_possible_grades.copy()
            matrix_genders = all_possible_genders.copy()
            
            # Initialize matrix with zeros
            grade_gender_matrix = {gender: {grade: 0 for grade in matrix_grades} for gender in matrix_genders}
            grade_totals = {grade: 0 for grade in matrix_grades}
            gender_totals = {gender: 0 for gender in matrix_genders}
            grand_total = 0
            
            # Populate matrix with ALL students (unfiltered)
            for student in all_student_data:
                gen = student['gender']
                grade = student['average_grade']
                
                if grade and grade in matrix_grades:
                    grade_gender_matrix[gen][grade] += 1
                    grade_totals[grade] += 1
                    gender_totals[gen] += 1
                    grand_total += 1
            
            # Add 'No Grade' category
            students_without_grade_all = len([s for s in all_student_data if not s['average_grade']])
            if students_without_grade_all > 0:
                matrix_grades_with_na = matrix_grades + ['No Grade']
                
                for gender in matrix_genders:
                    grade_gender_matrix[gender]['No Grade'] = 0
                grade_totals['No Grade'] = 0
                
                for student in all_student_data:
                    gen = student['gender']
                    grade = student['average_grade']
                    if not grade:
                        grade_gender_matrix[gen]['No Grade'] += 1
                        grade_totals['No Grade'] += 1
                        gender_totals[gen] += 1
                        grand_total += 1
                
                matrix_grades = matrix_grades_with_na
            
            unique_columns = matrix_grades
            matrix_data = grade_gender_matrix
            column_totals = grade_totals
            matrix_name = "Grade × Gender Cross-Analysis"
            
        else:
            # ---------------------------------------------------------
            # DIVISION × GENDER CROSS-ANALYSIS MATRIX - FOR SECONDARY
            # ---------------------------------------------------------
            
            unique_columns = all_possible_divisions.copy()
            matrix_genders = all_possible_genders.copy()
            
            # Initialize matrix with zeros
            division_gender_matrix = {gender: {div: 0 for div in unique_columns} for gender in matrix_genders}
            division_totals = {div: 0 for div in unique_columns}
            gender_totals = {gender: 0 for gender in matrix_genders}
            grand_total = 0
            
            # Populate matrix with ALL students (unfiltered)
            for student in all_student_data:
                gen = student['gender']
                div = student['division_display']
                
                if div and div != 'Not Assigned' and div in unique_columns:
                    division_gender_matrix[gen][div] += 1
                    division_totals[div] += 1
                    gender_totals[gen] += 1
                    grand_total += 1
            
            # Add 'Not Assigned' category
            students_without_division_all = len([s for s in all_student_data if not s['division_code']])
            if students_without_division_all > 0:
                unique_columns_with_na = unique_columns + ['Not Assigned']
                
                for gender in matrix_genders:
                    division_gender_matrix[gender]['Not Assigned'] = 0
                division_totals['Not Assigned'] = 0
                
                for student in all_student_data:
                    gen = student['gender']
                    div = student['division_display']
                    if not div or div == 'Not Assigned':
                        division_gender_matrix[gen]['Not Assigned'] += 1
                        division_totals['Not Assigned'] += 1
                        gender_totals[gen] += 1
                        grand_total += 1
                
                unique_columns = unique_columns_with_na
            
            matrix_data = division_gender_matrix
            column_totals = division_totals
            matrix_name = "Division × Gender Cross-Analysis"

        # ============= TOP/BOTTOM PERFORMERS (ALL STUDENTS) =============
        top_performers = sorted(
            [s for s in all_student_data if s['rank']],
            key=lambda x: x['rank']
        )[:10]

        bottom_performers = sorted(
            [s for s in all_student_data if s['rank']],
            key=lambda x: x['rank'],
            reverse=True
        )[:10]

        print(f"[ANALYSIS_PDF] Top performers: {len(top_performers)}")
        print(f"[ANALYSIS_PDF] Bottom performers: {len(bottom_performers)}")

        # ============= SECTION FILTERING =============
        # This is for PDF section selection, not the same as filter parameters
        if section == 'top_performers':
            display_data = top_performers
        elif section == 'bottom_performers':
            display_data = bottom_performers
        elif section == 'divisions_only' and not is_primary_nursery:
            display_data = [s for s in filtered_student_data if s['division_code']]
        elif section == 'grades_only' and is_primary_nursery:
            display_data = [s for s in filtered_student_data if s['average_grade']]
        elif section == 'genders_only':
            display_data = filtered_student_data
        elif section == 'detailed_list':
            display_data = filtered_student_data
        else:
            display_data = filtered_student_data

        # ============= PREPARE CONTEXT =============
        context = {
            # Exam session info
            'exam_session': exam_session,
            'is_primary_nursery': is_primary_nursery,
            'generated_date': timezone.now(),
            'generated_by': request.user.get_full_name() or request.user.username,
            'school_name': getattr(settings, 'SCHOOL_NAME', 'School Management System'),
            'school_address': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'is_pdf': True,
            
            # Section info
            'section': section,
            'section_title': {
                'full': 'Full Analysis Report',
                'overview': 'Overview Report',
                'divisions': 'Division Analysis',
                'genders': 'Gender Analysis',
                'rankings': 'Rankings Analysis',
                'top_performers': 'Top Performers Report',
                'bottom_performers': 'Bottom Performers Report',
                'divisions_only': 'Division Analysis Report',
                'grades_only': 'Grade Analysis Report',
                'genders_only': 'Gender Analysis Report',
                'detailed_list': 'Detailed Student List',
                'cross_analysis': matrix_name,
            }.get(section, 'Analysis Report'),
            
            # FILTERED DATA - for detailed lists and section-specific views
            'student_data': display_data,
            'total_students': len(filtered_student_data),
            'total_students_all': len(all_student_data),
            
            # Division/Grade stats
            'students_with_division': students_with_division_filtered if not is_primary_nursery else 0,
            'students_with_division_all': len([s for s in all_student_data if s['division_code']]) if not is_primary_nursery else 0,
            'students_without_division': students_without_division_filtered if not is_primary_nursery else 0,
            'students_without_division_all': len([s for s in all_student_data if not s['division_code']]) if not is_primary_nursery else 0,
            
            # Distributions (filtered)
            'division_distribution': division_distribution_filtered if not is_primary_nursery else {},
            'grade_distribution': grade_distribution_filtered if is_primary_nursery else {},
            'gender_distribution': gender_distribution_filtered,
            
            # Averages (filtered)
            'gender_averages': gender_averages_filtered,
            'division_averages': division_averages_filtered if not is_primary_nursery else {},
            
            # ALL STUDENTS DATA - for cross-analysis and top/bottom performers
            'all_student_data': all_student_data,
            'top_performers': top_performers,
            'bottom_performers': bottom_performers,
            
            # Cross-analysis matrix (ALL students)
            'matrix_data': matrix_data,
            'matrix_columns': unique_columns,
            'column_totals': column_totals,
            'gender_totals': gender_totals,
            'grand_total': grand_total,
            'matrix_name': matrix_name,
            
            # Education level specific data
            'all_possible_grades': all_possible_grades if is_primary_nursery else [],
            'grade_columns': unique_columns if is_primary_nursery else [],
            'grade_totals': grade_totals if is_primary_nursery else {},
            'grade_gender_matrix': grade_gender_matrix if is_primary_nursery else {},
            
            'division_columns': unique_columns if not is_primary_nursery else [],
            'division_totals': column_totals if not is_primary_nursery else {},
            'division_gender_matrix': matrix_data if not is_primary_nursery else {},
            
            # Filter values (for displaying current filters in PDF)
            'division_filter': division_filter if not is_primary_nursery else '',
            'grade_filter': grade_filter if is_primary_nursery else '',
            'gender_filter': gender_filter,
            'rank_filter': rank_filter,
            'top_n': top_n,
            'bottom_n': bottom_n,
            
            # Division mapping
            'division_code_to_display': division_code_to_display if not is_primary_nursery else {},
            'division_display_to_code': division_display_to_code if not is_primary_nursery else {},
        }

        # ============= TEMPLATE SELECTION =============
        template_mapping = {
            'full': 'admin/results/analysis_pdf_full.html',
            'overview': 'admin/results/analysis_pdf_section.html',
            'divisions': 'admin/results/analysis_pdf_section.html',
            'genders': 'admin/results/analysis_pdf_section.html',
            'rankings': 'admin/results/analysis_pdf_section.html',
            'top_performers': 'admin/results/analysis_pdf_performers.html',
            'bottom_performers': 'admin/results/analysis_pdf_performers.html',
            'divisions_only': 'admin/results/analysis_pdf_distribution.html',
            'grades_only': 'admin/results/analysis_pdf_distribution.html',
            'genders_only': 'admin/results/analysis_pdf_distribution.html',
            'detailed_list': 'admin/results/analysis_pdf_detailed.html',
            'cross_analysis': 'admin/results/analysis_pdf_distribution.html',
        }
        
        template_name = template_mapping.get(section, 'admin/results/analysis_pdf_section.html')
        print(f"[ANALYSIS_PDF] Using template: {template_name}")

        # ============= GENERATE PDF =============
        html_string = render_to_string(template_name, context)
        
        print("[ANALYSIS_PDF] Generating PDF...")
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = html.write_pdf()

        # Create filename with filter information
        filename_parts = [f"Analysis_{exam_session.name.replace(' ', '_')}"]
        
        if not is_primary_nursery and division_filter:
            filename_parts.append(f"div_{division_filter.replace(' ', '_')}")
        if is_primary_nursery and grade_filter:
            filename_parts.append(f"grade_{grade_filter}")
        if gender_filter:
            filename_parts.append(f"gender_{gender_filter}")
        if rank_filter:
            filename_parts.append(f"{rank_filter}_{top_n if rank_filter == 'top' else bottom_n}")
        
        filename_parts.append(timezone.now().strftime('%Y%m%d'))
        filename = "_".join(filename_parts) + ".pdf"
        filename = filename.replace('/', '_')

        print(f"[ANALYSIS_PDF] PDF generated successfully: {filename}")

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        print(f"[ANALYSIS_PDF][ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error generating PDF report: {str(e)}")
        return redirect('exam_session_analysis_view', exam_session_id=exam_session_id)






@login_required
def session_subject_analysis_view(request, exam_session_id):
    """
    Session-based subject performance analysis.
    Subject is selected via filter, not URL parameter.
    """
    # Get exam session with related data
    exam_session = get_object_or_404(
        ExamSession.objects.select_related(
            'exam_type',
            'academic_year',
            'term',
            'class_level',
            'class_level__educational_level',
            'stream_class'
        ),
        id=exam_session_id
    )
    
    # Get filter parameters
    subject_id = request.GET.get('subject_id', '')
    grade_filter = request.GET.get('grade_filter', '')
    gender_filter = request.GET.get('gender', '')
    rank_filter = request.GET.get('rank_filter', '')
    top_n = request.GET.get('top_n', '10')
    bottom_n = request.GET.get('bottom_n', '10')
    
    # Convert to integers where applicable
    try:
        top_n = int(top_n) if top_n else 10
    except ValueError:
        top_n = 10
    
    try:
        bottom_n = int(bottom_n) if bottom_n else 10
    except ValueError:
        bottom_n = 10
    
    # Get all active subjects for this educational level
    all_subjects = Subject.objects.filter(
        educational_level=exam_session.class_level.educational_level,
        is_active=True
    ).order_by('name')
    
    # Get the selected subject (default to first subject if none selected)
    selected_subject = None
    if subject_id:
        try:
            selected_subject = all_subjects.get(id=subject_id)
        except Subject.DoesNotExist:
            selected_subject = all_subjects.first()
    else:
        selected_subject = all_subjects.first()
    
    # If no subjects exist, return early
    if not selected_subject:
        context = {
            'exam_session': exam_session,
            'all_subjects': all_subjects,
            'selected_subject': None,
            'no_subjects': True,
            'page_title': f'Subject Analysis - {exam_session.name}',
        }
        return render(request, 'admin/results/session_subject_analysis.html', context)
    
    # Get students based on stream
    student_filters = {
        'class_level': exam_session.class_level,
        'is_active': True
    }
    
    if exam_session.stream_class:
        student_filters['stream_class'] = exam_session.stream_class
    
    students = Student.objects.filter(**student_filters).order_by('first_name', 'last_name')
    total_students = students.count()
    
    # Get results for the selected subject
    results = StudentResult.objects.filter(
        exam_session=exam_session,
        subject=selected_subject
    ).select_related('student')
    
    results_by_student = {r.student_id: r for r in results}
    
    # ============================================
    # CALCULATE POSITIONS DYNAMICALLY USING HELPER FUNCTION
    # ============================================
    
    # Get position map using helper function
    position_map = calculate_subject_positions(students, results_by_student)
    
    # ============================================
    # DATA STRUCTURES INITIALIZATION
    # ============================================
    
    # Get all possible grades from GradingScale
    grading_scales = GradingScale.objects.filter(
        education_level=exam_session.class_level.educational_level
    ).order_by('min_mark')
    all_grades = [scale.grade for scale in grading_scales]
    
    # Gender tracking
    all_genders = ['Male', 'Female', 'Other']
    
    # Initialize ALL STUDENTS data structures (unfiltered)
    all_student_data = []
    all_students_with_marks = 0
    all_total_marks_sum = 0
    all_marks_list = []
    all_gender_stats = {}
    all_grade_gender_matrix = {}
    all_grade_counts = {}
    
    # Initialize all gender stats
    for gender in all_genders:
        all_gender_stats[gender] = {
            'total': 0,
            'with_marks': 0,
            'marks_sum': 0,
            'average': 0
        }
    
    # Initialize all grade-gender matrix
    for gender in all_genders:
        all_grade_gender_matrix[gender] = {}
        for grade in all_grades:
            all_grade_gender_matrix[gender][grade] = 0
        all_grade_gender_matrix[gender]['No Grade'] = 0
    
    # Process each student for ALL DATA
    for student in students:
        gender = student.get_gender_display() or 'Other'
        if gender not in all_gender_stats:
            all_gender_stats[gender] = {'total': 0, 'with_marks': 0, 'marks_sum': 0, 'average': 0}
        if gender not in all_grade_gender_matrix:
            all_grade_gender_matrix[gender] = {}
            for grade in all_grades:
                all_grade_gender_matrix[gender][grade] = 0
            all_grade_gender_matrix[gender]['No Grade'] = 0
        
        all_gender_stats[gender]['total'] += 1
        
        result = results_by_student.get(student.id)
        marks = result.marks_obtained if result else None
        percentage = result.percentage if result else None
        grade = result.grade if result else None
        grade_point = result.grade_point if result else None
        # Get position from dynamically calculated map
        position = position_map.get(student.id)
        
        # Convert marks to float for calculations
        marks_float = None
        if marks is not None:
            try:
                marks_float = float(marks)
            except (TypeError, ValueError):
                marks_float = None
        
        if marks_float is not None:
            all_students_with_marks += 1
            all_total_marks_sum += marks_float
            all_marks_list.append(marks_float)
            
            all_gender_stats[gender]['with_marks'] += 1
            all_gender_stats[gender]['marks_sum'] += marks_float
            
            if grade:
                all_grade_counts[grade] = all_grade_counts.get(grade, 0) + 1
                all_grade_gender_matrix[gender][grade] = all_grade_gender_matrix[gender].get(grade, 0) + 1
            else:
                all_grade_gender_matrix[gender]['No Grade'] += 1
        
        # Add to all_student_data regardless of marks
        all_student_data.append({
            'id': student.id,
            'registration_number': student.registration_number or f"S{student.id:04d}",
            'full_name': student.full_name,
            'gender': gender,
            'marks': marks_float,
            'percentage': float(percentage) if percentage else None,
            'grade': grade,
            'grade_point': float(grade_point) if grade_point else None,
            'position': position,
            'has_marks': marks_float is not None
        })
    
    # Calculate all gender averages
    for gender in all_gender_stats:
        if all_gender_stats[gender]['with_marks'] > 0:
            all_gender_stats[gender]['average'] = round(
                all_gender_stats[gender]['marks_sum'] / all_gender_stats[gender]['with_marks'], 2
            )
    
    # ============================================
    # FILTERED DATA PROCESSING
    # ============================================
    
    # Start with a copy of all student data
    filtered_student_data = []
    for student in all_student_data:
        # Only include students with marks in filtered data
        if student['has_marks']:
            filtered_student_data.append(student.copy())
    
    # Apply grade filter
    if grade_filter:
        if grade_filter == 'No Grade':
            filtered_student_data = [s for s in filtered_student_data if not s['grade']]
        else:
            filtered_student_data = [s for s in filtered_student_data if s['grade'] == grade_filter]
    
    # Apply gender filter
    if gender_filter:
        filtered_student_data = [
            s for s in filtered_student_data 
            if s['gender'] and s['gender'].lower() == gender_filter.lower()
        ]
    
    # Apply rank filter - this must be applied AFTER grade and gender filters
    if rank_filter == 'top':
        # Get students with position from already filtered data
        students_with_position = [s for s in filtered_student_data if s['position']]
        students_with_position.sort(key=lambda x: x['position'])
        filtered_student_data = students_with_position[:min(top_n, len(students_with_position))]
    elif rank_filter == 'bottom':
        students_with_position = [s for s in filtered_student_data if s['position']]
        students_with_position.sort(key=lambda x: x['position'], reverse=True)
        filtered_student_data = students_with_position[:min(bottom_n, len(students_with_position))]
    
    # ============================================
    # CALCULATE FILTERED STATISTICS
    # ============================================
    
    filtered_students_with_marks = len(filtered_student_data)
    filtered_total_marks_sum = 0
    filtered_marks_list = []
    filtered_gender_stats = {}
    
    # Initialize filtered gender stats
    for gender in all_genders:
        filtered_gender_stats[gender] = {
            'total': 0,
            'with_marks': 0,
            'marks_sum': 0,
            'average': 0
        }
    
    # Process filtered data
    for student in filtered_student_data:
        gender = student['gender']
        if gender not in filtered_gender_stats:
            filtered_gender_stats[gender] = {'total': 0, 'with_marks': 0, 'marks_sum': 0, 'average': 0}
        
        filtered_gender_stats[gender]['total'] += 1
        filtered_gender_stats[gender]['with_marks'] += 1
        
        if student['marks'] is not None:
            filtered_total_marks_sum += student['marks']
            filtered_marks_list.append(student['marks'])
            filtered_gender_stats[gender]['marks_sum'] += student['marks']
    
    # Calculate filtered gender averages
    for gender in filtered_gender_stats:
        if filtered_gender_stats[gender]['with_marks'] > 0:
            filtered_gender_stats[gender]['average'] = round(
                filtered_gender_stats[gender]['marks_sum'] / filtered_gender_stats[gender]['with_marks'], 2
            )
    
    # Calculate filtered statistics
    filtered_statistics = {
        'total_students': len(filtered_student_data),
        'students_with_marks': filtered_students_with_marks,
        'students_without_marks': 0,  # All students in filtered data have marks
        'percentage_completed': 100.0 if filtered_students_with_marks > 0 else 0,
        'average_marks': filtered_total_marks_sum / filtered_students_with_marks if filtered_students_with_marks > 0 else 0,
        'highest_marks': max(filtered_marks_list) if filtered_marks_list else 0,
        'lowest_marks': min(filtered_marks_list) if filtered_marks_list else 0,
        'pass_rate': calculate_pass_rate(filtered_marks_list),
        'median_marks': calculate_median(filtered_marks_list),
        'std_deviation': calculate_std_deviation(filtered_marks_list),
    }
    
    # ============================================
    # OVERALL STATISTICS (ALL STUDENTS)
    # ============================================
    
    overall_statistics = {
        'total_students': total_students,
        'students_with_marks': all_students_with_marks,
        'students_without_marks': total_students - all_students_with_marks,
        'percentage_completed': (all_students_with_marks / total_students * 100) if total_students > 0 else 0,
        'average_marks': all_total_marks_sum / all_students_with_marks if all_students_with_marks > 0 else 0,
        'highest_marks': max(all_marks_list) if all_marks_list else 0,
        'lowest_marks': min(all_marks_list) if all_marks_list else 0,
        'pass_rate': calculate_pass_rate(all_marks_list),
        'median_marks': calculate_median(all_marks_list),
        'std_deviation': calculate_std_deviation(all_marks_list),
    }
    
    # ============================================
    # TOP/BOTTOM PERFORMERS FROM FILTERED DATA
    # ============================================
    
    # Get top performers from filtered data
    top_performers_filtered = []
    students_with_pos_filtered = [s for s in filtered_student_data if s['position']]
    students_with_pos_filtered.sort(key=lambda x: x['position'])
    
    # Limit to 10 or less
    for student in students_with_pos_filtered[:10]:
        top_performers_filtered.append(student)
    
    # Get bottom performers from filtered data
    bottom_performers_filtered = []
    students_with_pos_filtered.sort(key=lambda x: x['position'], reverse=True)
    for student in students_with_pos_filtered[:10]:
        bottom_performers_filtered.append(student)
    
    # ============================================
    # MATRIX DATA (FROM ALL STUDENTS - UNFILTERED)
    # ============================================
    
    # Calculate grade totals for matrix
    grade_totals = {}
    for grade in all_grades + ['No Grade']:
        grade_totals[grade] = 0
        for gender in all_grade_gender_matrix:
            grade_totals[grade] += all_grade_gender_matrix[gender].get(grade, 0)
    
    # Calculate gender totals for matrix
    gender_totals = {}
    for gender in all_grade_gender_matrix:
        gender_totals[gender] = sum(all_grade_gender_matrix[gender].values())
    
    grand_total = sum(gender_totals.values())
    
    # ============================================
    # GRADE DISTRIBUTION DATA (FROM ALL STUDENTS)
    # ============================================
    
    # Get grade distribution for filter dropdown (from all data)
    grade_distribution_all = {}
    for student in all_student_data:
        if student['grade'] and student['has_marks']:
            grade_distribution_all[student['grade']] = grade_distribution_all.get(student['grade'], 0) + 1
    
    # Prepare grade distribution list for summary table (from all data)
    grade_distribution_list = []
    for scale in grading_scales:
        grade_data = {
            'grade': scale.grade,
            'description': scale.get_grade_display().split(' - ')[1] if ' - ' in scale.get_grade_display() else scale.description,
            'range': f"{scale.min_mark}-{scale.max_mark}",
            'count': all_grade_counts.get(scale.grade, 0),
            'percentage': (all_grade_counts.get(scale.grade, 0) / all_students_with_marks * 100) if all_students_with_marks > 0 else 0,
            'male_count': all_grade_gender_matrix.get('Male', {}).get(scale.grade, 0),
            'female_count': all_grade_gender_matrix.get('Female', {}).get(scale.grade, 0),
            'other_count': all_grade_gender_matrix.get('Other', {}).get(scale.grade, 0),
        }
        grade_distribution_list.append(grade_data)
    
    # ============================================
    # SUBJECT COMPARISON DATA
    # ============================================
    
    # Get subject comparison data (overall)
    subject_comparison = get_subject_comparison_data(exam_session)
    
    # Get subject ranking
    subject_ranking = {
        'position': get_subject_ranking(exam_session, selected_subject),
        'total_subjects': len(subject_comparison)
    }
    
    # Get subject statistics for each subject (for subject cards)
    subject_stats = {}
    for subject in all_subjects:
        subject_results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject,
            marks_obtained__isnull=False
        )
        marks_list_subj = []
        for r in subject_results:
            if r.marks_obtained is not None:
                try:
                    marks_list_subj.append(float(r.marks_obtained))
                except (TypeError, ValueError):
                    continue
        
        subject_stats[subject.id] = {
            'students_with_marks': len(marks_list_subj),
            'average_marks': sum(marks_list_subj) / len(marks_list_subj) if marks_list_subj else 0,
            'pass_rate': calculate_pass_rate(marks_list_subj),
        }
    
    # ============================================
    # CONTEXT PREPARATION
    # ============================================
    
    context = {
        # Core objects
        'exam_session': exam_session,
        'all_subjects': all_subjects,
        'selected_subject': selected_subject,
        'subject_id': selected_subject.id,
        
        # FILTERED DATA - used for detailed view and top/bottom performers
        'student_data': filtered_student_data,  # This is for the detailed table
        'filtered_student_data': filtered_student_data,  # Alias for clarity
        'filtered_statistics': filtered_statistics,  # Statistics for filtered data
        'filtered_gender_statistics': filtered_gender_stats,  # Gender stats for filtered data
        'top_performers': top_performers_filtered,  # Top performers from filtered data
        'bottom_performers': bottom_performers_filtered,  # Bottom performers from filtered data
        
        # ALL DATA - used for matrix and overall summaries
        'all_student_data': all_student_data,
        'total_students': total_students,
        'students_with_marks': all_students_with_marks,
        'students_without_marks': total_students - all_students_with_marks,
        'statistics': filtered_statistics if (grade_filter or gender_filter or rank_filter) else overall_statistics,  # Show filtered stats when filters applied
        'overall_statistics': overall_statistics,
        'gender_statistics': all_gender_stats,  # Overall gender stats for filter dropdown
        
        # Grade-Gender Matrix (ALWAYS from all data - unfiltered)
        'grade_gender_matrix': all_grade_gender_matrix,
        'matrix_grades': all_grades + (['No Grade'] if any(all_grade_gender_matrix[g]['No Grade'] > 0 for g in all_grade_gender_matrix) else []),
        'grade_totals': grade_totals,
        'gender_totals': gender_totals,
        'grand_total': grand_total,
        
        # Grade distribution (ALWAYS from all data)
        'grade_distribution_all': grade_distribution_all,
        'grade_distribution_list': grade_distribution_list,
        
        # Subject comparison (ALWAYS from all data)
        'subject_comparison': subject_comparison,
        'subject_ranking': subject_ranking,
        'subject_stats': subject_stats,
        
        # Filter values
        'subject_id': subject_id,
        'grade_filter': grade_filter,
        'gender_filter': gender_filter,
        'rank_filter': rank_filter,
        'top_n': str(top_n),
        'bottom_n': str(bottom_n),
        
        # Filter status flags
        'has_filters': bool(grade_filter or gender_filter or rank_filter),
        
        # Filter options
        'rank_filter_options': [
            ('', 'All Students'),
            ('top', 'Top N Students'),
            ('bottom', 'Bottom N Students'),
        ],
        
        'page_title': f'Subject Performance Analysis - {exam_session.name}',
    }
    
    return render(request, 'admin/results/session_subject_analysis.html', context)


# ============================================
# HELPER FUNCTIONS FOR POSITION CALCULATION
# ============================================

def calculate_subject_positions(students, results_by_student):
    """
    Calculate unique positions for students based on their marks in a subject.
    Returns a dictionary mapping student_id to position (1, 2, 3, etc.)
    No two students share the same position - ties are broken by additional criteria.
    """
    # Collect students with valid marks
    students_with_marks = []
    
    for student in students:
        result = results_by_student.get(student.id)
        if result and result.marks_obtained is not None:
            try:
                marks_float = float(result.marks_obtained)
                students_with_marks.append({
                    'student_id': student.id,
                    'marks': marks_float,
                    'registration_number': student.registration_number or f"S{student.id:04d}",
                    'full_name': student.full_name,
                    'result': result
                })
            except (TypeError, ValueError):
                continue
    
    # Sort by marks (descending) and then by tie-breakers
    students_with_marks.sort(key=lambda x: (
        -x['marks'],  # Primary: Higher marks first
        x['registration_number'],  # Secondary: Registration number (alphabetical)
        x['full_name']  # Tertiary: Full name (alphabetical)
    ))
    
    # Assign unique positions
    position_map = {}
    for position, item in enumerate(students_with_marks, start=1):
        position_map[item['student_id']] = position
    
    return position_map


def get_student_position_in_subject(student_id, subject_id, exam_session_id):
    """
    Helper function to get a specific student's position in a subject.
    Useful for AJAX calls or detailed views.
    """
    try:
        # Get the student
        student = Student.objects.get(id=student_id)
        
        # Get all students in the same class/stream
        student_filters = {
            'class_level_id': student.class_level_id,
            'is_active': True
        }
        if student.stream_class:
            student_filters['stream_class'] = student.stream_class
        
        students = Student.objects.filter(**student_filters)
        
        # Get results for the subject
        results = StudentResult.objects.filter(
            exam_session_id=exam_session_id,
            subject_id=subject_id
        ).select_related('student')
        
        results_by_student = {r.student_id: r for r in results}
        
        # Calculate positions
        position_map = calculate_subject_positions(students, results_by_student)
        
        return position_map.get(student_id)
        
    except (Student.DoesNotExist, StudentResult.DoesNotExist):
        return None


def get_subject_rankings_list(exam_session_id, subject_id):
    """
    Helper function to get a complete ranking list for a subject.
    Returns a list of students with their marks and positions.
    """
    try:
        exam_session = ExamSession.objects.get(id=exam_session_id)
        
        # Get students
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }
        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class
        
        students = Student.objects.filter(**student_filters)
        
        # Get results
        results = StudentResult.objects.filter(
            exam_session_id=exam_session_id,
            subject_id=subject_id
        ).select_related('student')
        
        results_by_student = {r.student_id: r for r in results}
        
        # Calculate positions
        position_map = calculate_subject_positions(students, results_by_student)
        
        # Build ranking list
        ranking_list = []
        for student in students:
            result = results_by_student.get(student.id)
            position = position_map.get(student.id)
            
            if result and result.marks_obtained is not None:
                ranking_list.append({
                    'position': position,
                    'student_id': student.id,
                    'registration_number': student.registration_number or f"S{student.id:04d}",
                    'full_name': student.full_name,
                    'marks': float(result.marks_obtained),
                    'grade': result.grade,
                    'percentage': float(result.percentage) if result.percentage else None,
                })
        
        # Sort by position
        ranking_list.sort(key=lambda x: x['position'])
        
        return ranking_list
        
    except ExamSession.DoesNotExist:
        return []
    

@login_required
def session_subject_analysis_pdf(request, exam_session_id):
    """
    Generate PDF report for session-based subject performance analysis.
    """
    try:
        # Get exam session
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type', 'academic_year', 'term', 'class_level',
                'class_level__educational_level', 'stream_class'
            ),
            id=exam_session_id
        )
        
        # Get filter parameters
        section = request.GET.get('section', 'full')
        subject_id = request.GET.get('subject_id', '')
        grade_filter = request.GET.get('grade_filter', '')
        gender_filter = request.GET.get('gender', '')
        rank_filter = request.GET.get('rank_filter', '')
        top_n = request.GET.get('top_n', '10')
        bottom_n = request.GET.get('bottom_n', '10')
        
        # Convert to integers where applicable
        try:
            top_n = int(top_n) if top_n else 10
        except ValueError:
            top_n = 10
        
        try:
            bottom_n = int(bottom_n) if bottom_n else 10
        except ValueError:
            bottom_n = 10
        
        # Get all subjects
        all_subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('name')
        
        # Get selected subject
        selected_subject = None
        if subject_id:
            try:
                selected_subject = all_subjects.get(id=subject_id)
            except Subject.DoesNotExist:
                selected_subject = all_subjects.first()
        else:
            selected_subject = all_subjects.first()
        
        if not selected_subject:
            messages.error(request, "No subjects available for this educational level.")
            return redirect('session_subject_analysis_view', exam_session_id=exam_session_id)
        
        # Get students
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }
        
        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class
        
        students = Student.objects.filter(**student_filters)
        total_students = students.count()
        
        # Get results for selected subject
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=selected_subject
        ).select_related('student')
        
        results_by_student = {r.student_id: r for r in results}
        
        # ============================================
        # CALCULATE POSITIONS DYNAMICALLY USING HELPER FUNCTION
        # ============================================
        position_map = calculate_subject_positions(students, results_by_student)
        
        # ============================================
        # DATA STRUCTURES INITIALIZATION
        # ============================================
        
        # Prepare data structures
        all_student_data = []
        students_with_marks = 0
        total_marks_sum = 0
        marks_list = []
        gender_stats = {}
        grade_gender_matrix = {}
        grade_counts = {}
        
        # Get grading scales
        grading_scales = GradingScale.objects.filter(
            education_level=exam_session.class_level.educational_level
        ).order_by('min_mark')
        all_grades = [scale.grade for scale in grading_scales]
        
        # Initialize
        all_genders = ['Male', 'Female', 'Other']
        for gender in all_genders:
            gender_stats[gender] = {'total': 0, 'with_marks': 0, 'marks_sum': 0, 'average': 0}
            grade_gender_matrix[gender] = {grade: 0 for grade in all_grades}
            grade_gender_matrix[gender]['No Grade'] = 0
        
        # Process students
        for student in students:
            gender = student.get_gender_display() or 'Other'
            if gender not in gender_stats:
                gender_stats[gender] = {'total': 0, 'with_marks': 0, 'marks_sum': 0, 'average': 0}
            if gender not in grade_gender_matrix:
                grade_gender_matrix[gender] = {grade: 0 for grade in all_grades}
                grade_gender_matrix[gender]['No Grade'] = 0
            
            gender_stats[gender]['total'] += 1
            
            result = results_by_student.get(student.id)
            marks = result.marks_obtained if result else None
            percentage = result.percentage if result else None
            grade = result.grade if result else None
            grade_point = result.grade_point if result else None
            # Get position from dynamically calculated map
            position = position_map.get(student.id)
            
            # Convert marks to float
            marks_float = None
            if marks is not None:
                try:
                    marks_float = float(marks)
                except (TypeError, ValueError):
                    marks_float = None
            
            if marks_float is not None:
                students_with_marks += 1
                total_marks_sum += marks_float
                marks_list.append(marks_float)
                
                gender_stats[gender]['with_marks'] += 1
                gender_stats[gender]['marks_sum'] += marks_float
                
                if grade:
                    grade_counts[grade] = grade_counts.get(grade, 0) + 1
                    grade_gender_matrix[gender][grade] += 1
                else:
                    grade_gender_matrix[gender]['No Grade'] += 1
            
            all_student_data.append({
                'id': student.id,
                'registration_number': student.registration_number or f"S{student.id:04d}",
                'full_name': student.full_name,
                'gender': gender,
                'marks': marks_float,
                'percentage': float(percentage) if percentage else None,
                'grade': grade,
                'grade_point': float(grade_point) if grade_point else None,
                'position': position,
                'has_marks': marks_float is not None
            })
        
        # Calculate averages
        for gender in gender_stats:
            if gender_stats[gender]['with_marks'] > 0:
                gender_stats[gender]['average'] = round(
                    gender_stats[gender]['marks_sum'] / gender_stats[gender]['with_marks'], 2
                )
        
        # ============================================
        # FILTERED DATA PROCESSING
        # ============================================
        
        # Start with a copy of all student data with marks
        filtered_data = [s.copy() for s in all_student_data if s['has_marks']]
        
        # Apply grade filter
        if grade_filter:
            if grade_filter == 'No Grade':
                filtered_data = [s for s in filtered_data if not s['grade']]
            else:
                filtered_data = [s for s in filtered_data if s['grade'] == grade_filter]
        
        # Apply gender filter
        if gender_filter:
            filtered_data = [
                s for s in filtered_data 
                if s['gender'] and s['gender'].lower() == gender_filter.lower()
            ]
        
        # Apply rank filter
        if rank_filter == 'top':
            students_with_position = [s for s in filtered_data if s['position']]
            students_with_position.sort(key=lambda x: x['position'])
            filtered_data = students_with_position[:min(top_n, len(students_with_position))]
        elif rank_filter == 'bottom':
            students_with_position = [s for s in filtered_data if s['position']]
            students_with_position.sort(key=lambda x: x['position'], reverse=True)
            filtered_data = students_with_position[:min(bottom_n, len(students_with_position))]
        
        # ============================================
        # TOP/BOTTOM PERFORMERS (FROM ALL DATA)
        # ============================================
        
        # Get top performers from all data
        top_performers = []
        students_with_pos_all = [s for s in all_student_data if s['position']]
        students_with_pos_all.sort(key=lambda x: x['position'])
        for student in students_with_pos_all[:10]:
            if student['has_marks']:
                top_performers.append(student)
        
        # Get bottom performers from all data
        bottom_performers = []
        students_with_pos_all.sort(key=lambda x: x['position'], reverse=True)
        for student in students_with_pos_all[:10]:
            if student['has_marks']:
                bottom_performers.append(student)
        
        # ============================================
        # CALCULATE STATISTICS
        # ============================================
        
        # Overall statistics
        overall_statistics = {
            'total_students': total_students,
            'students_with_marks': students_with_marks,
            'students_without_marks': total_students - students_with_marks,
            'percentage_completed': (students_with_marks / total_students * 100) if total_students > 0 else 0,
            'average_marks': total_marks_sum / students_with_marks if students_with_marks > 0 else 0,
            'highest_marks': max(marks_list) if marks_list else 0,
            'lowest_marks': min(marks_list) if marks_list else 0,
            'pass_rate': calculate_pass_rate(marks_list),
            'median_marks': calculate_median(marks_list),
            'std_deviation': calculate_std_deviation(marks_list),
        }
        
        # Filtered statistics (if filters applied)
        filtered_statistics = None
        if grade_filter or gender_filter or rank_filter:
            filtered_marks_list = [s['marks'] for s in filtered_data if s['marks'] is not None]
            filtered_statistics = {
                'total_students': len(filtered_data),
                'students_with_marks': len([s for s in filtered_data if s['marks'] is not None]),
                'students_without_marks': 0,
                'percentage_completed': 100.0 if filtered_data else 0,
                'average_marks': sum(filtered_marks_list) / len(filtered_marks_list) if filtered_marks_list else 0,
                'highest_marks': max(filtered_marks_list) if filtered_marks_list else 0,
                'lowest_marks': min(filtered_marks_list) if filtered_marks_list else 0,
                'pass_rate': calculate_pass_rate(filtered_marks_list),
                'median_marks': calculate_median(filtered_marks_list),
                'std_deviation': calculate_std_deviation(filtered_marks_list),
            }
        
        # ============================================
        # MATRIX DATA
        # ============================================
        
        # Calculate matrix totals
        grade_totals = {}
        for grade in all_grades + ['No Grade']:
            grade_totals[grade] = 0
            for gender in grade_gender_matrix:
                grade_totals[grade] += grade_gender_matrix[gender].get(grade, 0)
        
        gender_totals = {}
        for gender in grade_gender_matrix:
            gender_totals[gender] = sum(grade_gender_matrix[gender].values())
        
        grand_total = sum(gender_totals.values())
        
        # ============================================
        # GRADE DISTRIBUTION
        # ============================================
        
        # Grade distribution list
        grade_distribution_list = []
        for scale in grading_scales:
            grade_data = {
                'grade': scale.grade,
                'description': scale.get_grade_display().split(' - ')[1] if ' - ' in scale.get_grade_display() else scale.description,
                'range': f"{scale.min_mark}-{scale.max_mark}",
                'count': grade_counts.get(scale.grade, 0),
                'percentage': (grade_counts.get(scale.grade, 0) / students_with_marks * 100) if students_with_marks > 0 else 0,
                'male_count': grade_gender_matrix.get('Male', {}).get(scale.grade, 0),
                'female_count': grade_gender_matrix.get('Female', {}).get(scale.grade, 0),
                'other_count': grade_gender_matrix.get('Other', {}).get(scale.grade, 0),
            }
            grade_distribution_list.append(grade_data)
        
        # ============================================
        # SUBJECT COMPARISON
        # ============================================
        
        subject_comparison = get_subject_comparison_data(exam_session)
        subject_ranking = {
            'position': get_subject_ranking(exam_session, selected_subject),
            'total_subjects': len(subject_comparison)
        }
        
        # ============================================
        # DETERMINE DISPLAY DATA BASED ON SECTION
        # ============================================
        
        display_data = filtered_data
        if section == 'top_performers':
            display_data = top_performers
        elif section == 'bottom_performers':
            display_data = bottom_performers
        elif section == 'matrix':
            display_data = all_student_data
        elif section == 'comparison':
            display_data = filtered_data
        elif section == 'detailed':
            display_data = filtered_data
        elif section == 'grades':
            display_data = filtered_data
        
        # Determine which statistics to use
        active_statistics = filtered_statistics if (grade_filter or gender_filter or rank_filter) else overall_statistics
        
        # ============================================
        # CONTEXT PREPARATION
        # ============================================
        
        # Prepare matrix grades list
        matrix_grades = all_grades + (['No Grade'] if any(grade_gender_matrix[g]['No Grade'] > 0 for g in grade_gender_matrix) else [])
        
        context = {
            # Core objects
            'exam_session': exam_session,
            'all_subjects': all_subjects,
            'selected_subject': selected_subject,
            'subject_id': selected_subject.id,
            
            # Data
            'student_data': display_data,
            'filtered_student_data': display_data,
            'all_student_data': all_student_data,
            
            # Statistics
            'statistics': active_statistics,
            'overall_statistics': overall_statistics,
            
            # Performers
            'top_performers': top_performers,
            'bottom_performers': bottom_performers,
            
            # Matrix data
            'grade_gender_matrix': grade_gender_matrix,
            'matrix_grades': matrix_grades,
            'grade_totals': grade_totals,
            'gender_totals': gender_totals,
            'grand_total': grand_total,
            
            # Gender statistics
            'gender_statistics': gender_stats,
            
            # Grade distribution
            'grade_distribution_list': grade_distribution_list,
            
            # Subject comparison
            'subject_comparison': subject_comparison,
            'subject_ranking': subject_ranking,
            
            # Counts
            'students_with_marks': students_with_marks,
            'total_students': total_students,
            
            # Section and filters
            'section': section,
            'subject_id': subject_id,
            'grade_filter': grade_filter,
            'gender_filter': gender_filter,
            'rank_filter': rank_filter,
            'top_n': str(top_n),
            'bottom_n': str(bottom_n),
            'has_filters': bool(grade_filter or gender_filter or rank_filter),
            
            # PDF metadata
            'generated_date': timezone.now(),
            'generated_by': request.user.get_full_name() or request.user.username,
            'school_name': getattr(settings, 'SCHOOL_NAME', 'School Management System'),
            'school_address': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'is_pdf': True,
        }
        
        # Select template based on section
        template_mapping = {
            'full': 'admin/results/session_subject_analysis_pdf_full.html',
            'overview': 'admin/results/session_subject_analysis_pdf_full.html',
            'matrix': 'admin/results/session_subject_analysis_pdf_full.html',
            'top_performers': 'admin/results/session_subject_analysis_pdf_full.html',
            'bottom_performers': 'admin/results/session_subject_analysis_pdf_full.html',
            'comparison': 'admin/results/session_subject_analysis_pdf_full.html',
            'detailed': 'admin/results/session_subject_analysis_pdf_full.html',
            'grades': 'admin/results/session_subject_analysis_pdf_full.html',
        }
        
        template_name = template_mapping.get(section, 'admin/results/session_subject_analysis_pdf_full.html')
        
        # Generate PDF
        html_string = render_to_string(template_name, context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = html.write_pdf()
        
        # Generate filename
        filename_parts = [
            f"Subject_{selected_subject.code}",
            exam_session.name.replace(' ', '_'),
            section.capitalize() if section != 'full' else 'Full_Report'
        ]
        
        if grade_filter:
            filename_parts.append(f"Grade_{grade_filter}")
        if gender_filter:
            filename_parts.append(f"Gender_{gender_filter}")
        if rank_filter:
            filename_parts.append(f"{rank_filter}_{top_n if rank_filter == 'top' else bottom_n}")
        
        filename_parts.append(timezone.now().strftime('%Y%m%d'))
        filename = "_".join(filename_parts) + ".pdf"
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating subject analysis PDF: {str(e)}")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
        messages.error(request, f"Error generating PDF report: {str(e)}")
        return redirect('session_subject_analysis_view', exam_session_id=exam_session_id)
    

@login_required
def ajax_subject_performance(request, exam_session_id, subject_id):
    """
    AJAX endpoint to get subject performance data for dynamic updates.
    """
    try:
        exam_session = get_object_or_404(ExamSession, id=exam_session_id)
        subject = get_object_or_404(Subject, id=subject_id)
        
        # Get results
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject,
            marks_obtained__isnull=False
        )
        
        # Calculate performance metrics - convert Decimal to float
        marks_list = []
        for r in results:
            if r.marks_obtained is not None:
                try:
                    marks_list.append(float(r.marks_obtained))
                except (TypeError, ValueError):
                    continue
        
        performance_data = {
            'subject_id': subject.id,
            'subject_code': subject.code,
            'subject_name': subject.name,
            'total_students': len(marks_list),
            'average_marks': sum(marks_list) / len(marks_list) if marks_list else 0,
            'highest_marks': max(marks_list) if marks_list else 0,
            'lowest_marks': min(marks_list) if marks_list else 0,
            'pass_rate': calculate_pass_rate(marks_list),
            'median_marks': calculate_median(marks_list),
            'std_deviation': calculate_std_deviation(marks_list),
        }
        
        # Grade distribution
        grade_counts = {}
        for result in results:
            if result.grade:
                grade_counts[result.grade] = grade_counts.get(result.grade, 0) + 1
        
        performance_data['grade_distribution'] = grade_counts
        
        return JsonResponse({
            'success': True,
            'data': performance_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
    

@login_required
def ajax_subject_comparison(request, exam_session_id):
    """
    AJAX endpoint to get comparison data for all subjects.
    """
    try:
        exam_session = get_object_or_404(ExamSession, id=exam_session_id)
        
        subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        )
        
        comparison_data = []
        
        for subject in subjects:
            results = StudentResult.objects.filter(
                exam_session=exam_session,
                subject=subject,
                marks_obtained__isnull=False
            )
            
            marks_list = [float(r.marks_obtained) for r in results]
            
            if marks_list:
                pass_count = sum(1 for m in marks_list if m >= 40)
                pass_rate = (pass_count / len(marks_list)) * 100 if marks_list else 0
                
                grade_a_count = results.filter(grade='A').count()
                grade_f_count = results.filter(grade='F').count()
                
                comparison_data.append({
                    'subject_id': subject.id,
                    'subject_code': subject.code,
                    'subject_name': subject.name,
                    'students_count': len(marks_list),
                    'average_marks': sum(marks_list) / len(marks_list),
                    'highest_marks': max(marks_list),
                    'lowest_marks': min(marks_list),
                    'pass_rate': pass_rate,
                    'grade_a_count': grade_a_count,
                    'grade_f_count': grade_f_count,
                })
        
        # Sort by average marks descending
        comparison_data.sort(key=lambda x: x['average_marks'], reverse=True)
        
        # Add ranking positions
        for idx, subject in enumerate(comparison_data, start=1):
            subject['ranking_position'] = idx
            subject['average_percentage'] = (subject['average_marks'] / 
                                           exam_session.exam_type.max_score * 100) if exam_session.exam_type.max_score else 0
        
        return JsonResponse({
            'success': True,
            'data': comparison_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def ajax_subject_list(request, exam_session_id):
    """
    AJAX endpoint to get list of subjects for dropdown.
    """
    try:
        exam_session = get_object_or_404(ExamSession, id=exam_session_id)
        
        subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('name')
        
        subject_list = []
        for subject in subjects:
            # Get basic stats for each subject
            results = StudentResult.objects.filter(
                exam_session=exam_session,
                subject=subject,
                marks_obtained__isnull=False
            )
            marks_list = [float(r.marks_obtained) for r in results]
            
            subject_list.append({
                'id': subject.id,
                'name': subject.name,
                'code': subject.code,
                'students_count': results.count(),
                'average_marks': sum(marks_list) / len(marks_list) if marks_list else 0,
            })
        
        return JsonResponse({
            'success': True,
            'subjects': subject_list
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ============= HELPER FUNCTIONS =============

def get_subject_comparison_data(exam_session):
    """
    Helper function to get comparison data for all subjects.
    Handles Decimal/float conversion properly.
    """
    subjects = Subject.objects.filter(
        educational_level=exam_session.class_level.educational_level,
        is_active=True
    )
    
    comparison_data = []
    
    for subject in subjects:
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject,
            marks_obtained__isnull=False
        )
        
        # Convert Decimal marks to float
        marks_list = []
        for r in results:
            if r.marks_obtained is not None:
                try:
                    marks_list.append(float(r.marks_obtained))
                except (TypeError, ValueError):
                    continue
        
        if marks_list:
            pass_count = sum(1 for m in marks_list if m >= 40)
            pass_rate = (pass_count / len(marks_list)) * 100 if marks_list else 0
            
            grade_a_count = results.filter(grade='A').count()
            grade_f_count = results.filter(grade='F').count()
            
            # Convert max_score to float to avoid Decimal/float division issues
            max_score = float(exam_session.exam_type.max_score) if exam_session.exam_type.max_score else 100.0
            
            comparison_data.append({
                'subject_id': subject.id,
                'subject_code': subject.code,
                'subject_name': subject.name,
                'students_count': len(marks_list),
                'average_marks': sum(marks_list) / len(marks_list),
                'highest_marks': max(marks_list),
                'lowest_marks': min(marks_list),
                'pass_rate': pass_rate,
                'grade_a_count': grade_a_count,
                'grade_f_count': grade_f_count,
                'average_percentage': (sum(marks_list) / len(marks_list) / max_score * 100),
            })
    
    # Sort by average marks descending
    comparison_data.sort(key=lambda x: x['average_marks'], reverse=True)
    
    # Add ranking positions
    for idx, subject in enumerate(comparison_data, start=1):
        subject['ranking_position'] = idx
    
    return comparison_data


def get_subject_ranking(exam_session, target_subject):
    """
    Helper function to get ranking position of a subject.
    """
    subjects = Subject.objects.filter(
        educational_level=exam_session.class_level.educational_level,
        is_active=True
    )
    
    subject_averages = []
    
    for subject in subjects:
        avg = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=subject,
            marks_obtained__isnull=False
        ).aggregate(avg_marks=Avg('marks_obtained'))['avg_marks']
        
        if avg:
            subject_averages.append({
                'subject_id': subject.id,
                'average': avg
            })
    
    subject_averages.sort(key=lambda x: x['average'], reverse=True)
    
    for idx, subject in enumerate(subject_averages, start=1):
        if subject['subject_id'] == target_subject.id:
            return idx
    
    return None


def calculate_pass_rate(marks_list):
    """
    Calculate pass rate based on 40% pass mark.
    Handles Decimal values by converting to float.
    """
    if not marks_list:
        return 0
    
    # Convert any Decimal values to float
    float_marks = []
    for m in marks_list:
        try:
            float_marks.append(float(m))
        except (TypeError, ValueError):
            continue
    
    if not float_marks:
        return 0
    
    pass_mark = 40
    passed = sum(1 for m in float_marks if m >= pass_mark)
    return (passed / len(float_marks)) * 100


def calculate_median(marks_list):
    """
    Calculate median of marks list.
    Handles Decimal values by converting to float.
    """
    if not marks_list:
        return 0
    
    # Convert any Decimal values to float
    float_marks = []
    for m in marks_list:
        try:
            float_marks.append(float(m))
        except (TypeError, ValueError):
            continue
    
    if not float_marks:
        return 0
    
    sorted_marks = sorted(float_marks)
    n = len(sorted_marks)
    
    if n % 2 == 0:
        return (sorted_marks[n//2 - 1] + sorted_marks[n//2]) / 2
    else:
        return sorted_marks[n//2]


def calculate_std_deviation(marks_list):
    """
    Calculate standard deviation of marks list.
    Handles Decimal values by converting to float.
    """
    if not marks_list or len(marks_list) < 2:
        return 0
    
    # Convert any Decimal values to float
    float_marks = []
    for m in marks_list:
        try:
            float_marks.append(float(m))
        except (TypeError, ValueError):
            continue
    
    if len(float_marks) < 2:
        return 0
    
    n = len(float_marks)
    mean = sum(float_marks) / n
    variance = sum((x - mean) ** 2 for x in float_marks) / n
    
    return variance ** 0.5


@login_required
def session_subject_matrix_analysis(request, exam_session_id):
    """
    Simplified subject analysis view showing only:
    1. Grade × Gender Cross-Analysis Matrix
    2. Detailed Student Analysis with filters (grade, marks range, gender, top/bottom N)
    """
    try:
        # Get exam session with related data
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type',
                'academic_year',
                'term',
                'class_level',
                'class_level__educational_level',
                'stream_class'
            ),
            id=exam_session_id
        )

        # Get filter parameters
        subject_id = request.GET.get('subject_id', '')
        grade_filter = request.GET.get('grade_filter', '')
        marks_min = request.GET.get('marks_min', '')
        marks_max = request.GET.get('marks_max', '')
        gender_filter = request.GET.get('gender', '')
        rank_filter = request.GET.get('rank_filter', '')
        top_n = request.GET.get('top_n', '10')
        bottom_n = request.GET.get('bottom_n', '10')

        # Convert to integers where applicable
        try:
            top_n = int(top_n) if top_n else 10
        except ValueError:
            top_n = 10

        try:
            bottom_n = int(bottom_n) if bottom_n else 10
        except ValueError:
            bottom_n = 10

        try:
            marks_min = float(marks_min) if marks_min else None
        except ValueError:
            marks_min = None

        try:
            marks_max = float(marks_max) if marks_max else None
        except ValueError:
            marks_max = None

        # Get all active subjects for this educational level
        all_subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('name')

        # Get the selected subject (default to first subject if none selected)
        selected_subject = None
        if subject_id:
            try:
                selected_subject = all_subjects.get(id=subject_id)
            except Subject.DoesNotExist:
                selected_subject = all_subjects.first()
        else:
            selected_subject = all_subjects.first()

        # If no subjects exist, return early
        if not selected_subject:
            context = {
                'exam_session': exam_session,
                'all_subjects': all_subjects,
                'selected_subject': None,
                'no_subjects': True,
                'page_title': f'Subject Matrix Analysis - {exam_session.name}',
            }
            return render(request, 'admin/results/session_subject_matrix_analysis.html', context)

        # Get students based on stream
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }

        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class

        students = Student.objects.filter(**student_filters)
        total_students = students.count()

        # Get results for the selected subject
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=selected_subject
        ).select_related('student')

        results_by_student = {r.student_id: r for r in results}

        # ============================================
        # CALCULATE POSITIONS DYNAMICALLY
        # ============================================
        position_map = calculate_subject_positions(students, results_by_student)

        # ============================================
        # DATA STRUCTURES INITIALIZATION
        # ============================================

        # Get all possible grades from GradingScale
        grading_scales = GradingScale.objects.filter(
            education_level=exam_session.class_level.educational_level
        ).order_by('min_mark')
        all_grades = [scale.grade for scale in grading_scales]

        # Gender tracking
        all_genders = ['Male', 'Female', 'Other']

        # Initialize ALL STUDENTS data structures (unfiltered)
        all_student_data = []
        students_with_marks = 0
        total_marks_sum = 0
        marks_list = []
        gender_stats = {}
        grade_gender_matrix = {}
        grade_counts = {}

        # Initialize gender stats
        for gender in all_genders:
            gender_stats[gender] = {
                'total': 0,
                'with_marks': 0,
                'marks_sum': 0,
                'average': 0
            }

        # Initialize grade-gender matrix
        for gender in all_genders:
            grade_gender_matrix[gender] = {}
            for grade in all_grades:
                grade_gender_matrix[gender][grade] = 0
            grade_gender_matrix[gender]['No Grade'] = 0

        # Process each student for ALL DATA
        for student in students:
            gender = student.get_gender_display() or 'Other'
            if gender not in gender_stats:
                gender_stats[gender] = {'total': 0, 'with_marks': 0, 'marks_sum': 0, 'average': 0}
            if gender not in grade_gender_matrix:
                grade_gender_matrix[gender] = {}
                for grade in all_grades:
                    grade_gender_matrix[gender][grade] = 0
                grade_gender_matrix[gender]['No Grade'] = 0

            gender_stats[gender]['total'] += 1

            result = results_by_student.get(student.id)
            marks = result.marks_obtained if result else None
            percentage = result.percentage if result else None
            grade = result.grade if result else None
            grade_point = result.grade_point if result else None
            position = position_map.get(student.id)

            # Convert marks to float
            marks_float = None
            if marks is not None:
                try:
                    marks_float = float(marks)
                except (TypeError, ValueError):
                    marks_float = None

            if marks_float is not None:
                students_with_marks += 1
                total_marks_sum += marks_float
                marks_list.append(marks_float)

                gender_stats[gender]['with_marks'] += 1
                gender_stats[gender]['marks_sum'] += marks_float

                if grade:
                    grade_counts[grade] = grade_counts.get(grade, 0) + 1
                    grade_gender_matrix[gender][grade] += 1
                else:
                    grade_gender_matrix[gender]['No Grade'] += 1

            all_student_data.append({
                'id': student.id,
                'registration_number': student.registration_number or f"S{student.id:04d}",
                'full_name': student.full_name,
                'gender': gender,
                'marks': marks_float,
                'percentage': float(percentage) if percentage else None,
                'grade': grade,
                'grade_point': float(grade_point) if grade_point else None,
                'position': position,
                'has_marks': marks_float is not None
            })

        # Calculate gender averages
        for gender in gender_stats:
            if gender_stats[gender]['with_marks'] > 0:
                gender_stats[gender]['average'] = round(
                    gender_stats[gender]['marks_sum'] / gender_stats[gender]['with_marks'], 2
                )

        # ============================================
        # FILTERED DATA PROCESSING
        # ============================================

        # Start with a copy of all student data with marks
        filtered_student_data = []
        for student in all_student_data:
            if student['has_marks']:
                filtered_student_data.append(student.copy())

        # Apply grade filter
        if grade_filter:
            if grade_filter == 'No Grade':
                filtered_student_data = [s for s in filtered_student_data if not s['grade']]
            else:
                filtered_student_data = [s for s in filtered_student_data if s['grade'] == grade_filter]

        # Apply marks range filter
        if marks_min is not None:
            filtered_student_data = [s for s in filtered_student_data if s['marks'] >= marks_min]
        if marks_max is not None:
            filtered_student_data = [s for s in filtered_student_data if s['marks'] <= marks_max]

        # Apply gender filter
        if gender_filter:
            filtered_student_data = [
                s for s in filtered_student_data 
                if s['gender'] and s['gender'].lower() == gender_filter.lower()
            ]

        # Apply rank filter
        if rank_filter == 'top':
            students_with_position = [s for s in filtered_student_data if s['position']]
            students_with_position.sort(key=lambda x: x['position'])
            filtered_student_data = students_with_position[:min(top_n, len(students_with_position))]
        elif rank_filter == 'bottom':
            students_with_position = [s for s in filtered_student_data if s['position']]
            students_with_position.sort(key=lambda x: x['position'], reverse=True)
            filtered_student_data = students_with_position[:min(bottom_n, len(students_with_position))]

        # ============================================
        # CALCULATE FILTERED STATISTICS
        # ============================================

        filtered_marks_list = [s['marks'] for s in filtered_student_data if s['marks'] is not None]
        
        statistics = {
            'total_students': len(filtered_student_data),
            'students_with_marks': len(filtered_student_data),
            'students_without_marks': 0,
            'percentage_completed': 100.0 if filtered_student_data else 0,
            'average_marks': sum(filtered_marks_list) / len(filtered_marks_list) if filtered_marks_list else 0,
            'highest_marks': max(filtered_marks_list) if filtered_marks_list else 0,
            'lowest_marks': min(filtered_marks_list) if filtered_marks_list else 0,
            'pass_rate': calculate_pass_rate(filtered_marks_list),
            'median_marks': calculate_median(filtered_marks_list),
            'std_deviation': calculate_std_deviation(filtered_marks_list),
        }

        # ============================================
        # MATRIX DATA (FROM ALL STUDENTS - UNFILTERED)
        # ============================================

        # Calculate grade totals for matrix
        grade_totals = {}
        for grade in all_grades + ['No Grade']:
            grade_totals[grade] = 0
            for gender in grade_gender_matrix:
                grade_totals[grade] += grade_gender_matrix[gender].get(grade, 0)

        # Calculate gender totals for matrix
        gender_totals = {}
        for gender in grade_gender_matrix:
            gender_totals[gender] = sum(grade_gender_matrix[gender].values())

        grand_total = sum(gender_totals.values())

        # Determine matrix grades (include No Grade if exists)
        matrix_grades = all_grades.copy()
        if any(grade_gender_matrix[g]['No Grade'] > 0 for g in grade_gender_matrix):
            matrix_grades.append('No Grade')

        # ============================================
        # PREPARE CONTEXT
        # ============================================

        context = {
            # Core objects
            'exam_session': exam_session,
            'all_subjects': all_subjects,
            'selected_subject': selected_subject,
            'subject_id': selected_subject.id,

            # Data
            'filtered_student_data': filtered_student_data,
            'all_student_data': all_student_data,

            # Statistics
            'statistics': statistics,
            'students_with_marks': students_with_marks,
            'total_students': total_students,

            # Matrix data
            'grade_gender_matrix': grade_gender_matrix,
            'matrix_grades': matrix_grades,
            'grade_totals': grade_totals,
            'gender_totals': gender_totals,
            'grand_total': grand_total,

            # Gender statistics
            'gender_statistics': gender_stats,

            # Filter values
            'subject_id': subject_id,
            'grade_filter': grade_filter,
            'marks_min': marks_min if marks_min is not None else '',
            'marks_max': marks_max if marks_max is not None else '',
            'gender_filter': gender_filter,
            'rank_filter': rank_filter,
            'top_n': str(top_n),
            'bottom_n': str(bottom_n),

            # Filter status flags
            'has_filters': bool(grade_filter or marks_min or marks_max or gender_filter or rank_filter),

            # Filter options
            'rank_filter_options': [
                ('', 'All Students'),
                ('top', 'Top N Students'),
                ('bottom', 'Bottom N Students'),
            ],
            'all_genders': all_genders,
            'all_grades': all_grades,

            'page_title': f'Subject Matrix Analysis - {exam_session.name}',
        }

        return render(request, 'admin/results/session_subject_matrix_analysis.html', context)

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in session_subject_matrix_analysis: {str(e)}")
        messages.error(request, f"Error loading subject matrix analysis: {str(e)}")
        return redirect('exam_session_analysis_view', exam_session_id=exam_session_id)    
    

@login_required
def session_subject_matrix_analysis_pdf(request, exam_session_id):
    """
    Generate PDF report for subject matrix analysis.
    """
    try:
        # Get exam session
        exam_session = get_object_or_404(
            ExamSession.objects.select_related(
                'exam_type', 'academic_year', 'term', 'class_level',
                'class_level__educational_level', 'stream_class'
            ),
            id=exam_session_id
        )

        # Get filter parameters
        subject_id = request.GET.get('subject_id', '')
        grade_filter = request.GET.get('grade_filter', '')
        marks_min = request.GET.get('marks_min', '')
        marks_max = request.GET.get('marks_max', '')
        gender_filter = request.GET.get('gender', '')
        rank_filter = request.GET.get('rank_filter', '')
        top_n = request.GET.get('top_n', '10')
        bottom_n = request.GET.get('bottom_n', '10')

        # Convert to integers where applicable
        try:
            top_n = int(top_n) if top_n else 10
        except ValueError:
            top_n = 10

        try:
            bottom_n = int(bottom_n) if bottom_n else 10
        except ValueError:
            bottom_n = 10

        try:
            marks_min = float(marks_min) if marks_min else None
        except ValueError:
            marks_min = None

        try:
            marks_max = float(marks_max) if marks_max else None
        except ValueError:
            marks_max = None

        # Get all subjects
        all_subjects = Subject.objects.filter(
            educational_level=exam_session.class_level.educational_level,
            is_active=True
        ).order_by('name')

        # Get selected subject
        selected_subject = None
        if subject_id:
            try:
                selected_subject = all_subjects.get(id=subject_id)
            except Subject.DoesNotExist:
                selected_subject = all_subjects.first()
        else:
            selected_subject = all_subjects.first()

        if not selected_subject:
            messages.error(request, "No subjects available for this educational level.")
            return redirect('session_subject_matrix_analysis', exam_session_id=exam_session_id)

        # Get students
        student_filters = {
            'class_level': exam_session.class_level,
            'is_active': True
        }

        if exam_session.stream_class:
            student_filters['stream_class'] = exam_session.stream_class

        students = Student.objects.filter(**student_filters)
        total_students = students.count()

        # Get results for selected subject
        results = StudentResult.objects.filter(
            exam_session=exam_session,
            subject=selected_subject
        ).select_related('student')

        results_by_student = {r.student_id: r for r in results}

        # Calculate positions
        position_map = calculate_subject_positions(students, results_by_student)

        # Get grading scales
        grading_scales = GradingScale.objects.filter(
            education_level=exam_session.class_level.educational_level
        ).order_by('min_mark')
        all_grades = [scale.grade for scale in grading_scales]

        # Initialize data structures
        all_genders = ['Male', 'Female', 'Other']
        all_student_data = []
        students_with_marks = 0
        grade_gender_matrix = {}
        grade_counts = {}

        # Initialize matrix
        for gender in all_genders:
            grade_gender_matrix[gender] = {}
            for grade in all_grades:
                grade_gender_matrix[gender][grade] = 0
            grade_gender_matrix[gender]['No Grade'] = 0

        # Process students
        for student in students:
            gender = student.get_gender_display() or 'Other'
            if gender not in grade_gender_matrix:
                grade_gender_matrix[gender] = {}
                for grade in all_grades:
                    grade_gender_matrix[gender][grade] = 0
                grade_gender_matrix[gender]['No Grade'] = 0

            result = results_by_student.get(student.id)
            marks = result.marks_obtained if result else None
            percentage = result.percentage if result else None
            grade = result.grade if result else None
            grade_point = result.grade_point if result else None
            position = position_map.get(student.id)

            marks_float = None
            if marks is not None:
                try:
                    marks_float = float(marks)
                except (TypeError, ValueError):
                    marks_float = None

            if marks_float is not None:
                students_with_marks += 1

                if grade:
                    grade_counts[grade] = grade_counts.get(grade, 0) + 1
                    grade_gender_matrix[gender][grade] += 1
                else:
                    grade_gender_matrix[gender]['No Grade'] += 1

            all_student_data.append({
                'id': student.id,
                'registration_number': student.registration_number or f"S{student.id:04d}",
                'full_name': student.full_name,
                'gender': gender,
                'marks': marks_float,
                'percentage': float(percentage) if percentage else None,
                'grade': grade,
                'grade_point': float(grade_point) if grade_point else None,
                'position': position,
                'has_marks': marks_float is not None
            })

        # Calculate matrix totals
        grade_totals = {}
        for grade in all_grades + ['No Grade']:
            grade_totals[grade] = 0
            for gender in grade_gender_matrix:
                grade_totals[grade] += grade_gender_matrix[gender].get(grade, 0)

        gender_totals = {}
        for gender in grade_gender_matrix:
            gender_totals[gender] = sum(grade_gender_matrix[gender].values())

        grand_total = sum(gender_totals.values())

        matrix_grades = all_grades.copy()
        if any(grade_gender_matrix[g]['No Grade'] > 0 for g in grade_gender_matrix):
            matrix_grades.append('No Grade')

        # Apply filters for filtered data
        filtered_data = [s.copy() for s in all_student_data if s['has_marks']]

        if grade_filter:
            if grade_filter == 'No Grade':
                filtered_data = [s for s in filtered_data if not s['grade']]
            else:
                filtered_data = [s for s in filtered_data if s['grade'] == grade_filter]

        if marks_min is not None:
            filtered_data = [s for s in filtered_data if s['marks'] >= marks_min]
        if marks_max is not None:
            filtered_data = [s for s in filtered_data if s['marks'] <= marks_max]

        if gender_filter:
            filtered_data = [
                s for s in filtered_data 
                if s['gender'] and s['gender'].lower() == gender_filter.lower()
            ]

        if rank_filter == 'top':
            students_with_position = [s for s in filtered_data if s['position']]
            students_with_position.sort(key=lambda x: x['position'])
            filtered_data = students_with_position[:min(top_n, len(students_with_position))]
        elif rank_filter == 'bottom':
            students_with_position = [s for s in filtered_data if s['position']]
            students_with_position.sort(key=lambda x: x['position'], reverse=True)
            filtered_data = students_with_position[:min(bottom_n, len(students_with_position))]

        # Calculate filtered statistics
        filtered_marks_list = [s['marks'] for s in filtered_data if s['marks'] is not None]
        statistics = {
            'average_marks': sum(filtered_marks_list) / len(filtered_marks_list) if filtered_marks_list else 0,
            'pass_rate': calculate_pass_rate(filtered_marks_list),
        }

        # Prepare context for PDF
        context = {
            'exam_session': exam_session,
            'selected_subject': selected_subject,
            'filtered_student_data': filtered_data,
            'statistics': statistics,
            'students_with_marks': students_with_marks,
            'total_students': total_students,
            'grade_gender_matrix': grade_gender_matrix,
            'matrix_grades': matrix_grades,
            'grade_totals': grade_totals,
            'gender_totals': gender_totals,
            'grand_total': grand_total,
            'grade_filter': grade_filter,
            'marks_min': marks_min,
            'marks_max': marks_max,
            'gender_filter': gender_filter,
            'rank_filter': rank_filter,
            'top_n': top_n,
            'bottom_n': bottom_n,
            'has_filters': bool(grade_filter or marks_min or marks_max or gender_filter or rank_filter),
            'generated_date': timezone.now(),
            'generated_by': request.user.get_full_name() or request.user.username,
            'school_name': getattr(settings, 'SCHOOL_NAME', 'School Management System'),
            'school_address': getattr(settings, 'SCHOOL_ADDRESS', ''),
            'is_pdf': True,
        }

        # Generate PDF
        html_string = render_to_string('admin/results/session_subject_matrix_analysis_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_file = html.write_pdf()

        # Generate filename
        filename_parts = [
            f"Matrix_{selected_subject.code}",
            exam_session.name.replace(' ', '_')
        ]

        if grade_filter:
            filename_parts.append(f"Grade_{grade_filter}")
        if marks_min is not None:
            filename_parts.append(f"Min{marks_min}")
        if marks_max is not None:
            filename_parts.append(f"Max{marks_max}")
        if gender_filter:
            filename_parts.append(f"Gender_{gender_filter}")
        if rank_filter:
            filename_parts.append(f"{rank_filter}_{top_n if rank_filter == 'top' else bottom_n}")

        filename_parts.append(timezone.now().strftime('%Y%m%d'))
        filename = "_".join(filename_parts) + ".pdf"

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating matrix analysis PDF: {str(e)}")
        messages.error(request, f"Error generating PDF report: {str(e)}")
        return redirect('session_subject_matrix_analysis', exam_session_id=exam_session_id)    
    
