# ============================================================================
# HOSTEL MANAGEMENT VIEWS
# ============================================================================

from decimal import Decimal
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
# ============================================================================
# STUDENT HOSTEL ALLOCATION VIEWS
# ============================================================================
from django.db.models.functions import Coalesce, ExtractMonth, ExtractYear
from django.urls import reverse
from django.utils import timezone
from django.db.models import Q, Sum, Count, F, Value, CharField, Case, When, DecimalField
from datetime import datetime
from accounts.models import GENDER_CHOICES
from core.models import AcademicYear, ClassLevel
from students.models import Bed, Hostel, HostelInstallmentPlan, HostelPayment, HostelPaymentTransaction, HostelRoom, Student, StudentHostelAllocation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import tempfile

# ============================================================================
# HOSTEL PAYMENTS EXPORT VIEWS
# ============================================================================

@login_required
def hostels_list(request):
    """
    Display hostels management page
    """
    hostels = Hostel.objects.all().order_by('name')
    
    # Get statistics
    total_hostels = hostels.count()
    active_hostels = hostels.filter(is_active=True).count()
    boys_hostels = hostels.filter(hostel_type='boys').count()
    girls_hostels = hostels.filter(hostel_type='girls').count()
    mixed_hostels = hostels.filter(hostel_type='mixed').count()
    
    # Calculate total capacity
    total_capacity = hostels.aggregate(total=models.Sum('max_students'))['total'] or 0
    
    # Get allocation counts (if you have allocations)
    # This will need to be updated based on your allocation model
    allocated_count = 0  # Placeholder
    
    context = {
        'hostels': hostels,
        'total_hostels': total_hostels,
        'active_hostels': active_hostels,
        'boys_hostels': boys_hostels,
        'girls_hostels': girls_hostels,
        'mixed_hostels': mixed_hostels,
        'total_capacity': total_capacity,
        'allocated_count': allocated_count,
        'hostel_types': Hostel.HOSTEL_TYPES,
        'payment_modes': Hostel.PAYMENT_MODE,
        'page_title': 'Hostel Management',
    }
    
    return render(request, 'admin/hostels/hostels_list.html', context)

@login_required
def hostel_students_list(request, hostel_id):
    """
    Display list of students allocated to a specific hostel
    """
    # Get the hostel
    hostel = get_object_or_404(
        Hostel.objects.prefetch_related('rooms', 'rooms__beds'),
        id=hostel_id,
        is_active=True
    )
    
    # Get all active allocations for this hostel
    allocations = StudentHostelAllocation.objects.filter(
        hostel=hostel,
        is_active=True
    ).select_related(
        'student',
        'student__class_level',
        'student__stream_class',
        'room',
        'bed',
        'academic_year'
    ).order_by('student__first_name', 'student__last_name')
    
    # Get statistics
    total_students = allocations.count()
    total_rooms = hostel.rooms.filter(is_active=True).count()
    total_beds = Bed.objects.filter(room__hostel=hostel).count()
    occupancy_rate = (total_students / hostel.max_students * 100) if hostel.max_students > 0 else 0
    
    # Group by room for room-wise distribution
    room_distribution = []
    for room in hostel.rooms.filter(is_active=True).order_by('room_number'):
        room_allocations = allocations.filter(room=room)
        room_beds = room.beds.count()
        room_distribution.append({
            'room': room,
            'student_count': room_allocations.count(),
            'bed_count': room_beds,
            'students': room_allocations[:5]  # Limit to 5 for preview
        })
    
    # Get recent allocations
    recent_allocations = allocations[:10]
    
    # Get payment summary for these students
    student_ids = allocations.values_list('student_id', flat=True)
    payment_summary = HostelPaymentTransaction.objects.filter(
        allocation__student_id__in=student_ids,
        allocation__hostel=hostel
    ).values('allocation__student_id').annotate(
        total_paid=Sum('amount')
    ).order_by('-total_paid')[:10]
    
    context = {
        'hostel': hostel,
        'allocations': allocations,
        'total_students': total_students,
        'total_rooms': total_rooms,
        'total_beds': total_beds,
        'occupancy_rate': round(occupancy_rate, 1),
        'available_spaces': hostel.max_students - total_students,
        'room_distribution': room_distribution,
        'recent_allocations': recent_allocations,
        'payment_summary': payment_summary,
        'page_title': f'Students in {hostel.name}',
    }
    
    return render(request, 'admin/hostels/hostel_students.html', context)


@login_required
def hostel_students_export(request, hostel_id):
    """
    Export student list for a hostel to Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    hostel = get_object_or_404(Hostel, id=hostel_id)
    
    allocations = StudentHostelAllocation.objects.filter(
        hostel=hostel,
        is_active=True
    ).select_related(
        'student',
        'student__class_level',
        'student__stream_class',
        'room',
        'bed'
    ).order_by('student__first_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{hostel.code} Students"
    
    # Add headers
    headers = ['Student Name', 'Registration #', 'Class', 'Stream', 'Room', 'Bed', 'Gender', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4e73df', end_color='4e73df', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row, allocation in enumerate(allocations, 2):
        ws.cell(row=row, column=1, value=allocation.student.full_name)
        ws.cell(row=row, column=2, value=allocation.student.registration_number or 'N/A')
        ws.cell(row=row, column=3, value=allocation.student.class_level.name if allocation.student.class_level else 'N/A')
        ws.cell(row=row, column=4, value=allocation.student.stream_class.stream_letter if allocation.student.stream_class else 'N/A')
        ws.cell(row=row, column=5, value=allocation.room.room_number if allocation.room else 'Not Assigned')
        ws.cell(row=row, column=6, value=allocation.bed.bed_number if allocation.bed else 'Not Assigned')
        ws.cell(row=row, column=7, value=allocation.student.get_gender_display() if allocation.student.gender else 'N/A')
        ws.cell(row=row, column=8, value='Active' if allocation.is_active else 'Inactive')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{hostel.code}_students_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@login_required
def hostels_crud(request):
    """
    Handle AJAX CRUD operations for hostels
    """
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_hostel(request)
            elif action == 'update':
                return update_hostel(request)
            elif action == 'toggle_status':
                return toggle_hostel_status(request)
            elif action == 'delete':
                return delete_hostel(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })


def create_hostel(request):
    """Create a new hostel"""
    # Get and validate required fields
    name = request.POST.get('name', '').strip()
    code = request.POST.get('code', '').strip().upper()
    hostel_type = request.POST.get('hostel_type', '').strip()
    max_students_str = request.POST.get('max_students', '').strip()
    total_fee_str = request.POST.get('total_fee', '').strip()
    payment_mode = request.POST.get('payment_mode', '').strip()
    installments_count_str = request.POST.get('installments_count', '1').strip()
    
    # Validate required fields
    required_fields = {
        'name': name,
        'code': code,
        'hostel_type': hostel_type,
        'max_students': max_students_str,
        'total_fee': total_fee_str,
        'payment_mode': payment_mode
    }
    
    for field_name, value in required_fields.items():
        if not value:
            return JsonResponse({
                'success': False,
                'message': f'{field_name.replace("_", " ").title()} is required.'
            })
    
    # Validate name and code
    if len(name) < 2:
        return JsonResponse({
            'success': False,
            'message': 'Hostel name must be at least 2 characters long.'
        })
    
    if len(name) > 100:
        return JsonResponse({
            'success': False,
            'message': 'Hostel name cannot exceed 100 characters.'
        })
    
    if len(code) < 2:
        return JsonResponse({
            'success': False,
            'message': 'Hostel code must be at least 2 characters long.'
        })
    
    if len(code) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Hostel code cannot exceed 20 characters.'
        })
    
    # Validate code format (alphanumeric with optional underscore)
    if not code.replace('_', '').isalnum():
        return JsonResponse({
            'success': False,
            'message': 'Hostel code can only contain letters, numbers, and underscores.'
        })
    
    # Validate max_students
    try:
        max_students = int(max_students_str)
        if max_students < 1:
            return JsonResponse({
                'success': False,
                'message': 'Maximum students must be at least 1.'
            })
        if max_students > 1000:
            return JsonResponse({
                'success': False,
                'message': 'Maximum students cannot exceed 1000.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Maximum students must be a valid number.'
        })
    
    # Validate total_fee
    try:
        total_fee = float(total_fee_str)
        if total_fee < 0:
            return JsonResponse({
                'success': False,
                'message': 'Total fee cannot be negative.'
            })
        if total_fee > 9999999.99:
            return JsonResponse({
                'success': False,
                'message': 'Total fee is too high.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Total fee must be a valid number.'
        })
    
    # Validate hostel_type
    valid_types = [choice[0] for choice in Hostel.HOSTEL_TYPES]
    if hostel_type not in valid_types:
        return JsonResponse({
            'success': False,
            'message': 'Invalid hostel type selected.'
        })
    
    # Validate payment_mode
    valid_modes = [choice[0] for choice in Hostel.PAYMENT_MODE]
    if payment_mode not in valid_modes:
        return JsonResponse({
            'success': False,
            'message': 'Invalid payment mode selected.'
        })
    
    # Validate installments_count
    try:
        installments_count = int(installments_count_str)
        if payment_mode == 'installments':
            if installments_count < 1:
                return JsonResponse({
                    'success': False,
                    'message': 'Installments count must be at least 1.'
                })
            if installments_count > 12:
                return JsonResponse({
                    'success': False,
                    'message': 'Installments count cannot exceed 12.'
                })
        else:
            installments_count = 1  # Set to default if not installments
    except ValueError:
        installments_count = 1
    
    # Check for duplicate name
    if Hostel.objects.filter(name__iexact=name).exists():
        return JsonResponse({
            'success': False,
            'message': f'A hostel with name "{name}" already exists.'
        })
    
    # Check for duplicate code
    if Hostel.objects.filter(code__iexact=code).exists():
        return JsonResponse({
            'success': False,
            'message': f'A hostel with code "{code}" already exists.'
        })
    
    # Get optional field
    is_active = request.POST.get('is_active') == 'on' or request.POST.get('is_active') == 'true'
    
    try:
        # Create the hostel
        hostel = Hostel.objects.create(
            name=name,
            code=code,
            hostel_type=hostel_type,
            max_students=max_students,
            total_fee=total_fee,
            payment_mode=payment_mode,
            installments_count=installments_count,
            is_active=is_active
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Hostel "{name}" created successfully.',
            'hostel': {
                'id': hostel.id,
                'name': hostel.name,
                'code': hostel.code,
                'hostel_type': hostel.hostel_type,
                'hostel_type_display': hostel.get_hostel_type_display(),
                'max_students': hostel.max_students,
                'total_fee': float(hostel.total_fee),
                'payment_mode': hostel.payment_mode,
                'payment_mode_display': hostel.get_payment_mode_display(),
                'installments_count': hostel.installments_count,
                'is_active': hostel.is_active,
                'created_at': hostel.created_at.strftime('%Y-%m-%d')
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'A hostel with this name or code already exists.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating hostel: {str(e)}'
        })


def update_hostel(request):
    """Update an existing hostel"""
    hostel_id = request.POST.get('id')
    if not hostel_id:
        return JsonResponse({
            'success': False,
            'message': 'Hostel ID is required.'
        })
    
    try:
        hostel = Hostel.objects.get(id=hostel_id)
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Hostel not found.'
        })
    
    # Get and validate required fields
    name = request.POST.get('name', '').strip()
    code = request.POST.get('code', '').strip().upper()
    hostel_type = request.POST.get('hostel_type', '').strip()
    max_students_str = request.POST.get('max_students', '').strip()
    total_fee_str = request.POST.get('total_fee', '').strip()
    payment_mode = request.POST.get('payment_mode', '').strip()
    installments_count_str = request.POST.get('installments_count', '1').strip()
    
    # Validate required fields
    required_fields = {
        'name': name,
        'code': code,
        'hostel_type': hostel_type,
        'max_students': max_students_str,
        'total_fee': total_fee_str,
        'payment_mode': payment_mode
    }
    
    for field_name, value in required_fields.items():
        if not value:
            return JsonResponse({
                'success': False,
                'message': f'{field_name.replace("_", " ").title()} is required.'
            })
    
    # Validate name and code
    if len(name) < 2:
        return JsonResponse({
            'success': False,
            'message': 'Hostel name must be at least 2 characters long.'
        })
    
    if len(name) > 100:
        return JsonResponse({
            'success': False,
            'message': 'Hostel name cannot exceed 100 characters.'
        })
    
    if len(code) < 2:
        return JsonResponse({
            'success': False,
            'message': 'Hostel code must be at least 2 characters long.'
        })
    
    if len(code) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Hostel code cannot exceed 20 characters.'
        })
    
    # Validate code format
    if not code.replace('_', '').isalnum():
        return JsonResponse({
            'success': False,
            'message': 'Hostel code can only contain letters, numbers, and underscores.'
        })
    
    # Validate max_students
    try:
        max_students = int(max_students_str)
        if max_students < 1:
            return JsonResponse({
                'success': False,
                'message': 'Maximum students must be at least 1.'
            })
        if max_students > 1000:
            return JsonResponse({
                'success': False,
                'message': 'Maximum students cannot exceed 1000.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Maximum students must be a valid number.'
        })
    
    # Validate total_fee
    try:
        total_fee = float(total_fee_str)
        if total_fee < 0:
            return JsonResponse({
                'success': False,
                'message': 'Total fee cannot be negative.'
            })
        if total_fee > 9999999.99:
            return JsonResponse({
                'success': False,
                'message': 'Total fee is too high.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Total fee must be a valid number.'
        })
    
    # Validate hostel_type
    valid_types = [choice[0] for choice in Hostel.HOSTEL_TYPES]
    if hostel_type not in valid_types:
        return JsonResponse({
            'success': False,
            'message': 'Invalid hostel type selected.'
        })
    
    # Validate payment_mode
    valid_modes = [choice[0] for choice in Hostel.PAYMENT_MODE]
    if payment_mode not in valid_modes:
        return JsonResponse({
            'success': False,
            'message': 'Invalid payment mode selected.'
        })
    
    # Validate installments_count
    try:
        installments_count = int(installments_count_str)
        if payment_mode == 'installments':
            if installments_count < 1:
                return JsonResponse({
                    'success': False,
                    'message': 'Installments count must be at least 1.'
                })
            if installments_count > 12:
                return JsonResponse({
                    'success': False,
                    'message': 'Installments count cannot exceed 12.'
                })
        else:
            installments_count = 1
    except ValueError:
        installments_count = 1
    
    # Check for duplicate name (excluding current)
    if Hostel.objects.filter(name__iexact=name).exclude(id=hostel.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'A hostel with name "{name}" already exists.'
        })
    
    # Check for duplicate code (excluding current)
    if Hostel.objects.filter(code__iexact=code).exclude(id=hostel.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'A hostel with code "{code}" already exists.'
        })
    
    # Get optional field
    is_active = request.POST.get('is_active') == 'on' or request.POST.get('is_active') == 'true'
    
    try:
        # Update the hostel
        hostel.name = name
        hostel.code = code
        hostel.hostel_type = hostel_type
        hostel.max_students = max_students
        hostel.total_fee = total_fee
        hostel.payment_mode = payment_mode
        hostel.installments_count = installments_count
        hostel.is_active = is_active
        hostel.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Hostel "{name}" updated successfully.',
            'hostel': {
                'id': hostel.id,
                'name': hostel.name,
                'code': hostel.code,
                'hostel_type': hostel.hostel_type,
                'hostel_type_display': hostel.get_hostel_type_display(),
                'max_students': hostel.max_students,
                'total_fee': float(hostel.total_fee),
                'payment_mode': hostel.payment_mode,
                'payment_mode_display': hostel.get_payment_mode_display(),
                'installments_count': hostel.installments_count,
                'is_active': hostel.is_active
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'A hostel with this name or code already exists.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating hostel: {str(e)}'
        })


def toggle_hostel_status(request):
    """Toggle hostel active/inactive status"""
    hostel_id = request.POST.get('id')
    if not hostel_id:
        return JsonResponse({
            'success': False,
            'message': 'Hostel ID is required.'
        })
    
    try:
        hostel = Hostel.objects.get(id=hostel_id)
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Hostel not found.'
        })
    
    try:
        # Toggle the status
        hostel.is_active = not hostel.is_active
        hostel.save()
        
        status_text = "activated" if hostel.is_active else "deactivated"
        
        return JsonResponse({
            'success': True,
            'message': f'Hostel "{hostel.name}" {status_text} successfully.',
            'is_active': hostel.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error toggling hostel status: {str(e)}'
        })


def delete_hostel(request):
    """Delete a hostel"""
    hostel_id = request.POST.get('id')
    if not hostel_id:
        return JsonResponse({
            'success': False,
            'message': 'Hostel ID is required.'
        })
    
    try:
        hostel = Hostel.objects.get(id=hostel_id)
        hostel_name = hostel.name
        
        # Check if hostel has any rooms
        if hostel.rooms.exists():
            return JsonResponse({
                'success': False,
                'message': f'Cannot delete hostel "{hostel_name}". It has associated rooms.'
            })
        
        # Check if hostel has any allocations
        if hasattr(hostel, 'studenthostelallocation_set') and hostel.studenthostelallocation_set.exists():
            return JsonResponse({
                'success': False,
                'message': f'Cannot delete hostel "{hostel_name}". It has student allocations.'
            })
        
        hostel.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Hostel "{hostel_name}" deleted successfully.',
            'id': hostel_id
        })
        
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Hostel not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting hostel: {str(e)}'
        })


# ============================================================================
# HOSTEL ROOM MANAGEMENT VIEWS
# ============================================================================

@login_required
def hostel_rooms_list(request):
    """
    Display hostel rooms management page
    """
    # Get all active hostels for filter dropdown
    hostels = Hostel.objects.filter(is_active=True).order_by('name')
    
    # Get all rooms with hostel info
    rooms = HostelRoom.objects.select_related('hostel').all().order_by('hostel__name', 'room_number')
    
    # Get statistics
    total_rooms = rooms.count()
    active_rooms = rooms.filter(is_active=True).count()
    total_capacity = rooms.aggregate(total=models.Sum('capacity'))['total'] or 0
    
    # Get occupied rooms count (if you have bed allocations)
    # This will need to be updated based on your bed model
    occupied_rooms = 0  # Placeholder
    
    # Get rooms per hostel for chart/data
    rooms_per_hostel = Hostel.objects.annotate(
        room_count=Count('rooms'),
        total_capacity=models.Sum('rooms__capacity')
    ).filter(room_count__gt=0).order_by('-room_count')
    
    context = {
        'rooms': rooms,
        'hostels': hostels,
        'total_rooms': total_rooms,
        'active_rooms': active_rooms,
        'occupied_rooms': occupied_rooms,
        'total_capacity': total_capacity,
        'rooms_per_hostel': rooms_per_hostel,
        'page_title': 'Hostel Rooms Management',
    }
    
    return render(request, 'admin/hostels/rooms_list.html', context)


@login_required
def hostel_rooms_by_hostel(request, hostel_id):
    """
    Get rooms for a specific hostel (AJAX endpoint for dropdowns)
    """
    try:
        hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
        rooms = HostelRoom.objects.filter(
            hostel=hostel,
            is_active=True
        ).order_by('room_number')
        
        room_list = []
        for room in rooms:
            # Get current number of beds in this room
            current_beds = Bed.objects.filter(room=room).count()
            available_beds = room.capacity - current_beds
            
            room_list.append({
                'id': room.id,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'current_beds': current_beds,
                'available_beds': available_beds,
                'display': f"Room {room.room_number} (Capacity: {room.capacity}, Available: {available_beds})"
            })
        
        return JsonResponse({
            'success': True,
            'rooms': room_list,
            'hostel_name': hostel.name
        })
        
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Hostel not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@login_required
def hostel_rooms_crud(request):
    """
    Handle AJAX CRUD operations for hostel rooms
    """
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_hostel_room(request)
            elif action == 'update':
                return update_hostel_room(request)
            elif action == 'toggle_status':
                return toggle_hostel_room_status(request)
            elif action == 'delete':
                return delete_hostel_room(request)
            elif action == 'bulk_create':
                return bulk_create_rooms(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })


def create_hostel_room(request):
    """Create a new hostel room"""
    # Get and validate required fields
    hostel_id = request.POST.get('hostel', '').strip()
    room_number = request.POST.get('room_number', '').strip().upper()
    capacity_str = request.POST.get('capacity', '').strip()
    
    if not hostel_id:
        return JsonResponse({
            'success': False,
            'message': 'Hostel selection is required.'
        })
    
    if not room_number:
        return JsonResponse({
            'success': False,
            'message': 'Room number is required.'
        })
    
    if not capacity_str:
        return JsonResponse({
            'success': False,
            'message': 'Capacity is required.'
        })
    
    # Validate room number
    if len(room_number) < 1:
        return JsonResponse({
            'success': False,
            'message': 'Room number must be at least 1 character long.'
        })
    
    if len(room_number) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Room number cannot exceed 20 characters.'
        })
    
    # Validate capacity
    try:
        capacity = int(capacity_str)
        if capacity < 1:
            return JsonResponse({
                'success': False,
                'message': 'Capacity must be at least 1.'
            })
        if capacity > 20:  # Assuming max 20 students per room
            return JsonResponse({
                'success': False,
                'message': 'Capacity cannot exceed 20 students per room.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Capacity must be a valid number.'
        })
    
    # Get hostel
    try:
        hostel = Hostel.objects.get(id=hostel_id, is_active=True)
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected hostel does not exist or is inactive.'
        })
    
    # Check for duplicate room number in the same hostel
    if HostelRoom.objects.filter(hostel=hostel, room_number__iexact=room_number).exists():
        return JsonResponse({
            'success': False,
            'message': f'Room "{room_number}" already exists in {hostel.name}.'
        })
    
    # Check if adding this room would exceed hostel capacity (optional)
    current_total_capacity = HostelRoom.objects.filter(hostel=hostel).aggregate(
        total=models.Sum('capacity')
    )['total'] or 0
    
    if current_total_capacity + capacity > hostel.max_students:
        return JsonResponse({
            'success': False,
            'message': f'Cannot add room. Total capacity would exceed hostel maximum of {hostel.max_students} students.'
        })
    
    # Get optional field
    is_active = request.POST.get('is_active') == 'on' or request.POST.get('is_active') == 'true'
    
    try:
        # Create the room
        room = HostelRoom.objects.create(
            hostel=hostel,
            room_number=room_number,
            capacity=capacity,
            is_active=is_active
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Room "{room_number}" created successfully in {hostel.name}.',
            'room': {
                'id': room.id,
                'hostel_id': room.hostel.id,
                'hostel_name': room.hostel.name,
                'hostel_code': room.hostel.code,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'is_active': room.is_active,
                'display': str(room)
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'Room "{room_number}" already exists in this hostel.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating room: {str(e)}'
        })


def update_hostel_room(request):
    """Update an existing hostel room"""
    room_id = request.POST.get('id')
    if not room_id:
        return JsonResponse({
            'success': False,
            'message': 'Room ID is required.'
        })
    
    try:
        room = HostelRoom.objects.select_related('hostel').get(id=room_id)
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Room not found.'
        })
    
    # Get and validate required fields
    hostel_id = request.POST.get('hostel', '').strip()
    room_number = request.POST.get('room_number', '').strip().upper()
    capacity_str = request.POST.get('capacity', '').strip()
    
    if not hostel_id or not room_number or not capacity_str:
        return JsonResponse({
            'success': False,
            'message': 'Hostel, room number, and capacity are required.'
        })
    
    # Validate room number
    if len(room_number) < 1:
        return JsonResponse({
            'success': False,
            'message': 'Room number must be at least 1 character long.'
        })
    
    if len(room_number) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Room number cannot exceed 20 characters.'
        })
    
    # Validate capacity
    try:
        capacity = int(capacity_str)
        if capacity < 1:
            return JsonResponse({
                'success': False,
                'message': 'Capacity must be at least 1.'
            })
        if capacity > 20:
            return JsonResponse({
                'success': False,
                'message': 'Capacity cannot exceed 20 students per room.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Capacity must be a valid number.'
        })
    
    # Get hostel
    try:
        hostel = Hostel.objects.get(id=hostel_id, is_active=True)
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected hostel does not exist or is inactive.'
        })
    
    # Check for duplicate room number in the same hostel (excluding current)
    if HostelRoom.objects.filter(
        hostel=hostel, 
        room_number__iexact=room_number
    ).exclude(id=room.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'Room "{room_number}" already exists in {hostel.name}.'
        })
    
    # Check if capacity reduction would affect existing beds/allocations
    if room.capacity > capacity:
        # Check if there are any beds or allocations that would be affected
        # This depends on your bed model - add appropriate checks
        pass
    
    # Check hostel capacity when changing hostel or capacity
    if room.hostel_id != hostel.id or room.capacity != capacity:
        # Calculate total capacity excluding current room
        other_rooms_capacity = HostelRoom.objects.filter(
            hostel=hostel
        ).exclude(id=room.id).aggregate(
            total=models.Sum('capacity')
        )['total'] or 0
        
        if other_rooms_capacity + capacity > hostel.max_students:
            return JsonResponse({
                'success': False,
                'message': f'Cannot update room. Total capacity would exceed hostel maximum of {hostel.max_students} students.'
            })
    
    # Get optional field
    is_active = request.POST.get('is_active') == 'on' or request.POST.get('is_active') == 'true'
    
    try:
        # Update the room
        room.hostel = hostel
        room.room_number = room_number
        room.capacity = capacity
        room.is_active = is_active
        room.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Room "{room_number}" updated successfully.',
            'room': {
                'id': room.id,
                'hostel_id': room.hostel.id,
                'hostel_name': room.hostel.name,
                'hostel_code': room.hostel.code,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'is_active': room.is_active,
                'display': str(room)
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'Room "{room_number}" already exists in this hostel.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating room: {str(e)}'
        })


def toggle_hostel_room_status(request):
    """Toggle hostel room active/inactive status"""
    room_id = request.POST.get('id')
    if not room_id:
        return JsonResponse({
            'success': False,
            'message': 'Room ID is required.'
        })
    
    try:
        room = HostelRoom.objects.select_related('hostel').get(id=room_id)
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Room not found.'
        })
    
    # Check if room has any beds/allocations before deactivating
    if room.is_active:
        # Check if room has any beds or allocations
        # This depends on your bed model
        has_beds = False  # Placeholder - update based on your Bed model
        if has_beds:
            return JsonResponse({
                'success': False,
                'message': f'Cannot deactivate room "{room.room_number}" because it has assigned beds.'
            })
    
    try:
        # Toggle the status
        room.is_active = not room.is_active
        room.save()
        
        status_text = "activated" if room.is_active else "deactivated"
        
        return JsonResponse({
            'success': True,
            'message': f'Room "{room.room_number}" in {room.hostel.name} {status_text} successfully.',
            'is_active': room.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error toggling room status: {str(e)}'
        })


def delete_hostel_room(request):
    """Delete a hostel room"""
    room_id = request.POST.get('id')
    if not room_id:
        return JsonResponse({
            'success': False,
            'message': 'Room ID is required.'
        })
    
    try:
        room = HostelRoom.objects.select_related('hostel').get(id=room_id)
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Room not found.'
        })
    
    # Check if room has any beds or allocations
    # This depends on your bed model
    has_beds = False  # Placeholder - update based on your Bed model
    
    if has_beds:
        return JsonResponse({
            'success': False,
            'message': f'Cannot delete room "{room.room_number}" because it has assigned beds. Remove beds first.'
        })
    
    room_info = f'"{room.room_number}" in {room.hostel.name}'
    
    try:
        room.delete()
        return JsonResponse({
            'success': True,
            'message': f'Room {room_info} deleted successfully.',
            'id': room_id
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting room: {str(e)}'
        })


def bulk_create_rooms(request):
    """Create multiple rooms at once for a hostel"""
    hostel_id = request.POST.get('hostel', '').strip()
    start_room = request.POST.get('start_room', '').strip().upper()
    end_room = request.POST.get('end_room', '').strip().upper()
    capacity_str = request.POST.get('capacity', '').strip()
    
    if not hostel_id:
        return JsonResponse({
            'success': False,
            'message': 'Hostel selection is required.'
        })
    
    if not start_room or not end_room:
        return JsonResponse({
            'success': False,
            'message': 'Start and end room numbers are required.'
        })
    
    if not capacity_str:
        return JsonResponse({
            'success': False,
            'message': 'Capacity is required.'
        })
    
    # Validate capacity
    try:
        capacity = int(capacity_str)
        if capacity < 1 or capacity > 20:
            return JsonResponse({
                'success': False,
                'message': 'Capacity must be between 1 and 20.'
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Capacity must be a valid number.'
        })
    
    # Get hostel
    try:
        hostel = Hostel.objects.get(id=hostel_id, is_active=True)
    except Hostel.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected hostel does not exist or is inactive.'
        })
    
    # Parse room range (assuming alphanumeric room numbers like 101, 102 or A1, A2)
    # This is a simple implementation - you might need more complex logic
    
    # For numeric rooms
    if start_room.isdigit() and end_room.isdigit():
        start_num = int(start_room)
        end_num = int(end_room)
        
        if start_num > end_num:
            return JsonResponse({
                'success': False,
                'message': 'Start room number must be less than or equal to end room number.'
            })
        
        if end_num - start_num > 50:
            return JsonResponse({
                'success': False,
                'message': 'Cannot create more than 50 rooms at once.'
            })
        
        room_numbers = [str(i).zfill(len(start_room)) for i in range(start_num, end_num + 1)]
    else:
        # For alphanumeric, just create the two specified rooms
        room_numbers = [start_room, end_room] if start_room != end_room else [start_room]
    
    created_count = 0
    failed_rooms = []
    
    with transaction.atomic():
        for room_num in room_numbers:
            try:
                # Check if room already exists
                if HostelRoom.objects.filter(hostel=hostel, room_number__iexact=room_num).exists():
                    failed_rooms.append(f"{room_num} (Already exists)")
                    continue
                
                # Check hostel capacity
                current_total = HostelRoom.objects.filter(hostel=hostel).aggregate(
                    total=models.Sum('capacity')
                )['total'] or 0
                
                if current_total + capacity > hostel.max_students:
                    failed_rooms.append(f"{room_num} (Would exceed hostel capacity)")
                    continue
                
                # Create room
                HostelRoom.objects.create(
                    hostel=hostel,
                    room_number=room_num,
                    capacity=capacity,
                    is_active=True
                )
                created_count += 1
                
            except Exception as e:
                failed_rooms.append(f"{room_num} ({str(e)})")
    
    message = f'Successfully created {created_count} room(s) in {hostel.name}.'
    if failed_rooms:
        message += f' Failed: {", ".join(failed_rooms[:5])}'
        if len(failed_rooms) > 5:
            message += f' and {len(failed_rooms) - 5} more.'
    
    return JsonResponse({
        'success': True,
        'message': message,
        'created_count': created_count,
        'failed_count': len(failed_rooms)
    })        


# ============================================================================
# BED MANAGEMENT VIEWS
# ============================================================================

@login_required
def beds_list(request):
    """
    Display beds management page
    """
    # Get all active hostels and rooms for filters
    hostels = Hostel.objects.filter(is_active=True).order_by('name')
    rooms = HostelRoom.objects.select_related('hostel').filter(is_active=True).order_by('hostel__name', 'room_number')
    
    # Get all beds with room and hostel info
    beds = Bed.objects.select_related(
        'room__hostel'
    ).all().order_by('room__hostel__name', 'room__room_number', 'bed_number')
    
    # Get statistics
    total_beds = beds.count()
    occupied_beds = beds.filter(is_occupied=True).count()
    available_beds = total_beds - occupied_beds
    occupancy_rate = round((occupied_beds / total_beds * 100), 1) if total_beds > 0 else 0
    
    # Get bed type distribution
    single_beds = beds.filter(bed_type='single').count()
    bunk_upper = beds.filter(bed_type='bunk_upper').count()
    bunk_lower = beds.filter(bed_type='bunk_lower').count()
    
    # Get beds per hostel
    beds_per_hostel = Hostel.objects.annotate(
        bed_count=Count('rooms__beds'),
        occupied_count=Count('rooms__beds', filter=Q(rooms__beds__is_occupied=True))
    ).filter(bed_count__gt=0).order_by('-bed_count')
    
    context = {
        'beds': beds,
        'hostels': hostels,
        'rooms': rooms,
        'total_beds': total_beds,
        'occupied_beds': occupied_beds,
        'available_beds': available_beds,
        'occupancy_rate': occupancy_rate,
        'single_beds': single_beds,
        'bunk_upper': bunk_upper,
        'bunk_lower': bunk_lower,
        'beds_per_hostel': beds_per_hostel,
        'bed_types': Bed.BED_TYPES,
        'page_title': 'Bed Management',
    }
    
    return render(request, 'admin/hostels/beds_list.html', context)


@login_required
def beds_by_room(request, room_id):
    """
    Get beds for a specific room (AJAX endpoint for dropdowns)
    """
    try:
        room = get_object_or_404(HostelRoom, id=room_id, is_active=True)
        beds = Bed.objects.filter(
            room=room
        ).order_by('bed_number')
        
        bed_list = []
        for bed in beds:
            bed_list.append({
                'id': bed.id,
                'bed_number': bed.bed_number,
                'bed_type': bed.bed_type,
                'bed_type_display': bed.get_bed_type_display(),
                'is_occupied': bed.is_occupied,
                'display': f"Bed {bed.bed_number} ({bed.get_bed_type_display()})" + (" - Occupied" if bed.is_occupied else "")
            })
        
        # Calculate room occupancy
        total_beds = beds.count()
        occupied_beds = beds.filter(is_occupied=True).count()
        available_beds = total_beds - occupied_beds
        
        return JsonResponse({
            'success': True,
            'beds': bed_list,
            'room_number': room.room_number,
            'hostel_name': room.hostel.name,
            'room_capacity': room.capacity,
            'total_beds': total_beds,
            'occupied_beds': occupied_beds,
            'available_beds': available_beds,
            'current_beds': total_beds
        })
        
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Room not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@login_required
def hostel_rooms_by_hostel(request, hostel_id):
    """
    Get rooms for a specific hostel (AJAX endpoint for dropdowns)
    Used by: const url = '{% url "admin_hostel_rooms_by_hostel" 0 %}'.replace('0', hostelId);
    """
    try:
        # Get the hostel with proper error handling
        if not hostel_id:
            return JsonResponse({
                'success': False,
                'message': 'Hostel ID is required.'
            })
        
        # Get hostel and verify it exists and is active
        try:
            hostel = Hostel.objects.get(id=hostel_id, is_active=True)
        except Hostel.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Hostel not found or is inactive.'
            })
        
        # Get all active rooms for this hostel
        rooms = HostelRoom.objects.filter(
            hostel=hostel,
            is_active=True
        ).order_by('room_number')
        
        room_list = []
        for room in rooms:
            # Get current number of beds in this room
            current_beds = Bed.objects.filter(room=room).count()
            occupied_beds = Bed.objects.filter(room=room, is_occupied=True).count()
            available_beds = room.capacity - current_beds
            vacant_beds = current_beds - occupied_beds
            
            # Create display text with useful information
            display_text = f"Room {room.room_number}"
            display_text += f" (Capacity: {room.capacity}"
            display_text += f", Beds: {current_beds}/{room.capacity}"
            if available_beds > 0:
                display_text += f", {available_beds} slots available"
            display_text += ")"
            
            room_list.append({
                'id': room.id,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'current_beds': current_beds,
                'occupied_beds': occupied_beds,
                'vacant_beds': vacant_beds,
                'available_beds': available_beds,
                'is_full': current_beds >= room.capacity,
                'has_available_slots': available_beds > 0,
                'display': display_text
            })
        
        # Get hostel statistics
        total_rooms = len(room_list)
        total_capacity = sum(room['capacity'] for room in room_list)
        total_beds = sum(room['current_beds'] for room in room_list)
        total_available = sum(room['available_beds'] for room in room_list)
        
        return JsonResponse({
            'success': True,
            'rooms': room_list,
            'hostel': {
                'id': hostel.id,
                'name': hostel.name,
                'code': hostel.code,
                'hostel_type': hostel.hostel_type,
                'hostel_type_display': hostel.get_hostel_type_display(),
                'max_students': hostel.max_students,
                'total_rooms': total_rooms,
                'total_capacity': total_capacity,
                'total_beds': total_beds,
                'total_available': total_available,
                'is_active': hostel.is_active
            },
            'statistics': {
                'total_rooms': total_rooms,
                'total_capacity': total_capacity,
                'total_beds': total_beds,
                'total_available': total_available,
                'occupancy_rate': round((total_beds / total_capacity * 100), 1) if total_capacity > 0 else 0
            }
        })
        
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in hostel_rooms_by_hostel: {str(e)}", exc_info=True)
        
        return JsonResponse({
            'success': False,
            'message': f'An error occurred while loading rooms: {str(e)}'
        })

@login_required
def beds_crud(request):
    """
    Handle AJAX CRUD operations for beds
    """
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_bed(request)
            elif action == 'update':
                return update_bed(request)
            elif action == 'toggle_occupancy':
                return toggle_bed_occupancy(request)
            elif action == 'delete':
                return delete_bed(request)
            elif action == 'bulk_create':
                return bulk_create_beds(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })


def create_bed(request):
    """Create a new bed in a room"""
    # Get and validate required fields
    room_id = request.POST.get('room', '').strip()
    bed_number = request.POST.get('bed_number', '').strip().upper()
    bed_type = request.POST.get('bed_type', '').strip()
    
    if not room_id:
        return JsonResponse({
            'success': False,
            'message': 'Room selection is required.'
        })
    
    if not bed_number:
        return JsonResponse({
            'success': False,
            'message': 'Bed number is required.'
        })
    
    if not bed_type:
        return JsonResponse({
            'success': False,
            'message': 'Bed type is required.'
        })
    
    # Validate bed number
    if len(bed_number) < 1:
        return JsonResponse({
            'success': False,
            'message': 'Bed number must be at least 1 character long.'
        })
    
    if len(bed_number) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Bed number cannot exceed 20 characters.'
        })
    
    # Validate bed number format (alphanumeric with optional hyphen)
    if not bed_number.replace('-', '').isalnum():
        return JsonResponse({
            'success': False,
            'message': 'Bed number can only contain letters, numbers, and hyphens.'
        })
    
    # Validate bed type
    valid_types = [choice[0] for choice in Bed.BED_TYPES]
    if bed_type not in valid_types:
        return JsonResponse({
            'success': False,
            'message': 'Invalid bed type selected.'
        })
    
    # Get room
    try:
        room = HostelRoom.objects.select_related('hostel').get(id=room_id, is_active=True)
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected room does not exist or is inactive.'
        })
    
    # Check for duplicate bed number in the same room
    if Bed.objects.filter(room=room, bed_number__iexact=bed_number).exists():
        return JsonResponse({
            'success': False,
            'message': f'Bed "{bed_number}" already exists in Room {room.room_number}.'
        })
    
    # Check room capacity
    current_beds_count = Bed.objects.filter(room=room).count()
    if current_beds_count >= room.capacity:
        return JsonResponse({
            'success': False,
            'message': f'Cannot add more beds. Room capacity is {room.capacity} beds.'
        })
    
    # Check if adding a bunk bed makes sense (optional validation)
    if bed_type in ['bunk_upper', 'bunk_lower']:
        # Check if the corresponding bunk bed exists or can be added
        other_bunk = 'bunk_lower' if bed_type == 'bunk_upper' else 'bunk_upper'
        # This is just a warning, not an error
        pass
    
    # Get optional field
    is_occupied = request.POST.get('is_occupied') == 'on' or request.POST.get('is_occupied') == 'true'
    
    try:
        # Create the bed
        bed = Bed.objects.create(
            room=room,
            bed_number=bed_number,
            bed_type=bed_type,
            is_occupied=is_occupied
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Bed "{bed_number}" created successfully in Room {room.room_number}.',
            'bed': {
                'id': bed.id,
                'room_id': bed.room.id,
                'room_number': bed.room.room_number,
                'hostel_id': bed.room.hostel.id,
                'hostel_name': bed.room.hostel.name,
                'hostel_code': bed.room.hostel.code,
                'bed_number': bed.bed_number,
                'bed_type': bed.bed_type,
                'bed_type_display': bed.get_bed_type_display(),
                'is_occupied': bed.is_occupied,
                'display': str(bed)
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'Bed "{bed_number}" already exists in this room.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating bed: {str(e)}'
        })


def update_bed(request):
    """Update an existing bed"""
    bed_id = request.POST.get('id')
    if not bed_id:
        return JsonResponse({
            'success': False,
            'message': 'Bed ID is required.'
        })
    
    try:
        bed = Bed.objects.select_related('room__hostel').get(id=bed_id)
    except Bed.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Bed not found.'
        })
    
    # Get and validate required fields
    room_id = request.POST.get('room', '').strip()
    bed_number = request.POST.get('bed_number', '').strip().upper()
    bed_type = request.POST.get('bed_type', '').strip()
    
    if not room_id or not bed_number or not bed_type:
        return JsonResponse({
            'success': False,
            'message': 'Room, bed number, and bed type are required.'
        })
    
    # Validate bed number
    if len(bed_number) < 1:
        return JsonResponse({
            'success': False,
            'message': 'Bed number must be at least 1 character long.'
        })
    
    if len(bed_number) > 20:
        return JsonResponse({
            'success': False,
            'message': 'Bed number cannot exceed 20 characters.'
        })
    
    # Validate bed number format
    if not bed_number.replace('-', '').isalnum():
        return JsonResponse({
            'success': False,
            'message': 'Bed number can only contain letters, numbers, and hyphens.'
        })
    
    # Validate bed type
    valid_types = [choice[0] for choice in Bed.BED_TYPES]
    if bed_type not in valid_types:
        return JsonResponse({
            'success': False,
            'message': 'Invalid bed type selected.'
        })
    
    # Get room
    try:
        room = HostelRoom.objects.select_related('hostel').get(id=room_id, is_active=True)
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected room does not exist or is inactive.'
        })
    
    # Check for duplicate bed number in the same room (excluding current)
    if Bed.objects.filter(
        room=room, 
        bed_number__iexact=bed_number
    ).exclude(id=bed.id).exists():
        return JsonResponse({
            'success': False,
            'message': f'Bed "{bed_number}" already exists in Room {room.room_number}.'
        })
    
    # Check room capacity when moving to different room
    if bed.room_id != room.id:
        current_beds_count = Bed.objects.filter(room=room).count()
        if current_beds_count >= room.capacity:
            return JsonResponse({
                'success': False,
                'message': f'Cannot move bed. Target room capacity is {room.capacity} beds.'
            })
    
    # Get optional field
    is_occupied = request.POST.get('is_occupied') == 'on' or request.POST.get('is_occupied') == 'true'
    
    try:
        # Update the bed
        bed.room = room
        bed.bed_number = bed_number
        bed.bed_type = bed_type
        bed.is_occupied = is_occupied
        bed.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Bed "{bed_number}" updated successfully.',
            'bed': {
                'id': bed.id,
                'room_id': bed.room.id,
                'room_number': bed.room.room_number,
                'hostel_id': bed.room.hostel.id,
                'hostel_name': bed.room.hostel.name,
                'hostel_code': bed.room.hostel.code,
                'bed_number': bed.bed_number,
                'bed_type': bed.bed_type,
                'bed_type_display': bed.get_bed_type_display(),
                'is_occupied': bed.is_occupied,
                'display': str(bed)
            }
        })
        
    except IntegrityError as e:
        if 'unique' in str(e).lower():
            return JsonResponse({
                'success': False,
                'message': f'Bed "{bed_number}" already exists in this room.'
            })
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating bed: {str(e)}'
        })


def toggle_bed_occupancy(request):
    """Toggle bed occupancy status"""
    bed_id = request.POST.get('id')
    if not bed_id:
        return JsonResponse({
            'success': False,
            'message': 'Bed ID is required.'
        })
    
    try:
        bed = Bed.objects.select_related('room__hostel').get(id=bed_id)
    except Bed.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Bed not found.'
        })
    
    try:
        # Toggle the occupancy
        bed.is_occupied = not bed.is_occupied
        bed.save()
        
        status_text = "occupied" if bed.is_occupied else "vacant"
        
        return JsonResponse({
            'success': True,
            'message': f'Bed "{bed.bed_number}" in Room {bed.room.room_number} marked as {status_text}.',
            'is_occupied': bed.is_occupied
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error toggling bed occupancy: {str(e)}'
        })


def delete_bed(request):
    """Delete a bed"""
    bed_id = request.POST.get('id')
    if not bed_id:
        return JsonResponse({
            'success': False,
            'message': 'Bed ID is required.'
        })
    
    try:
        bed = Bed.objects.select_related('room__hostel').get(id=bed_id)
    except Bed.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Bed not found.'
        })
    
    # Check if bed is occupied
    if bed.is_occupied:
        return JsonResponse({
            'success': False,
            'message': f'Cannot delete occupied bed "{bed.bed_number}". Please vacate it first.'
        })
    
    # Check if bed has any allocations (if you have allocation model)
    # This depends on your allocation model
    has_allocations = False  # Placeholder
    
    if has_allocations:
        return JsonResponse({
            'success': False,
            'message': f'Cannot delete bed "{bed.bed_number}" because it has student allocations.'
        })
    
    bed_info = f'Bed "{bed.bed_number}" in Room {bed.room.room_number} ({bed.room.hostel.name})'
    
    try:
        bed.delete()
        return JsonResponse({
            'success': True,
            'message': f'{bed_info} deleted successfully.',
            'id': bed_id
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting bed: {str(e)}'
        })


def bulk_create_beds(request):
    """Create multiple beds at once for a room"""
    room_id = request.POST.get('room', '').strip()
    start_bed = request.POST.get('start_bed', '').strip().upper()
    end_bed = request.POST.get('end_bed', '').strip().upper()
    bed_type = request.POST.get('bed_type', '').strip()
    
    if not room_id:
        return JsonResponse({
            'success': False,
            'message': 'Room selection is required.'
        })
    
    if not start_bed or not end_bed:
        return JsonResponse({
            'success': False,
            'message': 'Start and end bed numbers are required.'
        })
    
    if not bed_type:
        return JsonResponse({
            'success': False,
            'message': 'Bed type is required.'
        })
    
    # Validate bed type
    valid_types = [choice[0] for choice in Bed.BED_TYPES]
    if bed_type not in valid_types:
        return JsonResponse({
            'success': False,
            'message': 'Invalid bed type selected.'
        })
    
    # Get room
    try:
        room = HostelRoom.objects.select_related('hostel').get(id=room_id, is_active=True)
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Selected room does not exist or is inactive.'
        })
    
    # Parse bed range (assuming numeric bed numbers)
    if start_bed.isdigit() and end_bed.isdigit():
        start_num = int(start_bed)
        end_num = int(end_bed)
        
        if start_num > end_num:
            return JsonResponse({
                'success': False,
                'message': 'Start bed number must be less than or equal to end bed number.'
            })
        
        if end_num - start_num > 20:
            return JsonResponse({
                'success': False,
                'message': 'Cannot create more than 20 beds at once.'
            })
        
        bed_numbers = [str(i).zfill(len(start_bed)) for i in range(start_num, end_num + 1)]
    else:
        # For alphanumeric, just create the two specified beds
        bed_numbers = [start_bed, end_bed] if start_bed != end_bed else [start_bed]
    
    # Check room capacity
    current_beds_count = Bed.objects.filter(room=room).count()
    available_slots = room.capacity - current_beds_count
    
    if len(bed_numbers) > available_slots:
        return JsonResponse({
            'success': False,
            'message': f'Cannot create {len(bed_numbers)} beds. Only {available_slots} slots available in this room.'
        })
    
    created_count = 0
    failed_beds = []
    
    with transaction.atomic():
        for bed_num in bed_numbers:
            try:
                # Check if bed already exists
                if Bed.objects.filter(room=room, bed_number__iexact=bed_num).exists():
                    failed_beds.append(f"{bed_num} (Already exists)")
                    continue
                
                # Create bed
                Bed.objects.create(
                    room=room,
                    bed_number=bed_num,
                    bed_type=bed_type,
                    is_occupied=False
                )
                created_count += 1
                
            except Exception as e:
                failed_beds.append(f"{bed_num} ({str(e)})")
    
    message = f'Successfully created {created_count} bed(s) in Room {room.room_number}.'
    if failed_beds:
        message += f' Failed: {", ".join(failed_beds[:5])}'
        if len(failed_beds) > 5:
            message += f' and {len(failed_beds) - 5} more.'
    
    return JsonResponse({
        'success': True,
        'message': message,
        'created_count': created_count,
        'failed_count': len(failed_beds)
    })    



@login_required
def student_hostel_allocations_list(request):
    """
    Display student hostel allocations management page
    """
    # Get all allocations with related data
    allocations = StudentHostelAllocation.objects.select_related(
        'student', 'hostel', 'room', 'bed', 'academic_year'
    ).all().order_by('-start_date', 'student__first_name')
    
    # Get active academic year
    active_academic_year = AcademicYear.objects.filter(is_active=True).first()
    
    # Get statistics
    total_allocations = allocations.count()
    active_allocations = allocations.filter(is_active=True).count()
    
    # Get students with no active allocation
    students_with_allocation = allocations.filter(is_active=True).values_list('student_id', flat=True)
    available_students = Student.objects.filter(
        is_active=True,
        status='active'
    ).exclude(id__in=students_with_allocation).count()
    
    # Get hostel occupancy
    hostels = Hostel.objects.filter(is_active=True)
    hostel_stats = []
    for hostel in hostels:
        allocated = allocations.filter(hostel=hostel, is_active=True).count()
        capacity = hostel.max_students
        occupancy_rate = round((allocated / capacity * 100), 1) if capacity > 0 else 0
        hostel_stats.append({
            'hostel': hostel,
            'allocated': allocated,
            'capacity': capacity,
            'available': capacity - allocated,
            'occupancy_rate': occupancy_rate
        })
    
    # Get recent allocations
    recent_allocations = allocations[:10]
    
    context = {
        'allocations': allocations,
        'recent_allocations': recent_allocations,
        'total_allocations': total_allocations,
        'active_allocations': active_allocations,
        'available_students': available_students,
        'hostel_stats': hostel_stats,
        'active_academic_year': active_academic_year,
        'hostels': Hostel.objects.filter(is_active=True),
        'academic_years': AcademicYear.objects.all().order_by('-start_date'),
        'page_title': 'Student Hostel Allocations',
    }
    
    return render(request, 'admin/hostels/student_allocations_list.html', context)



# ============================================================================
# SINGLE STUDENT ALLOCATION VIEWS
# ============================================================================

@login_required
def student_allocation_create(request):
    """
    Create a new student hostel allocation (single student)
    """
    if request.method == 'POST':
        try:
            # Get form data
            student_id = request.POST.get('student')
            hostel_id = request.POST.get('hostel')
            room_id = request.POST.get('room')
            bed_id = request.POST.get('bed')
            academic_year_id = request.POST.get('academic_year')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')

            # Validate required fields
            if not all([student_id, hostel_id, academic_year_id, start_date]):
                return JsonResponse({
                    'success': False,
                    'message': 'Please fill all required fields.'
                })

            # Get related objects
            student = get_object_or_404(Student, id=student_id, is_active=True)
            hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
            academic_year = get_object_or_404(AcademicYear, id=academic_year_id)

            # Check if student already has active allocation
            if StudentHostelAllocation.objects.filter(student=student, is_active=True).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'{student.full_name} already has an active allocation.'
                })

            # Check hostel capacity
            current_allocations = StudentHostelAllocation.objects.filter(
                hostel=hostel, is_active=True
            ).count()
            
            if current_allocations >= hostel.max_students:
                return JsonResponse({
                    'success': False,
                    'message': f'{hostel.name} has reached maximum capacity ({hostel.max_students} students).'
                })

            # Validate room if provided
            room = None
            if room_id:
                room = get_object_or_404(HostelRoom, id=room_id, hostel=hostel, is_active=True)
                
                # Check room capacity
                room_allocations = StudentHostelAllocation.objects.filter(
                    room=room, is_active=True
                ).count()
                
                if room_allocations >= room.capacity:
                    return JsonResponse({
                        'success': False,
                        'message': f'Room {room.room_number} is full.'
                    })

            # Validate bed if provided
            bed = None
            if bed_id:
                if not room:
                    return JsonResponse({
                        'success': False,
                        'message': 'Room must be selected when assigning a bed.'
                    })
                
                bed = get_object_or_404(Bed, id=bed_id, room=room, is_occupied=False)
                
                # Check if bed is already allocated
                if StudentHostelAllocation.objects.filter(bed=bed, is_active=True).exists():
                    return JsonResponse({
                        'success': False,
                        'message': f'Bed {bed.bed_number} is already occupied.'
                    })

            # Parse dates
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                if end_date:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if end_date <= start_date:
                        return JsonResponse({
                            'success': False,
                            'message': 'End date must be after start date.'
                        })
                else:
                    end_date = academic_year.end_date
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format.'
                })

            # Create allocation
            with transaction.atomic():
                allocation = StudentHostelAllocation.objects.create(
                    student=student,
                    hostel=hostel,
                    room=room,
                    bed=bed,
                    academic_year=academic_year,
                    start_date=start_date,
                    end_date=end_date,
                    is_active=True
                )

                # Mark bed as occupied if assigned
                if bed:
                    bed.is_occupied = True
                    bed.save()

            messages.success(request, f'Successfully allocated {student.full_name} to {hostel.name}.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': 'list'})
            return redirect('admin_student_allocations_list')

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)})
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_student_allocation_create')

    # GET request - show form with enhanced context data
    # Get available students (without active allocations)
    students_with_allocation = StudentHostelAllocation.objects.filter(
        is_active=True
    ).values_list('student_id', flat=True)
    
    available_students = Student.objects.filter(
        is_active=True,
        status='active'
    ).exclude(id__in=students_with_allocation).select_related(
        'class_level', 'stream_class'
    ).order_by('first_name', 'last_name')
    
    # Get hostels with current allocation counts and available spaces
    hostels = Hostel.objects.filter(is_active=True).prefetch_related(
        'rooms', 'rooms__beds'
    ).order_by('name')
    
    # Annotate hostels with current allocations
    hostel_list = []
    for hostel in hostels:
        current_allocations = StudentHostelAllocation.objects.filter(
            hostel=hostel, is_active=True
        ).count()
        
        # Make hostel object have the attributes as properties
        hostel.current_allocations = current_allocations
        hostel.available_spaces = hostel.max_students - current_allocations
        
        hostel_list.append(hostel)
    
    # Get academic years
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    
    # Get class levels for potential filtering
    classes = ClassLevel.objects.all().order_by('name')
    
    context = {
        'title': 'New Student Allocation',
        'students': available_students,  # Available students for dropdown
        'available_students': available_students.count(),  # Count for badge
        'hostels': hostel_list,  # Hostels with allocation data
        'academic_years': academic_years,
        'classes': classes,
        'page_title': 'New Allocation',
        'current_date': timezone.now().date().isoformat(),  # For default start date
    }
    
    return render(request, 'admin/hostels/allocation_form.html', context)


@login_required
def student_allocation_edit(request, pk):
    """
    Edit an existing allocation with enhanced context data
    """
    # Get allocation with related data
    allocation = get_object_or_404(
        StudentHostelAllocation.objects.select_related(
            'student', 'hostel', 'room', 'bed', 'academic_year'
        ),
        id=pk
    )

    if request.method == 'POST':
        try:
            # Get form data
            room_id = request.POST.get('room')
            bed_id = request.POST.get('bed')
            end_date = request.POST.get('end_date')
            is_active = request.POST.get('is_active') == 'on'

            # Validate room if changing
            room = allocation.room
            if room_id and str(allocation.room_id) != room_id:
                room = get_object_or_404(HostelRoom, id=room_id, hostel=allocation.hostel, is_active=True)
                
                # Check room capacity
                room_allocations = StudentHostelAllocation.objects.filter(
                    room=room, is_active=True
                ).exclude(id=allocation.id).count()
                
                if room_allocations >= room.capacity:
                    return JsonResponse({
                        'success': False,
                        'message': f'Room {room.room_number} has reached maximum capacity.'
                    })

            # Validate bed if changing
            bed = allocation.bed
            if bed_id and str(allocation.bed_id) != bed_id:
                if not room:
                    return JsonResponse({
                        'success': False,
                        'message': 'Room must be selected when assigning a bed.'
                    })
                
                bed = get_object_or_404(Bed, id=bed_id, room=room, is_occupied=False)
                
                # Check if bed is already allocated
                if StudentHostelAllocation.objects.filter(bed=bed, is_active=True).exclude(id=allocation.id).exists():
                    return JsonResponse({
                        'success': False,
                        'message': f'Bed {bed.bed_number} is already occupied.'
                    })

            # Validate end date
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if end_date <= allocation.start_date:
                        return JsonResponse({
                            'success': False,
                            'message': 'End date must be after start date.'
                        })
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'message': 'Invalid date format.'
                    })
            else:
                end_date = None

            # Update allocation
            with transaction.atomic():
                # Free old bed if changing
                if bed and allocation.bed and allocation.bed.id != bed.id:
                    allocation.bed.is_occupied = False
                    allocation.bed.save()
                    
                    # Mark new bed as occupied if allocation is active
                    if is_active:
                        bed.is_occupied = True
                        bed.save()

                allocation.room = room
                allocation.bed = bed
                allocation.end_date = end_date
                
                # Handle status change
                if allocation.is_active != is_active:
                    allocation.is_active = is_active
                    if allocation.bed:
                        allocation.bed.is_occupied = is_active
                        allocation.bed.save()
                    
                    if not is_active and not allocation.end_date:
                        allocation.end_date = timezone.now().date()

                allocation.save()

            messages.success(request, f'Allocation updated successfully.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': 'list'})
            return redirect('admin_student_allocations_list')

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)})
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_student_allocation_edit', pk=pk)

    # GET request - show form with enhanced context data
    # Get available rooms for this hostel with current allocation counts
    hostel_rooms = HostelRoom.objects.filter(
        hostel=allocation.hostel,
        is_active=True
    ).annotate(
        current_allocations=Count(
            'studenthostelallocation',
            filter=Q(studenthostelallocation__is_active=True) & ~Q(studenthostelallocation__id=allocation.id)
        )
    ).order_by('room_number')

    # Add available spaces as a property
    for room in hostel_rooms:
        room.available_spaces = room.capacity - room.current_allocations

    # Get all hostels for potential change (if needed in future)
    hostels = Hostel.objects.filter(is_active=True).annotate(
        current_allocations=Count(
            'studenthostelallocation',
            filter=Q(studenthostelallocation__is_active=True)
        )
    ).order_by('name')

    # Calculate hostel available spaces
    for hostel in hostels:
        hostel.available_spaces = hostel.max_students - hostel.current_allocations

    # Get academic years
    academic_years = AcademicYear.objects.all().order_by('-start_date')

    # Get class levels for potential filtering
    classes = ClassLevel.objects.all().order_by('name')

    # Get available beds for the current room if room is assigned
    available_beds = []
    if allocation.room:
        occupied_bed_ids = StudentHostelAllocation.objects.filter(
            bed__isnull=False,
            is_active=True
        ).exclude(id=allocation.id).values_list('bed_id', flat=True)

        available_beds = Bed.objects.filter(
            room=allocation.room,
            is_occupied=False
        ).exclude(id__in=occupied_bed_ids).order_by('bed_number')

    context = {
        'title': 'Edit Allocation',
        'allocation': allocation,
        'student': allocation.student,
        'hostels': hostels,
        'hostel_rooms': hostel_rooms,
        'available_beds': available_beds,
        'academic_years': academic_years,
        'classes': classes,
        'page_title': f'Edit Allocation - {allocation.student.full_name}',
        'current_date': timezone.now().date().isoformat(),
    }
    
    return render(request, 'admin/hostels/allocation_edit.html', context)


@login_required
def student_allocation_detail(request, pk):
    """
    Display detailed information about a specific allocation with enhanced data
    """
    # Get allocation with all related data
    allocation = get_object_or_404(
        StudentHostelAllocation.objects.select_related(
            'student',
            'student__class_level',
            'student__stream_class',
            'hostel',
            'room',
            'bed',
            'academic_year'
        ),
        id=pk
    )
    
    # Get payment transactions for this allocation
    payment_transactions = HostelPaymentTransaction.objects.filter(
        allocation=allocation
    ).select_related(
        'installment_payment__installment_plan'
    ).order_by('-payment_date', '-created_at')
    
    # Get installment payments if applicable
    installment_payments = []
    if allocation.hostel.payment_mode == 'installments':
        installment_payments = HostelPayment.objects.filter(
            allocation=allocation
        ).select_related(
            'installment_plan'
        ).order_by('installment_plan__installment_number')
    
    # Calculate financial summaries
    total_paid = payment_transactions.aggregate(total=Sum('amount'))['total'] or 0
    total_fee = allocation.total_fee
    balance = total_fee - total_paid
    
    # Get payment statistics
    payment_stats = {
        'total_transactions': payment_transactions.count(),
        'last_payment_date': payment_transactions.first().payment_date if payment_transactions.exists() else None,
        'last_payment_amount': payment_transactions.first().amount if payment_transactions.exists() else None,
        'payment_methods': payment_transactions.values_list('payment_method', flat=True).distinct(),
    }
    
    # Get student's other allocations (historical)
    other_allocations = StudentHostelAllocation.objects.filter(
        student=allocation.student
    ).exclude(
        id=allocation.id
    ).select_related(
        'hostel', 'academic_year'
    ).order_by('-start_date')[:5]
    
    context = {
        'allocation': allocation,
        'payment_transactions': payment_transactions,
        'installment_payments': installment_payments,
        'total_paid': total_paid,
        'total_fee': total_fee,
        'balance': balance,
        'payment_stats': payment_stats,
        'other_allocations': other_allocations,
        'page_title': f'Allocation Details - {allocation.student.full_name}',
        'is_fully_paid': balance <= 0,
        'payment_percentage': (total_paid / total_fee * 100) if total_fee > 0 else 0,
    }
    
    return render(request, 'admin/hostels/allocation_detail.html', context)


@login_required
def student_allocation_delete(request):
    """
    Delete an allocation (AJAX only)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

    allocation_id = request.POST.get('id')
    if not allocation_id:
        return JsonResponse({'success': False, 'message': 'Allocation ID is required.'})

    try:
        allocation = StudentHostelAllocation.objects.select_related('bed').get(id=allocation_id)
        
        with transaction.atomic():
            # Free the bed if occupied
            if allocation.bed and allocation.bed.is_occupied:
                allocation.bed.is_occupied = False
                allocation.bed.save()
            
            student_name = allocation.student.full_name
            hostel_name = allocation.hostel.name
            allocation.delete()

        return JsonResponse({
            'success': True,
            'message': f'Allocation for {student_name} deleted successfully.'
        })

    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Allocation not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def student_allocation_toggle(request):
    """
    Toggle allocation active status (AJAX only)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

    allocation_id = request.POST.get('id')
    if not allocation_id:
        return JsonResponse({'success': False, 'message': 'Allocation ID is required.'})

    try:
        allocation = StudentHostelAllocation.objects.select_related('bed').get(id=allocation_id)
        
        with transaction.atomic():
            # Toggle status
            allocation.is_active = not allocation.is_active
            
            # Update bed occupancy
            if allocation.bed:
                allocation.bed.is_occupied = allocation.is_active
                allocation.bed.save()
            
            # Set end date if deactivating
            if not allocation.is_active and not allocation.end_date:
                allocation.end_date = timezone.now().date()
            
            allocation.save()

        status_text = "activated" if allocation.is_active else "deactivated"
        return JsonResponse({
            'success': True,
            'message': f'Allocation {status_text} successfully.',
            'is_active': allocation.is_active
        })

    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Allocation not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ============================================================================
# BULK OPERATIONS VIEWS
# ============================================================================

@login_required
def student_allocation_bulk(request):
    """
    Bulk operations page - select multiple students for bulk actions
    """
    context = {
        'hostels': Hostel.objects.filter(is_active=True),
        'academic_years': AcademicYear.objects.all().order_by('-start_date'),
        'classes': ClassLevel.objects.all(),
        'page_title': 'Bulk Operations'
    }
    return render(request, 'admin/hostels/bulk_operations.html', context)


@login_required
def student_allocation_bulk_create(request):
    """
    Create multiple allocations at once (AJAX only)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

    # Get form data
    student_ids = request.POST.getlist('students[]')
    hostel_id = request.POST.get('hostel')
    academic_year_id = request.POST.get('academic_year')
    start_date = request.POST.get('start_date')

    # Validate input
    if not student_ids:
        return JsonResponse({'success': False, 'message': 'No students selected.'})
    
    if not hostel_id:
        return JsonResponse({'success': False, 'message': 'Hostel is required.'})
    
    if not academic_year_id:
        return JsonResponse({'success': False, 'message': 'Academic year is required.'})
    
    if not start_date:
        return JsonResponse({'success': False, 'message': 'Start date is required.'})

    try:
        # Get related objects
        hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

        # Check hostel capacity
        current_allocations = StudentHostelAllocation.objects.filter(
            hostel=hostel, is_active=True
        ).count()
        
        available_slots = hostel.max_students - current_allocations
        
        if len(student_ids) > available_slots:
            return JsonResponse({
                'success': False,
                'message': f'Cannot allocate {len(student_ids)} students. Only {available_slots} slots available.'
            })

        # Process each student
        created_count = 0
        failed_students = []

        with transaction.atomic():
            for student_id in student_ids:
                try:
                    student = Student.objects.get(id=student_id, is_active=True)
                    
                    # Check if student already has allocation
                    if StudentHostelAllocation.objects.filter(student=student, is_active=True).exists():
                        failed_students.append(f"{student.full_name} (Already allocated)")
                        continue
                    
                    # Create allocation
                    StudentHostelAllocation.objects.create(
                        student=student,
                        hostel=hostel,
                        academic_year=academic_year,
                        start_date=start_date,
                        is_active=True
                    )
                    created_count += 1
                    
                except Student.DoesNotExist:
                    failed_students.append(f"Student ID {student_id} (Not found)")
                except Exception as e:
                    failed_students.append(f"Student ID {student_id} ({str(e)})")

        # Prepare response message
        message = f'Successfully allocated {created_count} student(s) to {hostel.name}.'
        if failed_students:
            message += f' Failed: {", ".join(failed_students[:5])}'
            if len(failed_students) > 5:
                message += f' and {len(failed_students) - 5} more.'

        return JsonResponse({
            'success': True,
            'message': message,
            'created_count': created_count,
            'failed_count': len(failed_students)
        })

    except Hostel.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hostel not found.'})
    except AcademicYear.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Academic year not found.'})
    except ValueError as e:
        return JsonResponse({'success': False, 'message': f'Invalid date format: {str(e)}'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def student_allocation_bulk_remove(request):
    """
    Remove multiple allocations (vacate students) at once (AJAX only)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

    student_ids = request.POST.getlist('students[]')
    
    if not student_ids:
        return JsonResponse({'success': False, 'message': 'No students selected.'})

    try:
        removed_count = 0
        failed_students = []

        with transaction.atomic():
            for student_id in student_ids:
                try:
                    # Find active allocation for this student
                    allocation = StudentHostelAllocation.objects.filter(
                        student_id=student_id,
                        is_active=True
                    ).select_related('student', 'bed').first()

                    if not allocation:
                        student = Student.objects.get(id=student_id)
                        failed_students.append(f"{student.full_name} (No active allocation)")
                        continue

                    # Vacate the allocation
                    allocation.is_active = False
                    allocation.end_date = timezone.now().date()
                    
                    # Free the bed
                    if allocation.bed:
                        allocation.bed.is_occupied = False
                        allocation.bed.save()
                    
                    allocation.save()
                    removed_count += 1

                except Student.DoesNotExist:
                    failed_students.append(f"Student ID {student_id} (Not found)")
                except Exception as e:
                    student = Student.objects.filter(id=student_id).first()
                    name = student.full_name if student else f"Student ID {student_id}"
                    failed_students.append(f"{name} ({str(e)})")

        # Prepare response message
        message = f'Successfully vacated {removed_count} student(s).'
        if failed_students:
            message += f' Failed: {", ".join(failed_students[:5])}'
            if len(failed_students) > 5:
                message += f' and {len(failed_students) - 5} more.'

        return JsonResponse({
            'success': True,
            'message': message,
            'removed_count': removed_count,
            'failed_count': len(failed_students)
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ============================================================================
# AJAX ENDPOINTS (Already exist but ensure they're compatible)
# ============================================================================


@login_required
def check_availability(request):
    """
    AJAX endpoint to check availability of hostel/room/bed combination
    """
    try:
        hostel_id = request.GET.get('hostel_id')
        room_id = request.GET.get('room_id')
        bed_id = request.GET.get('bed_id')
        
        response = {'success': True, 'available': True, 'messages': []}

        # Check hostel capacity
        if hostel_id:
            hostel = Hostel.objects.get(id=hostel_id)
            current = StudentHostelAllocation.objects.filter(
                hostel=hostel, is_active=True
            ).count()
            
            if current >= hostel.max_students:
                response['available'] = False
                response['messages'].append(f"Hostel is full")

        # Check room availability
        if room_id and response['available']:
            room = HostelRoom.objects.get(id=room_id)
            current = StudentHostelAllocation.objects.filter(
                room=room, is_active=True
            ).count()
            
            if current >= room.capacity:
                response['available'] = False
                response['messages'].append(f"Room is full")

        # Check bed availability
        if bed_id and response['available']:
            bed = Bed.objects.get(id=bed_id)
            if bed.is_occupied:
                response['available'] = False
                response['messages'].append(f"Bed is occupied")

        return JsonResponse(response)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@login_required
def get_available_students(request):
    """
    AJAX endpoint to get students available for allocation
    Returns data in format compatible with Select2 and bulk allocation table
    Includes detailed student information for the datatable
    """
    try:
        # Get parameters
        search = request.GET.get('search', '').strip()
        gender = request.GET.get('gender', '')
        class_level_id = request.GET.get('class_level', '')
        
        # Get students with no active allocation
        students_with_allocation = StudentHostelAllocation.objects.filter(
            is_active=True
        ).values_list('student_id', flat=True)
        
        # Base queryset - only students without active allocations
        students = Student.objects.filter(
            is_active=True,
            status='active'
        ).exclude(id__in=students_with_allocation).select_related(
            'class_level', 'stream_class'
        )
        
        # Apply search filter using individual fields
        if search:
            students = students.filter(
                Q(first_name__icontains=search) |
                Q(middle_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(registration_number__icontains=search) |
                Q(examination_number__icontains=search)
            )
        
        # Apply additional filters if provided
        if gender:
            students = students.filter(gender=gender)
        
        if class_level_id:
            students = students.filter(class_level_id=class_level_id)
        
        # Get total count before limiting
        total_count = students.count()
        
        # Limit results for performance (but get all for bulk modal)
        # For bulk modal we want all available students, for Select2 we limit
        if request.GET.get('source') == 'bulk':
            # Get all students for bulk modal
            students = students[:500]  # Reasonable limit
        else:
            # Limit for Select2 dropdown
            students = students[:50]
        
        student_list = []
        for student in students:
            # Create display text for Select2
            display_text = f"{student.full_name} ({student.registration_number})"
            if student.class_level:
                display_text += f" - {student.class_level.name}"
                if student.stream_class:
                    display_text += f" {student.stream_class.stream_letter}"
            
            # Get gender display value
            gender_display = dict(GENDER_CHOICES).get(student.gender, 'N/A') if student.gender else 'N/A'
            
            student_list.append({
                'id': student.id,
                'text': display_text,
                'full_name': student.full_name,
                'first_name': student.first_name,
                'middle_name': student.middle_name,
                'last_name': student.last_name,
                'registration_number': student.registration_number,
                'examination_number': student.examination_number,
                'gender': gender_display,
                'gender_code': student.gender,
                'class_level': student.class_level.name if student.class_level else 'N/A',
                'class_level_id': student.class_level.id if student.class_level else None,
                'stream': student.stream_class.stream_letter if student.stream_class else 'N/A',
                'stream_id': student.stream_class.id if student.stream_class else None,
                'admission_year': student.admission_year,
                'status': student.status
            })
        
        # Prepare response based on source
        response = {
            'results': student_list,
            'pagination': {
                'more': False,
                'total': total_count
            }
        }
        
        # Add statistics if requested for bulk modal
        if request.GET.get('include_stats') == 'true':
            # Calculate gender statistics
            gender_stats = Student.objects.filter(
                is_active=True,
                status='active'
            ).exclude(id__in=students_with_allocation).values('gender').annotate(
                count=Count('id')
            ).order_by('gender')
            
            # Calculate class level statistics
            class_stats = Student.objects.filter(
                is_active=True,
                status='active'
            ).exclude(id__in=students_with_allocation).values(
                'class_level__id', 'class_level__name'
            ).annotate(
                count=Count('id')
            ).order_by('-count')[:10]
            
            response['statistics'] = {
                'total_available': total_count,
                'gender_breakdown': [
                    {
                        'gender': dict(GENDER_CHOICES).get(stat['gender'], 'Unknown'),
                        'count': stat['count']
                    } for stat in gender_stats if stat['gender']
                ],
                'class_breakdown': [
                    {
                        'class_id': stat['class_level__id'],
                        'class_name': stat['class_level__name'],
                        'count': stat['count']
                    } for stat in class_stats if stat['class_level__id']
                ]
            }
        
        return JsonResponse(response)
        
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in get_available_students: {str(e)}", exc_info=True)
        
        return JsonResponse({
            'results': [],
            'error': str(e),
            'pagination': {'more': False, 'total': 0}
        }, status=500)


@login_required
def get_available_rooms(request, hostel_id):
    """
    AJAX endpoint to get available rooms in a hostel for allocation
    """
    try:
        hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
        
        # Get all active rooms in this hostel
        rooms = HostelRoom.objects.filter(
            hostel=hostel,
            is_active=True
        ).order_by('room_number')
        
        room_list = []
        for room in rooms:
            # Count current beds and allocations
            total_beds = Bed.objects.filter(room=room).count()
            occupied_beds = Bed.objects.filter(room=room, is_occupied=True).count()
            
            # Get active allocations in this room
            active_allocations = StudentHostelAllocation.objects.filter(
                room=room,
                is_active=True
            ).count()
            
            # Calculate available spaces
            available_spaces = room.capacity - active_allocations
            
            room_list.append({
                'id': room.id,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'total_beds': total_beds,
                'occupied_beds': occupied_beds,
                'active_allocations': active_allocations,
                'available_spaces': available_spaces,
                'has_available_spaces': available_spaces > 0,
                'display': f"Room {room.room_number} - {active_allocations}/{room.capacity} students"
            })
        
        return JsonResponse({
            'success': True,
            'rooms': room_list,
            'hostel_name': hostel.name
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@login_required
def get_available_beds(request, room_id):
    """
    AJAX endpoint to get available beds in a room for allocation
    """
    try:
        room = get_object_or_404(HostelRoom, id=room_id, is_active=True)
        
        # Get beds that are not occupied and not allocated
        occupied_bed_ids = StudentHostelAllocation.objects.filter(
            bed__isnull=False,
            is_active=True
        ).values_list('bed_id', flat=True)
        
        available_beds = Bed.objects.filter(
            room=room,
            is_occupied=False
        ).exclude(id__in=occupied_bed_ids).order_by('bed_number')
        
        bed_list = []
        for bed in available_beds:
            bed_list.append({
                'id': bed.id,
                'bed_number': bed.bed_number,
                'bed_type': bed.bed_type,
                'bed_type_display': bed.get_bed_type_display(),
                'display': f"Bed {bed.bed_number} ({bed.get_bed_type_display()})"
            })
        
        # Get current allocations count
        current_allocations = StudentHostelAllocation.objects.filter(
            room=room,
            is_active=True
        ).count()
        
        return JsonResponse({
            'success': True,
            'beds': bed_list,
            'room_number': room.room_number,
            'current_allocations': current_allocations,
            'room_capacity': room.capacity,
            'available_spaces': room.capacity - current_allocations,
            'has_available_beds': len(bed_list) > 0
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })




@login_required
def student_hostel_allocations_crud(request):
    """
    Handle AJAX CRUD operations for student hostel allocations
    """
    if request.method == 'POST':
        action = request.POST.get('action', '').lower()
        
        try:
            if action == 'create':
                return create_allocation(request)
            elif action == 'update':
                return update_allocation(request)
            elif action == 'toggle_status':
                return toggle_allocation_status(request)
            elif action == 'delete':
                return delete_allocation(request)
            elif action == 'bulk_create':
                return bulk_create_allocations(request)
            elif action == 'vacate':
                return vacate_allocation(request)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid action specified.'
                })
                
        except ValidationError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'POST request required.'
    })


def create_allocation(request):
    """Create a new student hostel allocation"""
    # Get required fields
    student_id = request.POST.get('student', '').strip()
    hostel_id = request.POST.get('hostel', '').strip()
    academic_year_id = request.POST.get('academic_year', '').strip()
    start_date = request.POST.get('start_date', '').strip()
    
    # Optional fields
    room_id = request.POST.get('room', '').strip()
    bed_id = request.POST.get('bed', '').strip()
    end_date = request.POST.get('end_date', '').strip()
    
    # Validate required fields
    if not student_id:
        return JsonResponse({'success': False, 'message': 'Student is required.'})
    if not hostel_id:
        return JsonResponse({'success': False, 'message': 'Hostel is required.'})
    if not academic_year_id:
        return JsonResponse({'success': False, 'message': 'Academic year is required.'})
    if not start_date:
        return JsonResponse({'success': False, 'message': 'Start date is required.'})
    
    # Validate student
    try:
        student = Student.objects.get(id=student_id, is_active=True)
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Student not found or inactive.'})
    
    # Check if student already has active allocation
    if StudentHostelAllocation.objects.filter(student=student, is_active=True).exists():
        return JsonResponse({
            'success': False, 
            'message': f'Student {student.full_name} already has an active allocation.'
        })
    
    # Validate hostel
    try:
        hostel = Hostel.objects.get(id=hostel_id, is_active=True)
    except Hostel.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Hostel not found or inactive.'})
    
    # Check hostel capacity
    current_allocations = StudentHostelAllocation.objects.filter(
        hostel=hostel, 
        is_active=True
    ).count()
    
    if current_allocations >= hostel.max_students:
        return JsonResponse({
            'success': False,
            'message': f'Hostel {hostel.name} has reached maximum capacity ({hostel.max_students} students).'
        })
    
    # Validate academic year
    try:
        academic_year = AcademicYear.objects.get(id=academic_year_id)
    except AcademicYear.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Academic year not found.'})
    
    # Validate room if provided
    room = None
    if room_id:
        try:
            room = HostelRoom.objects.get(id=room_id, hostel=hostel, is_active=True)
            
            # Check room capacity
            room_allocations = StudentHostelAllocation.objects.filter(
                room=room,
                is_active=True
            ).count()
            
            if room_allocations >= room.capacity:
                return JsonResponse({
                    'success': False,
                    'message': f'Room {room.room_number} has reached maximum capacity.'
                })
                
        except HostelRoom.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Room not found or does not belong to selected hostel.'})
    
    # Validate bed if provided
    bed = None
    if bed_id:
        if not room:
            return JsonResponse({
                'success': False, 
                'message': 'Room must be selected when assigning a specific bed.'
            })
        
        try:
            bed = Bed.objects.get(id=bed_id, room=room, is_occupied=False)
            
            # Check if bed is already allocated
            if StudentHostelAllocation.objects.filter(bed=bed, is_active=True).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Bed {bed.bed_number} is already allocated.'
                })
                
            # Mark bed as occupied
            bed.is_occupied = True
            bed.save()
            
        except Bed.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Bed not found or already occupied.'})
    
    # Validate dates
    try:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid start date format.'})
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            if end_date <= start_date:
                return JsonResponse({
                    'success': False, 
                    'message': 'End date must be after start date.'
                })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid end date format.'})
    else:
        # Set default end date to end of academic year if not provided
        if academic_year.end_date:
            end_date = academic_year.end_date
        else:
            end_date = None
    
    try:
        # Create allocation
        allocation = StudentHostelAllocation.objects.create(
            student=student,
            hostel=hostel,
            room=room,
            bed=bed,
            academic_year=academic_year,
            start_date=start_date,
            end_date=end_date,
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully allocated {student.full_name} to {hostel.name}.',
            'allocation': {
                'id': allocation.id,
                'student': student.full_name,
                'student_id': student.id,
                'hostel': hostel.name,
                'hostel_id': hostel.id,
                'room': room.room_number if room else None,
                'bed': bed.bed_number if bed else None,
                'academic_year': str(academic_year),
                'start_date': allocation.start_date.strftime('%Y-%m-%d'),
                'end_date': allocation.end_date.strftime('%Y-%m-%d') if allocation.end_date else None,
                'is_active': allocation.is_active
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error creating allocation: {str(e)}'
        })


def update_allocation(request):
    """Update an existing allocation"""
    allocation_id = request.POST.get('id')
    if not allocation_id:
        return JsonResponse({'success': False, 'message': 'Allocation ID is required.'})
    
    try:
        allocation = StudentHostelAllocation.objects.select_related(
            'student', 'hostel', 'room', 'bed', 'academic_year'
        ).get(id=allocation_id)
    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Allocation not found.'})
    
    # Get fields
    room_id = request.POST.get('room', '').strip()
    bed_id = request.POST.get('bed', '').strip()
    end_date = request.POST.get('end_date', '').strip()
    is_active = request.POST.get('is_active') == 'on' or request.POST.get('is_active') == 'true'
    
    # Validate room if changing
    room = allocation.room
    if room_id and str(allocation.room_id) != room_id:
        try:
            room = HostelRoom.objects.get(id=room_id, hostel=allocation.hostel, is_active=True)
            
            # Check room capacity
            room_allocations = StudentHostelAllocation.objects.filter(
                room=room,
                is_active=True
            ).exclude(id=allocation.id).count()
            
            if room_allocations >= room.capacity:
                return JsonResponse({
                    'success': False,
                    'message': f'Room {room.room_number} has reached maximum capacity.'
                })
                
        except HostelRoom.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Room not found.'})
    
    # Validate bed if changing
    bed = allocation.bed
    if bed_id and str(allocation.bed_id) != bed_id:
        if not room:
            return JsonResponse({
                'success': False, 
                'message': 'Room must be selected when assigning a specific bed.'
            })
        
        try:
            bed = Bed.objects.get(id=bed_id, room=room, is_occupied=False)
            
            # Check if bed is already allocated
            if StudentHostelAllocation.objects.filter(bed=bed, is_active=True).exclude(id=allocation.id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Bed {bed.bed_number} is already allocated.'
                })
                
            # Free old bed if exists
            if allocation.bed:
                allocation.bed.is_occupied = False
                allocation.bed.save()
                
            # Mark new bed as occupied
            bed.is_occupied = True
            bed.save()
            
        except Bed.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Bed not found or already occupied.'})
    
    # Validate end date
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            if end_date <= allocation.start_date:
                return JsonResponse({
                    'success': False, 
                    'message': 'End date must be after start date.'
                })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid end date format.'})
    else:
        end_date = None
    
    try:
        # Update allocation
        if room:
            allocation.room = room
        if bed:
            allocation.bed = bed
        if end_date:
            allocation.end_date = end_date
        
        # Handle status change
        if allocation.is_active != is_active:
            allocation.is_active = is_active
            if allocation.bed:
                allocation.bed.is_occupied = is_active
                allocation.bed.save()
        
        allocation.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Allocation for {allocation.student.full_name} updated successfully.',
            'allocation': {
                'id': allocation.id,
                'is_active': allocation.is_active
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating allocation: {str(e)}'
        })


def toggle_allocation_status(request):
    """Toggle allocation active/inactive status"""
    allocation_id = request.POST.get('id')
    if not allocation_id:
        return JsonResponse({'success': False, 'message': 'Allocation ID is required.'})
    
    try:
        allocation = StudentHostelAllocation.objects.select_related('bed').get(id=allocation_id)
    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Allocation not found.'})
    
    try:
        # Toggle status
        allocation.is_active = not allocation.is_active
        
        # Update bed occupancy
        if allocation.bed:
            allocation.bed.is_occupied = allocation.is_active
            allocation.bed.save()
        
        # If deactivating, set end date to now
        if not allocation.is_active and not allocation.end_date:
            allocation.end_date = timezone.now().date()
        
        allocation.save()
        
        status_text = "activated" if allocation.is_active else "deactivated"
        
        return JsonResponse({
            'success': True,
            'message': f'Allocation for {allocation.student.full_name} {status_text} successfully.',
            'is_active': allocation.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error toggling allocation status: {str(e)}'
        })


def vacate_allocation(request):
    """Vacate a student allocation (mark as inactive with end date)"""
    allocation_id = request.POST.get('id')
    if not allocation_id:
        return JsonResponse({'success': False, 'message': 'Allocation ID is required.'})
    
    try:
        allocation = StudentHostelAllocation.objects.select_related('bed').get(id=allocation_id)
    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Allocation not found.'})
    
    if not allocation.is_active:
        return JsonResponse({'success': False, 'message': 'Allocation is already inactive.'})
    
    try:
        # Deactivate allocation
        allocation.is_active = False
        allocation.end_date = timezone.now().date()
        
        # Free the bed
        if allocation.bed:
            allocation.bed.is_occupied = False
            allocation.bed.save()
        
        allocation.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully vacated {allocation.student.full_name} from {allocation.hostel.name}.',
            'end_date': allocation.end_date.strftime('%Y-%m-%d')
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error vacating allocation: {str(e)}'
        })


def delete_allocation(request):
    """Delete an allocation (permanent)"""
    allocation_id = request.POST.get('id')
    if not allocation_id:
        return JsonResponse({'success': False, 'message': 'Allocation ID is required.'})
    
    try:
        allocation = StudentHostelAllocation.objects.select_related('bed').get(id=allocation_id)
    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Allocation not found.'})
    
    # Free the bed if occupied
    if allocation.bed and allocation.bed.is_occupied:
        allocation.bed.is_occupied = False
        allocation.bed.save()
    
    student_name = allocation.student.full_name
    hostel_name = allocation.hostel.name
    
    try:
        allocation.delete()
        return JsonResponse({
            'success': True,
            'message': f'Allocation for {student_name} to {hostel_name} deleted successfully.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting allocation: {str(e)}'
        })


def bulk_create_allocations(request):
    """Create multiple allocations at once"""
    student_ids = request.POST.getlist('students[]')
    hostel_id = request.POST.get('hostel', '').strip()
    academic_year_id = request.POST.get('academic_year', '').strip()
    start_date = request.POST.get('start_date', '').strip()
    
    if not student_ids:
        return JsonResponse({'success': False, 'message': 'No students selected.'})
    
    if not hostel_id:
        return JsonResponse({'success': False, 'message': 'Hostel is required.'})
    
    if not academic_year_id:
        return JsonResponse({'success': False, 'message': 'Academic year is required.'})
    
    if not start_date:
        return JsonResponse({'success': False, 'message': 'Start date is required.'})
    
    try:
        hostel = Hostel.objects.get(id=hostel_id, is_active=True)
        academic_year = AcademicYear.objects.get(id=academic_year_id)
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    except (Hostel.DoesNotExist, AcademicYear.DoesNotExist, ValueError) as e:
        return JsonResponse({'success': False, 'message': f'Invalid data: {str(e)}'})
    
    # Check hostel capacity
    current_allocations = StudentHostelAllocation.objects.filter(
        hostel=hostel,
        is_active=True
    ).count()
    
    available_slots = hostel.max_students - current_allocations
    
    if len(student_ids) > available_slots:
        return JsonResponse({
            'success': False,
            'message': f'Cannot allocate {len(student_ids)} students. Only {available_slots} slots available in {hostel.name}.'
        })
    
    created_count = 0
    failed_students = []
    
    with transaction.atomic():
        for student_id in student_ids:
            try:
                student = Student.objects.get(id=student_id, is_active=True)
                
                # Check if student already has allocation
                if StudentHostelAllocation.objects.filter(student=student, is_active=True).exists():
                    failed_students.append(f"{student.full_name} (Already allocated)")
                    continue
                
                # Create allocation
                StudentHostelAllocation.objects.create(
                    student=student,
                    hostel=hostel,
                    academic_year=academic_year,
                    start_date=start_date,
                    is_active=True
                )
                created_count += 1
                
            except Exception as e:
                failed_students.append(f"Student ID {student_id} ({str(e)})")
    
    message = f'Successfully allocated {created_count} student(s) to {hostel.name}.'
    if failed_students:
        message += f' Failed: {", ".join(failed_students[:5])}'
        if len(failed_students) > 5:
            message += f' and {len(failed_students) - 5} more.'
    
    return JsonResponse({
        'success': True,
        'message': message,
        'created_count': created_count,
        'failed_count': len(failed_students)
    })


@login_required
def allocation_details(request, allocation_id):
    """
    Get detailed information about a specific allocation (AJAX)
    """
    try:
        allocation = StudentHostelAllocation.objects.select_related(
            'student', 'hostel', 'room', 'bed', 'academic_year'
        ).get(id=allocation_id)
        
        # Get payment information
        payments = allocation.payments.all().order_by('-payment_date')
        
        total_paid = allocation.total_paid
        balance = allocation.balance
        
        data = {
            'id': allocation.id,
            'student': {
                'id': allocation.student.id,
                'full_name': allocation.student.full_name,
                'registration_number': allocation.student.registration_number,
                'gender': allocation.student.gender,
                'class_level': allocation.student.class_level.name if allocation.student.class_level else None,
                'stream': allocation.student.stream_class.stream_letter if allocation.student.stream_class else None,
            },
            'hostel': {
                'id': allocation.hostel.id,
                'name': allocation.hostel.name,
                'code': allocation.hostel.code,
                'total_fee': float(allocation.hostel.total_fee),
            },
            'room': {
                'id': allocation.room.id if allocation.room else None,
                'room_number': allocation.room.room_number if allocation.room else None,
            } if allocation.room else None,
            'bed': {
                'id': allocation.bed.id if allocation.bed else None,
                'bed_number': allocation.bed.bed_number if allocation.bed else None,
                'bed_type': allocation.bed.bed_type if allocation.bed else None,
            } if allocation.bed else None,
            'academic_year': str(allocation.academic_year),
            'start_date': allocation.start_date.strftime('%Y-%m-%d'),
            'end_date': allocation.end_date.strftime('%Y-%m-%d') if allocation.end_date else None,
            'is_active': allocation.is_active,
            'financial': {
                'total_fee': float(allocation.total_fee),
                'total_paid': float(total_paid),
                'balance': float(balance),
                'payment_status': 'Paid' if balance <= 0 else 'Partial' if total_paid > 0 else 'Unpaid',
            },
            'payments': [
                {
                    'id': p.id,
                    'receipt_number': p.receipt_number,
                    'amount_paid': float(p.amount_paid),
                    'payment_date': p.payment_date.strftime('%Y-%m-%d'),
                    'installment': p.installment_plan.installment_number if p.installment_plan else None,
                    'status': p.status,
                }
                for p in payments
            ]
        }
        
        return JsonResponse({
            'success': True,
            'allocation': data
        })
        
    except StudentHostelAllocation.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Allocation not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })    

# ============================================================================
# HOSTEL INSTALLMENT PLAN VIEWS
# ============================================================================

@login_required
def hostel_installment_plans_list(request):
    """
    Display hostel installment plans management page
    """
    # Get all installment plans with related hostel data
    installment_plans = HostelInstallmentPlan.objects.select_related(
        'hostel'
    ).all().order_by('hostel__name', 'installment_number')
    
    # Get statistics
    total_plans = installment_plans.count()
    total_hostels_with_plans = installment_plans.values('hostel').distinct().count()
    
    # Get hostels for filter
    hostels = Hostel.objects.filter(is_active=True).order_by('name')
    
    # Get plans per hostel
    plans_per_hostel = []
    for hostel in hostels:
        plans = installment_plans.filter(hostel=hostel)
        if plans.exists():
            total_amount = plans.aggregate(total=models.Sum('amount'))['total'] or 0
            plans_per_hostel.append({
                'hostel': hostel,
                'plan_count': plans.count(),
                'total_amount': total_amount,
                'installments': list(plans)
            })
    
    context = {
        'installment_plans': installment_plans,
        'total_plans': total_plans,
        'total_hostels_with_plans': total_hostels_with_plans,
        'hostels': hostels,
        'plans_per_hostel': plans_per_hostel,
        'page_title': 'Hostel Installment Plans',
    }
    
    return render(request, 'admin/hostels/installment_plans_list.html', context)


@login_required
def hostel_installment_plan_create(request):
    """
    Create a new hostel installment plan
    """
    if request.method == 'POST':
        try:
            # Get form data
            hostel_id = request.POST.get('hostel')
            installment_number = request.POST.get('installment_number')
            amount = request.POST.get('amount')
            start_month = request.POST.get('start_month')
            start_day = request.POST.get('start_day')
            end_month = request.POST.get('end_month')
            end_day = request.POST.get('end_day')

            # Validate required fields
            if not all([hostel_id, installment_number, amount, start_month, start_day, end_month, end_day]):
                return JsonResponse({
                    'success': False,
                    'message': 'Please fill all required fields.'
                })

            # Get hostel
            hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)

            # Validate installment number
            try:
                installment_number = int(installment_number)
                if installment_number < 1 or installment_number > 4:
                    return JsonResponse({
                        'success': False,
                        'message': 'Installment number must be between 1 and 4.'
                    })
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid installment number.'
                })

            # Check if installment number already exists for this hostel
            if HostelInstallmentPlan.objects.filter(hostel=hostel, installment_number=installment_number).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Installment {installment_number} already exists for {hostel.name}.'
                })

            # Validate amount
            try:
                amount = float(amount)
                if amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'Amount must be greater than 0.'
                    })
                if amount > 9999999.99:
                    return JsonResponse({
                        'success': False,
                        'message': 'Amount is too high.'
                    })
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid amount format.'
                })

            # ============================================
            # ENHANCED: Total amount validation against hostel fee
            # ============================================
            
            # Get all existing installment plans for this hostel
            existing_plans = HostelInstallmentPlan.objects.filter(
                hostel=hostel
            ).order_by('installment_number')
            
            # Calculate existing total - ensure it's a float
            existing_total = float(existing_plans.aggregate(total=models.Sum('amount'))['total'] or 0)
            
            # Calculate new total if this plan is added
            new_total = existing_total + amount
            
            # Check if hostel has total_fee attribute
            if hasattr(hostel, 'total_fee') and hostel.total_fee:
                # Convert hostel.total_fee to float for comparison
                hostel_fee = float(hostel.total_fee)
                
                # Check if new total exceeds hostel fee
                if new_total > hostel_fee:
                    # Calculate remaining allowed amount
                    remaining_allowed = hostel_fee - existing_total
                    
                    return JsonResponse({
                        'success': False,
                        'message': (
                            f'Total installment amount (TSh {new_total:,.2f}) exceeds hostel fee '
                            f'(TSh {hostel_fee:,.2f}).\n'
                            f'Current total from existing installments: TSh {existing_total:,.2f}\n'
                            f'Maximum allowed for this installment: TSh {remaining_allowed:,.2f}'
                        )
                    })
                
                # Check if new total is exactly equal to hostel fee (optional - could be a warning)
                if new_total == hostel_fee:
                    # This is fine - total matches exactly
                    pass
            else:
                # If hostel doesn't have total_fee, skip this validation
                pass

            # Validate dates
            try:
                start_month = int(start_month)
                start_day = int(start_day)
                end_month = int(end_month)
                end_day = int(end_day)
                
                if not (1 <= start_month <= 12 and 1 <= end_month <= 12):
                    return JsonResponse({
                        'success': False,
                        'message': 'Month must be between 1 and 12.'
                    })
                
                if not (1 <= start_day <= 31 and 1 <= end_day <= 31):
                    return JsonResponse({
                        'success': False,
                        'message': 'Day must be between 1 and 31.'
                    })
                
                # Validate day based on month
                def validate_day(month, day):
                    days_in_month = {
                        1: 31, 2: 29, 3: 31, 4: 30, 5: 31, 6: 30,
                        7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
                    }
                    if day > days_in_month.get(month, 31):
                        return False
                    return True
                
                if not validate_day(start_month, start_day):
                    return JsonResponse({
                        'success': False,
                        'message': f'Invalid start date: Month {start_month} does not have day {start_day}.'
                    })
                
                if not validate_day(end_month, end_day):
                    return JsonResponse({
                        'success': False,
                        'message': f'Invalid end date: Month {end_month} does not have day {end_day}.'
                    })
                
                # Convert to comparable values for validation
                def date_to_value(month, day):
                    return month * 100 + day
                
                start_value = date_to_value(start_month, start_day)
                end_value = date_to_value(end_month, end_day)
                
                # Validate that end date is after start date
                if end_value <= start_value:
                    return JsonResponse({
                        'success': False,
                        'message': 'End date must be after start date.'
                    })
                
                # ============================================
                # Validate against existing installment plans
                # ============================================
                
                # Check for overlapping dates with existing plans
                for existing_plan in existing_plans:
                    existing_start = date_to_value(existing_plan.start_month, existing_plan.start_day)
                    existing_end = date_to_value(existing_plan.end_month, existing_plan.end_day)
                    
                    # Check for any overlap
                    # Overlap occurs if new plan starts before existing plan ends AND ends after existing plan starts
                    if start_value <= existing_end and end_value >= existing_start:
                        return JsonResponse({
                            'success': False,
                            'message': f'Date range overlaps with Installment {existing_plan.installment_number} ({existing_plan.start_day:02d}/{existing_plan.start_month:02d} - {existing_plan.end_day:02d}/{existing_plan.end_month:02d}).'
                        })
                
                # Special validation for consecutive installments
                # If this is installment 2,3,4, ensure it comes after the previous installment
                if installment_number > 1:
                    previous_plan = HostelInstallmentPlan.objects.filter(
                        hostel=hostel,
                        installment_number=installment_number - 1
                    ).first()
                    
                    if previous_plan:
                        previous_end = date_to_value(previous_plan.end_month, previous_plan.end_day)
                        
                        # New plan should start after previous plan ends (allow for gap)
                        if start_value <= previous_end:
                            return JsonResponse({
                                'success': False,
                                'message': f'Installment {installment_number} must start after Installment {previous_plan.installment_number} ends. Previous installment ends on {previous_plan.end_day:02d}/{previous_plan.end_month:02d}.'
                            })
                
                # Check if this is the first installment (should start from beginning of payment cycle)
                if installment_number == 1:
                    # Optionally ensure first installment starts from month 1 (January)
                    # This is optional - you can remove if not needed
                    if start_month != 1 or start_day != 1:
                        # You can either enforce or just warn
                        # For now, we'll allow flexibility but could add warning
                        pass
                
                # Check if last installment (4) should end at year end (December 31)
                if installment_number == 4:
                    # Optional validation for last installment
                    if end_month != 12 or end_day != 31:
                        # You could add a warning or enforce
                        pass
                
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date values.'
                })

            # Create installment plan (convert amount back to Decimal for storage)
            from decimal import Decimal
            with transaction.atomic():
                installment_plan = HostelInstallmentPlan.objects.create(
                    hostel=hostel,
                    installment_number=installment_number,
                    amount=Decimal(str(amount)),  # Convert float to Decimal for storage
                    start_month=start_month,
                    start_day=start_day,
                    end_month=end_month,
                    end_day=end_day
                )

            messages.success(request, f'Installment plan created successfully for {hostel.name}.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': 'list'})
            return redirect('admin_hostel_installment_plans_list')

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)})
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_hostel_installment_plan_create')

    # GET request - show form
    hostels = Hostel.objects.filter(is_active=True).annotate(
        current_plans=Count('installment_plans')
    ).order_by('name')
    
    context = {
        'title': 'New Installment Plan',
        'hostels': hostels,
        'months': range(1, 13),
        'days': range(1, 32),
        'page_title': 'Create Installment Plan',
    }
    
    return render(request, 'admin/hostels/installment_plan_form.html', context)


@login_required
def hostel_installment_plan_edit(request, pk):
    """
    Edit an existing installment plan
    """
    installment_plan = get_object_or_404(
        HostelInstallmentPlan.objects.select_related('hostel'),
        id=pk
    )

    if request.method == 'POST':
        try:
            # Get form data
            amount = request.POST.get('amount')
            start_month = request.POST.get('start_month')
            start_day = request.POST.get('start_day')
            end_month = request.POST.get('end_month')
            end_day = request.POST.get('end_day')

            # Validate required fields
            if not all([amount, start_month, start_day, end_month, end_day]):
                return JsonResponse({
                    'success': False,
                    'message': 'Please fill all required fields.'
                })

            # Validate amount
            try:
                amount = float(amount)
                if amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'message': 'Amount must be greater than 0.'
                    })
                if amount > 9999999.99:
                    return JsonResponse({
                        'success': False,
                        'message': 'Amount is too high.'
                    })
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid amount format.'
                })

            # ============================================
            # ENHANCED: Total amount validation against hostel fee (for edit)
            # ============================================
            
            # Get all other existing installment plans for this hostel (excluding current)
            other_plans = HostelInstallmentPlan.objects.filter(
                hostel=installment_plan.hostel
            ).exclude(id=installment_plan.id)
            
            # Calculate total from other plans - convert to float
            other_total = float(other_plans.aggregate(total=models.Sum('amount'))['total'] or 0)
            
            # Calculate new total if this plan is updated
            new_total = other_total + amount
            
            # Check if hostel has total_fee attribute
            if hasattr(installment_plan.hostel, 'total_fee') and installment_plan.hostel.total_fee:
                # Convert hostel.total_fee to float
                hostel_fee = float(installment_plan.hostel.total_fee)
                
                # Check if new total exceeds hostel fee
                if new_total > hostel_fee:
                    # Calculate remaining allowed amount
                    remaining_allowed = hostel_fee - other_total
                    
                    return JsonResponse({
                        'success': False,
                        'message': (
                            f'Total installment amount (TSh {new_total:,.2f}) exceeds hostel fee '
                            f'(TSh {hostel_fee:,.2f}).\n'
                            f'Current total from other installments: TSh {other_total:,.2f}\n'
                            f'Maximum allowed for this installment: TSh {remaining_allowed:,.2f}'
                        )
                    })
                
                # Check if new total is exactly equal to hostel fee (optional - could be a warning)
                if new_total == hostel_fee:
                    # This is fine - total matches exactly
                    pass
            else:
                # If hostel doesn't have total_fee, skip this validation
                pass

            # Validate dates
            try:
                start_month = int(start_month)
                start_day = int(start_day)
                end_month = int(end_month)
                end_day = int(end_day)
                
                if not (1 <= start_month <= 12 and 1 <= end_month <= 12):
                    return JsonResponse({
                        'success': False,
                        'message': 'Month must be between 1 and 12.'
                    })
                
                if not (1 <= start_day <= 31 and 1 <= end_day <= 31):
                    return JsonResponse({
                        'success': False,
                        'message': 'Day must be between 1 and 31.'
                    })
                
                # Validate day based on month
                def validate_day(month, day):
                    days_in_month = {
                        1: 31, 2: 29, 3: 31, 4: 30, 5: 31, 6: 30,
                        7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
                    }
                    if day > days_in_month.get(month, 31):
                        return False
                    return True
                
                if not validate_day(start_month, start_day):
                    return JsonResponse({
                        'success': False,
                        'message': f'Invalid start date: Month {start_month} does not have day {start_day}.'
                    })
                
                if not validate_day(end_month, end_day):
                    return JsonResponse({
                        'success': False,
                        'message': f'Invalid end date: Month {end_month} does not have day {end_day}.'
                    })
                
                # Convert to comparable values
                def date_to_value(month, day):
                    return month * 100 + day
                
                start_value = date_to_value(start_month, start_day)
                end_value = date_to_value(end_month, end_day)
                
                # Validate that end date is after start date
                if end_value <= start_value:
                    return JsonResponse({
                        'success': False,
                        'message': 'End date must be after start date.'
                    })
                
                # ============================================
                # Validate against other installment plans (excluding current)
                # ============================================
                
                # Check for overlapping dates with other plans
                for other_plan in other_plans:
                    other_start = date_to_value(other_plan.start_month, other_plan.start_day)
                    other_end = date_to_value(other_plan.end_month, other_plan.end_day)
                    
                    # Check for any overlap
                    if start_value <= other_end and end_value >= other_start:
                        return JsonResponse({
                            'success': False,
                            'message': f'Date range overlaps with Installment {other_plan.installment_number} ({other_plan.start_day:02d}/{other_plan.start_month:02d} - {other_plan.end_day:02d}/{other_plan.end_month:02d}).'
                        })
                
                # Check sequence with previous installment
                if installment_plan.installment_number > 1:
                    previous_plan = HostelInstallmentPlan.objects.filter(
                        hostel=installment_plan.hostel,
                        installment_number=installment_plan.installment_number - 1
                    ).first()
                    
                    if previous_plan:
                        previous_end = date_to_value(previous_plan.end_month, previous_plan.end_day)
                        
                        if start_value <= previous_end:
                            return JsonResponse({
                                'success': False,
                                'message': f'This installment must start after Installment {previous_plan.installment_number} ends (on {previous_plan.end_day:02d}/{previous_plan.end_month:02d}).'
                            })
                
                # Check sequence with next installment
                if installment_plan.installment_number < 4:
                    next_plan = HostelInstallmentPlan.objects.filter(
                        hostel=installment_plan.hostel,
                        installment_number=installment_plan.installment_number + 1
                    ).first()
                    
                    if next_plan:
                        next_start = date_to_value(next_plan.start_month, next_plan.start_day)
                        
                        if end_value >= next_start:
                            return JsonResponse({
                                'success': False,
                                'message': f'This installment must end before Installment {next_plan.installment_number} starts (on {next_plan.start_day:02d}/{next_plan.start_month:02d}).'
                            })
                
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date values.'
                })

            # Update installment plan (convert amount back to Decimal for storage)
            from decimal import Decimal
            with transaction.atomic():
                installment_plan.amount = Decimal(str(amount))
                installment_plan.start_month = start_month
                installment_plan.start_day = start_day
                installment_plan.end_month = end_month
                installment_plan.end_day = end_day
                installment_plan.save()

            messages.success(request, 'Installment plan updated successfully.')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': 'list'})
            return redirect('admin_hostel_installment_plans_list')

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)})
            messages.error(request, f'Error: {str(e)}')
            return redirect('admin_hostel_installment_plan_edit', pk=pk)

    # GET request - show form
    # Get total amount from all installments for this hostel to display
    all_plans = HostelInstallmentPlan.objects.filter(hostel=installment_plan.hostel)
    total_installment_amount = float(all_plans.aggregate(total=models.Sum('amount'))['total'] or 0)
    
    context = {
        'title': f'Edit Installment {installment_plan.installment_number} - {installment_plan.hostel.name}',
        'installment_plan': installment_plan,
        'hostel': installment_plan.hostel,
        'total_installment_amount': total_installment_amount,
        'remaining_amount': (float(installment_plan.hostel.total_fee) - total_installment_amount) if installment_plan.hostel.total_fee else None,
        'months': range(1, 13),
        'days': range(1, 32),
        'page_title': 'Edit Installment Plan',
    }
    
    return render(request, 'admin/hostels/installment_plan_edit.html', context)


@login_required
def hostel_installment_plan_detail(request, pk):
    """
    Production-safe installment analytics view
    """
    from django.db.models import Sum, Count, Value, F, Case, When, CharField, DecimalField
    from django.db.models.functions import Coalesce, ExtractYear, ExtractMonth
    from decimal import Decimal

    # -------------------------------------------------
    # 1️⃣ Get Installment Plan
    # -------------------------------------------------
    installment_plan = get_object_or_404(
        HostelInstallmentPlan.objects.select_related('hostel'),
        id=pk
    )

    ZERO_DECIMAL = Value(
        Decimal("0.00"),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    # -------------------------------------------------
    # 2️⃣ Installment Payments (DB-level status)
    # -------------------------------------------------
    installment_payments = (
        HostelPayment.objects
        .filter(installment_plan=installment_plan)
        .select_related(
            'allocation',
            'allocation__student',
            'allocation__student__class_level',
            'allocation__student__stream_class',
            'allocation__hostel',
            'installment_plan'
        )
        .annotate(
            total_paid_db=Coalesce(
                Sum('transactions__amount'),
                ZERO_DECIMAL
            ),
            required_amount_db=F('installment_plan__amount')
        )
        .annotate(
            status_db=Case(
                When(
                    total_paid_db__gte=F('required_amount_db'),
                    then=Value('paid')
                ),
                default=Value('partial'),
                output_field=CharField()
            )
        )
        .order_by('-created_at')
    )

    # -------------------------------------------------
    # 3️⃣ Payment Transactions
    # -------------------------------------------------
    payment_transactions = (
        HostelPaymentTransaction.objects
        .filter(
            installment_payment__installment_plan=installment_plan
        )
        .select_related(
            'installment_payment__installment_plan',
            'allocation__student'
        )
        .order_by('-payment_date', '-created_at')
    )

    # -------------------------------------------------
    # 4️⃣ Global Statistics
    # -------------------------------------------------
    totals = payment_transactions.aggregate(
        total_paid=Coalesce(Sum('amount'), ZERO_DECIMAL),
        total_transactions=Count('id')
    )

    total_paid_amount = totals['total_paid']
    total_transactions = totals['total_transactions']

    students_with_plan = (
        installment_payments
        .values('allocation__student')
        .distinct()
        .count()
    )

    fully_paid_count = installment_payments.filter(
        status_db='paid'
    ).count()

    partial_paid_count = installment_payments.filter(
        status_db='partial'
    ).count()

    # -------------------------------------------------
    # 5️⃣ Collection Efficiency (Decimal Safe)
    # -------------------------------------------------
    installment_count = installment_payments.count()

    expected_total = (
        Decimal(installment_count) *
        installment_plan.amount
    )

    if expected_total > 0:
        collection_efficiency = (
            (total_paid_amount / expected_total) * Decimal("100")
        )
    else:
        collection_efficiency = Decimal("0.00")

    # -------------------------------------------------
    # 6️⃣ Monthly Summary (No Field Conflict)
    # -------------------------------------------------
    monthly_summary = (
        payment_transactions
        .annotate(
            payment_year=ExtractYear('payment_date'),
            payment_month=ExtractMonth('payment_date')
        )
        .values('payment_year', 'payment_month')
        .annotate(
            total=Coalesce(Sum('amount'), ZERO_DECIMAL),
            count=Count('id')
        )
        .order_by('payment_year', 'payment_month')
    )

    # -------------------------------------------------
    # 7️⃣ Top Paying Students (Fixed - using actual field names)
    # -------------------------------------------------
    top_students = (
        installment_payments
        .values(
            'allocation__student__id',
            'allocation__student__first_name',
            'allocation__student__middle_name',
            'allocation__student__last_name',
            'allocation__student__registration_number'
        )
        .annotate(
            total_paid=Coalesce(
                Sum('transactions__amount'),
                ZERO_DECIMAL
            ),
            payment_count=Count('transactions')
        )
        .order_by('-total_paid')[:10]
    )
    
    # Convert to list with full_name property
    top_students_list = []
    for student in top_students:
        # Construct full name from components
        first = student.get('allocation__student__first_name', '')
        middle = student.get('allocation__student__middle_name', '')
        last = student.get('allocation__student__last_name', '')
        
        full_name = f"{first} {middle} {last}".strip()
        if not full_name:
            full_name = "Unknown"
            
        top_students_list.append({
            'id': student['allocation__student__id'],
            'full_name': full_name,
            'first_name': first,
            'middle_name': middle,
            'last_name': last,
            'registration_number': student.get('allocation__student__registration_number', 'N/A'),
            'total_paid': student['total_paid'],
            'payment_count': student['payment_count']
        })

    # -------------------------------------------------
    # 8️⃣ Context
    # -------------------------------------------------
    context = {
        'installment_plan': installment_plan,
        'installment_payments': installment_payments,
        'payment_transactions': payment_transactions[:20],
        'total_paid_amount': total_paid_amount,
        'total_transactions': total_transactions,
        'students_with_plan': students_with_plan,
        'fully_paid_count': fully_paid_count,
        'partial_paid_count': partial_paid_count,
        'collection_efficiency': round(collection_efficiency, 2),
        'monthly_summary': monthly_summary,
        'top_students': top_students_list,  # Use the processed list
        'page_title': (
            f'Installment '
            f'{installment_plan.installment_number} - '
            f'{installment_plan.hostel.name}'
        ),
    }

    return render(
        request,
        'admin/hostels/installment_plan_detail.html',
        context
    )


@login_required
def hostel_installment_plan_delete(request):
    """
    Delete an installment plan (AJAX only)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

    plan_id = request.POST.get('id')
    if not plan_id:
        return JsonResponse({'success': False, 'message': 'Plan ID is required.'})

    try:
        installment_plan = HostelInstallmentPlan.objects.select_related('hostel').get(id=plan_id)
        
        # Check if plan has any payments
        if installment_plan.student_payments.exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete plan with existing payments.'
            })
        
        plan_info = f'Installment {installment_plan.installment_number} for {installment_plan.hostel.name}'
        installment_plan.delete()

        return JsonResponse({
            'success': True,
            'message': f'{plan_info} deleted successfully.'
        })

    except HostelInstallmentPlan.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Installment plan not found.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def get_installment_plans_by_hostel(request, hostel_id):
    """
    AJAX endpoint to get installment plans for a specific hostel
    """
    try:
        hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
        plans = HostelInstallmentPlan.objects.filter(
            hostel=hostel
        ).order_by('installment_number')
        
        plan_list = []
        for plan in plans:
            plan_list.append({
                'id': plan.id,
                'installment_number': plan.installment_number,
                'amount': float(plan.amount),
                'amount_display': f"TSh {plan.amount:,.2f}",
                'start_date': f"{plan.start_day:02d}/{plan.start_month:02d}",
                'end_date': f"{plan.end_day:02d}/{plan.end_month:02d}",
                'display': f"Installment {plan.installment_number} - TSh {plan.amount:,.2f} ({plan.start_day:02d}/{plan.start_month:02d} to {plan.end_day:02d}/{plan.end_month:02d})"
            })
        
        return JsonResponse({
            'success': True,
            'plans': plan_list,
            'hostel_name': hostel.name,
            'total_plans': len(plan_list),
            'total_amount': sum(plan['amount'] for plan in plan_list)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })        


# views.py (add these to your existing hostel views)



# ============================================================================
# HOSTEL PAYMENT MANAGEMENT VIEWS
# ============================================================================

@login_required
def hostel_payments_list(request):
    """
    Display all hostel payments with filtering options
    """
    # Get all payment transactions
    transactions = HostelPaymentTransaction.objects.select_related(
        'allocation__student',
        'allocation__hostel',
        'installment_payment__installment_plan'
    ).all().order_by('-payment_date', '-created_at')            
    
    # Get statistics
    total_transactions = transactions.count()
    total_amount = transactions.aggregate(total=Sum('amount'))['total'] or 0
    
    # Get unique allocations for filter
    allocations = StudentHostelAllocation.objects.filter(
        is_active=True
    ).select_related('student', 'hostel').order_by('student__first_name')
    
    # Get hostels for filter
    hostels = Hostel.objects.filter(is_active=True).order_by('name')
    
    # Get payment types for filter
    payment_types = HostelPaymentTransaction.PAYMENT_TYPE
    
    context = {
        'transactions': transactions[:50],  # Limit to recent 50 for performance
        'total_transactions': total_transactions,
        'total_amount': total_amount,
        'allocations': allocations,
        'hostels': hostels,
        'payment_types': payment_types,
        'page_title': 'Hostel Payments',
    }
    
    return render(request, 'admin/hostels/payments_list.html', context)


@login_required
def hostel_payment_create(request):
    """
    Create a new hostel payment
    """
    # Get students with active allocations
    allocations = StudentHostelAllocation.objects.filter(
        is_active=True
    ).select_related('student', 'hostel').order_by('student__first_name')
    
    context = {
        'allocations': allocations,
        'page_title': 'Record Hostel Payment',
        'today': timezone.now().date().isoformat(),
    }
    
    return render(request, 'admin/hostels/payment_form.html', context)


@login_required
def hostel_payment_detail(request, pk):
    """
    Display detailed information about a payment transaction
    """
    transaction = get_object_or_404(
        HostelPaymentTransaction.objects.select_related(
            'allocation__student',
            'allocation__hostel',
            'installment_payment__installment_plan'
        ),
        id=pk
    )
    
    # Get related transactions if this is an installment payment
    related_transactions = []
    if transaction.installment_payment:
        related_transactions = HostelPaymentTransaction.objects.filter(
            installment_payment=transaction.installment_payment
        ).exclude(id=transaction.id).order_by('-payment_date')
    
    context = {
        'transaction': transaction,
        'related_transactions': related_transactions,
        'page_title': f'Payment Details - {transaction.receipt_number}',
    }
    
    return render(request, 'admin/hostels/payment_detail.html', context)


@login_required
def student_payment_history(request, student_id):
    """
    Display payment history for a specific student
    """
    student = get_object_or_404(Student, id=student_id)
    
    # Get all allocations for this student
    allocations = StudentHostelAllocation.objects.filter(
        student=student
    ).select_related('hostel', 'academic_year').order_by('-start_date')
    
    # Get all transactions for this student
    transactions = HostelPaymentTransaction.objects.filter(
        allocation__student=student
    ).select_related(
        'allocation__hostel',
        'installment_payment__installment_plan'
    ).order_by('-payment_date')
    
    # Calculate totals
    total_paid = transactions.aggregate(total=Sum('amount'))['total'] or 0
    
    # Get current balance for active allocation
    current_allocation = allocations.filter(is_active=True).first()
    current_balance = current_allocation.balance if current_allocation else 0
    
    context = {
        'student': student,
        'allocations': allocations,
        'transactions': transactions,
        'total_paid': total_paid,
        'current_allocation': current_allocation,
        'current_balance': current_balance,
        'page_title': f'Payment History - {student.full_name}',
    }
    
    return render(request, 'admin/hostels/student_payment_history.html', context)


@login_required
def allocation_payments(request, allocation_id):
    """
    Display payments for a specific allocation with enhanced analytics and filtering
    """
    # Get allocation with all related data
    allocation = get_object_or_404(
        StudentHostelAllocation.objects.select_related(
            'student',
            'student__class_level',
            'student__stream_class',
            'hostel',
            'room',
            'bed',
            'academic_year'
        ),
        id=allocation_id
    )
    
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    payment_type = request.GET.get('payment_type')
    payment_method = request.GET.get('payment_method')
    
    # Base queryset for transactions
    transactions_qs = HostelPaymentTransaction.objects.filter(
        allocation=allocation
    ).select_related(
        'installment_payment__installment_plan'
    ).order_by('-payment_date', '-created_at')
    
    # Apply filters
    if date_from:
        transactions_qs = transactions_qs.filter(payment_date__gte=date_from)
    if date_to:
        transactions_qs = transactions_qs.filter(payment_date__lte=date_to)
    if payment_type:
        transactions_qs = transactions_qs.filter(payment_type=payment_type)
    if payment_method:
        transactions_qs = transactions_qs.filter(payment_method=payment_method)
    
    transactions = transactions_qs
    
    # Get installment payments if applicable
    installment_payments = []
    if allocation.hostel.payment_mode == 'installments':
        installment_payments = HostelPayment.objects.filter(
            allocation=allocation
        ).select_related('installment_plan').order_by('installment_plan__installment_number')
        
        # Calculate detailed installment stats
        for inst in installment_payments:
            inst.paid_percentage = (inst.total_paid / inst.required_amount * 100) if inst.required_amount > 0 else 0
    
    # Calculate totals and statistics
    total_paid = transactions.aggregate(total=Sum('amount'))['total'] or 0
    total_fee = allocation.total_fee
    balance = total_fee - total_paid
    payment_percentage = (total_paid / total_fee * 100) if total_fee > 0 else 0
    
    # Payment statistics by type and method
    stats_by_type = transactions.values('payment_type').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_type')
    
    stats_by_method = transactions.values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_method')
    
    # Monthly payment summary (for charts)
    monthly_summary = transactions.extra(
        select={'month': "EXTRACT(month FROM payment_date)"},
        where={"payment_date__isnull": False}
    ).values('month').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('month')
    
    # Get recent activity
    recent_activity = transactions[:5]
    
    # Get payment reminders/upcoming installments
    upcoming_installments = []
    if allocation.hostel.payment_mode == 'installments' and allocation.is_active:
        current_date = timezone.now().date()
        for inst in installment_payments:
            if inst.remaining_amount > 0:
                # Calculate days until next payment (simplified - you might want more sophisticated logic)
                upcoming_installments.append({
                    'installment': inst,
                    'days_until': 30,  # Placeholder - implement actual date calculation
                    'amount_due': inst.remaining_amount
                })
    
    context = {
        'allocation': allocation,
        'transactions': transactions,
        'installment_payments': installment_payments,
        'total_paid': total_paid,
        'total_fee': total_fee,
        'balance': balance,
        'payment_percentage': payment_percentage,
        'stats_by_type': stats_by_type,
        'stats_by_method': stats_by_method,
        'monthly_summary': monthly_summary,
        'recent_activity': recent_activity,
        'upcoming_installments': upcoming_installments,
        'is_fully_paid': balance <= 0,
        'payment_status': 'Fully Paid' if balance <= 0 else 'Partial' if total_paid > 0 else 'Unpaid',
        'page_title': f'Payments - {allocation.student.full_name}',
        'filter_params': {
            'date_from': date_from,
            'date_to': date_to,
            'payment_type': payment_type,
            'payment_method': payment_method,
        }
    }
    
    return render(request, 'admin/hostels/allocation_payments.html', context)


@login_required
def single_transaction_payment_export_pdf(request, transaction_id):
    """
    Generate a PDF receipt for a single payment transaction using WeasyPrint
    """
    # Get the transaction with all related data
    transaction = get_object_or_404(
        HostelPaymentTransaction.objects.select_related(
            'allocation',
            'allocation__student',
            'allocation__student__class_level',
            'allocation__student__stream_class',
            'allocation__hostel',
            'allocation__room',
            'allocation__bed',
            'allocation__academic_year',
            'installment_payment__installment_plan'
        ),
        id=transaction_id
    )
    
    # Get allocation for easy access
    allocation = transaction.allocation
    
    # Calculate payment summary
    total_paid_to_date = HostelPaymentTransaction.objects.filter(
        allocation=allocation
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    remaining_balance = allocation.total_fee - total_paid_to_date
    
    # Get related transactions for this allocation (excluding current)
    related_transactions = HostelPaymentTransaction.objects.filter(
        allocation=allocation
    ).exclude(
        id=transaction.id
    ).select_related(
        'installment_payment__installment_plan'
    ).order_by('-payment_date')[:5]  # Limit to 5 most recent
    
    # Get payment method display
    payment_method_display = dict(HostelPaymentTransaction.PAYMENT_METHOD).get(
        transaction.payment_method, transaction.payment_method
    )
    
    # Get payment type display
    payment_type_display = dict(HostelPaymentTransaction.PAYMENT_TYPE).get(
        transaction.payment_type, transaction.payment_type
    )
    
    # Prepare context for PDF
    context = {
        'transaction': transaction,
        'allocation': allocation,
        'student': allocation.student,
        'hostel': allocation.hostel,
        'total_paid_to_date': total_paid_to_date,
        'remaining_balance': remaining_balance,
        'related_transactions': related_transactions,
        'payment_method_display': payment_method_display,
        'payment_type_display': payment_type_display,
        'generated_at': timezone.now(),
        'generated_by': request.user.get_full_name() or request.user.username,
        'is_fully_paid': remaining_balance <= 0,
        'request': request,
    }
    
    # Render HTML template
    html_string = render_to_string('admin/hostels/single_transaction_receipt_pdf.html', context)
    
    # Generate PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"receipt_{transaction.receipt_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Write PDF to response
    html.write_pdf(response)
    
    return response

@login_required
def allocation_payments_export_pdf(request, allocation_id):
    """
    Export allocation payments to PDF using WeasyPrint
    """
    # Get allocation with all related data
    allocation = get_object_or_404(
        StudentHostelAllocation.objects.select_related(
            'student',
            'student__class_level',
            'student__stream_class',
            'hostel',
            'room',
            'bed',
            'academic_year'
        ),
        id=allocation_id
    )
    
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    payment_type = request.GET.get('payment_type')
    payment_method = request.GET.get('payment_method')
    
    # Base queryset for transactions
    transactions_qs = HostelPaymentTransaction.objects.filter(
        allocation=allocation
    ).select_related(
        'installment_payment__installment_plan'
    ).order_by('-payment_date', '-created_at')
    
    # Apply filters
    if date_from:
        transactions_qs = transactions_qs.filter(payment_date__gte=date_from)
    if date_to:
        transactions_qs = transactions_qs.filter(payment_date__lte=date_to)
    if payment_type:
        transactions_qs = transactions_qs.filter(payment_type=payment_type)
    if payment_method:
        transactions_qs = transactions_qs.filter(payment_method=payment_method)
    
    transactions = transactions_qs
    
    # Get installment payments if applicable
    installment_payments = []
    if allocation.hostel.payment_mode == 'installments':
        installment_payments = HostelPayment.objects.filter(
            allocation=allocation
        ).select_related('installment_plan').order_by('installment_plan__installment_number')
        
        # Calculate detailed installment stats
        for inst in installment_payments:
            inst.paid_percentage = (inst.total_paid / inst.required_amount * 100) if inst.required_amount > 0 else 0
    
    # Calculate totals and statistics
    total_paid = transactions.aggregate(total=Sum('amount'))['total'] or 0
    total_fee = allocation.total_fee
    balance = total_fee - total_paid
    payment_percentage = (total_paid / total_fee * 100) if total_fee > 0 else 0
    
    # Payment statistics by type and method
    stats_by_type = transactions.values('payment_type').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_type')
    
    stats_by_method = transactions.values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_method')
    
    # Prepare context for PDF
    context = {
        'allocation': allocation,
        'transactions': transactions,
        'installment_payments': installment_payments,
        'total_paid': total_paid,
        'total_fee': total_fee,
        'balance': balance,
        'payment_percentage': payment_percentage,
        'stats_by_type': stats_by_type,
        'stats_by_method': stats_by_method,
        'is_fully_paid': balance <= 0,
        'payment_status': 'Fully Paid' if balance <= 0 else 'Partial' if total_paid > 0 else 'Unpaid',
        'generated_at': timezone.now(),
        'filter_params': {
            'date_from': date_from,
            'date_to': date_to,
            'payment_type': payment_type,
            'payment_method': payment_method,
        },
        'request': request,
    }
    
    # Render HTML template
    html_string = render_to_string('admin/hostels/allocation_payments_pdf.html', context)
    
    # Generate PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"payments_{allocation.student.registration_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Write PDF to response
    html.write_pdf(response)
    
    return response


# ============================================================================
# CORE PAYMENT PROCESSING LOGIC
# ============================================================================

# Updated PaymentProcessor class with payment method and transaction number support
class PaymentProcessor:
    """
    Handles the complex payment logic for hostel fees
    """
    
    def __init__(self, allocation, amount, payment_type='installment', 
                 payment_method='cash', transaction_number=None,
                 month=None, year=None):
        self.allocation = allocation
        self.amount = Decimal(str(amount))
        self.payment_type = payment_type
        self.payment_method = payment_method
        self.transaction_number = transaction_number
        self.month = month
        self.year = year
        self.student = allocation.student
        self.hostel = allocation.hostel
        
        # Check if student is in final year/level
        self.is_final_year = self._check_final_year()
    
    def _check_final_year(self):
        """
        Determine if student is in their final academic year/level
        """
        if not self.student.class_level:
            return False
        
        # Get the highest class in this educational level
        highest_class = ClassLevel.objects.filter(
            educational_level=self.student.class_level.educational_level,
            is_active=True
        ).order_by('-order').first()
        
        if not highest_class:
            return False
        
        return self.student.class_level.id == highest_class.id
    
    def _get_outstanding_installments(self):
        """
        Get all unpaid or partially paid installments in order
        """
        # Get all installment plans for this hostel
        all_plans = HostelInstallmentPlan.objects.filter(
            hostel=self.hostel
        ).order_by('installment_number')
        
        outstanding = []
        for plan in all_plans:
            # Get or create installment payment record
            installment_payment, created = HostelPayment.objects.get_or_create(
                allocation=self.allocation,
                installment_plan=plan,
                defaults={}
            )
            
            # Calculate paid amount
            paid = installment_payment.transactions.aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            if paid < plan.amount:
                outstanding.append({
                    'plan': plan,
                    'installment_payment': installment_payment,
                    'paid': paid,
                    'remaining': plan.amount - paid
                })
        
        return outstanding
    
    @transaction.atomic
    def process_payment(self):
        """
        Main payment processing logic
        """
        try:
            if self.payment_type == 'installments':
                return self._process_installment_payment()
            elif self.payment_type == 'monthly':
                return self._process_monthly_payment()
            elif self.payment_type == 'yearly':
                return self._process_yearly_payment()
            else:
                raise ValidationError("Invalid payment type")
        except Exception as e:
            # Ensure we always return a dict with message
            return {
                'success': False,
                'message': str(e)
            }
    
    def _create_transaction(self, amount, **kwargs):
        """
        Helper method to create a payment transaction with common fields
        """
        return HostelPaymentTransaction.objects.create(
            allocation=self.allocation,
            amount=amount,
            payment_type=self.payment_type,
            payment_method=self.payment_method,
            transaction_number=self.transaction_number,
            receipt_number=self._generate_receipt_number(),
            payment_date=timezone.now().date(),
            **kwargs
        )
    
    def _process_installment_payment(self):
        """
        Process installment payment with redistribution logic
        Shows detailed distribution of payment across installments
        """
        remaining_amount = self.amount
        processed_transactions = []
        distribution_details = []
        outstanding = self._get_outstanding_installments()
        
        # First, process all outstanding installments in order
        for item in outstanding:
            if remaining_amount <= 0:
                break
            
            # Calculate amount to allocate to this installment
            amount_to_pay = min(remaining_amount, item['remaining'])
            
            # Create transaction
            transaction = self._create_transaction(
                amount=amount_to_pay,
                installment_payment=item['installment_payment']
            )
            
            processed_transactions.append(transaction)
            
            # Record distribution details
            distribution_details.append({
                'installment_number': item['plan'].installment_number,
                'amount_paid': float(amount_to_pay),
                'previous_remaining': float(item['remaining']),
                'new_remaining': float(item['remaining'] - amount_to_pay),
                'status': 'Fully Paid' if (item['remaining'] - amount_to_pay) <= 0 else 'Partially Paid'
            })
            
            remaining_amount -= amount_to_pay
        
        # Calculate summary statistics
        total_paid = self.amount - remaining_amount
        installments_covered = len(processed_transactions)
        total_installments = len(outstanding) + (0 if not outstanding else 0)
        
        # Create summary message
        summary = self._create_distribution_summary(
            distribution_details, 
            total_paid, 
            remaining_amount,
            installments_covered,
            total_installments
        )
        
        # Handle remaining amount after processing installments
        if remaining_amount > 0:
            # All installments are fully paid
            if self.is_final_year:
                # Final year - refund excess
                refund = self._create_refund(remaining_amount, "Excess payment after all installments")
                return {
                    'success': True,
                    'transactions': processed_transactions,
                    'distribution': distribution_details,
                    'summary': summary,
                    'refund': refund,
                    'message': (
                        f"Payment of TSh {float(self.amount):,.2f} processed.\n"
                        f"{summary}\n"
                        f"All installments are now fully paid.\n"
                        f"Excess TSh {float(remaining_amount):,.2f} will be refunded."
                    )
                }
            else:
                # Not final year - store as balance for next year
                balance = self._store_as_balance(remaining_amount)
                return {
                    'success': True,
                    'transactions': processed_transactions,
                    'distribution': distribution_details,
                    'summary': summary,
                    'balance': balance,
                    'message': (
                        f"Payment of TSh {float(self.amount):,.2f} processed.\n"
                        f"{summary}\n"
                        f"All installments are now fully paid.\n"
                        f"Remaining TSh {float(remaining_amount):,.2f} stored as balance for next year."
                    )
                }
        else:
            # Payment exactly covered some installments
            if not outstanding:
                # All installments were already paid (shouldn't happen here)
                return {
                    'success': True,
                    'transactions': processed_transactions,
                    'distribution': distribution_details,
                    'summary': summary,
                    'message': f"Payment of TSh {float(self.amount):,.2f} processed.\n{summary}"
                }
            else:
                # Partial payment of current outstanding installments
                next_outstanding = outstanding[len(processed_transactions):] if len(processed_transactions) < len(outstanding) else []
                
                next_installment_info = ""
                if next_outstanding:
                    next_installment = next_outstanding[0]
                    next_installment_info = (
                        f"\nNext pending: Installment #{next_installment['plan'].installment_number} "
                        f"requires TSh {float(next_installment['remaining']):,.2f}"
                    )
                
                return {
                    'success': True,
                    'transactions': processed_transactions,
                    'distribution': distribution_details,
                    'summary': summary,
                    'message': (
                        f"Payment of TSh {float(self.amount):,.2f} processed.\n"
                        f"{summary}"
                        f"{next_installment_info}"
                    )
                }

    def _create_distribution_summary(self, distribution, total_paid, remaining, covered_count, total_outstanding):
        """
        Create a human-readable summary of payment distribution
        """
        summary_lines = []
        summary_lines.append(f"Total amount paid: TSh {float(total_paid):,.2f}")
        summary_lines.append(f"Installments affected: {covered_count}")
        
        # Detailed breakdown
        summary_lines.append("\nDistribution breakdown:")
        for dist in distribution:
            status_icon = "✅" if dist['status'] == 'Fully Paid' else "⚠️"
            summary_lines.append(
                f"  {status_icon} Installment #{dist['installment_number']}: "
                f"Paid TSh {dist['amount_paid']:,.2f} | "
                f"Remaining: TSh {dist['new_remaining']:,.2f}"
            )
        
        if remaining > 0:
            summary_lines.append(f"\n💰 Remaining amount to be handled: TSh {float(remaining):,.2f}")
        
        return "\n".join(summary_lines)
    
    def _process_monthly_payment(self):
        """
        Process monthly payment
        """
        if not self.month or not self.year:
            return {
                'success': False,
                'message': 'Month and year required for monthly payment'
            }
        
        monthly_fee = self.hostel.total_fee / 12
        
        # Check if payment exceeds monthly fee
        if self.amount > monthly_fee:
            excess = self.amount - monthly_fee
            
            # Process the monthly portion
            transaction = self._create_transaction(
                amount=monthly_fee,
                month=self.month,
                year=self.year
            )
            
            # Handle excess
            if self.is_final_year:
                return {
                    'success': True,
                    'transaction': transaction,
                    'refund': self._create_refund(excess, "Excess monthly payment in final year"),
                    'message': f'Monthly payment processed. Excess TSh {excess} refunded.'
                }
            else:
                return {
                    'success': True,
                    'transaction': transaction,
                    'balance': self._store_as_balance(excess),
                    'message': f'Monthly payment processed. TSh {excess} stored as balance.'
                }
        else:
            transaction = self._create_transaction(
                amount=self.amount,
                month=self.month,
                year=self.year
            )
            
            # Check if payment is less than monthly fee (partial payment)
            if self.amount < monthly_fee:
                return {
                    'success': True,
                    'transaction': transaction,
                    'message': f'Partial monthly payment of TSh {self.amount} recorded.'
                }
            else:
                return {
                    'success': True,
                    'transaction': transaction,
                    'message': 'Monthly payment processed successfully.'
                }
    
    def _process_yearly_payment(self):
        """
        Process yearly payment
        """
        yearly_fee = self.hostel.total_fee
        
        if self.amount > yearly_fee:
            excess = self.amount - yearly_fee
            
            # Process yearly portion
            transaction = self._create_transaction(amount=yearly_fee)
            
            # Handle excess
            if self.is_final_year:
                return {
                    'success': True,
                    'transaction': transaction,
                    'refund': self._create_refund(excess, "Excess yearly payment in final year"),
                    'message': f'Yearly payment processed. Excess TSh {excess} refunded.'
                }
            else:
                return {
                    'success': True,
                    'transaction': transaction,
                    'balance': self._store_as_balance(excess),
                    'message': f'Yearly payment processed. TSh {excess} stored as balance.'
                }
        else:
            transaction = self._create_transaction(amount=self.amount)
            
            # Check if payment is less than yearly fee (partial payment)
            if self.amount < yearly_fee:
                return {
                    'success': True,
                    'transaction': transaction,
                    'message': f'Partial yearly payment of TSh {self.amount} recorded.'
                }
            else:
                return {
                    'success': True,
                    'transaction': transaction,
                    'message': 'Yearly payment processed successfully.'
                }
    
    def _create_refund(self, amount, reason):
        """
        Create a refund record
        """
        return {
            'amount': float(amount),
            'reason': reason,
            'student': self.student.full_name,
            'date': timezone.now().date().isoformat()
        }
    
    def _store_as_balance(self, amount):
        """
        Store amount as balance for future use
        """
        return {
            'amount': float(amount),
            'student': self.student.full_name,
            'allocation': self.allocation.id,
            'date': timezone.now().date().isoformat()
        }
    
    def _generate_receipt_number(self):
        """
        Generate a unique receipt number
        Format: RCP/YYYYMMDD/XXXX
        """
        today = timezone.now()
        date_str = today.strftime('%Y%m%d')
        
        # Get count of transactions today
        count_today = HostelPaymentTransaction.objects.filter(
            payment_date=today.date()
        ).count()
        
        return f"RCP/{date_str}/{count_today + 1:04d}"


@login_required
def process_payment(request):
    """
    AJAX endpoint to process a payment
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        # Get form data
        allocation_id = request.POST.get('allocation_id')
        amount = request.POST.get('amount')
        payment_type = request.POST.get('payment_type')
        payment_method = request.POST.get('payment_method')
        transaction_number = request.POST.get('transaction_number', '').strip()
        month = request.POST.get('month')
        year = request.POST.get('year')
        
        # Validate required fields
        if not allocation_id:
            return JsonResponse({'success': False, 'message': 'Student allocation is required'})
        
        if not amount:
            return JsonResponse({'success': False, 'message': 'Amount is required'})
        
        if not payment_type:
            return JsonResponse({'success': False, 'message': 'Payment type is required'})
        
        if not payment_method:
            return JsonResponse({'success': False, 'message': 'Payment method is required'})
        
        # Validate transaction number for non-cash payments
        if payment_method != 'cash' and not transaction_number:
            return JsonResponse({
                'success': False, 
                'message': f'Transaction number is required for {dict(HostelPaymentTransaction.PAYMENT_METHOD).get(payment_method, payment_method)} payments'
            })
        
        # Check if transaction number already exists (if provided)
        if transaction_number:
            existing_transaction = HostelPaymentTransaction.objects.filter(
                transaction_number=transaction_number
            ).first()
            if existing_transaction:
                return JsonResponse({
                    'success': False,
                    'message': f'Transaction number "{transaction_number}" already exists in the system'
                })
        
        # Validate amount
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return JsonResponse({'success': False, 'message': 'Amount must be greater than 0'})
        except:
            return JsonResponse({'success': False, 'message': 'Invalid amount format'})
        
        # Get allocation
        allocation = get_object_or_404(
            StudentHostelAllocation.objects.select_related(
                'student', 'hostel', 'student__class_level'
            ),
            id=allocation_id,
            is_active=True
        )
        
        # Validate payment type matches hostel payment mode
        if payment_type != allocation.hostel.payment_mode and payment_type != 'installment':
            if allocation.hostel.payment_mode == 'installments' and payment_type != 'installment':
                return JsonResponse({
                    'success': False,
                    'message': f'This hostel uses installment payments. Please select "Installment" payment type.'
                })
        
        # Validate monthly payment requirements
        if payment_type == 'monthly':
            if not month or not year:
                return JsonResponse({
                    'success': False,
                    'message': 'Month and year are required for monthly payments'
                })
            
            # Check if payment for this month already exists
            existing_monthly = HostelPaymentTransaction.objects.filter(
                allocation=allocation,
                payment_type='monthly',
                month=month,
                year=year
            ).exists()
            
            if existing_monthly:
                return JsonResponse({
                    'success': False,
                    'message': f'A payment for {month}/{year} already exists for this student'
                })
        
        # Process payment with auto-generated receipt number
        processor = PaymentProcessor(
            allocation=allocation,
            amount=amount,
            payment_type=payment_type,
            payment_method=payment_method,
            transaction_number=transaction_number if transaction_number else None,
            month=int(month) if month else None,
            year=int(year) if year else None
        )
        
        result = processor.process_payment()
        
        # Ensure result has a message
        if 'message' not in result:
            result['message'] = 'Payment processed successfully'
        
        # Convert any model instances to serializable data
        serialized_result = serialize_payment_result(result)
        
        return JsonResponse({
            'success': True,
            'message': result['message'],
            'data': serialized_result
        })
        
    except ValidationError as e:
        return JsonResponse({'success': False, 'message': str(e)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': f'Error processing payment: {str(e)}'})


def serialize_payment_result(result):
    """
    Helper function to convert model instances to serializable dictionaries
    """
    serialized = {}
    
    for key, value in result.items():
        if key == 'transactions' and value:
            # Convert list of transaction objects to dictionaries
            serialized[key] = []
            for transaction in value:
                if hasattr(transaction, 'id'):  # Check if it's a model instance
                    serialized[key].append({
                        'id': transaction.id,
                        'amount': float(transaction.amount),
                        'receipt_number': transaction.receipt_number,
                        'payment_type': transaction.payment_type,
                        'payment_method': transaction.payment_method,
                        'transaction_number': transaction.transaction_number,
                        'payment_date': transaction.payment_date.isoformat() if transaction.payment_date else None,
                    })
                else:
                    serialized[key].append(value)
        elif key == 'transaction' and value and hasattr(value, 'id'):
            # Convert single transaction object to dictionary
            serialized[key] = {
                'id': value.id,
                'amount': float(value.amount),
                'receipt_number': value.receipt_number,
                'payment_type': value.payment_type,
                'payment_method': value.payment_method,
                'transaction_number': value.transaction_number,
                'payment_date': value.payment_date.isoformat() if value.payment_date else None,
            }
        elif key == 'refund' and value:
            # Refund is already a dict, ensure amounts are float
            serialized[key] = value
            if 'amount' in value:
                serialized[key]['amount'] = float(value['amount'])
        elif key == 'balance' and value:
            # Balance is already a dict, ensure amounts are float
            serialized[key] = value
            if 'amount' in value:
                serialized[key]['amount'] = float(value['amount'])
        else:
            serialized[key] = value
    
    return serialized


@login_required
def process_refund(request):
    """
    Process a refund for a student
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        student_id = request.POST.get('student_id')
        amount = request.POST.get('amount')
        reason = request.POST.get('reason', '')
        
        if not student_id or not amount:
            return JsonResponse({'success': False, 'message': 'Student and amount are required'})
        
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return JsonResponse({'success': False, 'message': 'Amount must be greater than 0'})
        except:
            return JsonResponse({'success': False, 'message': 'Invalid amount format'})
        
        student = get_object_or_404(Student, id=student_id)
        
        # Process refund logic here
        # You might want to create a Refund model
        
        return JsonResponse({
            'success': True,
            'message': f'Refund of TSh {amount} processed for {student.full_name}',
            'refund': {
                'student': student.full_name,
                'amount': float(amount),
                'reason': reason,
                'date': timezone.now().date().isoformat()
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing refund: {str(e)}'})


@login_required
def student_balance_info(request, student_id):
    """
    Get balance information for a student
    """
    try:
        student = get_object_or_404(Student, id=student_id)
        
        # Get active allocation
        allocation = StudentHostelAllocation.objects.filter(
            student=student,
            is_active=True
        ).select_related('hostel').first()
        
        if not allocation:
            return JsonResponse({
                'success': False,
                'message': 'Student has no active hostel allocation'
            })
        
        # Calculate totals
        total_paid = allocation.payment_transactions.aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        balance = allocation.total_fee - total_paid
        
        # Get installment information if applicable
        installments = []
        if allocation.hostel.payment_mode == 'installments':
            plans = HostelInstallmentPlan.objects.filter(
                hostel=allocation.hostel
            ).order_by('installment_number')
            
            for plan in plans:
                installment_payment, created = HostelPayment.objects.get_or_create(
                    allocation=allocation,
                    installment_plan=plan
                )
                
                paid = installment_payment.transactions.aggregate(
                    total=Sum('amount')
                )['total'] or 0
                
                installments.append({
                    'number': plan.installment_number,
                    'required': float(plan.amount),
                    'paid': float(paid),
                    'remaining': float(plan.amount - paid),
                    'status': 'paid' if paid >= plan.amount else 'partial' if paid > 0 else 'unpaid'
                })
        
        return JsonResponse({
            'success': True,
            'data': {
                'student': student.full_name,
                'allocation_id': allocation.id,
                'hostel': allocation.hostel.name,
                'total_fee': float(allocation.total_fee),
                'total_paid': float(total_paid),
                'balance': float(balance),
                'payment_mode': allocation.hostel.payment_mode,
                'installments': installments if installments else None
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })        




@login_required
def hostel_payments_export_excel(request):
    """
    Export hostel payments to Excel using OpenPyXL with enhanced styling
    """
    # Get filter parameters
    filters = get_payment_filters_from_request(request)
    
    # Get filtered transactions
    transactions = get_filtered_payment_transactions(filters)
    
    # Calculate totals for summary
    total_amount = transactions.aggregate(total=Sum('amount'))['total'] or 0
    total_transactions = transactions.count()
    hostels_count = transactions.values('allocation__hostel').distinct().count()
    average_payment = total_amount / total_transactions if total_transactions > 0 else 0
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hostel Payments"
    
    # ============================================
    # Define Styles
    # ============================================
    
    # School Header Styles
    school_title_font = Font(name='Arial', size=18, bold=True, color='4e73df')
    school_subtitle_font = Font(name='Arial', size=11, bold=False, color='666666')
    school_info_font = Font(name='Arial', size=10, bold=False, color='333333')
    
    # Section Header Styles
    section_header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    section_header_fill = PatternFill(start_color='4e73df', end_color='4e73df', fill_type='solid')
    section_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    section_header_border = Border(
        left=Side(style='medium', color='4e73df'),
        right=Side(style='medium', color='4e73df'),
        top=Side(style='medium', color='4e73df'),
        bottom=Side(style='medium', color='4e73df')
    )
    
    # Summary Section Styles
    summary_label_font = Font(name='Arial', size=10, bold=True, color='4e73df')
    summary_value_font = Font(name='Arial', size=12, bold=True, color='333333')
    summary_cell_border = Border(
        left=Side(style='thin', color='4e73df'),
        right=Side(style='thin', color='4e73df'),
        top=Side(style='thin', color='4e73df'),
        bottom=Side(style='thin', color='4e73df')
    )
    
    # Filter Section Styles
    filter_header_font = Font(name='Arial', size=10, bold=True, color='856404')
    filter_text_font = Font(name='Arial', size=9, color='856404')
    filter_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
    filter_border = Border(
        left=Side(style='thin', color='ffeeba'),
        right=Side(style='thin', color='ffeeba'),
        top=Side(style='thin', color='ffeeba'),
        bottom=Side(style='thin', color='ffeeba')
    )
    
    # Table Header Styles
    table_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    table_header_fill = PatternFill(start_color='4e73df', end_color='4e73df', fill_type='solid')
    table_header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    table_header_border = Border(
        left=Side(style='medium', color='224abe'),
        right=Side(style='medium', color='224abe'),
        top=Side(style='medium', color='224abe'),
        bottom=Side(style='medium', color='224abe')
    )
    
    # Cell Styles
    cell_font = Font(name='Arial', size=9)
    cell_border = Border(
        left=Side(style='thin', color='d1d3e2'),
        right=Side(style='thin', color='d1d3e2'),
        top=Side(style='thin', color='d1d3e2'),
        bottom=Side(style='thin', color='d1d3e2')
    )
    
    # Amount Cell Styles
    amount_font = Font(name='Arial', size=9, bold=True)
    amount_alignment = Alignment(horizontal='right', vertical='center')
    positive_font = Font(name='Arial', size=9, bold=True, color='e74a3b')
    zero_font = Font(name='Arial', size=9, bold=True, color='1cc88a')
    
    # Footer Styles
    footer_font = Font(name='Arial', size=8, italic=True, color='666666')
    
    # ============================================
    # Build School Header (Rows 1-5)
    # ============================================
    
    # School Logo/Name
    ws.merge_cells('A1:O1')
    school_title = ws.cell(row=1, column=1, value="MOUNT KILIMANJARO UNIVERSITY")
    school_title.font = school_title_font
    school_title.alignment = Alignment(horizontal='center', vertical='center')
    
    # School Subtitle
    ws.merge_cells('A2:O2')
    school_subtitle = ws.cell(row=2, column=1, value="Hostel Management System - Payment Report")
    school_subtitle.font = school_subtitle_font
    school_subtitle.alignment = Alignment(horizontal='center', vertical='center')
    
    # School Contact Info
    ws.merge_cells('A3:O3')
    contact_info = ws.cell(row=3, column=1, value="P.O. Box 1234, Moshi, Tanzania | Tel: +255 123 456 789 | Email: finance@mkuniversity.ac.tz")
    contact_info.font = school_info_font
    contact_info.alignment = Alignment(horizontal='center', vertical='center')
    
    # Generation Info
    ws.merge_cells('A4:O4')
    user_name = request.user.get_full_name() if request.user.get_full_name() else request.user.username
    gen_info = ws.cell(row=4, column=1, value=f"Generated on: {timezone.now().strftime('%d/%m/%Y %H:%M:%S')} | Generated by: {user_name}")
    gen_info.font = school_info_font
    gen_info.alignment = Alignment(horizontal='center', vertical='center')
    
    # Empty row for spacing
    ws.merge_cells('A5:O5')
    ws.cell(row=5, column=1, value="")
    
    # ============================================
    # Summary Statistics Section (Rows 6-11)
    # ============================================
    
    # Summary Header
    ws.merge_cells('A6:O6')
    summary_header = ws.cell(row=6, column=1, value="SUMMARY STATISTICS")
    summary_header.font = section_header_font
    summary_header.fill = section_header_fill
    summary_header.alignment = section_header_alignment
    summary_header.border = section_header_border
    
    # Statistics Grid
    stats = [
        ('Total Transactions', total_transactions),
        ('Total Amount', f'TSh {total_amount:,.2f}'),
        ('Average Payment', f'TSh {average_payment:,.2f}'),
        ('Hostels Covered', hostels_count),
        ('Date Range', f"{filters.get('date_from', 'Earliest')} to {filters.get('date_to', 'Latest')}"),
        ('Payment Types', ', '.join(transactions.values_list('payment_type', flat=True).distinct()) or 'All'),
    ]
    
    stat_row = 7
    for i, (label, value) in enumerate(stats):
        col = (i % 3) * 5 + 1  # Distribute across 3 columns (1, 6, 11)
        
        # Label
        label_cell = ws.cell(row=stat_row + (i//3)*2, column=col, value=f"{label}:")
        label_cell.font = summary_label_font
        label_cell.border = summary_cell_border
        
        # Value
        value_cell = ws.cell(row=stat_row + (i//3)*2 + 1, column=col, value=value)
        value_cell.font = summary_value_font
        value_cell.border = summary_cell_border
        
        # Add empty cells to maintain grid structure (prevents merging issues)
        for c in range(col + 1, min(col + 5, 16)):
            if c <= 15:  # Ensure we don't go beyond column O
                empty_cell = ws.cell(row=stat_row + (i//3)*2, column=c, value="")
                empty_cell.border = summary_cell_border
                empty_cell = ws.cell(row=stat_row + (i//3)*2 + 1, column=c, value="")
                empty_cell.border = summary_cell_border
    
    # Empty row after summary
    current_row = stat_row + 4
    
    # ============================================
    # Applied Filters Section (if any)
    # ============================================
    
    if filters:
        ws.merge_cells(f'A{current_row}:O{current_row}')
        filter_header = ws.cell(row=current_row, column=1, value="APPLIED FILTERS")
        filter_header.font = filter_header_font
        filter_header.fill = filter_fill
        filter_header.alignment = Alignment(horizontal='center', vertical='center')
        filter_header.border = filter_border
        current_row += 1
        
        filter_text = []
        if filters.get('hostel_id'): 
            hostel = Hostel.objects.filter(id=filters['hostel_id']).first()
            filter_text.append(f"Hostel: {hostel.name if hostel else filters['hostel_id']}")
        if filters.get('payment_type'): 
            payment_type_display = dict(HostelPaymentTransaction.PAYMENT_TYPE).get(filters['payment_type'], filters['payment_type'])
            filter_text.append(f"Payment Type: {payment_type_display}")
        if filters.get('payment_method'): 
            method_display = dict(HostelPaymentTransaction.PAYMENT_METHOD).get(filters['payment_method'], filters['payment_method'])
            filter_text.append(f"Payment Method: {method_display}")
        if filters.get('balance_status'): 
            status_display = 'Fully Paid' if filters['balance_status'] == 'paid' else 'Partial Payment'
            filter_text.append(f"Balance Status: {status_display}")
        if filters.get('date_from'): filter_text.append(f"From: {filters['date_from']}")
        if filters.get('date_to'): filter_text.append(f"To: {filters['date_to']}")
        if filters.get('min_amount'): filter_text.append(f"Min Amount: TSh {float(filters['min_amount']):,.2f}")
        if filters.get('max_amount'): filter_text.append(f"Max Amount: TSh {float(filters['max_amount']):,.2f}")
        
        ws.merge_cells(f'A{current_row}:O{current_row}')
        filter_cell = ws.cell(row=current_row, column=1, value=" | ".join(filter_text))
        filter_cell.font = filter_text_font
        filter_cell.fill = filter_fill
        filter_cell.alignment = Alignment(horizontal='left', vertical='center')
        filter_cell.border = filter_border
        current_row += 1
        
        # Empty row after filters
        ws.merge_cells(f'A{current_row}:O{current_row}')
        ws.cell(row=current_row, column=1, value="")
        current_row += 1
    else:
        # Show no filters message
        ws.merge_cells(f'A{current_row}:O{current_row}')
        no_filter = ws.cell(row=current_row, column=1, value="NO FILTERS APPLIED - Showing all transactions")
        no_filter.font = filter_text_font
        no_filter.fill = filter_fill
        no_filter.alignment = Alignment(horizontal='center', vertical='center')
        no_filter.border = filter_border
        current_row += 2
    
    # ============================================
    # Payment Data Table
    # ============================================
    
    # Table Header
    headers = [
        'Receipt #', 'Student Name', 'Reg No', 'Class', 'Stream',
        'Hostel', 'Code', 'Amount (TSh)', 'Remaining (TSh)',
        'Payment Type', 'Method', 'Reference',
        'Payment Date', 'Created At', 'Status'
    ]
    
    header_row = current_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = table_header_alignment
        cell.border = table_header_border
    
    # Data rows
    for row_num, transaction in enumerate(transactions, header_row + 1):
        allocation = transaction.allocation
        student = allocation.student
        hostel = allocation.hostel
        
        # Calculate remaining amount
        if transaction.payment_type == 'installments' and transaction.installment_payment:
            remaining = transaction.installment_payment.remaining_amount
        else:
            remaining = allocation.balance
        
        # Determine status
        if remaining <= 0:
            status = 'FULLY PAID'
            status_color = '1cc88a'
        elif transaction.amount > 0:
            status = 'PARTIAL'
            status_color = 'e74a3b'
        else:
            status = 'UNPAID'
            status_color = '6c757d'
        
        row_data = [
            transaction.receipt_number,
            student.full_name,
            student.registration_number or 'N/A',
            student.class_level.name if student.class_level else 'N/A',
            student.stream_class.stream_letter if student.stream_class else 'N/A',
            hostel.name,
            hostel.code,
            float(transaction.amount),
            float(remaining),
            transaction.get_payment_type_display().upper(),
            transaction.get_payment_method_display().upper(),
            transaction.transaction_number or 'N/A',
            transaction.payment_date.strftime('%d/%m/%Y') if transaction.payment_date else 'N/A',
            transaction.created_at.strftime('%d/%m/%Y %H:%M'),
            status
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = cell_font
            cell.border = cell_border
            
            # Apply specific styling
            if col_num in [8, 9]:  # Amount columns
                cell.font = amount_font
                cell.alignment = amount_alignment
                cell.number_format = '#,##0.00'
                
                # Color code remaining amount
                if col_num == 9:
                    if remaining <= 0:
                        cell.font = zero_font
                    else:
                        cell.font = positive_font
            elif col_num in [13, 14]:  # Date columns
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_num == 15:  # Status column
                cell.font = Font(name='Arial', size=9, bold=True, color=status_color)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # ============================================
    # Totals Row
    # ============================================
    
    if transactions.exists():
        total_row = header_row + len(transactions) + 1
        
        # Total label
        total_label_cell = ws.cell(row=total_row, column=6, value="GRAND TOTAL:")
        total_label_cell.font = Font(bold=True)
        total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_label_cell.border = cell_border
        
        # Total amount
        total_amount_cell = ws.cell(row=total_row, column=8, value=float(total_amount))
        total_amount_cell.font = Font(bold=True, size=10)
        total_amount_cell.alignment = amount_alignment
        total_amount_cell.number_format = '#,##0.00'
        total_amount_cell.border = cell_border
        
        # Total transactions count
        total_count_cell = ws.cell(row=total_row, column=15, value=f"Count: {total_transactions}")
        total_count_cell.font = Font(bold=True, size=10)
        total_count_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_count_cell.border = cell_border
        
        footer_row = total_row + 2
    else:
        # No transactions - show message
        no_data_row = header_row + 1
        ws.merge_cells(f'A{no_data_row}:O{no_data_row}')
        no_data_cell = ws.cell(row=no_data_row, column=1, value="No payment transactions found matching the criteria.")
        no_data_cell.font = Font(italic=True, color='666666')
        no_data_cell.alignment = Alignment(horizontal='center', vertical='center')
        no_data_cell.border = cell_border
        
        footer_row = no_data_row + 2
    
    # ============================================
    # Footer Section
    # ============================================
    
    # Report generation note
    ws.merge_cells(f'A{footer_row}:O{footer_row}')
    footer_note = ws.cell(row=footer_row, column=1, 
                          value="This is a computer-generated report. No signature is required.")
    footer_note.font = footer_font
    footer_note.alignment = Alignment(horizontal='center', vertical='center')
    
    # Confidentiality notice
    ws.merge_cells(f'A{footer_row + 1}:O{footer_row + 1}')
    confidential = ws.cell(row=footer_row + 1, column=1, 
                          value="CONFIDENTIAL: This report contains financial information and is intended for authorized personnel only.")
    confidential.font = footer_font
    confidential.alignment = Alignment(horizontal='center', vertical='center')
    
    # ============================================
    # Auto-adjust Column Widths
    # ============================================
    
    # Fix for MergedCell issue - iterate through columns properly
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Iterate through rows in this column
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        # Set width with limits
        adjusted_width = min(max_length + 4, 60)  # Max 60 characters
        if column_letter in ['B', 'F']:  # Student Name and Hostel columns get more space
            adjusted_width = min(max_length + 8, 70)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # ============================================
    # Freeze Panes
    # ============================================
    
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    
    # ============================================
    # Add Data Validation and Filters
    # ============================================
    
    if transactions.exists():
        # Add autofilter to table headers
        ws.auto_filter.ref = f"A{header_row}:O{header_row + len(transactions)}"
    
    # ============================================
    # Create Response
    # ============================================
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"hostel_payments_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@login_required
def hostel_payments_export_pdf(request):
    """
    Export hostel payments to PDF using WeasyPrint
    """
    # Get filter parameters
    filters = get_payment_filters_from_request(request)
    
    # Get filtered transactions
    transactions = get_filtered_payment_transactions(filters)
    
    # Calculate totals
    total_amount = transactions.aggregate(total=Sum('amount'))['total'] or 0
    total_transactions = transactions.count()
    
    # Get unique hostels count
    hostels_count = transactions.values('allocation__hostel').distinct().count()
    
    # Prepare data for template
    payment_data = []
    for transaction in transactions:
        allocation = transaction.allocation
        student = allocation.student
        hostel = allocation.hostel
        
        # Calculate remaining amount
        if transaction.payment_type == 'installments' and transaction.installment_payment:
            remaining = transaction.installment_payment.remaining_amount
        else:
            remaining = allocation.balance
        
        payment_data.append({
            'receipt_number': transaction.receipt_number,
            'student_name': student.full_name,
            'registration_number': student.registration_number or 'N/A',
            'class_name': student.class_level.name if student.class_level else 'N/A',
            'stream': student.stream_class.stream_letter if student.stream_class else 'N/A',
            'hostel_name': hostel.name,
            'hostel_code': hostel.code,
            'amount': transaction.amount,
            'remaining': remaining,
            'payment_type': transaction.get_payment_type_display(),
            'payment_method': transaction.get_payment_method_display(),
            'transaction_number': transaction.transaction_number or 'N/A',
            'payment_date': transaction.payment_date.strftime('%d/%m/%Y') if transaction.payment_date else 'N/A',
            'created_at': transaction.created_at.strftime('%d/%m/%Y %H:%M'),
        })
    
    # Render HTML template
    html_string = render_to_string('admin/hostels/payments_export_pdf.html', {
        'payments': payment_data,
        'total_amount': total_amount,
        'total_transactions': total_transactions,
        'hostels_count': hostels_count,
        'generated_at': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
        'filters': filters,
        'request': request,
    })
    
    # Generate PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=hostel_payments_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    html.write_pdf(response)
    return response


def get_payment_filters_from_request(request):
    """
    Extract payment filters from request GET parameters
    """
    filters = {}
    
    # Get filter parameters
    hostel_id = request.GET.get('hostel')
    payment_type = request.GET.get('payment_type')
    payment_method = request.GET.get('payment_method')
    balance_status = request.GET.get('balance_status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    min_amount = request.GET.get('min_amount')
    max_amount = request.GET.get('max_amount')
    
    if hostel_id:
        filters['hostel_id'] = hostel_id
    
    if payment_type:
        filters['payment_type'] = payment_type
    
    if payment_method:
        filters['payment_method'] = payment_method
    
    if balance_status:
        filters['balance_status'] = balance_status
    
    if date_from:
        filters['date_from'] = date_from
    
    if date_to:
        filters['date_to'] = date_to
    
    if min_amount:
        filters['min_amount'] = min_amount
    
    if max_amount:
        filters['max_amount'] = max_amount
    
    return filters


def get_filtered_payment_transactions(filters):
    """
    Get filtered payment transactions based on filters
    """
    queryset = HostelPaymentTransaction.objects.select_related(
        'allocation__student',
        'allocation__student__class_level',
        'allocation__student__stream_class',
        'allocation__hostel',
        'installment_payment__installment_plan'
    ).all()
    
    # Apply filters
    if filters.get('hostel_id'):
        queryset = queryset.filter(allocation__hostel_id=filters['hostel_id'])
    
    if filters.get('payment_type'):
        queryset = queryset.filter(payment_type=filters['payment_type'])
    
    if filters.get('payment_method'):
        queryset = queryset.filter(payment_method=filters['payment_method'])
    
    if filters.get('date_from'):
        queryset = queryset.filter(payment_date__gte=filters['date_from'])
    
    if filters.get('date_to'):
        queryset = queryset.filter(payment_date__lte=filters['date_to'])
    
    if filters.get('min_amount'):
        queryset = queryset.filter(amount__gte=float(filters['min_amount']))
    
    if filters.get('max_amount'):
        queryset = queryset.filter(amount__lte=float(filters['max_amount']))
    
    # Balance status filter (post-processing)
    if filters.get('balance_status'):
        filtered_ids = []
        for transaction in queryset:
            if transaction.payment_type == 'installments' and transaction.installment_payment:
                remaining = transaction.installment_payment.remaining_amount
            else:
                remaining = transaction.allocation.balance
            
            if filters['balance_status'] == 'paid' and remaining <= 0:
                filtered_ids.append(transaction.id)
            elif filters['balance_status'] == 'partial' and remaining > 0:
                filtered_ids.append(transaction.id)
        
        queryset = queryset.filter(id__in=filtered_ids)
    
    return queryset.order_by('-payment_date', '-created_at')


@login_required
def hostel_payment_receipt_pdf(request, pk):
    """
    Generate a PDF receipt for a payment transaction using WeasyPrint
    """
    # Get the transaction
    transaction = get_object_or_404(
        HostelPaymentTransaction.objects.select_related(
            'allocation__student',
            'allocation__student__class_level',
            'allocation__student__stream_class',
            'allocation__hostel',
            'allocation__room',
            'allocation__bed',
            'installment_payment__installment_plan'
        ),
        id=pk
    )
    
    # Get related transactions if this is an installment payment
    related_transactions = []
    if transaction.installment_payment:
        related_transactions = HostelPaymentTransaction.objects.filter(
            installment_payment=transaction.installment_payment
        ).exclude(id=transaction.id).select_related(
            'installment_payment__installment_plan'
        ).order_by('-payment_date')
    
    # Calculate financial summary
    total_paid = transaction.allocation.total_paid
    total_fee = transaction.allocation.total_fee
    balance = total_fee - total_paid
    
    # Prepare context
    context = {
        'transaction': transaction,
        'related_transactions': related_transactions,
        'total_paid': total_paid,
        'total_fee': total_fee,
        'balance': balance,
        'generated_at': timezone.now(),
        'request': request,
    }
    
    # Render HTML template
    html_string = render_to_string('admin/hostels/payment_receipt_pdf.html', context)
    
    # Generate PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"receipt_{transaction.receipt_number}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Write PDF to response
    html.write_pdf(response)
    
    return response    



# ============================================================================
# SINGLE HOSTEL ROOMS LIST VIEW
# ============================================================================

@login_required
def single_hostel_rooms_list(request):
    """
    Display rooms for a specific hostel with enhanced details
    """
    # Get hostel ID from query parameters
    hostel_id = request.GET.get('hostel')
    if not hostel_id:
        messages.error(request, 'Hostel ID is required.')
        return redirect('admin_hostels_list')
    
    # Get the hostel
    hostel = get_object_or_404(
        Hostel.objects.prefetch_related('rooms', 'rooms__beds'),
        id=hostel_id,
        is_active=True
    )
    
    # Get all rooms for this hostel with bed statistics
    rooms = HostelRoom.objects.filter(
        hostel=hostel,
        is_active=True
    ).annotate(
        total_beds=Count('beds'),
        occupied_beds=Count('beds', filter=Q(beds__is_occupied=True)),
        available_beds=Count('beds', filter=Q(beds__is_occupied=False))
    ).order_by('room_number')
    
    # Calculate statistics
    total_rooms = rooms.count()
    total_capacity = rooms.aggregate(total=Sum('capacity'))['total'] or 0
    total_beds = rooms.aggregate(total=Sum('total_beds'))['total'] or 0
    occupied_beds = rooms.aggregate(total=Sum('occupied_beds'))['total'] or 0
    available_beds = total_beds - occupied_beds
    
    # Get room occupancy rates
    for room in rooms:
        room.occupancy_rate = (room.occupied_beds / room.capacity * 100) if room.capacity > 0 else 0
        room.bed_occupancy_rate = (room.occupied_beds / room.total_beds * 100) if room.total_beds > 0 else 0
    
    context = {
        'hostel': hostel,
        'rooms': rooms,
        'total_rooms': total_rooms,
        'total_capacity': total_capacity,
        'total_beds': total_beds,
        'occupied_beds': occupied_beds,
        'available_beds': available_beds,
        'page_title': f'Rooms in {hostel.name}',
    }
    
    return render(request, 'admin/hostels/single_hostel_rooms.html', context)


# ============================================================================
# SINGLE HOSTEL PAYMENTS LIST VIEW
# ============================================================================

@login_required
def hostel_student_payments_list(request):
    """
    Display all payments for students in a specific hostel
    """
    # Get hostel ID from query parameters
    hostel_id = request.GET.get('hostel')
    if not hostel_id:
        messages.error(request, 'Hostel ID is required.')
        return redirect('admin_hostels_list')
    
    # Get the hostel
    hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
    
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    payment_type = request.GET.get('payment_type')
    payment_method = request.GET.get('payment_method')
    student_id = request.GET.get('student')
    class_level_id = request.GET.get('class_level')
    min_amount = request.GET.get('min_amount')
    max_amount = request.GET.get('max_amount')
    
    # Base queryset for payments in this hostel
    transactions = HostelPaymentTransaction.objects.filter(
        allocation__hostel=hostel
    ).select_related(
        'allocation',
        'allocation__student',
        'allocation__student__class_level',
        'allocation__student__stream_class',
        'allocation__room',
        'allocation__bed',
        'installment_payment',
        'installment_payment__installment_plan'
    ).order_by('-payment_date', '-created_at')
    
    # Apply filters
    if date_from:
        transactions = transactions.filter(payment_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(payment_date__lte=date_to)
    if payment_type:
        transactions = transactions.filter(payment_type=payment_type)
    if payment_method:
        transactions = transactions.filter(payment_method=payment_method)
    if student_id:
        transactions = transactions.filter(allocation__student_id=student_id)
    if class_level_id:
        transactions = transactions.filter(allocation__student__class_level_id=class_level_id)
    if min_amount:
        try:
            transactions = transactions.filter(amount__gte=float(min_amount))
        except ValueError:
            pass
    if max_amount:
        try:
            transactions = transactions.filter(amount__lte=float(max_amount))
        except ValueError:
            pass
    
    # Calculate statistics
    total_transactions = transactions.count()
    total_amount = transactions.aggregate(total=Sum('amount'))['total'] or 0
    average_payment = total_amount / total_transactions if total_transactions > 0 else 0
    
    # Payment type breakdown
    payment_type_breakdown = transactions.values('payment_type').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_type')
    
    # Payment method breakdown
    payment_method_breakdown = transactions.values('payment_method').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('payment_method')
    
    # Monthly summary for charts - using different annotation name to avoid conflict
    from django.db.models.functions import ExtractMonth
    monthly_summary = transactions.filter(
        payment_date__isnull=False
    ).annotate(
        payment_month=ExtractMonth('payment_date')
    ).values('payment_month').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('payment_month')
    
    # Get students in this hostel for filter dropdown
    students_in_hostel = Student.objects.filter(
        hostel_allocations__hostel=hostel,
        hostel_allocations__is_active=True
    ).distinct().order_by('first_name', 'last_name')
    
    # FIXED: Get class levels for filter dropdown - using correct related name 'students' (plural)
    class_levels = ClassLevel.objects.filter(
        students__hostel_allocations__hostel=hostel,  # Changed from 'student' to 'students'
        students__hostel_allocations__is_active=True
    ).distinct().order_by('name')
    
    # Format monthly summary for template
    formatted_monthly_summary = []
    month_names = {
        1: 'January', 2: 'February', 3: 'March', 4: 'April',
        5: 'May', 6: 'June', 7: 'July', 8: 'August',
        9: 'September', 10: 'October', 11: 'November', 12: 'December'
    }
    
    for item in monthly_summary:
        if item['payment_month']:  # Check if month exists
            formatted_monthly_summary.append({
                'month_number': item['payment_month'],
                'month_name': month_names.get(item['payment_month'], f'Month {item["payment_month"]}'),
                'total': item['total'],
                'count': item['count']
            })
    
    # Also get all class levels for the filter dropdown (fallback)
    all_class_levels = ClassLevel.objects.filter(is_active=True).order_by('name')
    
    context = {
        'hostel': hostel,
        'transactions': transactions[:100],  # Limit for performance
        'total_transactions': total_transactions,
        'total_amount': total_amount,
        'average_payment': average_payment,
        'payment_type_breakdown': payment_type_breakdown,
        'payment_method_breakdown': payment_method_breakdown,
        'monthly_summary': formatted_monthly_summary,
        'students': students_in_hostel,
        'class_levels': class_levels if class_levels.exists() else all_class_levels,  # Fallback if no filtered results
        'payment_types': HostelPaymentTransaction.PAYMENT_TYPE,
        'payment_methods': HostelPaymentTransaction.PAYMENT_METHOD,
        'filter_params': {
            'date_from': date_from,
            'date_to': date_to,
            'payment_type': payment_type,
            'payment_method': payment_method,
            'student_id': student_id,
            'class_level_id': class_level_id,
            'min_amount': min_amount,
            'max_amount': max_amount,
        },
        'page_title': f'Payments - {hostel.name}',
    }
    
    return render(request, 'admin/hostels/hostel_payments_list.html', context)


# ============================================================================
# SINGLE HOSTEL INSTALLMENT PLANS LIST VIEW
# ============================================================================

from decimal import Decimal

@login_required
def hostel_installment_list(request):
    """
    Display all installment plans for a specific hostel
    """
    # Get hostel ID from query parameters
    hostel_id = request.GET.get('hostel')
    if not hostel_id:
        messages.error(request, 'Hostel ID is required.')
        return redirect('admin_hostels_list')
    
    # Get the hostel
    hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)
    
    # Get all installment plans for this hostel
    installment_plans = HostelInstallmentPlan.objects.filter(
        hostel=hostel
    ).order_by('installment_number')
    
    # Calculate statistics
    from django.db.models import Sum, Count
    total_plans = installment_plans.count()
    total_amount = installment_plans.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Get payment statistics for each installment plan
    for plan in installment_plans:
        # Get all payments for this installment plan in this hostel
        payments = HostelPayment.objects.filter(
            installment_plan=plan,
            allocation__hostel=hostel
        ).prefetch_related('transactions')
        
        # Count total students with this installment plan
        total_students = payments.values('allocation__student').distinct().count()
        
        # Calculate total collected amount across all payments
        total_collected = payments.aggregate(
            total=Sum('transactions__amount')
        )['total'] or Decimal('0')
        
        # Count fully paid, partially paid, and unpaid
        fully_paid = 0
        partially_paid = 0
        unpaid = 0
        
        for payment in payments:
            # Calculate total paid for this payment by summing transactions
            paid_amount = payment.transactions.aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            required_amount = payment.installment_plan.amount
            
            if paid_amount >= required_amount:
                fully_paid += 1
            elif paid_amount > 0:
                partially_paid += 1
            else:
                unpaid += 1
        
        # Calculate collection rate using Decimal
        if total_students > 0 and plan.amount > 0:
            expected_total = plan.amount * total_students
            collection_rate = (total_collected / expected_total * Decimal('100'))
        else:
            collection_rate = Decimal('0')
        
        plan.payment_stats = {
            'total_students': total_students,
            'total_collected': float(total_collected),  # Convert for template if needed
            'fully_paid': fully_paid,
            'partially_paid': partially_paid,
            'unpaid': unpaid,
            'collection_rate': float(collection_rate),  # Convert for template
        }
    
    # Check if total matches hostel fee - use Decimal comparison
    if hostel.total_fee:
        matches_hostel_fee = abs(total_amount - hostel.total_fee) < Decimal('0.01')
        difference = hostel.total_fee - total_amount
    else:
        matches_hostel_fee = False
        difference = Decimal('0')
    
    context = {
        'hostel': hostel,
        'installment_plans': installment_plans,
        'total_plans': total_plans,
        'total_amount': total_amount,
        'matches_hostel_fee': matches_hostel_fee,
        'difference': difference,
        'page_title': f'Installment Plans - {hostel.name}',
    }
    
    return render(request, 'admin/hostels/hostel_installment_plans.html', context)

# ============================================================================
# SINGLE HOSTEL STUDENT ALLOCATION CREATE VIEW
# ============================================================================

@login_required
def hostel_student_allocation(request):
    """
    Create a new student allocation specifically for a hostel
    This redirects to the bulk operations page with pre-selected hostel
    """
    # Get hostel ID from query parameters
    hostel_id = request.GET.get('hostel')
    
    if hostel_id:
        # Verify hostel exists
        try:
            hostel = Hostel.objects.get(id=hostel_id, is_active=True)
            # Redirect to the bulk operations page with hostel pre-selected
            return redirect(f"{reverse('admin_student_allocation_bulk')}?hostel={hostel_id}")
        except Hostel.DoesNotExist:
            messages.error(request, 'Hostel not found.')
            return redirect('admin_hostels_list')
    else:
        # If no hostel specified, go to regular bulk operations page
        return redirect('admin_student_allocation_bulk')

# ============================================================================
# AJAX ENDPOINTS FOR THESE VIEWS
# ============================================================================

@login_required
def get_hostel_room_details(request, room_id):
    """
    AJAX endpoint to get detailed information about a specific room
    """
    try:
        room = get_object_or_404(
            HostelRoom.objects.select_related('hostel').prefetch_related('beds'),
            id=room_id,
            is_active=True
        )
        
        # Get bed details
        beds = room.beds.all().order_by('bed_number')
        bed_list = []
        for bed in beds:
            bed_list.append({
                'id': bed.id,
                'bed_number': bed.bed_number,
                'bed_type': bed.get_bed_type_display(),
                'is_occupied': bed.is_occupied,
                'student_name': bed.current_allocation.student.full_name if bed.is_occupied and hasattr(bed, 'current_allocation') else None
            })
        
        # Get current allocations in this room
        current_allocations = StudentHostelAllocation.objects.filter(
            room=room,
            is_active=True
        ).select_related('student').order_by('bed__bed_number')
        
        allocation_list = []
        for allocation in current_allocations:
            allocation_list.append({
                'id': allocation.id,
                'student_name': allocation.student.full_name,
                'student_reg': allocation.student.registration_number,
                'bed_number': allocation.bed.bed_number if allocation.bed else 'Not Assigned',
                'payment_status': allocation.payment_status
            })
        
        return JsonResponse({
            'success': True,
            'room': {
                'id': room.id,
                'room_number': room.room_number,
                'capacity': room.capacity,
                'total_beds': len(bed_list),
                'occupied_beds': sum(1 for bed in bed_list if bed['is_occupied']),
                'available_beds': room.capacity - len(bed_list),
                'beds': bed_list,
                'current_allocations': allocation_list,
                'hostel_name': room.hostel.name,
                'hostel_code': room.hostel.code,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })


@login_required
def get_hostel_payment_summary(request, hostel_id):
    """
    AJAX endpoint to get payment summary for a hostel
    """
    try:
        hostel = get_object_or_404(Hostel, id=hostel_id)
        
        # Get all active allocations in this hostel
        allocations = StudentHostelAllocation.objects.filter(
            hostel=hostel,
            is_active=True
        )
        
        # Calculate payment statistics
        total_students = allocations.count()
        
        # Get all transactions for this hostel
        transactions = HostelPaymentTransaction.objects.filter(
            allocation__hostel=hostel
        )
        
        total_collected = transactions.aggregate(total=Sum('amount'))['total'] or 0
        total_expected = allocations.aggregate(total=Sum('total_fee'))['total'] or 0
        
        # Payment status breakdown
        paid_students = 0
        partial_students = 0
        unpaid_students = 0
        
        for allocation in allocations:
            if allocation.balance <= 0:
                paid_students += 1
            elif allocation.total_paid > 0:
                partial_students += 1
            else:
                unpaid_students += 1
        
        return JsonResponse({
            'success': True,
            'summary': {
                'total_students': total_students,
                'total_collected': float(total_collected),
                'total_expected': float(total_expected),
                'outstanding': float(total_expected - total_collected),
                'collection_rate': (total_collected / total_expected * 100) if total_expected > 0 else 0,
                'paid_students': paid_students,
                'partial_students': partial_students,
                'unpaid_students': unpaid_students,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })    


# ============================================================================
# HOSTEL ROOM CREATE VIEW
# ============================================================================

@login_required
def hostel_room_create(request):
    """
    Create a new room in a specific hostel
    """
    # Get hostel ID from query parameters
    hostel_id = request.GET.get('hostel')
    if not hostel_id:
        messages.error(request, 'Hostel ID is required.')
        return redirect('admin_hostels_list')
    
    # Get the hostel
    hostel = get_object_or_404(Hostel, id=hostel_id, is_active=True)

    if request.method == 'POST':
        try:
            # Get form data
            room_number = request.POST.get('room_number', '').strip().upper()
            capacity_str = request.POST.get('capacity', '').strip()
            is_active = request.POST.get('is_active') == 'on'

            # Validate required fields
            if not room_number:
                return JsonResponse({
                    'success': False,
                    'message': 'Room number is required.'
                })

            if not capacity_str:
                return JsonResponse({
                    'success': False,
                    'message': 'Capacity is required.'
                })

            # Validate room number
            if len(room_number) < 1:
                return JsonResponse({
                    'success': False,
                    'message': 'Room number must be at least 1 character long.'
                })

            if len(room_number) > 20:
                return JsonResponse({
                    'success': False,
                    'message': 'Room number cannot exceed 20 characters.'
                })

            # Validate room number format (alphanumeric with optional hyphen)
            if not room_number.replace('-', '').replace(' ', '').isalnum():
                return JsonResponse({
                    'success': False,
                    'message': 'Room number can only contain letters, numbers, spaces, and hyphens.'
                })

            # Validate capacity
            try:
                capacity = int(capacity_str)
                if capacity < 1:
                    return JsonResponse({
                        'success': False,
                        'message': 'Capacity must be at least 1.'
                    })
                if capacity > 20:  # Assuming max 20 students per room
                    return JsonResponse({
                        'success': False,
                        'message': 'Capacity cannot exceed 20 students per room.'
                    })
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Capacity must be a valid number.'
                })

            # Check for duplicate room number in the same hostel
            if HostelRoom.objects.filter(hostel=hostel, room_number__iexact=room_number).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Room "{room_number}" already exists in {hostel.name}.'
                })

            # Check hostel capacity
            current_total_capacity = HostelRoom.objects.filter(
                hostel=hostel, is_active=True
            ).aggregate(total=Sum('capacity'))['total'] or 0

            if current_total_capacity + capacity > hostel.max_students:
                return JsonResponse({
                    'success': False,
                    'message': (
                        f'Cannot add room. Total capacity would exceed hostel maximum '
                        f'of {hostel.max_students} students. '
                        f'Current capacity: {current_total_capacity}, '
                        f'Remaining: {hostel.max_students - current_total_capacity}'
                    )
                })

            # Create the room
            with transaction.atomic():
                room = HostelRoom.objects.create(
                    hostel=hostel,
                    room_number=room_number,
                    capacity=capacity,
                    is_active=is_active
                )

            messages.success(request, f'Room "{room_number}" created successfully in {hostel.name}.')

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Room "{room_number}" created successfully.',
                    'redirect_url': reverse('admin_single_hostel_rooms_list') + f'?hostel={hostel.id}'
                })
            return redirect(f"{reverse('admin_single_hostel_rooms_list')}?hostel={hostel.id}")

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error creating room: {str(e)}'
                })
            messages.error(request, f'Error creating room: {str(e)}')
            return redirect(f"{reverse('admin_hostel_room_create')}?hostel={hostel.id}")

    # GET request - show form
    # Calculate available capacity
    current_total_capacity = HostelRoom.objects.filter(
        hostel=hostel, is_active=True
    ).aggregate(total=Sum('capacity'))['total'] or 0
    
    available_capacity = hostel.max_students - current_total_capacity
    
    context = {
        'hostel': hostel,
        'available_capacity': available_capacity,
        'current_total_capacity': current_total_capacity,
        'page_title': f'Add Room to {hostel.name}',
        'form_title': 'Add New Room',
        'submit_text': 'Create Room',
        'cancel_url': f"{reverse('admin_single_hostel_rooms_list')}?hostel={hostel.id}",
    }
    
    return render(request, 'admin/hostels/room_form.html', context)


# ============================================================================
# HOSTEL ROOM EDIT VIEW
# ============================================================================

@login_required
def hostel_room_edit(request, room_id):
    """
    Edit an existing hostel room
    """
    # Get the room with related hostel
    room = get_object_or_404(
        HostelRoom.objects.select_related('hostel'),
        id=room_id
    )
    
    hostel = room.hostel

    if request.method == 'POST':
        try:
            # Get form data
            room_number = request.POST.get('room_number', '').strip().upper()
            capacity_str = request.POST.get('capacity', '').strip()
            is_active = request.POST.get('is_active') == 'on'

            # Validate required fields
            if not room_number:
                return JsonResponse({
                    'success': False,
                    'message': 'Room number is required.'
                })

            if not capacity_str:
                return JsonResponse({
                    'success': False,
                    'message': 'Capacity is required.'
                })

            # Validate room number
            if len(room_number) < 1:
                return JsonResponse({
                    'success': False,
                    'message': 'Room number must be at least 1 character long.'
                })

            if len(room_number) > 20:
                return JsonResponse({
                    'success': False,
                    'message': 'Room number cannot exceed 20 characters.'
                })

            # Validate room number format
            if not room_number.replace('-', '').replace(' ', '').isalnum():
                return JsonResponse({
                    'success': False,
                    'message': 'Room number can only contain letters, numbers, spaces, and hyphens.'
                })

            # Validate capacity
            try:
                capacity = int(capacity_str)
                if capacity < 1:
                    return JsonResponse({
                        'success': False,
                        'message': 'Capacity must be at least 1.'
                    })
                if capacity > 20:
                    return JsonResponse({
                        'success': False,
                        'message': 'Capacity cannot exceed 20 students per room.'
                    })
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Capacity must be a valid number.'
                })

            # Check for duplicate room number in the same hostel (excluding current)
            if HostelRoom.objects.filter(
                hostel=hostel, 
                room_number__iexact=room_number
            ).exclude(id=room.id).exists():
                return JsonResponse({
                    'success': False,
                    'message': f'Room "{room_number}" already exists in {hostel.name}.'
                })

            # Check hostel capacity when capacity is being increased
            if capacity > room.capacity:
                # Calculate total capacity excluding current room
                other_rooms_capacity = HostelRoom.objects.filter(
                    hostel=hostel,
                    is_active=True
                ).exclude(id=room.id).aggregate(total=Sum('capacity'))['total'] or 0

                if other_rooms_capacity + capacity > hostel.max_students:
                    return JsonResponse({
                        'success': False,
                        'message': (
                            f'Cannot increase capacity. Total would exceed hostel maximum '
                            f'of {hostel.max_students} students. '
                            f'Current total: {other_rooms_capacity}, '
                            f'Maximum allowed for this room: {hostel.max_students - other_rooms_capacity}'
                        )
                    })

            # Check if capacity reduction would affect existing beds/allocations
            if capacity < room.capacity:
                # Get current beds count
                current_beds = Bed.objects.filter(room=room).count()
                
                if current_beds > capacity:
                    return JsonResponse({
                        'success': False,
                        'message': (
                            f'Cannot reduce capacity to {capacity}. '
                            f'Room already has {current_beds} beds installed. '
                            f'Please remove excess beds first.'
                        )
                    })

                # Check if there are any allocations that would exceed new capacity
                current_allocations = StudentHostelAllocation.objects.filter(
                    room=room,
                    is_active=True
                ).count()
                
                if current_allocations > capacity:
                    return JsonResponse({
                        'success': False,
                        'message': (
                            f'Cannot reduce capacity to {capacity}. '
                            f'Room currently has {current_allocations} students allocated. '
                            f'Please relocate students first.'
                        )
                    })

            # Check if deactivating room with active allocations
            if not is_active and room.is_active:
                active_allocations = StudentHostelAllocation.objects.filter(
                    room=room,
                    is_active=True
                ).exists()
                
                if active_allocations:
                    return JsonResponse({
                        'success': False,
                        'message': 'Cannot deactivate room with active student allocations.'
                    })

            # Update the room
            with transaction.atomic():
                room.room_number = room_number
                room.capacity = capacity
                room.is_active = is_active
                room.save()

            messages.success(request, f'Room "{room_number}" updated successfully.')

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Room "{room_number}" updated successfully.',
                    'redirect_url': reverse('admin_single_hostel_rooms_list') + f'?hostel={hostel.id}'
                })
            return redirect(f"{reverse('admin_single_hostel_rooms_list')}?hostel={hostel.id}")

        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'Error updating room: {str(e)}'
                })
            messages.error(request, f'Error updating room: {str(e)}')
            return redirect(reverse('admin_hostel_room_edit', args=[room.id]))

    # GET request - show form
    # Get statistics for this room
    total_beds = Bed.objects.filter(room=room).count()
    occupied_beds = Bed.objects.filter(room=room, is_occupied=True).count()
    available_beds = total_beds - occupied_beds
    
    active_allocations = StudentHostelAllocation.objects.filter(
        room=room,
        is_active=True
    ).count()
    
    # Calculate hostel capacity usage
    total_hostel_capacity = HostelRoom.objects.filter(
        hostel=hostel, is_active=True
    ).aggregate(total=Sum('capacity'))['total'] or 0
    
    available_hostel_capacity = hostel.max_students - total_hostel_capacity + room.capacity
    
    context = {
        'room': room,
        'hostel': hostel,
        'total_beds': total_beds,
        'occupied_beds': occupied_beds,
        'available_beds': available_beds,
        'active_allocations': active_allocations,
        'available_hostel_capacity': available_hostel_capacity,
        'page_title': f'Edit Room {room.room_number}',
        'form_title': f'Edit Room {room.room_number}',
        'submit_text': 'Update Room',
        'cancel_url': f"{reverse('admin_single_hostel_rooms_list')}?hostel={hostel.id}",
        'is_edit': True,
    }
    
    return render(request, 'admin/hostels/room_form.html', context)


# ============================================================================
# HOSTEL ROOM DETAIL VIEW
# ============================================================================

@login_required
def hostel_room_detail(request, room_id):
    """
    Display detailed information about a specific room
    """
    # Get room with all related data
    room = get_object_or_404(
        HostelRoom.objects.select_related('hostel').prefetch_related(
            'beds',
            'studenthostelallocation_set',
            'studenthostelallocation_set__student',
            'studenthostelallocation_set__student__class_level',
            'studenthostelallocation_set__student__stream_class'
        ),
        id=room_id
    )
    
    hostel = room.hostel
    
    # Get bed details
    beds = room.beds.all().order_by('bed_number')
    
    # Get active allocations in this room
    active_allocations = StudentHostelAllocation.objects.filter(
        room=room,
        is_active=True
    ).select_related(
        'student',
        'student__class_level',
        'student__stream_class',
        'bed'
    ).order_by('bed__bed_number')
    
    # Get allocation history (inactive)
    allocation_history = StudentHostelAllocation.objects.filter(
        room=room,
        is_active=False
    ).select_related(
        'student',
        'student__class_level',
        'student__stream_class',
        'bed'
    ).order_by('-end_date', '-start_date')[:20]  # Limit to last 20
    
    # Calculate statistics
    total_beds = beds.count()
    occupied_beds = beds.filter(is_occupied=True).count()
    available_beds = total_beds - occupied_beds
    
    bed_occupancy_rate = (occupied_beds / total_beds * 100) if total_beds > 0 else 0
    room_occupancy_rate = (active_allocations.count() / room.capacity * 100) if room.capacity > 0 else 0
    
    # Bed type breakdown
    bed_types = {}
    for bed in beds:
        bed_type = bed.get_bed_type_display()
        if bed_type not in bed_types:
            bed_types[bed_type] = {'total': 0, 'occupied': 0}
        bed_types[bed_type]['total'] += 1
        if bed.is_occupied:
            bed_types[bed_type]['occupied'] += 1
    
    # Get payment statistics for students in this room
    student_ids = active_allocations.values_list('student_id', flat=True)
    payment_stats = HostelPaymentTransaction.objects.filter(
        allocation__student_id__in=student_ids,
        allocation__room=room
    ).aggregate(
        total_paid=Sum('amount'),
        total_transactions=Count('id')
    )
    
    total_paid = payment_stats['total_paid'] or 0
    total_expected = active_allocations.aggregate(total=Sum('total_fee'))['total'] or 0
    outstanding = total_expected - total_paid
    
    context = {
        'room': room,
        'hostel': hostel,
        'beds': beds,
        'active_allocations': active_allocations,
        'allocation_history': allocation_history,
        'total_beds': total_beds,
        'occupied_beds': occupied_beds,
        'available_beds': available_beds,
        'bed_occupancy_rate': round(bed_occupancy_rate, 1),
        'room_occupancy_rate': round(room_occupancy_rate, 1),
        'bed_types': bed_types,
        'total_paid': total_paid,
        'total_expected': total_expected,
        'outstanding': outstanding,
        'collection_rate': (total_paid / total_expected * 100) if total_expected > 0 else 0,
        'page_title': f'Room {room.room_number} Details',
    }
    
    return render(request, 'admin/hostels/room_detail.html', context)


# Add these to your views.py

@login_required
def api_check_room_number(request):
    """
    API endpoint to check if a room number already exists in a hostel
    """
    try:
        hostel_id = request.GET.get('hostel_id')
        room_number = request.GET.get('room_number')
        exclude_id = request.GET.get('exclude_id')
        
        if not hostel_id or not room_number:
            return JsonResponse({'exists': False, 'error': 'Missing parameters'})
        
        # Build query
        query = HostelRoom.objects.filter(
            hostel_id=hostel_id,
            room_number__iexact=room_number
        )
        
        # Exclude current room if editing
        if exclude_id and exclude_id != 'null':
            query = query.exclude(id=exclude_id)
        
        exists = query.exists()
        
        return JsonResponse({'exists': exists})
        
    except Exception as e:
        return JsonResponse({'exists': False, 'error': str(e)})


@login_required
def api_delete_room(request, room_id):
    """
    API endpoint to delete a room
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'})
    
    try:
        room = get_object_or_404(HostelRoom, id=room_id)
        
        # Check if room has any active allocations
        active_allocations = StudentHostelAllocation.objects.filter(
            room=room,
            is_active=True
        ).exists()
        
        if active_allocations:
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete room with active student allocations.'
            })
        
        # Check if room has any beds
        has_beds = Bed.objects.filter(room=room).exists()
        
        # Store room info for message
        room_info = f"{room.room_number} in {room.hostel.name}"
        
        # Delete the room (cascade will delete beds if any)
        room.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Room {room_info} deleted successfully.'
        })
        
    except HostelRoom.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Room not found.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting room: {str(e)}'
        })    